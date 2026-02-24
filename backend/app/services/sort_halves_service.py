"""Business logic for the 'Sort Halves' workflow."""

from __future__ import annotations

from copy import copy
from io import BytesIO
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def get_workbook_sheet_names(file_bytes: bytes) -> List[str]:
    """Return sorted worksheet names from workbook bytes."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    return sorted(wb.sheetnames)


def process_sort_halves(source_bytes: bytes, destination_bytes: bytes) -> Tuple[bytes, Dict[str, Tuple[int, int]]]:
    """Transfer rows from source workbook to destination workbook sheets.

    Logic mirrors desktop SortHalvesThread:
    - read source sheet: "Четверти" if exists, else second sheet, else first
    - group rows by tournament (column B)
    - for each tournament matching destination sheet, append row[2:14]
    - preserve style from destination template row (row 2 if exists)
    - return updated workbook bytes and summary {sheet: (inserted, normative)}
    """
    src_io = BytesIO(source_bytes)
    dst_io = BytesIO(destination_bytes)

    src_wb = load_workbook(src_io, read_only=True, data_only=True)
    dst_wb = load_workbook(dst_io)

    if "Четверти" in src_wb.sheetnames:
        src_ws = src_wb["Четверти"]
    elif len(src_wb.worksheets) > 1:
        src_ws = src_wb.worksheets[1]
    else:
        src_ws = src_wb.worksheets[0]

    grouped: Dict[str, List[List[object]]] = {}
    for row in src_ws.iter_rows(values_only=True):
        if not row or len(row) < 4:
            continue
        tournament = row[1]
        if not tournament:
            continue
        sheet_name = str(tournament).strip()
        if not sheet_name or sheet_name not in dst_wb.sheetnames:
            continue
        data = list(row[2:14])
        grouped.setdefault(sheet_name, []).append(data)

    games_summary: Dict[str, Tuple[int, int]] = {}

    for sheet_name, rows_data in grouped.items():
        dest_ws = dst_wb[sheet_name]

        template_row = 2 if dest_ws.max_row >= 2 else 1
        max_len = len(rows_data[0]) if rows_data else 0
        template_styles = {}
        for col_idx in range(1, max_len + 1):
            tmpl_cell = dest_ws.cell(row=template_row, column=col_idx)
            template_styles[col_idx] = copy(tmpl_cell._style)

        insert_row = dest_ws.max_row + 1
        for i, cell in enumerate(dest_ws["A"][1:], start=2):
            value = cell.value
            if (value is None or str(value).strip() == "") and not isinstance(cell, MergedCell):
                row_has_merged = False
                for col_idx in range(1, max_len + 1):
                    candidate = dest_ws.cell(row=i, column=col_idx)
                    if isinstance(candidate, MergedCell):
                        row_has_merged = True
                        break
                if not row_has_merged:
                    insert_row = i
                    break

        inserted_count = 0
        for data in rows_data:
            data_len = len(data)

            while True:
                row_has_merged = False
                for col_idx in range(1, data_len + 1):
                    candidate = dest_ws.cell(row=insert_row, column=col_idx)
                    if isinstance(candidate, MergedCell):
                        row_has_merged = True
                        break
                if row_has_merged:
                    insert_row += 1
                else:
                    break

            for col_idx, val in enumerate(data, start=1):
                dest_cell = dest_ws.cell(row=insert_row, column=col_idx)
                if col_idx in template_styles:
                    dest_cell._style = copy(template_styles[col_idx])
                dest_cell.value = val

            inserted_count += 1
            insert_row += 1

        unique_teams = set()
        for row_vals in dest_ws.iter_rows(min_row=2, values_only=True):
            if not row_vals:
                continue
            t1 = row_vals[0]
            t2 = row_vals[1] if len(row_vals) > 1 else None
            if isinstance(t1, str) and t1.strip():
                unique_teams.add(t1.strip())
            if isinstance(t2, str) and t2.strip():
                unique_teams.add(t2.strip())
        normative = len(unique_teams) // 2
        games_summary[sheet_name] = (inserted_count, normative)

    out_io = BytesIO()
    dst_wb.save(out_io)
    out_io.seek(0)
    return out_io.read(), games_summary
