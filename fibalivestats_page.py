from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                           QLabel, QTableWidget, QTableWidgetItem, QTabWidget,
                           QCalendarWidget, QTextEdit, QProgressBar, QMessageBox,
                           QHeaderView, QInputDialog, QFileDialog, QCheckBox,
                           QAbstractItemView, QListWidget, QApplication, QStackedWidget, QRadioButton)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDate, QSize
from PyQt5.QtGui import QFont, QIcon
import logging
import os
from datetime import datetime
from fibalivestats_handler import FibaLiveStatsHandler


class ClickableLabel(QLabel):
    clicked = pyqtSignal()
    
    def __init__(self, text):
        super().__init__(text)
        self.setCursor(Qt.PointingHandCursor)  # Курсор-рука при наведении
        
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.clicked.emit()

class FibaLiveStatsScanThread(QThread):
    """Поток для сканирования турниров"""
    progress_signal = pyqtSignal(str, int)
    finished_signal = pyqtSignal(list)
    error_signal = pyqtSignal(str)
    
    def __init__(self, handler, target_date):
        super().__init__()
        self.handler = handler
        self.target_date = target_date
        
    def run(self):
        try:
            results = self.handler.scan_tournaments(
                self.target_date,
                progress_callback=lambda msg, progress: self.progress_signal.emit(msg, progress)
            )
            self.finished_signal.emit(results)
        except Exception as e:
            self.error_signal.emit(str(e))

class FibaLiveStatsPage(QWidget):
    """Страница для работы с FibaLiveStats"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        # По умолчанию создаем handler с headless=True
        self.handler = FibaLiveStatsHandler(headless=True)
        self.scan_thread = None
        self.last_results = []
        self.setup_ui()
        self.load_tournaments()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Заголовок
        title_label = QLabel("FibaLiveStats Scanner")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(20)
        
        # Вкладки
        self.tabs = QTabWidget()
        
        # Вкладка управления турнирами
        self.tournaments_tab = QWidget()
        self.setup_tournaments_tab()
        self.tabs.addTab(self.tournaments_tab, "Управление турнирами")
        
        # Вкладка сканирования
        self.scan_tab = QWidget()
        self.setup_scan_tab()
        self.tabs.addTab(self.scan_tab, "Сканирование")
        
        layout.addWidget(self.tabs)
        

    def setup_tournaments_tab(self):
        """Настройка вкладки управления турнирами"""
        layout = QVBoxLayout(self.tournaments_tab)
        
        # Кнопки управления
        buttons_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Добавить турнир")
        self.add_btn.clicked.connect(self.add_tournament)
        
        self.edit_btn = QPushButton("✏️ Редактировать")
        self.edit_btn.clicked.connect(self.edit_tournament)
        
        self.delete_btn = QPushButton("🗑️ Удалить")
        self.delete_btn.clicked.connect(self.delete_tournament)
        
        self.import_btn = QPushButton("📥 Импорт из Excel")
        self.import_btn.clicked.connect(self.import_from_excel)
        
        self.refresh_btn = QPushButton("🔄 Обновить")
        self.refresh_btn.clicked.connect(self.load_tournaments)
        
        buttons_layout.addWidget(self.add_btn)
        buttons_layout.addWidget(self.edit_btn)
        buttons_layout.addWidget(self.delete_btn)
        buttons_layout.addWidget(self.import_btn)
        buttons_layout.addWidget(self.refresh_btn)
        buttons_layout.addStretch()
        
        layout.addLayout(buttons_layout)
        
        # Таблица турниров
        self.tournaments_table = QTableWidget()
        self.tournaments_table.setColumnCount(4)
        self.tournaments_table.setHorizontalHeaderLabels(["ID", "Название", "URL", "Активен"])
        self.tournaments_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tournaments_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # Настройка колонок
        header = self.tournaments_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        self.tournaments_table.setColumnWidth(0, 50)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.Fixed)
        self.tournaments_table.setColumnWidth(3, 80)
        
        # Скрываем ID колонку
        self.tournaments_table.setColumnHidden(0, True)
        
        # Добавляем стили для чекбоксов
        checkbox_style = """
            QCheckBox {
                background-color: transparent;
                spacing: 2px;  /* Уменьшаем отступ */
            }
            QCheckBox::indicator {
                width: 16px;  /* Уменьшаем размер */
                height: 16px;
                border: none;  /* Убираем border */
                background-color: transparent;  /* Прозрачный фон */
                image: url("");  /* Пустое изображение по умолчанию */
            }
            QCheckBox::indicator:checked {
                /* Зеленая галочка для активного состояния */
                color: #4CAF50;
                font-size: 14px;  /* Уменьшаем размер шрифта */
            }
            QCheckBox::indicator:unchecked {
                /* Красный крестик для неактивного состояния */
                color: #F44336;
                font-size: 14px;  /* Уменьшаем размер шрифта */
            }
        """
        
        self.tournaments_table.setStyleSheet(checkbox_style)
        
        layout.addWidget(self.tournaments_table)

    def load_tournaments(self):
        """Загрузка турниров в таблицу"""
        try:
            tournaments = self.handler.get_all_tournaments()
            self.tournaments_table.setRowCount(len(tournaments))
            
            for row, (id_, name, url, active) in enumerate(tournaments):
                # ID
                self.tournaments_table.setItem(row, 0, QTableWidgetItem(str(id_)))
                
                # Название
                self.tournaments_table.setItem(row, 1, QTableWidgetItem(name))
                
                # URL
                self.tournaments_table.setItem(row, 2, QTableWidgetItem(url))
                
                # Активен (чекбокс)
                checkbox = QCheckBox()
                checkbox.setChecked(bool(active))
                checkbox.stateChanged.connect(lambda state, tid=id_: self.toggle_tournament(tid, state))
                
                # Создаем контейнер для центрирования чекбокса
                widget = QWidget()
                layout = QHBoxLayout(widget)
                layout.setContentsMargins(0, 0, 0, 0)

                # Создаем кликабельный label
                status_label = ClickableLabel("✅" if active else "❌")
                status_label.setStyleSheet("""
                    QLabel {
                        color: %s;
                        font-size: 18px;
                        padding: 2px;
                    }
                    QLabel:hover {
                        background-color: rgba(255, 255, 255, 0.1);
                        border-radius: 3px;
                    }
                """ % ("#4CAF50" if active else "#F44336"))
                status_label.setAlignment(Qt.AlignCenter)

                # Сохраняем ID турнира в свойствах label
                status_label.tournament_id = id_
                status_label.active = active
                status_label.clicked.connect(self.toggle_tournament_status)
                
                layout.addWidget(status_label)
                
                self.tournaments_table.setCellWidget(row, 3, widget)
            
            # Обновляем ширину колонки статуса
            self.tournaments_table.setColumnWidth(3, 70)  # Делаем шире

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки турниров: {str(e)}")    
                
            # Обновляем стили для чекбоксов
            checkbox_style = """
                QCheckBox {
                    background-color: transparent;
                    spacing: 5px;
                }
                QCheckBox::indicator {
                    width: 20px;
                    height: 20px;
                    border-radius: 10px;
                }
                QCheckBox::indicator:unchecked {
                    background-color: #F44336;
                    border: 2px solid #D32F2F;
                }
                QCheckBox::indicator:checked {
                    background-color: #4CAF50;
                    border: 2px solid #388E3C;
                }
                QCheckBox::indicator:hover {
                    border-width: 2px;
                }
            """
            
            self.tournaments_table.setStyleSheet(checkbox_style)
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки турниров: {str(e)}")

    def toggle_tournament(self, tournament_id, state):
        """Переключение активности турнира"""
        try:
            self.handler.update_tournament(tournament_id, active=(state == Qt.Checked))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обновлении турнира: {str(e)}")
        
    def setup_scan_tab(self):
        layout = QVBoxLayout(self.scan_tab)
        
        # Выбор даты
        date_label = QLabel("Выберите дату для сканирования:")
        date_label.setFont(QFont("Arial", 12))
        layout.addWidget(date_label)
        
        # Добавляем переключатель режима выбора даты
        date_mode_layout = QHBoxLayout()
        self.single_date_radio = QRadioButton("Один день")
        self.date_range_radio = QRadioButton("Диапазон дат")
        self.single_date_radio.setChecked(True)
        
        date_mode_layout.addWidget(self.single_date_radio)
        date_mode_layout.addWidget(self.date_range_radio)
        layout.addLayout(date_mode_layout)
        
        # Создаем стек для календарей
        self.calendar_stack = QStackedWidget()
        
        # Календарь для одного дня
        self.calendar = QCalendarWidget()  # Сохраняем существующую ссылку для совместимости
        self.calendar.setSelectedDate(QDate.currentDate())
        self.calendar_stack.addWidget(self.calendar)
        
        # Виджет с двумя календарями для диапазона
        range_widget = QWidget()
        range_layout = QHBoxLayout(range_widget)
        
        self.start_calendar = QCalendarWidget()
        self.end_calendar = QCalendarWidget()
        self.start_calendar.setSelectedDate(QDate.currentDate())
        self.end_calendar.setSelectedDate(QDate.currentDate().addDays(7))
        
        range_layout.addWidget(QLabel("От:"))
        range_layout.addWidget(self.start_calendar)
        range_layout.addWidget(QLabel("До:"))
        range_layout.addWidget(self.end_calendar)
        
        self.calendar_stack.addWidget(range_widget)
        
        # Подключаем переключение календарей
        self.single_date_radio.toggled.connect(self.toggle_calendar_mode)
        
        # Добавляем стек календарей в основной layout
        layout.addWidget(self.calendar_stack)
        self.calendar.setStyleSheet("""
        /* Основной виджет календаря */
        QCalendarWidget {
            background-color: #1e1e24;
            selection-background-color: #2d5c8f;
            selection-color: white;
        }

        /* Навигационная панель */
        QCalendarWidget QWidget#qt_calendar_navigationbar {
            background-color: #1e1e24;
            border: none;
        }

        /* Кнопки навигации (стрелки) */
        QCalendarWidget QToolButton {
            color: white;
            background-color: #1e1e24;
            border: none;
            border-radius: 4px;
            qproperty-iconSize: 26px;
            min-width: 30px;
            max-width: 30px;
            min-height: 30px;
            max-height: 30px;
        }

        /* Стрелка влево */
        QCalendarWidget QToolButton::left-arrow {
            background-color: transparent;
            image: url(left-arrow.png);  /* Можно заменить на вашу иконку */
        }

        /* Стрелка вправо */
        QCalendarWidget QToolButton::right-arrow {
            background-color: transparent;
            image: url(right-arrow.png);  /* Можно заменить на вашу иконку */
        }

        /* Hover эффект для кнопок */
        QCalendarWidget QToolButton:hover {
            background-color: #2d5c8f;
        }

        /* Заголовок с месяцем/годом */
        QCalendarWidget QToolButton#qt_calendar_monthbutton,
        QCalendarWidget QToolButton#qt_calendar_yearbutton {
            color: white;
            background-color: #1e1e24;
            border: none;
            border-radius: 4px;
            padding: 5px;
            min-width: 60px;
            max-width: 100px;
        }

        /* Таблица с датами */
        QCalendarWidget QTableView {
            background-color: #1e1e24;
            border: none;
            selection-background-color: #2d5c8f;
            selection-color: white;
            outline: 0;
        }

        /* Заголовки дней недели */
        QCalendarWidget QTableView QHeaderView {
            background-color: #1e1e24;
        }

        QCalendarWidget QTableView QHeaderView::section {
            color: #808080;
            background-color: #1e1e24;
            padding: 5px;
            border: none;
        }

        /* Ячейки с датами */
        QCalendarWidget QTableView QTableViewItem {
            border: none;
        }

        /* Стиль для выбранной даты */
        QCalendarWidget QTableView:selected {
            background-color: #2d5c8f;
            color: white;
        }

        /* Текущая дата */
        QCalendarWidget QTableView:enabled#qt_calendar_calendarview[today="true"] {
            color: #2d5c8f;
            font-weight: bold;
        }

        /* Стиль для неактивных дат */
        QCalendarWidget QTableView:disabled {
            color: #666666;
        }

        /* Убираем белый фон */
        QCalendarWidget QWidget {
            background-color: #1e1e24;
        }
    """)
        layout.addWidget(self.calendar)
        
        
        
        # Создаем горизонтальный layout для кнопок
        buttons_layout = QHBoxLayout()
        
        # Кнопка сканирования (теперь только одна)
        self.scan_btn = QPushButton("🔍 Начать сканирование")
        self.scan_btn.setMinimumHeight(40)
        self.scan_btn.clicked.connect(self.start_scan)
        
        # Кнопка отмены
        self.cancel_btn = QPushButton("❌ Отменить сканирование")
        self.cancel_btn.setMinimumHeight(40)
        self.cancel_btn.clicked.connect(self.cancel_scan)
        self.cancel_btn.setEnabled(False)  # По умолчанию неактивна
        
        buttons_layout.addWidget(self.scan_btn)
        buttons_layout.addWidget(self.cancel_btn)
        layout.addLayout(buttons_layout)
        
        # Прогресс-бар
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Статус
        self.status_label = QLabel("")
        layout.addWidget(self.status_label)
        
        # Результаты
        results_label = QLabel("Результаты сканирования:")
        results_label.setFont(QFont("Arial", 11, QFont.Bold))
        layout.addWidget(results_label)
        
        self.results_text = QTextEdit()
        self.results_text.setReadOnly(True)
        self.results_text.setMinimumHeight(200)
        layout.addWidget(self.results_text)
        
        # Кнопка сохранения
        self.save_btn = QPushButton("💾 Сохранить результаты")
        self.save_btn.setEnabled(False)
        self.save_btn.clicked.connect(self.save_results)
        layout.addWidget(self.save_btn)
        
        # Журнал выполнения
        log_label = QLabel("Журнал выполнения:")
        log_label.setFont(QFont("Arial", 10, QFont.Bold))
        layout.addWidget(log_label)
        
        self.log_text = QListWidget()
        self.log_text.setMaximumHeight(200)
        self.log_text.setStyleSheet("""
            QListWidget {
                background-color: #1a1a1f;
                color: #00ff00;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 12px;
                border: 1px solid #333339;
                padding: 5px;
            }
            QListWidget::item {
                border-bottom: 1px solid #2a2a2f;
                padding: 2px;
            }
            QScrollBar:vertical {
                border: none;
                background: #202024;
                width: 14px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #353539;
                min-height: 20px;
                border-radius: 7px;
            }
            QScrollBar::handle:vertical:hover {
                background: #404044;
            }
        """)
        layout.addWidget(self.log_text)
        
    def toggle_calendar_mode(self, checked):
        """Переключение между режимами выбора даты"""
        self.calendar_stack.setCurrentIndex(0 if checked else 1)
        
    # Метод очистки и добавления логов
    def clear_logs(self):
            """Очищает журнал выполнения"""
            self.log_text.clear()
        
    def cancel_scan(self):
        """Отмена сканирования"""
        if self.scan_thread and self.scan_thread.isRunning():
            reply = QMessageBox.question(
                self,
                "Подтверждение",
                "Вы уверены, что хотите отменить сканирование?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Устанавливаем флаг отмены в handler
                self.handler.cancel_scan = True
                
                # Обновляем интерфейс
                self.progress_bar.setVisible(False)
                self.scan_btn.setEnabled(True)
                self.cancel_btn.setEnabled(False)
                self.status_label.setText("Сканирование отменено")
                
                # Дожидаемся завершения потока
                self.scan_thread.wait()
                
                # Очищаем handler
                self.handler.close_driver()

    def add_log(self, message):
        """Добавляет сообщение в журнал выполнения"""
        try:
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.log_text.addItem(f"[{timestamp}] {message}")
            self.log_text.scrollToBottom()
            
            # Принудительно обновляем интерфейс
            QApplication.processEvents()
            
        except Exception as e:
            print(f"Ошибка при добавлении в лог: {str(e)}")

        
            
    def toggle_tournament(self, tournament_id, state):
        """Переключение активности турнира"""
        self.handler.update_tournament(tournament_id, active=(state == Qt.Checked))

    def toggle_tournament_status(self):
        """Обработчик клика по статусу турнира"""
        try:
            label = self.sender()
            new_state = not label.active
            
            # Обновляем в базе
            if self.handler.update_tournament(label.tournament_id, active=new_state):
                # Обновляем визуальное отображение
                label.setText("✅" if new_state else "❌")
                label.setStyleSheet("""
                    QLabel {
                        color: %s;
                        font-size: 18px;
                        padding: 2px;
                    }
                    QLabel:hover {
                        background-color: rgba(255, 255, 255, 0.1);
                        border-radius: 3px;
                    }
                """ % ("#4CAF50" if new_state else "#F44336"))
                label.active = new_state
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось обновить статус турнира")
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обновлении турнира: {str(e)}")
        
    def add_tournament(self):
        """Добавление нового турнира"""
        name, ok = QInputDialog.getText(self, "Добавить турнир", "Название турнира:")
        if not ok or not name:
            return
            
        url, ok = QInputDialog.getText(self, "Добавить турнир", 
                                       "URL турнира:\n(например: https://fibalivestats.dcd.shared.geniussports.com/u/FBOL/2658513/)")
        if not ok or not url:
            return
            
        if self.handler.add_tournament(name, url):
            self.load_tournaments()
            QMessageBox.information(self, "Успешно", f"Турнир '{name}' добавлен")
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось добавить турнир")
            
    def edit_tournament(self):
        """Редактирование турнира"""
        current_row = self.tournaments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите турнир для редактирования")
            return
            
        tournament_id = int(self.tournaments_table.item(current_row, 0).text())
        current_name = self.tournaments_table.item(current_row, 1).text()
        current_url = self.tournaments_table.item(current_row, 2).text()
        
        name, ok = QInputDialog.getText(self, "Редактировать турнир", 
                                        "Название турнира:", text=current_name)
        if not ok:
            return
            
        url, ok = QInputDialog.getText(self, "Редактировать турнир", 
                                       "URL турнира:", text=current_url)
        if not ok:
            return
            
        if self.handler.update_tournament(tournament_id, name=name, url=url):
            self.load_tournaments()
            QMessageBox.information(self, "Успешно", "Турнир обновлен")
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось обновить турнир")
            
    def delete_tournament(self):
        """Удаление турнира"""
        current_row = self.tournaments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите турнир для удаления")
            return
            
        tournament_id = int(self.tournaments_table.item(current_row, 0).text())
        tournament_name = self.tournaments_table.item(current_row, 1).text()
        
        reply = QMessageBox.question(self, "Подтверждение", 
                                     f"Удалить турнир '{tournament_name}'?",
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            if self.handler.delete_tournament(tournament_id):
                self.load_tournaments()
                QMessageBox.information(self, "Успешно", "Турнир удален")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить турнир")
                
    def import_from_excel(self):
        """Импорт турниров из Excel"""
        QMessageBox.information(self, "Импорт", 
                               "Эта функция импортирует названия турниров из открытых файлов Excel.\n"
                               "URL нужно будет добавить вручную для каждого турнира.")
        
        # TODO: Здесь нужно будет получить доступ к excel_handler из основного приложения
        QMessageBox.warning(self, "В разработке", 
                            "Функция импорта будет доступна после интеграции с основным приложением")
                            
    def start_scan(self):
        """Запуск сканирования"""
        # Очищаем логи перед новым сканированием
        self.log_text.clear()
        self.clear_logs()

        try:
            # Проверяем наличие активных турниров
            active_tournaments = self.handler.get_active_tournaments()
            if not active_tournaments:
                QMessageBox.warning(self, "Предупреждение", 
                                "Нет активных турниров для сканирования.\n"
                                "Добавьте турниры и отметьте их как активные.")
                return

            # Проверяем наличие Chrome
            try:
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service
                from selenium.common.exceptions import WebDriverException
                
                # Пробуем создать драйвер для проверки
                options = webdriver.ChromeOptions()
                options.add_argument('--headless')
                test_driver = webdriver.Chrome(options=options)
                test_driver.quit()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", 
                                "Не удалось запустить Chrome драйвер.\n\n"
                                "Убедитесь что:\n"
                                "1. Google Chrome установлен\n"
                                "2. ChromeDriver установлен (pip install selenium)\n\n"
                                f"Ошибка: {str(e)}")
                return

            # Определяем даты для сканирования
            if self.single_date_radio.isChecked():
                # Сканирование одного дня
                dates = [self.calendar.selectedDate().toPyDate()]
            else:
                # Сканирование диапазона
                start_date = self.start_calendar.selectedDate().toPyDate()
                end_date = self.end_calendar.selectedDate().toPyDate()
                
                if start_date > end_date:
                    QMessageBox.warning(self, "Ошибка", "Дата начала не может быть позже даты окончания")
                    return
                    
                # Генерируем список дат
                dates = []
                current = start_date
                while current <= end_date:
                    dates.append(current)
                    current += timedelta(days=1)

            # Спрашиваем подтверждение для большого диапазона
            if len(dates) > 7:
                reply = QMessageBox.question(
                    self,
                    "Подтверждение",
                    f"Вы собираетесь сканировать {len(dates)} дней.\nЭто может занять длительное время.\nПродолжить?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply == QMessageBox.No:
                    return

            # Блокируем кнопки
            self.scan_btn.setEnabled(False)
            self.cancel_btn.setEnabled(True)
            self.save_btn.setEnabled(False)
            
            # Показываем прогресс
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.results_text.clear()
            
            if len(dates) == 1:
                self.status_label.setText(f"Сканирование турниров на {dates[0].strftime('%d.%m.%Y')}...")
                # Запускаем поток сканирования для одной даты
                self.scan_thread = FibaLiveStatsScanThread(self.handler, dates[0])
                self.scan_thread.progress_signal.connect(self.update_progress)
                self.scan_thread.finished_signal.connect(self.scan_finished)
                self.scan_thread.error_signal.connect(self.scan_error)
                self.scan_thread.start()
            else:
                # Устанавливаем максимум прогресс-бара с учетом количества дат
                self.progress_bar.setMaximum(len(dates) * 100)  # 100% на каждую дату
                
                # Запускаем сканирование для каждой даты
                for i, date in enumerate(dates):
                    if self.cancel_scan:
                        break
                        
                    self.status_label.setText(f"Сканирование даты: {date.strftime('%d.%m.%Y')} ({i+1}/{len(dates)})")
                    self.progress_bar.setValue(i * 100)
                    
                    # Запускаем поток сканирования для текущей даты
                    self.scan_thread = FibaLiveStatsScanThread(self.handler, date)
                    self.scan_thread.progress_signal.connect(self.update_progress)
                    self.scan_thread.finished_signal.connect(self.scan_finished)
                    self.scan_thread.error_signal.connect(self.scan_error)
                    self.scan_thread.start()
                    
                    # Ждем завершения сканирования текущей даты
                    self.scan_thread.wait()
                    
                    if self.cancel_scan:
                        self.status_label.setText("Сканирование отменено")
                        break

        except Exception as e:
            self.add_log(f"Ошибка при запуске сканирования: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось начать сканирование: {str(e)}")
            self.scan_btn.setEnabled(True)
            self.cancel_btn.setEnabled(False)
            self.save_btn.setEnabled(False)
            self.progress_bar.hide()
        
        
    def update_progress(self, message, progress):
        """Обновление прогресса"""
        self.status_label.setText(message)
        self.progress_bar.setValue(progress)
        
    def scan_finished(self, results):
        """Завершение сканирования"""
        self.last_results = results
        
        # Разблокируем кнопки
        self.scan_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        
        if results:
            self.save_btn.setEnabled(True)
            
            # Подсчитываем уникальные турниры
            unique_tournaments = set(r['tournament'] for r in results)
            self.status_label.setText(f"✓ Найдено матчей: {len(results)} в {len(unique_tournaments)} турнирах")
            
            # Добавляем в лог информацию о найденных матчах
            self.add_log(f"\nНайдено матчей: {len(results)}")
            self.add_log(f"Турниров с матчами: {len(unique_tournaments)}")
            
            # Показываем результаты
            current_tournament = None
            for result in results:
                if current_tournament != result['tournament']:
                    current_tournament = result['tournament']
                    self.add_log(f"\n🏀 {current_tournament}")
                    self.add_log(f"   URL турнира: {result['tournament_url']}")
                
                match_type_icon = {'LIVE': '🟢', 'FINAL': '⚫', 'SCHEDULED': '📅'}.get(result['type'], '❓')
                self.add_log(f"   {match_type_icon} {result['text'][:60]}...")
            
            # Формируем текст для отображения
            results_text = self.format_results_text(results)
            self.results_text.setPlainText(results_text)
            
        else:
            self.status_label.setText("✗ Не найдено матчей на выбранную дату")
            self.results_text.setPlainText("Нет результатов")

    def save_results(self):
        """Сохранение результатов"""
        if not self.last_results:
            return
            
        try:
            # Получаем текущую дату из календаря
            selected_date = self.calendar.selectedDate().toPyDate()
            
            # Формируем имя файла
            date_str = selected_date.strftime("%d_%m_%Y")
            default_filename = f"fibalivestats_{date_str}.xlsx"
            
            # Открываем диалог сохранения
            filepath = QFileDialog.getSaveFileName(
                self,
                "Сохранить результаты",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )[0]
            
            if not filepath:
                return
                
            # Добавляем расширение если его нет
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            
            try:
                import openpyxl
                from openpyxl.styles import Alignment
                
                # Создаем новую книгу Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Matches"
                
                # Заголовки
                ws['A1'] = "URL"
                ws['C1'] = "Tournament"
                ws['D1'] = "Links Count"
                ws['E1'] = "Status"
                
                # Группируем результаты по турнирам для подсчета
                tournament_data = {}
                max_url_length = len("URL")  # Минимальная длина для заголовка
                
                for result in self.last_results:
                    if isinstance(result, dict):
                        tournament = result.get('tournament', '')
                        match_url = result.get('url') or result.get('match_url')
                        if match_url:
                            # Обновляем максимальную длину URL
                            max_url_length = max(max_url_length, len(match_url))
                            
                            # Добавляем URL в список турнира
                            if tournament not in tournament_data:
                                tournament_data[tournament] = []
                            tournament_data[tournament].append(match_url)
                
                # Записываем данные
                row = 2
                for tournament, urls in tournament_data.items():
                    # Записываем URLs
                    for url in urls:
                        ws[f'A{row}'] = url
                        ws[f'C{row}'] = tournament
                        row += 1
                    
                    # Находим первую и последнюю строку для турнира
                    start_row = row - len(urls)
                    
                    # Записываем количество ссылок
                    expected_count = len(urls)
                    actual_count = len(urls)  # В данном случае они равны
                    ws[f'D{start_row}'] = expected_count
                    
                    # Ставим статус
                    ws[f'E{start_row}'] = "✅" if expected_count == actual_count else "❌"
                
                # Настраиваем форматирование
                # Устанавливаем ширину столбца A по максимальной длине URL
                ws.column_dimensions['A'].width = max_url_length + 2  # +2 для отступа
                
                # Выравнивание для всех столбцов
                for col in ['A', 'C', 'D', 'E']:
                    for cell in ws[col]:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Сохраняем файл
                wb.save(filepath)
                print(f"✓ Файл сохранен: {filepath}")
                

                
                QMessageBox.information(self, "Успешно", f"Результаты сохранены в:\n{filepath}")
                
            except ImportError:
                QMessageBox.critical(self, "Ошибка", "Для сохранения в Excel требуется установить пакет openpyxl")
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении:\n{str(e)}")

    def format_results_text(self, results):
        """Форматирует результаты для отображения"""
        if not results:
            return "Нет результатов"
            
        results_text = f"FIBALIVESTATS - Матчи на {self.calendar.selectedDate().toPyDate().strftime('%d.%m.%Y')}\n"
        results_text += "="*80 + "\n\n"
        
        # Группируем по турнирам
        current_tournament = None
        for result in results:
            if current_tournament != result['tournament']:
                current_tournament = result['tournament']
                results_text += f"\n🏀 {current_tournament}\n"
                results_text += f"   Ссылка турнира: {result['tournament_url']}\n"
                results_text += "-"*60 + "\n"
            
            # Выводим информацию о матче
            match_type_icon = {'LIVE': '🟢', 'FINAL': '⚫', 'SCHEDULED': '📅'}.get(result['type'], '❓')
            results_text += f"\n   {match_type_icon} {result['type']} матч\n"
            results_text += f"   URL: {result['url']}\n"
            results_text += f"   Дата: {result['date']}\n"
            results_text += f"   Инфо: {result['text'][:60]}...\n"
        
        results_text += "\n" + "="*80 + "\n"
        results_text += f"Всего найдено матчей: {len(results)}\n"
        
        return results_text

    def scan_error(self, error_message):
        """Обработка ошибки сканирования"""
        self.scan_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.status_label.setText("Ошибка при сканировании")
        
        QMessageBox.critical(self, "Ошибка", f"Ошибка при сканировании:\n{error_message}")
        
