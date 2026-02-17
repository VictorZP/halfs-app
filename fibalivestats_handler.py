import logging
import re
import os
import sqlite3
import json
from contextlib import contextmanager
from db_connection import db_connect
import tkinter as tk
from tkinter import filedialog
import subprocess
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import partial
from typing import Optional
import time
import asyncio
import aiohttp
from concurrent.futures import ThreadPoolExecutor
from abc import ABC, abstractmethod
from queue import Queue
import threading
from datetime import datetime, date, timedelta
"""
This module contains the implementation of the FibaLiveStatsHandler which is
responsible for scanning tournaments hosted on the FIBA Live Stats platform.

Originally this handler attempted to orchestrate multiple third‑party parsers
(for example for the Mexican LNBP) via a `MultiSourceMatchFinder`.  However
that design proved brittle because the foreign parsers were not actually
related to the FIBA service.  To simplify the architecture and make each
source independent, the universal coordination of parsers has been moved
out of this file.  The FibaLiveStatsHandler now focuses solely on
operations specific to the FIBA Live Stats website: managing the browser
driver, scanning tournament pages, parsing match cards, and extracting
dates and statuses.

Any code relating to other leagues or generic parser orchestration has been
removed from this module.  If you wish to parse another website (e.g.
LNBP), use the appropriate parser from its own module and coordinate
parsers in the application entrypoint (see `scan_router.py` or your
main script).
"""

# NOTE: We no longer import parsers from the non‑existent `parsers` package.
# Any site‑specific parsers should live in their own modules and be invoked
# separately by the caller.  FibaLiveStatsHandler does not register or
# manage them.

# The FibaLiveStatsParser remains available via match_parsers.py for
# backwards‑compatibility, but it is no longer imported here.  Instead,
# the FibaLiveStatsHandler can parse tournaments directly via its own
# methods.

from match_parsers import FibaLiveStatsParser  # type: ignore




def retry_on_error(func):
    """Декоратор для повторных попыток при ошибках"""
    def wrapper(self, *args, **kwargs):
        max_attempts = 3
        delay = 2  # начальная задержка в секундах
        
        for attempt in range(max_attempts):
            try:
                return func(self, *args, **kwargs)
            except Exception as e:
                if attempt == max_attempts - 1:  # последняя попытка
                    raise
                
                self.add_log(f"Попытка {attempt + 1} не удалась: {str(e)}")
                self.add_log(f"Повторная попытка через {delay} сек...")
                
                # Пробуем переподключиться к драйверу
                try:
                    self.driver.refresh()
                except:
                    self.init_driver()
                
                time.sleep(delay)
                delay *= 2  # увеличиваем задержку
    
    return wrapper

class BaseMatchParser(ABC):
    """Базовый абстрактный класс для парсеров сайтов с матчами"""
    
    def __init__(self, handler):
        self.handler = handler
        self.driver = handler.driver

    @abstractmethod
    def validate_url(self, url: str) -> bool:
        """Проверяет, подходит ли URL для данного парсера"""
        pass
        
    @abstractmethod
    async def parse_matches(self, url: str, date: datetime) -> list:
        """Парсит матчи с сайта для указанной даты"""
        pass

    def format_date(self, date: datetime) -> str:
        """Форматирует дату для запроса к сайту"""
        return date.strftime("%Y-%m-%d")

class MultiSourceMatchFinder:
    """Менеджер для работы с разными источниками данных о матчах"""
    
    def __init__(self):
        self.parsers = []
        self.results = []
        
    def register_parser(self, parser: BaseMatchParser):
        """Регистрирует новый парсер"""
        self.parsers.append(parser)
        
    async def find_matches(self, date: datetime, progress_callback=None) -> list:
        """Ищет матчи во всех источниках параллельно"""
        tasks = []
        
        for parser in self.parsers:
            if hasattr(parser, 'base_url'):
                task = asyncio.create_task(parser.parse_matches(parser.base_url, date))
                tasks.append(task)
                
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        all_matches = []
        for parser_results in results:
            if isinstance(parser_results, list):
                all_matches.extend(parser_results)
                
        return all_matches
        
class FibaLiveStatsHandler:
    """Обработчик для работы с FibaLiveStats"""
    
    def __init__(self, headless=True):
        # Инициализация логирования
        self.setup_logging()
        
        # ВАЖНО: Сначала устанавливаем headless
        self.headless = headless
        
        self.db_path = os.path.join(
            os.path.expanduser("~"), 
            "AppData", 
            "Local", 
            "ExcelAnalyzer",
            "fibalivestats.db"
        )
        
        # Создаем директорию если её нет
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        
        # Инициализируем базу данных
        self.init_database()
        
        # Паттерны для распознавания дат
        self.date_patterns = [
            (r'\d{2}/\d{2}/\d{4}', '%d/%m/%Y'),  # dd/mm/yyyy
            (r'\d{2}/\d{2}/\d{4}', '%m/%d/%Y'),  # mm/dd/yyyy  
            (r'\d{2}\.\d{2}\.\d{4}', '%d.%m.%Y'),  # dd.mm.yyyy
            (r'\d{2}\.\d{2}\.\d{4}', '%m.%d.%Y'),  # mm.dd.yyyy
            (r'\d{2}/\d{2}/\d{2}', '%d/%m/%y'),  # dd/mm/yy
            (r'\d{2}\.\d{2}\.\d{2}', '%d.%m.%y'),  # dd.mm.yy
        ]
        
        self.driver = None
        self.cancel_scan = False
        
        from match_parsers import FibaLiveStatsParser
        self.parser = FibaLiveStatsParser(self)
        self.visited_urls = set()  # Добавляем множество для отслеживания посещенных URL

        # NOTE: We no longer use a multi‑source match finder here.  If you need
        # to aggregate results from multiple parsers, orchestrate them in
        # your application code (e.g. via scan_router.get_matches_by_url).
        
    # --- Методы инициализации и настройки ---


    def process_match(self, match, target_date):
        """Обработка отдельного матча"""
        try:
            original_url = self.driver.current_url
            
            # Все матчи проверяем
            if match['type'] in ['LIVE', 'FINAL']:
                print(f"  {'🟢' if match['type'] == 'LIVE' else '⚫'} {match['type']} матч - проверяем дату")
            elif match['type'] == 'SCHEDULED':
                print(f"  📅 Запланированный матч - проверяем дату")
            else:
                return None

            # Перемещаем элемент если нужно
            if match['y'] < 0 or match['y'] > self.driver.execute_script("return window.innerHeight"):
                print(f"  Перемещаем элемент (Y={match['y']:.0f})...")
                
                js_fix = f"""
                var selector = '{match['selector']}';
                var elements = document.querySelectorAll(selector);
                var element = elements[{match['element_index']}];
                
                if (element) {{
                    element.style.position = 'fixed';
                    element.style.top = '150px';
                    element.style.left = '50%';
                    element.style.transform = 'translateX(-50%)';
                    element.style.zIndex = '99999';
                    element.style.backgroundColor = 'white';
                    element.style.border = '2px solid blue';
                    
                    var rect = element.getBoundingClientRect();
                    return {{x: rect.x + rect.width/2, y: rect.y + rect.height/2}};
                }}
                return null;
                """
                
                new_pos = self.driver.execute_script(js_fix)
                if new_pos:
                    print(f"  ✓ Элемент перемещен")
                    match['x'] = new_pos['x']
                    match['y'] = new_pos['y']
            
            # Кликаем по элементу
            js_click = f"""
            var selector = '{match['selector']}';
            var elements = document.querySelectorAll(selector);
            var element = elements[{match['element_index']}];
            
            if (element) {{
                var link = element.querySelector('a');
                if (link) {{
                    link.click();
                    return 'link_clicked';
                }}
                element.click();
                return 'element_clicked';
            }}
            return 'not_clicked';
            """
            
            click_result = self.driver.execute_script(js_click)
            print(f"  Клик: {click_result}")
            
            time.sleep(1.5)
            new_url = self.driver.current_url
            
            # Проверяем URL и дату
            if new_url != original_url and "/u/" in new_url:
                try:
                    if new_url in self.visited_urls:
                        print(f"  ⚠️ URL уже был обработан: {new_url}")
                        self.driver.back()
                        time.sleep(0.5)
                        return None
                        
                    self.visited_urls.add(new_url)
                    print(f"  ✓ Переход выполнен: {new_url}")
                    
                    # Получаем текст страницы
                    page_text = self.driver.find_element(By.TAG_NAME, "body").text
                    date_found = False
                    
                    # Проверяем дату для всех типов матчей
                    date_found, found_date = self.check_date_match(page_text, target_date)
                    
                    if not date_found:
                        print(f"  ✗ Дата {target_date.strftime('%d.%m.%Y')} НЕ найдена - пропускаем матч")
                        self.driver.back()
                        time.sleep(0.5)
                        return None
                    
                    print(f"  ✓ Матч подтвержден на дату {found_date}")
                    match['date'] = found_date
                    
                    # Добавляем результат только если дата совпадает
                    result = {
                        'url': new_url,
                        'text': match['text'],
                        'date': match.get('date', target_date.strftime('%d/%m/%Y')),
                        'type': match['type']
                    }
                    
                    print("  ✓✓✓ Матч добавлен!")
                    self.driver.back()
                    time.sleep(0.5)
                    return result
                    
                except Exception as inner_e:
                    print(f"  ✗ Ошибка при проверке даты: {str(inner_e)}")
                    self.driver.back()
                    time.sleep(0.5)
                    return None
                        
            return None
                
        except Exception as e:
            print(f"  ✗ Ошибка обработки матча: {str(e)}")
            try:
                self.driver.back()
                time.sleep(0.5)
            except:
                pass
            return None
    
    def process_match_threaded(self, match_data):
        """Обработка SCHEDULED матча в отдельном потоке"""
        try:
            match, target_date = match_data
            
            # Для SCHEDULED матчей проверяем дату прямо в карточке
            if match['type'] == 'SCHEDULED':
                print(f"  📅 Запланированный матч - проверяем дату в карточке")
                date_found, found_date = self.check_date_match(match['text'], target_date)
                
                if not date_found:
                    print(f"  ✗ Дата {target_date.strftime('%d.%m.%Y')} НЕ найдена в карточке")
                    return None
                
                print(f"  ✓ Матч подтвержден на дату {found_date}")
                match['date'] = found_date
                return {
                    'url': match['url'],
                    'text': match['text'],
                    'date': match.get('date', target_date.strftime('%d/%m/%Y')),
                    'type': match['type']
                }
                
            return None
                
        except Exception as e:
            print(f"Ошибка в потоке: {str(e)}")
            return None

    def process_matches_parallel(self, all_matches, target_date, max_workers=4):
        """Параллельная обработка матчей"""
        matches_found = []
        
        self.add_log(f"\nОбработка матчей...")
        self.add_log(f"Всего матчей для проверки: {len(all_matches)}")

        # Разделяем матчи по типам
        scheduled_matches = [m for m in all_matches if m['type'] == 'SCHEDULED']
        other_matches = [m for m in all_matches if m['type'] in ['LIVE', 'FINAL']]
        
        self.add_log(f"SCHEDULED матчей: {len(scheduled_matches)}")
        self.add_log(f"LIVE/FINAL матчей: {len(other_matches)}")

        # Обрабатываем SCHEDULED матчи параллельно
        if scheduled_matches:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_match = {
                    executor.submit(self.process_match_threaded, (match, target_date)): match 
                    for match in scheduled_matches
                }
                
                for future in as_completed(future_to_match):
                    try:
                        result = future.result()
                        if result:
                            matches_found.append(result)
                    except Exception as e:
                        print(f"✗ Ошибка обработки SCHEDULED матча: {str(e)}")
        
        # Обрабатываем LIVE и FINAL матчи последовательно через основной драйвер
        for match in other_matches:
            try:
                result = self.process_match(match, target_date)
                if result:
                    matches_found.append(result)
            except Exception as e:
                print(f"✗ Ошибка обработки {match['type']} матча: {str(e)}")
        
        print(f"ДЕБАГ: Конец process_matches_parallel")
        print(f"ДЕБАГ: Найдено матчей: {len(matches_found)}")
        print(f"ДЕБАГ: Матчи: {matches_found}")

        return matches_found

    def setup_logging(self):
        """Настройка логирования"""
        try:
            # Создаем директорию для логов если её нет
            log_dir = os.path.join(
                os.path.expanduser("~"),
                "AppData",
                "Local",
                "ExcelAnalyzer",
                "logs"
            )
            os.makedirs(log_dir, exist_ok=True)
            
            log_file = os.path.join(log_dir, "fibalivestats.log")
            
            # Инициализируем self.logger перед проверкой
            if not hasattr(self, 'logger'):
                self.logger = logging.getLogger('fibalivestats')
            
            # Пробуем закрыть все хендлеры логгера если они есть
            if self.logger.handlers:
                for handler in self.logger.handlers[:]:
                    handler.close()
                    self.logger.removeHandler(handler)
            
            # Устанавливаем уровень логирования
            self.logger.setLevel(logging.DEBUG)
            
            # Создаем форматтер
            formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            
            try:
                # Пробуем открыть файл в режиме дополнения
                file_handler = logging.FileHandler(log_file, mode='a', encoding='utf-8')
                file_handler.setFormatter(formatter)
                self.logger.addHandler(file_handler)
                
                # Добавляем вывод в консоль
                console_handler = logging.StreamHandler()
                console_handler.setFormatter(formatter)
                self.logger.addHandler(console_handler)
                
            except Exception as e:
                print(f"Ошибка при настройке файлового логирования: {str(e)}")
                # Если не удалось настроить файловый лог, используем только консоль
                console_handler = logging.StreamHandler()
                console_handler.setFormatter(formatter)
                self.logger.addHandler(console_handler)
            
            self.logger.info("Логирование инициализировано")
            
        except Exception as e:
            print(f"Критическая ошибка при настройке логирования: {str(e)}")
            # Создаем базовый логгер с минимальной конфигурацией
            self.logger = logging.getLogger('fibalivestats')
            self.logger.setLevel(logging.DEBUG)
            
            # Создаем форматтер здесь, чтобы он был доступен
            formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            
            console_handler = logging.StreamHandler()
            console_handler.setFormatter(formatter)
            self.logger.addHandler(console_handler)

    _SCHEMA = 'fibalivestats'

    @contextmanager
    def _connect(self):
        with db_connect(schema=self._SCHEMA, sqlite_path=self.db_path) as conn:
            yield conn

    def init_database(self):
        """Инициализация базы данных"""
        with self._connect() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tournaments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    url TEXT NOT NULL,
                    active INTEGER DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()

    def add_log(self, message, level="INFO"):
        """Добавление сообщения в лог"""
        if level.upper() == "ERROR":
            logging.error(message)
        elif level.upper() == "WARNING":
            logging.warning(message)
        else:
            logging.info(message)
        print(message)
    
    def init_driver(self, headless=None):
        if self.driver:
            return
                
        if headless is None:
            headless = getattr(self, 'headless', True)
        
        options = Options()
        
        if headless:
            options.add_argument('--headless=new')
        
        # Добавляем настройки для обхода обнаружения автоматизации
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # Добавляем правдоподобный user-agent
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36')
        
        # Другие настройки остаются как есть
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-software-rasterizer')
        options.add_argument('--disable-extensions')
        
        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
            
            # Скрываем факт использования WebDriver
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            self.driver.set_page_load_timeout(30)
            self.driver.implicitly_wait(5)
            
            logging.info("Chrome драйвер успешно инициализирован")
        except Exception as e:
            logging.error(f"Ошибка при инициализации Chrome драйвера: {str(e)}")
            raise


    def clear_browser_cache(self):
        """Очистка кэша браузера"""
        try:
            if self.driver:
                self.driver.execute_script("window.localStorage.clear();")
                self.driver.execute_script("window.sessionStorage.clear();")
                self.driver.delete_all_cookies()
                logging.info("✓ Кэш браузера очищен")
        except Exception as e:
            logging.error(f"Ошибка при очистке кэша: {str(e)}")

    def close_driver(self):
        """Закрытие драйвера"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None

    # --- Методы для работы с БД ---

    def add_tournament(self, name, url, active=True):
        """Добавление нового турнира"""
        try:
            with self._connect() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO tournaments (name, url, active) 
                    VALUES (?, ?, ?)
                ''', (name, url, 1 if active else 0))
                conn.commit()
            return True
        except Exception as e:
            if 'UNIQUE' in str(e).upper() or 'duplicate' in str(e).lower():
                logging.warning(f"Турнир '{name}' уже существует")
            else:
                logging.error(f"Ошибка при добавлении турнира: {str(e)}")
            return False

    def get_all_tournaments(self):
        """Получение всех турниров"""
        with self._connect() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT id, name, url, active FROM tournaments ORDER BY name')
            return cursor.fetchall()

    def get_active_tournaments(self):
        """Получение активных турниров"""
        with self._connect() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT name, url FROM tournaments WHERE active = 1 ORDER BY name')
            return cursor.fetchall()

    def update_tournament(self, tournament_id, name=None, url=None, active=None):
        """Обновление турнира"""
        try:
            with self._connect() as conn:
                cursor = conn.cursor()
                updates = []
                params = []
                if name is not None:
                    updates.append("name = ?")
                    params.append(name)
                if url is not None:
                    updates.append("url = ?")
                    params.append(url)
                if active is not None:
                    updates.append("active = ?")
                    params.append(1 if active else 0)
                if updates:
                    updates.append("updated_at = CURRENT_TIMESTAMP")
                    params.append(tournament_id)
                    query = f"UPDATE tournaments SET {', '.join(updates)} WHERE id = ?"
                    cursor.execute(query, params)
                    conn.commit()
            return True
        except Exception as e:
            logging.error(f"Ошибка при обновлении турнира: {str(e)}")
            return False

    def delete_tournament(self, tournament_id):
        """Удаление турнира"""
        try:
            with self._connect() as conn:
                cursor = conn.cursor()
                cursor.execute('DELETE FROM tournaments WHERE id = ?', (tournament_id,))
                conn.commit()
            return True
        except Exception as e:
            logging.error(f"Ошибка при удалении турнира: {str(e)}")
            return False

    # --- Методы проверки и валидации ---

    def check_url_availability(self, url: str, max_retries: int = 3) -> bool:
        """Проверка доступности URL с повторными попытками"""
        for attempt in range(max_retries):
            try:
                self.driver.get(url)
                WebDriverWait(self.driver, 10).until(
                    lambda d: d.execute_script('return document.readyState') == 'complete'
                )
                return True
            except Exception as e:
                if attempt == max_retries - 1:
                    logging.error(f"URL недоступен после {max_retries} попыток: {url}")
                    return False
                logging.warning(f"Попытка {attempt + 1}: Ошибка доступа к {url}")
                time.sleep(2 ** attempt)
        return False

    def normalize_tournament_url(self, url):
        """НЕ нормализуем URL - оставляем ID матча как есть"""
        return url.rstrip('/') + '/' if not url.endswith('/') else url

    def check_date_match(self, page_text: str, target_date: datetime.date) -> tuple[bool, str]:
        try:
            print(f"\nПроверяем текст на наличие даты {target_date.strftime('%d.%m.%Y')}:")
            print(f"Анализируемый текст (первые 200 символов): {page_text[:200]}")
            
            date_patterns = [
                # Матч начался/начнется (Salto inicial)
                (r'Salto\s+inicial:\s*(\d{1,2}[/.]\d{1,2}[/.]\d{2,4})', 'game_start'),
                # Время начала
                (r'Start\s+time:\s*(\d{1,2}[/.]\d{1,2}[/.]\d{2,4})', 'game_start'),
                # Дата игры
                (r'Game\s+Date:\s*(\d{1,2}[/.]\d{1,2}[/.]\d{2,4})', 'game_date'),
                # Просто дата
                (r'Date:\s*(\d{1,2}[/.]\d{1,2}[/.]\d{2,4})', 'date'),
                # Время по GMT
                (r'GMT\s+(\d{1,2}[/.]\d{1,2}[/.]\d{2,4})', 'gmt'),

                # Новые паттерны для SCHEDULED матчей
                (r'(\d{2}/\d{2}/\d{4})\s*\d{2}:\d{2}\s*[ap]m', 'scheduled'),  # 22/08/2025 09:00 am
                (r'(\d{2}\.\d{2}\.\d{4})\s*\d{2}:\d{2}\s*[ap]m', 'scheduled'),  # 22.08.2025 09:00 am
                (r'(\d{2}-\d{2}-\d{4})\s*\d{2}:\d{2}\s*[ap]m', 'scheduled'),  # 22-08-2025 09:00 am

                # Общие паттерны для дат
                (r'(\d{2}/\d{2}/\d{4})', 'date'),  # 22/08/2025
                (r'(\d{2}\.\d{2}\.\d{4})', 'date'),  # 22.08.2025
                (r'(\d{2}-\d{2}-\d{4})', 'date'),  # 22-08-2025
            ]


            found_dates = []
            
            for pattern, pattern_type in date_patterns:
                matches = re.finditer(pattern, page_text, re.IGNORECASE)
                for match in matches:
                    try:
                        date_str = match.group(1).strip()
                        date_str = date_str.replace('.', '/')
                        
                        # Парсим компоненты даты
                        parts = date_str.split('/')
                        if len(parts) == 3:
                            day = int(parts[0])
                            month = int(parts[1])
                            year = int(parts[2])
                            
                            # Добавляем 2000 к двузначному году
                            if year < 100:
                                year = 2000 + year
                                
                            try:
                                match_date = datetime(year, month, day).date()
                                found_dates.append({
                                    'date': match_date,
                                    'type': pattern_type,
                                    'original': date_str
                                })
                                print(f"Найдена дата: {date_str} ({pattern_type})")
                            except ValueError:
                                continue
                                
                    except Exception as e:
                        print(f"Ошибка при разборе даты: {str(e)}")
                        continue


            # Если нашли даты, проверяем их
            if found_dates:
                # Сначала ищем дату игры (game_start или game_date)
                game_dates = [d for d in found_dates if d['type'] in ['game_start', 'game_date']]
                if game_dates:
                    # Проверяем только даты начала игры
                    for date_info in game_dates:
                        if date_info['date'] == target_date:
                            print(f"✓ Найдена дата игры: {date_info['original']}")
                            return True, date_info['original']
                    print(f"✗ Дата игры не совпадает с целевой")
                    return False, ""
                
                # Если нет специфичных дат игры, проверяем все найденные даты
                for date_info in found_dates:
                    if date_info['date'] == target_date:
                        print(f"✓ Найдена подходящая дата: {date_info['original']}")
                        return True, date_info['original']

            print(f"✗ Подходящая дата не найдена")
            return False, ""

            

        except Exception as e:
            print(f"Ошибка при проверке даты: {str(e)}")
            return False, ""

    def check_date_in_text(self, text, target_date):
        """Проверяет наличие даты в тексте"""
        day = target_date.day
        month = target_date.month
        year = target_date.year
        
        # Создаем все возможные форматы даты
        date_formats = [
            # Американский формат MM/DD/YYYY
            f"{month}/{day}/{year}",
            f"{month:02d}/{day:02d}/{year}",
            f"{month}/{day}/{str(year)[2:]}",
            f"{month:02d}/{day:02d}/{str(year)[2:]}",
            
            # Европейский формат DD/MM/YYYY  
            f"{day}/{month}/{year}",
            f"{day:02d}/{month:02d}/{year}",
            f"{day}/{month}/{str(year)[2:]}",
            f"{day:02d}/{month:02d}/{str(year)[2:]}",
            
            # Форматы с точками
            f"{day}.{month}.{year}",
            f"{day:02d}.{month:02d}.{year}",
            
            # Форматы с дефисами
            f"{year}-{month:02d}-{day:02d}",
            f"{day:02d}-{month:02d}-{year}",
        ]
        
        for date_format in date_formats:
            if date_format in text:
                logging.info(f"✓ Найдена дата в формате: {date_format}")
                return True
        
        return False

    # --- Методы для работы с элементами страницы ---

    @retry_on_error
    def get_match_url(self, card_element) -> Optional[str]:
        """Получает URL матча из карточки с повторными попытками"""
        try:
            # Кликаем по карточке
            self.safe_click(card_element)
            time.sleep(2)
            
            # Получаем URL
            current_url = self.driver.current_url
            
            # Проверяем что это URL матча
            if '/u/' in current_url:
                return current_url
                
            return None
            
        except Exception as e:
            print(f"Ошибка при получении URL матча: {str(e)}")
            raise

    def safe_click(self, element):
        """Безопасный клик по элементу с прокруткой"""
        try:
            # Прокручиваем к элементу
            self.driver.execute_script("arguments[0].scrollIntoView(true);", element)
            time.sleep(0.5)
            
            try:
                # Пробуем обычный клик
                element.click()
            except:
                # Если не получилось - через JavaScript
                self.driver.execute_script("arguments[0].click();", element)
                
        except Exception as e:
            print(f"Ошибка при клике: {str(e)}")
            raise

    def click_show_more_button(self):
        """Нажимает кнопку показа других матчей"""
        try:
            # Универсальные селекторы
            button_selectors = [
                # Кнопки с текстом на разных языках
                "//*[contains(text(), 'MOSTRAR')]",
                "//*[contains(text(), 'SHOW')]",
                "//*[contains(text(), 'VER')]",
                "//*[contains(text(), 'MORE')]",
                "//*[contains(text(), 'OTROS')]",
                "//*[contains(text(), 'OTHER')]",
                "//*[contains(text(), 'OCULTAR')]",  # Скрыть (если уже открыто)
                
                # Кнопки со стрелками
                "button[class*='arrow']",
                "button[class*='expand']",
                "button[class*='toggle']",
                "div[class*='show-more']",
                
                # В правом верхнем углу
                ".header button",
                "header button",
                "[class*='top'] button",
            ]
            
            for selector in button_selectors:
                try:
                    if selector.startswith("//"):
                        elements = self.driver.find_elements(By.XPATH, selector)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        if element.is_displayed():
                            try:
                                element.click()
                                logging.info(f"✓ Кнопка нажата")
                                return True
                            except:
                                self.driver.execute_script("arguments[0].click();", element)
                                logging.info(f"✓ Кнопка нажата через JS")
                                return True
                except:
                    continue
            
            return False
            
        except Exception as e:
            logging.error(f"Ошибка при нажатии кнопки: {str(e)}")
            return False

    def find_match_cards(self):
        """Находит карточки матчей в верхней панели"""
        try:
            # Селекторы для карточек матчей
            card_selectors = [
                # Карточки в верхней панели
                ".top-bar > div",
                ".matches-bar > div",
                ".games-bar > div",
                "header div[class*='match']",
                "header div[class*='game']",
                
                # Общие селекторы
                "div[class*='match-card']",
                "div[class*='game-card']",
                "div[class*='event-card']",
                
                # Карточки с определенным фоном
                "div[style*='background']",
                
                # По структуре (карточки обычно содержат команды и счет)
                "div:has(> div):has(> div)",
            ]
            
            all_cards = []
            
            for selector in card_selectors:
                try:
                    if selector.startswith("div:has"):
                        # Специальная обработка для :has селектора
                        elements = self.driver.find_elements(By.XPATH, "//div[div and position() < 10]")
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    
                    for element in elements:
                        # Фильтруем только видимые элементы в верхней части
                        if element.is_displayed() and element.location['y'] < 300:
                            text = element.text.strip()
                            # Проверяем, что это похоже на карточку матча
                            if text and (
                                any(x in text for x in ['vs', 'LIVE', 'FINAL', ':', '-', '/', '\\']) or
                                any(char.isdigit() for char in text)
                            ):
                                all_cards.append(element)
                except:
                    continue
            
            # Убираем дубликаты
            unique_cards = []
            seen_texts = set()
            for card in all_cards:
                card_text = card.text.strip()
                if card_text and card_text not in seen_texts:
                    unique_cards.append(card)
                    seen_texts.add(card_text)
            
            logging.info(f"Найдено уникальных карточек: {len(unique_cards)}")
            return unique_cards
            
        except Exception as e:
            logging.error(f"Ошибка при поиске карточек: {str(e)}")
            return []

    # --- Основные методы сканирования ---

    @retry_on_error
    def check_tournaments(self, tournaments, date_to_check):
        logging.info(f"\n{'='*60}")
        logging.info(f"Начало проверки турниров")
        logging.info(f"Дата проверки: {date_to_check.strftime('%d.%m.%Y')}")
        logging.info(f"Количество турниров: {len(tournaments)}")
        logging.info(f"{'='*60}\n")

        """Проверяет список турниров на указанную дату"""
        self.cancel_scan = False
        results = []
        
        print(f"\n🏀 ПРОВЕРКА ТУРНИРОВ FIBALIVESTATS")
        print(f"📅 Дата для проверки: {date_to_check.strftime('%d.%m.%Y')}")
        print(f"📋 Количество турниров: {len(tournaments)}")
        print("="*60)
        
        for i, (name, url) in enumerate(tournaments, 1):
            if self.cancel_scan:
                print("\n🛑 Сканирование отменено пользователем")
                break
            
            # Используем парсер для получения матчей
            matches = self.parser.get_matches(url, date_to_check)
            
            if matches:
                for match in matches:
                    results.append({
                        'tournament': name,
                        'tournament_url': url,
                        'match_url': match['url'],
                        'match_type': match['type'],
                        'match_date': match['date'],
                        'match_info': f"{match['teams']['home']} vs {match['teams']['away']}"
                    })
                logging.info(f"✓ Турнир {name} - найдено матчей: {len(matches)}")
            else:
                print(f"✗ Турнир {name} - матчей нет")
        
        # Добавьте эти строки перед выводом результатов
        print("\nОтладка matches_found:")
        print(f"Тип: {type(matches_found)}")
        print(f"Значение: {matches_found}")
        print(f"Длина: {len(matches_found) if matches_found else 'None'}")

        # Затем уже вывод результатов
        print("\n" + "="*60)
        if matches_found:  # matches_found содержит найденные матчи
            print(f"✅ НАЙДЕНО МАТЧЕЙ НА {target_date.strftime('%d.%m.%Y')}: {len(matches_found)}")

        # Добавляем итоговое сообщение
        print("\n" + "="*60)
        if results:
            print(f"✅ Всего найдено матчей: {len(results)}")
        else:
            print(f"❌ Матчи на {date_to_check.strftime('%d.%m.%Y')} не найдены")
        print("="*60)
        
        return results

    def check_date_in_page(self, url: str, target_date: datetime) -> list:
        """Проверяет наличие матчей на указанную дату на странице"""
        try:
            # Обрабатываем только адреса FIBA Live Stats.  Если URL относится к
            # другому домену (например, lnbp.mx), мы не пытаемся анализировать
            # страницу внутри этого обработчика и возвращаем пустой список.  Это
            # предотвращает ошибку с неинициализированной переменной `all_matches`
            # и упрощает разделение обязанностей между разными парсерами.
            low = (url or '').lower()
            if 'fibalivestats' not in low:
                # Возвращаем пустой список, чтобы вызывающий код мог решить,
                # какой парсер использовать.  None здесь может привести к
                # ошибкам, поэтому используем [] как безопасный возврат.
                return []

            if self.cancel_scan:
                self.add_log("Сканирование отменено пользователем")
                return []
        
      
            
            # Получаем текущее время в UTC
            current_utc = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            self.add_log(f"Проверка URL: {url}")
            self.add_log(f"Дата проверки: {current_utc}")
            
           
            # Очищаем кэш перед загрузкой новой страницы
            self.clear_browser_cache()
            
            # Проверяем доступность URL
            if not self.check_url_availability(url):
                self.add_log(f"✗ URL недоступен: {url}")
                return None
            
            self.add_log(f"Поиск матчей на {target_date.strftime('%d.%m.%Y')}...")

            print(f"\n{'='*60}")
            print(f"ПРОВЕРКА ТУРНИРА")
            print(f"URL: {url}")
            print(f"Дата: {target_date.strftime('%d.%m.%Y')}")
            print(f"{'='*60}\n")
            
            # Проверяем на ошибку 403
            if "403 ERROR" in self.driver.page_source:
                self.add_log("⚠️ Получена ошибка 403 - пробуем перезагрузить страницу")
                
                # Очищаем куки и кэш
                self.driver.delete_all_cookies()
                self.driver.execute_script("window.localStorage.clear();")
                self.driver.execute_script("window.sessionStorage.clear();")
                
                # Пробуем загрузить страницу снова
                self.driver.get(url)
                time.sleep(5)
                
                if "403 ERROR" in self.driver.page_source:
                    self.add_log("❌ Страница недоступна (ошибка 403)")
                    return None

            # Определяем даты для проверки
            today = date.today()
            yesterday = today - timedelta(days=1)
            tomorrow = today + timedelta(days=1)
            
            # Проверяем все возможные даты
            check_dates = []
            
            if target_date == today:
                check_dates.extend([today, yesterday])  # Проверяем сегодня и вчера
                print("📅 Ищем матчи на СЕГОДНЯ (включая вчерашние)")
            elif target_date == yesterday:
                check_dates.extend([yesterday, today])  # Проверяем вчера и сегодня
                print("📅 Ищем матчи на ВЧЕРА (включая сегодняшние)")
            else:
                # Для конкретной даты проверяем её саму и соседние даты
                check_dates.extend([
                    target_date - timedelta(days=1),
                    target_date,
                    target_date + timedelta(days=1)
                ])
                print(f"📅 Ищем матчи около даты {target_date.strftime('%d.%m.%Y')}")

            # Создаем форматы дат для всех проверяемых дат
            date_formats = []
            for check_date in check_dates:
                date_formats.extend([
                    # DD/MM/YYYY
                    f"{check_date.day}/{check_date.month}/{check_date.year}",
                    f"{check_date.day:02d}/{check_date.month:02d}/{check_date.year}",
                    f"{check_date.day}/{check_date.month:02d}/{check_date.year}",
                    f"{check_date.day:02d}/{check_date.month}/{check_date.year}",
                    
                    # DD/MM/YY 
                    f"{check_date.day}/{check_date.month}/{str(check_date.year)[2:]}",
                    f"{check_date.day:02d}/{check_date.month:02d}/{str(check_date.year)[2:]}",
                    
                    # D/M/YY (без ведущих нулей)
                    f"{check_date.day}/{check_date.month}/{str(check_date.year)[2:]}",
                    
                    # DD.MM.YYYY
                    f"{check_date.day}.{check_date.month}.{check_date.year}",
                    f"{check_date.day:02d}.{check_date.month:02d}.{check_date.year}",
                    f"{check_date.day}.{check_date.month:02d}.{check_date.year}",
                    f"{check_date.day:02d}.{check_date.month}.{check_date.year}",
                    
                    # D.M.YYYY (без ведущих нулей)
                    f"{check_date.day}.{check_date.month}.{check_date.year}",
                    
                    # D/M/YYYY (без ведущих нулей)
                    f"{check_date.day}/{check_date.month}/{check_date.year}",
                ])
            
            # Удаляем дубликаты
            date_formats = list(set(date_formats))
            print(f"Форматы дат для поиска: {date_formats}")

            # Анализ страницы перед поиском матчей
            print("\nСостояние страницы:")
            print(f"URL: {self.driver.current_url}")
            print(f"Заголовок: {self.driver.title}")
            print("Видимые элементы с текстом:")
            elements = self.driver.find_elements(By.CSS_SELECTOR, ".og-match-block, .topBar > div")
            for elem in elements:
                if elem.is_displayed():
                    try:
                        text = elem.text.strip()
                        if text:
                            print(f"- {text[:100]}")
                    except:
                        continue

            # Определяем, нужно ли искать кнопку "SHOW OTHER GAMES"
            if "fibalivestats" in url.lower():
                # Только для FibaLiveStats ищем кнопку
                try:
                    show_games_button = self.driver.find_element(By.XPATH, "//*[contains(text(), 'SHOW OTHER GAMES')]")
                    print("\nНайдена кнопка 'SHOW OTHER GAMES' - пытаемся открыть список матчей")
                    show_games_button.click()
                    time.sleep(3)
                except:
                    print("\nКнопка 'SHOW OTHER GAMES' не найдена на FibaLiveStats")
            else:
                print("\nПропускаем поиск кнопки 'SHOW OTHER GAMES' для не-FibaLiveStats сайта")

            # Определяем тип сайта и выбираем соответствующий метод поиска матчей
            if "lnbp.mx" in url.lower():
                # JavaScript для поиска матчей LNBP
                js_code = """
                function findLNBPMatches() {
                    var matches = [];
                    
                    // Находим таблицу статистики
                    var rows = document.querySelectorAll('table tr');
                    
                    rows.forEach(function(row, index) {
                        try {
                            // Пропускаем заголовок таблицы
                            if (index === 0) return;
                            
                            // Получаем все ячейки строки
                            var cells = row.querySelectorAll('td');
                            
                            if (cells.length >= 2) {
                                var team1 = cells[0] ? cells[0].textContent.trim() : '';
                                var team2 = cells[1] ? cells[1].textContent.trim() : '';
                                
                                if (team1 && team2) {
                                    matches.push({
                                        text: team1 + ' vs ' + team2,
                                        type: 'SCHEDULED',
                                        teams: [team1, team2],
                                        url: row.querySelector('a') ? row.querySelector('a').href : null
                                    });
                                }
                            }
                        } catch (e) {
                            console.error('Error processing row:', e);
                        }
                    });
                    
                    return matches;
                }
                return findLNBPMatches();
                """
            else:
                js_code = """
                function findAllMatches() {
                    try {
                        var results = [];
                        console.log('Начинаем поиск матчей...');
                        
                        // Храним уже найденные URL для избежания дубликатов
                        var foundUrls = new Set();
                        
                        // Оптимизированный список селекторов
                        var selectors = [
                            '.og-match-block',
                            '.og-game-block',
                            '.top-bar > div[class*="match"]',
                            '.matches-bar > div[class*="match"]'
                        ];
                        
                        console.log('Используем селекторы:', selectors.join(', '));
                        
                        // Маркеры для определения типа матча
                        var liveMarkers = ['VIVO', 'LIVE', 'PERIODO', 'PERIOD'];
                        var finalMarkers = ['FINAL', 'FIN'];
                        
                        selectors.forEach(function(selector) {
                            var elements = document.querySelectorAll(selector);
                            console.log('Найдено элементов для селектора ' + selector + ':', elements.length);
                            
                            elements.forEach(function(element, index) {
                                try {
                                    // Проверяем видимость и размеры элемента
                                    if (!element.offsetWidth || !element.offsetHeight) {
                                        return;
                                    }
                                    
                                    // Получаем ссылку из элемента
                                    var link = element.querySelector('a');
                                    if (!link || !link.href) {
                                        return;
                                    }
                                    
                                    // Проверяем не обрабатывали ли мы уже этот URL
                                    if (foundUrls.has(link.href)) {
                                        return;
                                    }
                                    
                                    var text = (element.innerText || element.textContent || '').trim();
                                    if (!text || text.length < 5) {
                                        return;
                                    }
                                    
                                    var rect = element.getBoundingClientRect();
                                    var type = 'UNKNOWN';
                                    
                                    // Определяем тип матча
                                    if (liveMarkers.some(marker => text.includes(marker))) {
                                        type = 'LIVE';
                                    } else if (finalMarkers.some(marker => text.includes(marker))) {
                                        type = 'FINAL';
                                    } else {
                                        type = 'SCHEDULED';
                                    }
                                    
                                    // Добавляем URL в множество обработанных
                                    foundUrls.add(link.href);
                                    
                                    // Добавляем матч в результаты
                                    results.push({
                                        text: text,
                                        type: type,
                                        url: link.href,
                                        x: rect.x + rect.width/2,
                                        y: rect.y + rect.height/2,
                                        width: rect.width,
                                        height: rect.height,
                                        element_index: index,
                                        selector: selector
                                    });
                                    
                                } catch (error) {
                                    console.error('Ошибка при обработке элемента:', error);
                                }
                            });
                        });
                        
                        // Удаляем возможные дубликаты по URL
                        results = Array.from(new Map(results.map(item => [item.url, item])).values());
                        
                        console.log('Всего найдено уникальных матчей:', results.length);
                        return results;
                        
                    } catch (error) {
                        console.error('Error in findAllMatches:', error);
                        console.error('Stack:', error.stack);
                        return [];
                    }
                }
                
                return findAllMatches();
                """

                # Выполняем поиск матчей
                all_matches = self.driver.execute_script(js_code)

                # Получаем логи браузера
                browser_logs = self.driver.get_log('browser')
                for log in browser_logs:
                    print(f"Browser Log: {log['message']}")
                
                if all_matches:
                    self.add_log(f"✓ Найдено потенциальных матчей: {len(all_matches)}")
                else:
                    print("✗ Матчи не найдены")
                    return None

            # Обрабатываем найденные матчи
            self.visited_urls.clear()
            try:
                matches_found = self.process_matches_parallel(all_matches, target_date)
                if getattr(self, 'cancel_scan', False):
                    self.add_log("Сканирование отменено пользователем")
                    return None
            except Exception as e:
                self.add_log(f"Ошибка при обработке матчей: {str(e)}", "ERROR")
                return None

            # Результаты (вне цикла for)
            print("\n" + "="*60)
            if matches_found:
                self.add_log(f"✅ Найдено матчей: {len(matches_found)}")
                for match in matches_found:
                    match_type_icon = {'LIVE': '🟢', 'FINAL': '⚫', 'SCHEDULED': '📅'}.get(match['type'], '❓')
                    self.add_log(f"   {match_type_icon} {match['text'][:60]}...")
            else:
                self.add_log("✗ Матчи не найдены")

            return matches_found
                
        except Exception as e:
            self.add_log(f"❌ Ошибка: {str(e)}")
            return None

    def scan_tournaments(self, target_date, progress_callback=None):
        """Сканирует активные турниры на указанную дату"""
        try:

            # Сбрасываем флаг отмены при начале сканирования
            self.cancel_scan = False

            # Получаем список активных турниров
            tournaments = self.get_active_tournaments()
            
            if not tournaments:
                logging.info("Нет активных турниров для сканирования")
                return []
            
            # Инициализация браузера
            if progress_callback:
                progress_callback("Инициализация браузера...", 0)
            
            self.init_driver(headless=self.headless)
            
            all_results = []
            total = len(tournaments)
            
            try:
                for i, (name, url) in enumerate(tournaments):
                    if self.cancel_scan:
                        self.add_log("Сканирование отменено пользователем")
                        break
                        
                    if progress_callback:
                        progress = int((i / total) * 100)
                        progress_callback(f"Проверяем турнир: {name}", progress)
                    
                    logging.info(f"\nПроверяем турнир: {name}")
                    logging.info(f"URL: {url}")
                    
                    # Получаем матчи турнира.  Если это турнир FibaLiveStats,
                    # используем собственный парсер.  Для других доменов
                    # (например, LNBP) вызываем общий маршрутизатор из scan_router.
                    low = (url or '').lower()
                    try:
                        if 'fibalivestats' in low:
                            matches = self.check_date_in_page(url, target_date)
                        else:
                            # импортируем лениво, чтобы избежать циклических зависимостей
                            try:
                                from scan_router import get_matches_by_url
                            except Exception:
                                get_matches_by_url = None
                            if get_matches_by_url:
                                matches = get_matches_by_url(self.driver, url, target_date, logger=self.add_log)
                            else:
                                # если маршрутизатор не найден, просто пропускаем
                                matches = []
                    except Exception as e:
                        self.add_log(f"Ошибка при получении матчей для {url}: {e}", "ERROR")
                        matches = []
                    
                    if matches:
                        # Добавляем информацию о турнире к каждому матчу
                        for match in matches:
                            match['tournament'] = name
                            match['tournament_url'] = url
                            all_results.append(match)
                        
                        self.add_log(f"✓ Турнир {name} - найдено матчей: {len(matches)}")
                    else:
                        self.add_log(f"✗ Турнир {name} - матчей нет")

                
                return all_results
                
            finally:
                self.close_driver()
                # Сбрасываем флаг отмены при завершении
                self.cancel_scan = False
                
        except Exception as e:
            self.add_log(f"Ошибка при сканировании турниров: {str(e)}", "ERROR")
            return []
    
    async def scan_all_sources(self, target_date: datetime) -> list:
        """
        Сканирует только источники FibaLiveStats.

        В предыдущей версии этот метод пытался собирать данные из нескольких
        источников (через MultiSourceMatchFinder).  Теперь вся логика
        мультисайтового сканирования вынесена в вызывающий код.  Этот метод
        возвращает только результаты сканирования FibaLiveStats, выполняя
        синхронный вызов `scan_tournaments`.

        Parameters
        ----------
        target_date : datetime
            Дата, на которую требуется найти матчи.

        Returns
        -------
        list
            Список найденных матчей FibaLiveStats.  Если произошла
            ошибка, возвращается пустой список.
        """
        try:
            return self.scan_tournaments(target_date)
        except Exception as e:
            self.add_log(f"Ошибка при сканировании FibaLiveStats: {str(e)}", "ERROR")
            return []
        
    def cancel_scanning(self):
        """Отмена процесса сканирования"""
        try:
            self.cancel_scan = True
            self.add_log("Сканирование отменено пользователем")
            # Закрываем драйвер
            self.close_driver()
        except Exception as e:
            self.add_log(f"Ошибка при отмене сканирования: {str(e)}", "ERROR")

    # --- Вспомогательные методы ---

    def save_results(self, urls, date):
        """Сохранение результатов в Excel с дополнительной информацией"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            
            # Формируем имя файла
            date_str = date.strftime("%d_%m_%Y")
            time_str = datetime.now().strftime("%H%M%S")
            default_filename = f"fibalivestats_{date_str}_{time_str}.xlsx"
            
            # Пробуем разные пути для сохранения
            possible_paths = [
                os.path.join(os.path.dirname(os.path.abspath(__file__))),  # Текущая директория скрипта
                os.path.join(os.path.expanduser("~"), "Documents"),  # Документы
                os.path.join(os.path.expanduser("~"), "Desktop"),  # Рабочий стол
                os.getcwd()  # Текущая рабочая директория
            ]
            
            last_error = None
            for save_dir in possible_paths:
                try:
                    os.makedirs(save_dir, exist_ok=True)
                    filepath = os.path.join(save_dir, default_filename)
                    
                    # Создаем новую книгу Excel
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Matches"
                    
                    # Стили для заголовков
                    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True)
                    
                    # Заголовки с форматированием
                    headers = [
                        "URL матча",
                        "Тип матча",
                        "Дата матча",
                        "Время сканирования"
                    ]
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # Записываем данные
                    row = 2
                    scan_time = datetime.now().strftime("%H:%M:%S")
                    
                    for url in urls:
                        if isinstance(url, dict):
                            # Если url - это словарь с дополнительной информацией
                            match_url = url.get('url') or url.get('match_url')
                            match_type = url.get('type', '')
                            match_date = url.get('date', date_str)
                            
                            if match_url:
                                ws.cell(row=row, column=1, value=match_url)
                                ws.cell(row=row, column=2, value=match_type)
                                ws.cell(row=row, column=3, value=match_date)
                                ws.cell(row=row, column=4, value=scan_time)
                                row += 1
                        elif isinstance(url, str):
                            # Если url - это просто строка
                            ws.cell(row=row, column=1, value=url)
                            ws.cell(row=row, column=4, value=scan_time)
                            row += 1
                    
                    # Автоматическая ширина столбцов
                    for column in ws.columns:
                        max_length = 0
                        column_letter = openpyxl.utils.get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Добавляем информацию о сканировании
                    ws2 = wb.create_sheet(title="Info")
                    
                    info_data = [
                        ["Дата сканирования", date.strftime("%d.%m.%Y")],
                        ["Время сканирования", scan_time],
                        ["Количество матчей", row - 2],
                        ["Пользователь", "danilaDanikov"]
                    ]
                    
                    for i, (key, value) in enumerate(info_data, 1):
                        ws2.cell(row=i, column=1, value=key)
                        ws2.cell(row=i, column=2, value=value)
                    
                    # Пробуем сохранить файл
                    wb.save(filepath)
                    print(f"✓ Файл сохранен: {filepath}")
                    
                    # Открываем проводник с выделенным файлом
                    try:
                        subprocess.Popen(f'explorer /select,"{filepath}"')
                    except:
                        pass  # Игнорируем ошибки при открытии проводника
                    
                    return filepath
                    
                except Exception as e:
                    last_error = e
                    print(f"⚠️ Не удалось сохранить в {save_dir}: {str(e)}")
                    continue
            
            # Если не удалось сохранить ни в одной директории
            if last_error:
                raise Exception(f"Не удалось сохранить файл ни в одной директории. Последняя ошибка: {str(last_error)}")
            
        except Exception as e:
            print(f"❌ Критическая ошибка при сохранении: {str(e)}")
            return None

    def save_debug_screenshot(self, name):
        """Сохраняет скриншот для отладки"""
        try:
            screenshot_dir = os.path.join(
                os.path.expanduser("~"),
                "AppData", "Local", "ExcelAnalyzer", "debug"
            )
            os.makedirs(screenshot_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_path = os.path.join(screenshot_dir, f"{timestamp}_{name}.png")
            
            self.driver.save_screenshot(screenshot_path)
            logging.info(f"Скриншот сохранен: {screenshot_path}")
            
            # Также сохраняем HTML
            html_path = screenshot_path.replace('.png', '.html')
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(self.driver.page_source)
            logging.info(f"HTML сохранен: {html_path}")
            
        except Exception as e:
            logging.error(f"Не удалось сохранить скриншот: {e}")

    def import_tournaments_from_excel(self, excel_handler):
        """Импорт турниров из Excel файлов"""
        imported = 0
        
        try:
            # Импорт из Ройки (Halfs Champs)
            if hasattr(excel_handler, 'data') and excel_handler.data:
                for tournament_name in excel_handler.data.keys():
                    # Генерируем заглушку URL для турнира
                    placeholder_url = f"https://fibalivestats.dcd.shared.geniussports.com/u/TOURNAMENT/{tournament_name}/"
                    if self.add_tournament(tournament_name, placeholder_url, active=False):
                        imported += 1
                        
            logging.info(f"Импортировано {imported} турниров")
            return imported
            
        except Exception as e:
            logging.error(f"Ошибка при импорте турниров: {str(e)}")
            return imported
        
    async def process_matches_async(self, matches, progress_callback=None):
        """Асинхронная обработка матчей"""
        tasks = []
        async with aiohttp.ClientSession() as session:
            for i, match in enumerate(matches):
                if progress_callback:
                    progress = int((i / len(matches)) * 100)
                    progress_callback(f"Обработка матча {i+1}/{len(matches)}", progress)
                
                task = asyncio.create_task(self.process_match_async(session, match))
                tasks.append(task)
            return await asyncio.gather(*tasks)
            
    async def process_match_async(self, session, match):
        """Асинхронная обработка отдельного матча"""
        try:
            async with session.get(match['url']) as response:
                html = await response.text()
                # Ваша логика обработки матча...
                return result
        except Exception as e:
            logging.error(f"Error processing match: {str(e)}")
            return None
        
class LNBPParser(BaseMatchParser):
    def __init__(self, handler):
        super().__init__(handler)
        self.base_url = "https://www.lnbp.mx/stats.html"

    def validate_url(self, url: str) -> bool:
        return "lnbp.mx" in url.lower()    


    async def parse_matches(self, url: str, target_date: datetime) -> list:
        matches = []
        try:
            self.add_log("Начало парсинга LNBP...")
            self.driver.get(url)
            
            # Ждем загрузку страницы
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            time.sleep(2)

            # JavaScript для поиска матчей
            js_code = """
            function findLNBPMatches() {
                var matches = [];
                try {
                    // Общие селекторы для матчей
                    var selectors = [
                        '.partido', '.match', '.game',           // Общие классы
                        '[class*="partido"]', '[class*="match"]', // Частичное совпадение
                        '.calendario-item', '.schedule-item',     // Элементы календаря
                        '.resultado-item', '.result-item',        // Элементы результатов
                        'div:has(> .team, > .equipo)',           // Контейнеры с командами
                        'div:has(> .score, > .resultado)'        // Контейнеры со счетом
                    ];

                    selectors.forEach(function(selector) {
                        document.querySelectorAll(selector).forEach(function(element) {
                            try {
                                if (!element.offsetWidth || !element.offsetHeight) return;

                                var text = element.innerText || element.textContent || '';
                                if (!text.trim()) return;

                                // Поиск команд
                                var teams = Array.from(element.querySelectorAll('.team, .equipo, [class*="team"], [class*="equipo"]'))
                                    .map(t => t.innerText.trim())
                                    .filter(t => t.length > 0);

                                // Если не нашли команды, пробуем разбить текст
                                if (teams.length < 2) {
                                    var possibleTeams = text.split(/vs\.?|contra|-/).map(t => t.trim());
                                    if (possibleTeams.length >= 2) {
                                        teams = [possibleTeams[0], possibleTeams[1]];
                                    }
                                }

                                if (teams.length >= 2) {
                                    // Определяем тип матча
                                    var type = 'SCHEDULED';
                                    if (text.toLowerCase().includes('vivo') || text.toLowerCase().includes('live')) {
                                        type = 'LIVE';
                                    } else if (text.toLowerCase().includes('final')) {
                                        type = 'FINAL';
                                    }

                                    // Ищем URL
                                    var link = element.querySelector('a');
                                    var url = link ? link.href : null;

                                    matches.push({
                                        teams: teams,
                                        type: type,
                                        url: url,
                                        text: text,
                                        rect: element.getBoundingClientRect()
                                    });
                                }
                            } catch (e) {
                                console.error('Error processing element:', e);
                            }
                        });
                    });

                    return matches;
                } catch (e) {
                    console.error('Error in findLNBPMatches:', e);
                    return [];
                }
            }
            return findLNBPMatches();
            """

            # Выполняем JavaScript
            found_matches = self.driver.execute_script(js_code)
            self.add_log(f"JavaScript нашел {len(found_matches)} потенциальных матчей")

            # Обрабатываем найденные матчи
            for match_data in found_matches:
                try:
                    if not match_data.get('url'):
                        continue

                    # Проверяем дату матча
                    match_text = match_data['text']
                    date_str = self._format_date_spanish(target_date)
                    
                    # Если дата найдена в тексте матча или это LIVE матч
                    if date_str.lower() in match_text.lower() or match_data['type'] == 'LIVE':
                        matches.append({
                            'url': match_data['url'],
                            'type': match_data['type'],
                            'teams': {
                                'home': match_data['teams'][0],
                                'away': match_data['teams'][1]
                            },
                            'text': f"{match_data['teams'][0]} vs {match_data['teams'][1]}",
                            'date': target_date.strftime('%d/%m/%Y'),
                            'source': 'LNBP'
                        })
                        self.add_log(f"Добавлен матч: {match_data['teams'][0]} vs {match_data['teams'][1]} ({match_data['type']})")

                except Exception as e:
                    self.add_log(f"Ошибка при обработке матча: {str(e)}")
                    continue

            self.add_log(f"Всего найдено матчей LNBP: {len(matches)}")
            return matches

        except Exception as e:
            self.add_log(f"Ошибка при парсинге LNBP: {str(e)}")
            return []

    def _format_date_spanish(self, date: datetime) -> str:
        """Форматирует дату в испанском формате"""
        months_es = {
            1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 
            5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
            9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
        }
        return f"{date.day} de {months_es[date.month]} de {date.year}"