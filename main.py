import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Load .env file BEFORE any other imports that might read DATABASE_URL
try:
    from dotenv import load_dotenv
    _env_path = os.path.join(
        os.path.dirname(sys.executable) if getattr(sys, 'frozen', False)
        else os.path.dirname(os.path.abspath(__file__)),
        '.env'
    )
    load_dotenv(_env_path, override=False)
except ImportError:
    pass

import json
import pandas as pd
import traceback
import logging
import shutil
from contextlib import contextmanager
from bets_monitor import BetsAPIMonitor
from telegram_notifier import TelegramNotifier
from config import TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID
from datetime import datetime

# Явный импорт halfs_database на верхнем уровне, чтобы PyInstaller
# гарантированно включил модуль в сборку .exe
try:
    from halfs_database import HalfsDatabase as _HalfsDatabase  # noqa: F401
except ImportError:
    _HalfsDatabase = None
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, QHBoxLayout,
                             QWidget, QFileDialog, QLabel, QLineEdit, QScrollArea,
                             QGridLayout, QMessageBox, QTabWidget, QProgressBar, QTableWidget,
                             QTableWidgetItem, QHeaderView, QMenu, QAction, QActionGroup,
                             QSizePolicy, QAbstractItemView, QCompleter, QListWidget,
                             QListWidgetItem, QStackedWidget, QFrame, QInputDialog, QDialog,
                             QCheckBox, QDialogButtonBox, QStyledItemDelegate, QCalendarWidget, QProgressDialog, QDialogButtonBox, QGroupBox, QPlainTextEdit, QComboBox, QSplitter,
                             QTextEdit, QSystemTrayIcon, QDateEdit)
from PyQt5.QtCore import (Qt, QThread, pyqtSignal, QSettings, QSize, QRect, QPoint, QEvent,
                          QStringListModel, QDate, QTimer)  # Добавлен QTimer для уведомлений
from PyQt5.QtGui import (QFont, QColor, QPalette, QIcon, QPixmap, QBrush, QPen, QPainter, QLinearGradient)
from tkinter import messagebox
import pandas as pd
import requests
import time
from typing import Dict, Tuple, List, Optional, Set
import math


# Импортируем наш новый модуль
from betsapi_handler import BetsAPIHandler
from fibalivestats_page import FibaLiveStatsPage

import sqlite3
from datetime import datetime
import logging

# Лог-файл рядом с exe/скриптом, а не в рабочей директории
_log_base = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
_log_file = os.path.join(_log_base, 'app.log')
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(_log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

from bets_notifier import BetsNotifier


def get_data_dir() -> str:
    """Папка для хранения баз данных — рядом с .exe / main.py."""
    if getattr(sys, 'frozen', False):
        # Запуск из .exe (PyInstaller)
        base = os.path.dirname(sys.executable)
    else:
        # Запуск из Python
        base = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base, "data")
    os.makedirs(data_dir, exist_ok=True)

    # --- Диагностика: записываем путь к data в лог рядом с exe/скриптом ---
    try:
        diag_path = os.path.join(base, "data_dir_debug.log")
        with open(diag_path, "a", encoding="utf-8") as f:
            import datetime as _dt
            f.write(f"[{_dt.datetime.now():%Y-%m-%d %H:%M:%S}] "
                    f"frozen={getattr(sys, 'frozen', False)} | "
                    f"sys.executable={sys.executable} | "
                    f"base={base} | "
                    f"data_dir={data_dir} | "
                    f"exists={os.path.isdir(data_dir)} | "
                    f"files={os.listdir(data_dir) if os.path.isdir(data_dir) else 'N/A'}\n")
    except Exception:
        pass

    return data_dir


def my_exception_hook(exctype, value, tb):
    # Записываем ошибку в файл на рабочем столе
    error_log_path = os.path.join(os.path.expanduser("~"), "Desktop", "error_log.txt")
    try:
        with open(error_log_path, "w", encoding="utf-8") as f:
            f.write(f"Тип: {exctype}\n")
            f.write(f"Значение: {value}\n")
            f.write(f"Трассировка:\n{''.join(traceback.format_tb(tb))}")
    except:
        pass
    
    # Логируем ошибку
    logging.critical(f"Необработанное исключение: {value}\n{''.join(traceback.format_tb(tb))}")
    
    # Показываем сообщение (без зависимости от PyQt5)
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, 
            f"Произошла ошибка: {value}\n\nДетали записаны в файл:\n{error_log_path}", 
            "Ошибка в приложении", 0)
    except:
        pass

# Устанавливаем перехватчик исключений
sys.excepthook = my_exception_hook

# После всех импортов, но перед классами
class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return None
    
class RoykaDatabase:
    """Класс для работы с базой данных раздела Ройка"""

    _SCHEMA = 'royka'

    def __init__(self):
        self.db_path = os.path.join(get_data_dir(), "royka.db")
        self.init_database()

    @contextmanager
    def _connect(self):
        """Unified connection: PostgreSQL or SQLite."""
        from db_connection import db_connect
        with db_connect(schema=self._SCHEMA, sqlite_path=self.db_path) as conn:
            yield conn

    def init_database(self):
        """Инициализация базы данных"""
        with self._connect() as conn:
            cursor = conn.cursor()
            
            # Создаем таблицу для матчей
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS matches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    tournament TEXT NOT NULL,
                    team_home TEXT NOT NULL,
                    team_away TEXT NOT NULL,
                    t1h REAL,
                    t2h REAL,
                    tim REAL NOT NULL,
                    deviation REAL,
                    kickoff REAL,
                    predict TEXT NOT NULL,
                    result REAL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Создаем индексы для быстрого поиска
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_tournament 
                ON matches(tournament)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_date 
                ON matches(date)
            """)
            
            conn.commit()
    
    def add_matches(self, matches_data):
        """Добавление матчей в базу"""
        with self._connect() as conn:
            cursor = conn.cursor()
            cursor.executemany("""
                INSERT INTO matches (
                    date, tournament, team_home, team_away,
                    t1h, t2h, tim, deviation, kickoff,
                    predict, result
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, matches_data)
            conn.commit()
    
    def get_statistics(self):
        """Получение статистики базы данных"""
        try:
            with self._connect() as conn:
                cursor = conn.cursor()
                
                # Получаем общее количество записей
                cursor.execute("SELECT COUNT(*) FROM matches")
                total_records = cursor.fetchone()[0]
                
                # Получаем количество уникальных турниров
                cursor.execute("SELECT COUNT(DISTINCT tournament) FROM matches")
                tournaments_count = cursor.fetchone()[0]
                
                # Получаем количество уникальных команд
                cursor.execute("""
                    SELECT COUNT(DISTINCT team) FROM (
                        SELECT team_home as team FROM matches
                        UNION
                        SELECT team_away as team FROM matches
                    )
                """)
                teams_count = cursor.fetchone()[0]
                
                # Получаем дату последнего обновления
                cursor.execute("SELECT MAX(date) FROM matches")
                last_update = cursor.fetchone()[0]
                
                return {
                    'total_records': total_records,
                    'tournaments_count': tournaments_count,
                    'teams_count': teams_count,
                    'last_update': last_update
                }
                
        except Exception as e:
            print(f"Error in get_statistics: {str(e)}")  # Добавляем вывод ошибки в консоль
            return {
                'total_records': 0,
                'tournaments_count': 0,
                'teams_count': 0,
                'last_update': None
            }
    
    def clear_database(self):
        """Очистка всех данных"""
        with self._connect() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM matches")
            conn.commit()
    
    def create_backup(self, backup_path=None):
        """Создание резервной копии базы"""
        if not backup_path:
            backup_dir = os.path.join(
                os.path.expanduser("~"),
                "AppData",
                "Local",
                "ExcelAnalyzer",
                "backups"
            )
            os.makedirs(backup_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(
                backup_dir,
                f"royka_backup_{timestamp}.db"
            )
        
        import shutil
        shutil.copy2(self.db_path, backup_path)
        return backup_path
    
    def restore_from_backup(self, backup_path):
        """Восстановление из резервной копии"""
        import shutil
        shutil.copy2(backup_path, self.db_path)

    def normalize_numeric_values(self):
        """Преобразует все числовые значения в базе в корректный формат"""
        try:
            with self._connect() as conn:
                cursor = conn.cursor()
                
                # Создаем временную таблицу с правильными типами данных
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS matches_temp (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        date TEXT NOT NULL,
                        tournament TEXT NOT NULL,
                        team_home TEXT NOT NULL,
                        team_away TEXT NOT NULL,
                        t1h REAL,
                        t2h REAL,
                        tim REAL NOT NULL,
                        deviation REAL,
                        kickoff REAL,
                        predict REAL,
                        result REAL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Копируем данные с преобразованием
                cursor.execute("""
                    INSERT INTO matches_temp (
                        date, tournament, team_home, team_away,
                        t1h, t2h, tim, deviation, kickoff, predict, result
                    )
                    SELECT 
                        date, tournament, team_home, team_away,
                        CAST(REPLACE(COALESCE(NULLIF(t1h, ''), '0'), ',', '.') AS REAL),
                        CAST(REPLACE(COALESCE(NULLIF(t2h, ''), '0'), ',', '.') AS REAL),
                        CAST(REPLACE(COALESCE(NULLIF(tim, ''), '0'), ',', '.') AS REAL),
                        CAST(REPLACE(COALESCE(NULLIF(deviation, ''), '0'), ',', '.') AS REAL),
                        CAST(REPLACE(COALESCE(NULLIF(kickoff, ''), '0'), ',', '.') AS REAL),
                        CAST(REPLACE(COALESCE(NULLIF(predict, ''), '0'), ',', '.') AS REAL),
                        CAST(REPLACE(COALESCE(NULLIF(result, ''), '0'), ',', '.') AS REAL)
                    FROM matches
                """)
                
                # Удаляем старую таблицу
                cursor.execute("DROP TABLE matches")
                
                # Переименовываем временную таблицу
                cursor.execute("ALTER TABLE matches_temp RENAME TO matches")
                
                # Создаем индексы заново
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS idx_tournament 
                    ON matches(tournament)
                """)
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS idx_date 
                    ON matches(date)
                """)
                
                # Получаем количество обработанных записей
                cursor.execute("SELECT COUNT(*) FROM matches")
                count = cursor.fetchone()[0]
                
                return count
                
        except Exception as e:
            logging.error(f"Ошибка при нормализации данных: {str(e)}")
            raise    

# Добавьте эти классы в main.py после импортов

class DateSelectionDialog(QDialog):
    # ... код класса ...
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_date = None
        self.setup_ui()
    
    def setup_ui(self):
        self.setWindowTitle("Выберите дату матчей")
        self.setMinimumSize(400, 400)
        
        layout = QVBoxLayout(self)
        
        # Заголовок
        title = QLabel("Выберите дату для поиска матчей:")
        title.setFont(QFont("Arial", 12))
        layout.addWidget(title)
        
        # Календарь с темной темой
        self.calendar = QCalendarWidget()
        self.calendar.setStyleSheet("""
            QCalendarWidget {
                background-color: #1e1e1e;
                color: #ffffff;
            }
            QCalendarWidget QToolButton {
                color: white;
                background-color: #2d2d2d;
                border: 1px solid #3d3d3d;
                border-radius: 3px;
                padding: 6px;
            }
            QCalendarWidget QToolButton:hover {
                background-color: #3d3d3d;
                border: 1px solid #4d4d4d;
            }
            QCalendarWidget QMenu {
                background-color: #2d2d2d;
                color: white;
                border: 1px solid #3d3d3d;
            }
            QCalendarWidget QSpinBox {
                color: white;
                background-color: #2d2d2d;
                border: 1px solid #3d3d3d;
                border-radius: 3px;
                padding: 3px;
            }
            QCalendarWidget QTableView {
                background-color: #1e1e1e;
                selection-background-color: #2d5c8f;
                selection-color: white;
                alternate-background-color: #232323;
            }
            QCalendarWidget QTableView:enabled {
                color: white;
            }image.png
            QCalendarWidget QTableView:disabled {
                color: #666666;
            }
            QCalendarWidget QTableView::item:hover {
                background-color: #2d5c8f;
            }
        """)
        
        self.calendar.setSelectedDate(QDate.currentDate())
        self.calendar.clicked.connect(self.on_date_selected)
        layout.addWidget(self.calendar)
    
    # Продолжение существующего кода...
        
        # Выбранная дата
        self.date_label = QLabel(f"Выбрана дата: {QDate.currentDate().toString('dd.MM.yyyy')}")
        self.date_label.setFont(QFont("Arial", 11, QFont.Bold))
        layout.addWidget(self.date_label)
        
        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def on_date_selected(self, date):
        """Обработка выбора даты"""
        self.date_label.setText(f"Выбрана дата: {date.toString('dd.MM.yyyy')}")
    
    def get_selected_date(self):
        """Возвращает выбранную дату в формате YYYYMMDD"""
        date = self.calendar.selectedDate()
        return date.toString("yyyyMMdd")


class MatchSelectionDialog(QDialog):
    """Диалог выбора матча из нескольких найденных"""
    
    def __init__(self, matches, team1, team2, parent=None):
        super().__init__(parent)
        self.matches = matches
        self.selected_match = None
        self.team1 = team1
        self.team2 = team2
        self.setup_ui()
    
    def setup_ui(self):
        self.setWindowTitle("Выберите матч")
        self.setMinimumSize(700, 400)
        
        layout = QVBoxLayout(self)
        
        # Заголовок
        title = QLabel(f"Найдено матчей: {len(self.matches)}\nИскали: {self.team1} vs {self.team2}")
        title.setFont(QFont("Arial", 12))
        layout.addWidget(title)
        
        # Список матчей
        self.matches_list = QListWidget()
        for i, match in enumerate(self.matches):
            match_text = f"{match['time']} | {match['home']} vs {match['away']}"
            if match['league']:
                match_text += f" | {match['league']}"
            
            # Добавляем индикатор live/upcoming
            if match['match_type'] == 'inplay':
                match_text = "🔴 LIVE | " + match_text
            else:
                match_text = "📅 " + match_text
            
            # Добавляем информацию о найденных командах
            if 'teams_found' in match:
                match_text += f"\n    Совпадение: {match['teams_found']}"
            
            # Добавляем уровень уверенности
            if 'confidence' in match:
                if match['confidence'] >= 100:
                    match_text += " ⭐⭐⭐"  # Обе команды найдены
                elif match['confidence'] >= 70:
                    match_text += " ⭐⭐"    # Одна команда найдена
                else:
                    match_text += " ⭐"      # Частичное совпадение
            
            self.matches_list.addItem(match_text)
        
        self.matches_list.setCurrentRow(0)
        layout.addWidget(self.matches_list)
        
        # Информационная подпись
        info_label = QLabel("⭐⭐⭐ - обе команды найдены | ⭐⭐ - одна команда найдена")
        info_label.setFont(QFont("Arial", 9))
        layout.addWidget(info_label)
        
        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept_selection)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def accept_selection(self):
        """Сохраняет выбранный матч"""
        current_row = self.matches_list.currentRow()
        if current_row >= 0:
            self.selected_match = self.matches[current_row]
        self.accept()


# Настройка логирования
log_dir = os.path.join(os.path.expanduser("~"), "AppData", "Local", "Temp")
os.makedirs(log_dir, exist_ok=True)
log_path = os.path.join(log_dir, "ExcelAnalyzer_main.log")
logging.basicConfig(
    filename=log_path,
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Функция для перехвата необработанных исключений
def my_exception_hook(exctype, value, tb):
    # Записываем ошибку в файл на рабочем столе
    error_log_path = os.path.join(os.path.expanduser("~"), "Desktop", "error_log.txt")
    try:
        with open(error_log_path, "w", encoding="utf-8") as f:
            f.write(f"Тип: {exctype}\n")
            f.write(f"Значение: {value}\n")
            f.write(f"Трассировка:\n{''.join(traceback.format_tb(tb))}")
    except:
        pass
    
    # Логируем ошибку
    logging.critical(f"Необработанное исключение: {value}\n{''.join(traceback.format_tb(tb))}")
    
    # Выводим в консоль
    print(f"КРИТИЧЕСКАЯ ОШИБКА: {value}")
    print("".join(traceback.format_tb(tb)))
    
    # Показываем сообщение (без зависимости от PyQt5)
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, 
            f"Произошла ошибка: {value}\n\nДетали записаны в файл:\n{error_log_path}", 
            "Ошибка в приложении", 0)
    except:
        pass

# Устанавливаем перехватчик исключений
sys.excepthook = my_exception_hook

def log_step(step_name):
    """Функция для логирования шагов инициализации"""
    msg = f"ШАГ: {step_name}"
    logging.info(msg)
    # Также записываем в отдельный файл для надежности
    try:
        with open(os.path.join(log_dir, "main_steps.txt"), "a", encoding="utf-8") as f:
            f.write(f"{datetime.now()} - {msg}\n")
    except Exception as e:
        logging.error(f"Не удалось записать в файл шагов: {str(e)}")

# Функция для определения путей к ресурсам
def resource_path(relative_path):
    """Функция для получения абсолютного пути к ресурсу"""
    try:
        # Записываем в лог для отладки
        with open(os.path.join(log_dir, "resource_path.log"), "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.now()}] Запрос ресурса: {relative_path}\n")
        
        # PyInstaller создает временную папку и хранит путь в _MEIPASS
        base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
        result_path = os.path.join(base_path, relative_path)
        
        # Логируем результат
        with open(os.path.join(log_dir, "resource_path.log"), "a", encoding="utf-8") as f:
            f.write(f"База: {base_path}\nРезультат: {result_path}\nСуществует: {os.path.exists(result_path)}\n")
        
        return result_path
    except Exception as e:
        # В случае ошибки возвращаем просто текущий путь
        logging.error(f"Ошибка в resource_path: {str(e)}")
        return os.path.join(os.path.abspath("."), relative_path)

class CustomTitleBar(QWidget):
    """Кастомная строка заголовка для окон"""
    
    def __init__(self, parent=None, title="", icon_path=None):
        super().__init__(parent)
        self.parent = parent
        self.title = title
        self.setFixedHeight(30)  # Высота заголовка
        self.pressing = False
        self.start = QPoint(0, 0)
        
        # Настройка внешнего вида
        self.setStyleSheet("""
            background-color: #1A1A20;  /* Темный фон заголовка */
            color: #999999;             /* Приглушенный текст */
        """)
        
        # Создаем layout для заголовка
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 0, 5, 0)
        
        # Иконка, если указана
        if icon_path:
            self.icon_label = QLabel()
            self.icon_label.setPixmap(QIcon(icon_path).pixmap(16, 16))
            layout.addWidget(self.icon_label)
            layout.addSpacing(5)
        
        # Текст заголовка
        self.title_label = QLabel(title)
        self.title_label.setStyleSheet("color: #999999; font-weight: normal;")
        self.title_label.setFont(QFont("Segoe UI", 9))
        layout.addWidget(self.title_label)
        layout.addStretch()
        
        # Кнопки управления окном
        self.min_btn = QPushButton("−")
        self.min_btn.setFixedSize(24, 24)
        self.min_btn.clicked.connect(self.minimize_window)
        
        self.max_btn = QPushButton("□")
        self.max_btn.setFixedSize(24, 24)
        self.max_btn.clicked.connect(self.maximize_window)
        
        self.close_btn = QPushButton("×")
        self.close_btn.setFixedSize(24, 24)
        self.close_btn.clicked.connect(self.close_window)
        
        # Стили для кнопок
        self.min_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #999999;
                border: none;
                font-family: Arial;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #363640;
                color: #CCCCCC;
            }
        """)
        
        self.max_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #999999;
                border: none;
                font-family: Arial;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #363640;
                color: #CCCCCC;
            }
        """)
        
        self.close_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #999999;
                border: none;
                font-family: Arial;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #E81123;
                color: white;
            }
        """)
        
        layout.addWidget(self.min_btn)
        layout.addWidget(self.max_btn)
        layout.addWidget(self.close_btn)
        
        self.setLayout(layout)
    
    def update_title(self, title):
        """Обновляет текст заголовка"""
        self.title_label.setText(title)
    
    def mousePressEvent(self, event):
        """Обработка нажатия мыши для перемещения окна"""
        if event.button() == Qt.LeftButton:
            self.pressing = True
            self.start = self.mapToGlobal(event.pos())
    
    def mouseMoveEvent(self, event):
        """Обработка перемещения мыши для движения окна"""
        if self.pressing:
            self.end = self.mapToGlobal(event.pos())
            movement = self.end - self.start
            
            # Перемещаем родительское окно
            if self.parent:
                parent_pos = self.parent.mapToGlobal(QPoint(0, 0))
                self.parent.move(parent_pos.x() + movement.x(),
                                parent_pos.y() + movement.y())
                self.start = self.end
    
    def mouseReleaseEvent(self, event):
        """Обработка отпускания кнопки мыши"""
        if event.button() == Qt.LeftButton:
            self.pressing = False
    
    def mouseDoubleClickEvent(self, event):
        """Обработка двойного клика для разворачивания окна"""
        if event.button() == Qt.LeftButton:
            self.maximize_window()
    
    def minimize_window(self):
        """Сворачивает окно"""
        if self.parent:
            self.parent.showMinimized()
    
    def maximize_window(self):
        """Разворачивает или восстанавливает размер окна"""
        if self.parent:
            if self.parent.isMaximized():
                self.parent.showNormal()
                self.max_btn.setText("□")
            else:
                self.parent.showMaximized()
                self.max_btn.setText("❐")
    
    def close_window(self):
        """Закрывает окно"""
        if self.parent:
            self.parent.close()

# Импортируем необходимые библиотеки для работы с Excel
try:
    import xlwings as xw
    logging.info("xlwings импортирован успешно")
except ImportError:
    xw = None
    logging.warning("xlwings не найден")

# Класс для работы с Excel через xlwings
class ExcelHandler:
    def __init__(self):
        self.halfs_file = None
        self.cyber_files = []
        self.available = xw is not None
    
    def set_halfs_file(self, file_path):
        """Установка файла Половины"""
        self.halfs_file = file_path
        return os.path.exists(file_path)
    
    def add_cyber_file(self, file_path):
        """Добавление файла Cyber"""
        if file_path not in self.cyber_files:
            self.cyber_files.append(file_path)
        return os.path.exists(file_path)
    
    def clear_cyber_files(self):
        """Очистка списка файлов Cyber"""
        self.cyber_files = []
    
    def get_opened_excel_files(self, file_type=None):
        """
        Получение списка открытых файлов Excel
        
        Args:
            file_type: 'cyber' для киберфайлов, 'halves' для половин, None для всех
        """
        if not self.available:
            return []
            
        try:
            open_files = []
            
            # Получаем ВСЕ запущенные приложения Excel, а не только активное
            for app in xw.apps:
                for book in app.books:
                    file_path = book.fullname
                    file_name = os.path.basename(file_path).lower()
                    
                    # Логируем для отладки
                    logging.info(f"Проверяем файл: {file_name} (полный путь: {file_path})")
                    
                    # Пропускаем временные файлы
                    if file_name.startswith('~') or file_name.startswith('$'):
                        continue
                    
                    # Фильтрация по типу файла
                    if file_type == 'cyber':
                        # Киберфайлы: начинаются с "Cyber" (регистронезависимо)
                        if file_name.startswith('cyber'):
                            # Проверяем, что файл еще не добавлен (избегаем дубликатов)
                            if file_path not in open_files:
                                open_files.append(file_path)
                                logging.info(f"  -> Добавлен как Cyber файл")
                    elif file_type == 'halves':
                        # Файлы половин: содержат "Половин" в названии
                        if 'половин' in file_name:
                            # Проверяем, что файл еще не добавлен (избегаем дубликатов)
                            if file_path not in open_files:
                                open_files.append(file_path)
                                logging.info(f"  -> Добавлен как файл Половин")
                    else:
                        # Без фильтра - добавляем все файлы
                        if file_path not in open_files:
                            open_files.append(file_path)
                            
            logging.info(f"Найдено файлов типа '{file_type}': {len(open_files)}")
            logging.info(f"Файлы: {[os.path.basename(f) for f in open_files]}")
            return open_files
            
        except Exception as e:
            logging.error(f"Ошибка при получении открытых файлов: {str(e)}")
            # Пробуем альтернативный способ через активное приложение
            try:
                app = xw.apps.active
                if app:
                    for book in app.books:
                        file_path = book.fullname
                        file_name = os.path.basename(file_path).lower()
                        
                        if file_name.startswith('~') or file_name.startswith('$'):
                            continue
                        
                        if file_type == 'cyber':
                            if file_name.startswith('cyber'):
                                if file_path not in open_files:
                                    open_files.append(file_path)
                        elif file_type == 'halves':
                            if 'половин' in file_name:
                                if file_path not in open_files:
                                    open_files.append(file_path)
                        else:
                            if file_path not in open_files:
                                open_files.append(file_path)
            except:
                pass
                
            return open_files
    
    def get_halfs_sheets(self):
        """Получение списка листов из файла Половины"""
        if not self.halfs_file or not self.available:
            return []
        
        try:
            # Проверяем, открыт ли файл уже
            app = xw.apps.active
            if app:
                for book in app.books:
                    if book.fullname.lower() == self.halfs_file.lower():
                        # Файл уже открыт
                        return [sheet.name for sheet in book.sheets]
            
            # Если файл не открыт, открываем его временно
            temp_app = xw.App(visible=False)
            try:
                wb = temp_app.books.open(self.halfs_file)
                sheets = [sheet.name for sheet in wb.sheets]
                wb.close()
                return sheets
            finally:
                temp_app.quit()
        except Exception as e:
            logging.error(f"Ошибка при получении листов: {str(e)}")
            return []
    
    def process_files(self, progress_callback, completed_callback, error_callback):
        """Основная функция обработки файлов и переноса данных"""
        if not self.halfs_file or not self.cyber_files or not self.available:
            error_callback("Не выбраны файлы или библиотека xlwings недоступна")
            return
        
        try:
            # Функция для безопасного получения приложения Excel
            # Функция для безопасного получения приложения Excel
            def get_excel_app():
                try:
                    # Пробуем получить список приложений
                    if xw.apps and len(xw.apps) > 0:
                        app = xw.apps[0]
                        
                        # Закрываем пустые книги (Book1, Книга1 и т.д.)
                        try:
                            for book in app.books:
                                if book.name in ['Book1', 'Книга1', 'Book1.xlsx', 'Книга1.xlsx'] and not book.saved:
                                    # Проверяем, что книга действительно пустая
                                    if len(book.sheets) == 1 and book.sheets[0].used_range.last_cell.row == 1:
                                        book.close()
                        except:
                            pass
                        
                        return app
                    else:
                        # Если Excel не запущен, создаем его невидимым
                        return xw.App(visible=False, add_book=False)
                except:
                    return xw.App(visible=False, add_book=False)
            
            # Функция для безопасного получения книги
            def get_workbook_safe(file_path, app=None):
                file_name = os.path.basename(file_path)
                
                # Сначала пробуем найти среди открытых книг
                try:
                    for check_app in xw.apps:
                        try:
                            for book in check_app.books:
                                try:
                                    # Проверяем по имени файла
                                    if book.name.lower() == file_name.lower():
                                        return book, True  # Книга, уже_открыта
                                except:
                                    pass
                                
                                try:
                                    # Проверяем по полному пути
                                    if book.fullname.lower() == file_path.lower():
                                        return book, True
                                except:
                                    pass
                        except:
                            continue
                except:
                    pass
                
                # Если не нашли, пробуем открыть
                if app is None:
                    app = get_excel_app()
                
                try:
                    return app.books.open(file_path), False  # Книга, не_была_открыта
                except:
                    # Если не удалось открыть, возвращаем None
                    return None, False
            
            # Получаем приложение Excel
            app = get_excel_app()
           
            try:
                # Подключаемся к файлу половины
                progress_callback(f"Подключение к файлу: {os.path.basename(self.halfs_file)}")
                halfs_wb, halfs_was_open = get_workbook_safe(self.halfs_file, app)
                
                if not halfs_wb:
                    error_callback(f"Не удалось открыть файл половины: {self.halfs_file}")
                    return
                
                progress_callback(f"✓ Подключились к файлу: {os.path.basename(self.halfs_file)}")
                
                # Получаем список листов
                try:
                    halfs_sheets = [sheet.name for sheet in halfs_wb.sheets]
                except:
                    # Если не удалось получить листы, пробуем переподключиться
                    app = get_excel_app()
                    halfs_wb, _ = get_workbook_safe(self.halfs_file, app)
                    if not halfs_wb:
                        error_callback("Потеряна связь с файлом половины")
                        return
                    halfs_sheets = [sheet.name for sheet in halfs_wb.sheets]
                
                total_matches = 0
                processed_matches = 0
                problem_files = []
                successfully_processed = []
                
                # Обрабатываем каждый файл Cyber
                for cyber_file in self.cyber_files:
                    cyber_file_name = os.path.basename(cyber_file)
                    progress_callback(f"Обработка: {cyber_file_name}")
                    
                    try:
                        # Получаем книгу Cyber
                        cyber_wb, cyber_was_open = get_workbook_safe(cyber_file, app)
                        
                        if not cyber_wb:
                            progress_callback(f"⚠️ Не удалось открыть: {cyber_file_name}")
                            problem_files.append(cyber_file_name)
                            continue
                        
                        try:
                            # Проверяем наличие листа Line
                            sheet_names = []
                            try:
                                for sheet in cyber_wb.sheets:
                                    sheet_names.append(sheet.name)
                            except:
                                # Переподключаемся если потеряли связь
                                app = get_excel_app()
                                cyber_wb, _ = get_workbook_safe(cyber_file, app)
                                if cyber_wb:
                                    for sheet in cyber_wb.sheets:
                                        sheet_names.append(sheet.name)
                            
                            if "Line" not in sheet_names:
                                progress_callback(f"⚠️ Нет листа Line в: {cyber_file_name}")
                                continue
                            
                            line_sheet = cyber_wb.sheets["Line"]
                            
                            # Определяем диапазон данных
                            try:
                                last_row = min(line_sheet.used_range.last_cell.row, 200)  # Ограничиваем 200 строками
                            except:
                                last_row = 100
                            
                            # Собираем матчи
                            matches = []
                            consecutive_errors = 0
                            
                            for row in range(5, last_row + 1):
                                if row % 16 == 0:  # Пропускаем разделители
                                    continue
                                
                                if consecutive_errors > 5:  # Прекращаем после 5 ошибок подряд
                                    break
                                
                                try:
                                    # Читаем данные
                                    tournament = line_sheet.range(f"C{row}").value
                                    team1 = line_sheet.range(f"D{row}").value
                                    team2 = line_sheet.range(f"E{row}").value
                                    total = line_sheet.range(f"G{row}").value
                                    
                                    if tournament and team1 and team2 and total:
                                        matches.append({
                                            "tournament": str(tournament),
                                            "team1": str(team1),
                                            "team2": str(team2),
                                            "total": float(total) if total else 0
                                        })
                                        consecutive_errors = 0
                                    
                                except Exception as e:
                                    consecutive_errors += 1
                                    if consecutive_errors == 1:  # Логируем только первую ошибку
                                        logging.warning(f"Ошибка чтения строки {row}: {str(e)}")
                            
                            if not matches:
                                progress_callback(f"⚠️ Нет данных в: {cyber_file_name}")
                                continue
                            
                            total_matches += len(matches)
                            progress_callback(f"✓ Найдено {len(matches)} матчей в: {cyber_file_name}")
                            
                            # Переносим данные
                            tournament_insert_rows = {}
                            cleared_tournaments = set()
                            matches_from_this_file = 0
                            
                            for match in matches:
                                tournament = match["tournament"]
                                
                                if tournament not in halfs_sheets:
                                    continue
                                
                                # Получаем лист турнира
                                try:
                                    tournament_sheet = halfs_wb.sheets[tournament]
                                except:
                                    # Переподключаемся если потеряли связь
                                    app = get_excel_app()
                                    halfs_wb, _ = get_workbook_safe(self.halfs_file, app)
                                    if not halfs_wb:
                                        error_callback("Потеряна связь с файлом половины")
                                        return
                                    tournament_sheet = halfs_wb.sheets[tournament]
                                
                                # Находим позицию для вставки
                                if tournament not in tournament_insert_rows:
                                    found = False
                                    for check_row in range(1, 200):
                                        try:
                                            if tournament_sheet.range(f"AD{check_row}").value == "Team 1":
                                                tournament_insert_rows[tournament] = check_row + 1
                                                found = True
                                                break
                                        except:
                                            pass
                                    
                                    if not found:
                                        continue
                                
                                # Вставляем данные
                                insert_row = tournament_insert_rows[tournament]

                                # --- очистка таблицы (один раз на турнир) ---
                                if tournament not in cleared_tournaments:
                                    # если под "Team 1" уже что-то есть — чистим вниз до пустой строки
                                    v1 = tournament_sheet.range(f"AD{insert_row}").value
                                    v2 = tournament_sheet.range(f"AE{insert_row}").value
                                    v3 = tournament_sheet.range(f"AF{insert_row}").value
                                    if v1 or v2 or v3:
                                        r = insert_row
                                        while True:
                                            a = tournament_sheet.range(f"AD{r}").value
                                            b = tournament_sheet.range(f"AE{r}").value
                                            c = tournament_sheet.range(f"AF{r}").value
                                            if not (a or b or c):
                                                break
                                            tournament_sheet.range(f"AD{r}:AF{r}").value = None
                                            r += 1
                                    cleared_tournaments.add(tournament)
                                # --- конец очистки ---
                                
                                try:
                                    tournament_sheet.range(f"AD{insert_row}").value = match["team1"]
                                    tournament_sheet.range(f"AE{insert_row}").value = match["team2"]
                                    tournament_sheet.range(f"AF{insert_row}").value = match["total"]
                                    
                                    tournament_insert_rows[tournament] = insert_row + 1
                                    processed_matches += 1
                                    matches_from_this_file += 1
                                    
                                    # Обновляем прогресс
                                    if processed_matches % 5 == 0:
                                        progress_value = int((processed_matches / total_matches) * 100) if total_matches > 0 else 0
                                        progress_callback(f"Прогресс: {processed_matches}/{total_matches} ({progress_value}%)", progress_value)
                                        
                                except Exception as e:
                                    logging.error(f"Ошибка вставки: {str(e)}")
                            
                            if matches_from_this_file > 0:
                                successfully_processed.append(f"{cyber_file_name} ({matches_from_this_file} матчей)")
                        
                        finally:
                            # Закрываем файл если мы его открывали
                            if not cyber_was_open and cyber_wb:
                                try:
                                    cyber_wb.close()
                                except:
                                    pass
                    
                    except Exception as e:
                        logging.error(f"Ошибка обработки {cyber_file_name}: {str(e)}")
                        problem_files.append(f"{cyber_file_name}: {str(e)[:50]}")
                
                # Сохраняем изменения
                progress_callback("Сохранение изменений...")
                try:
                    halfs_wb.save()
                    progress_callback("✓ Изменения сохранены")
                except Exception as e:
                    progress_callback(f"⚠️ Ошибка сохранения: {str(e)}")
                
                # Закрываем файл половины если мы его открывали
                if not halfs_was_open and halfs_wb:
                    try:
                        halfs_wb.close()
                    except:
                        pass
                
                # Формируем итоговый отчет
                result_msg = f"✅ Обработка завершена\n\n"
                result_msg += f"📊 Статистика:\n"
                result_msg += f"• Обработано матчей: {processed_matches}\n"
                result_msg += f"• Успешно обработано файлов: {len(successfully_processed)}\n"
                
                if successfully_processed:
                    result_msg += f"\n✓ Успешные файлы:\n"
                    for sf in successfully_processed:
                        result_msg += f"  • {sf}\n"
                
                if problem_files:
                    result_msg += f"\n⚠️ Проблемные файлы ({len(problem_files)}):\n"
                    for pf in problem_files:
                        result_msg += f"  • {pf}\n"
                
                completed_callback(result_msg)
            
            finally:
                pass
                
        except Exception as e:
            error_msg = f"Критическая ошибка: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            error_callback(error_msg)


# Класс потока для обработки файлов Excel
class ExcelProcessorThread(QThread):
    progress_signal = pyqtSignal(str, int)
    finished_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)
    
    def __init__(self, excel_handler):
        super().__init__()
        self.excel_handler = excel_handler
    
    def run(self):
        try:
            self.excel_handler.process_files(
                progress_callback=self.update_progress,
                completed_callback=self.processing_finished,
                error_callback=self.processing_error
            )
        except Exception as e:
            self.error_signal.emit(f"Критическая ошибка: {str(e)}")
    
    def update_progress(self, message, progress=None):
        if progress is not None:
            self.progress_signal.emit(message, progress)
        else:
            self.progress_signal.emit(message, -1)  # -1 означает не обновлять прогресс-бар
    
    def processing_finished(self, message):
        self.finished_signal.emit(message)
    
    def processing_error(self, error_message):
        self.error_signal.emit(error_message)


# Класс потока для обработки Excel в фоне (Для ройки)
class RoykaProcessorThread(QThread):
    """Thread for processing Excel file in background"""
    progress_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(object)
    error_signal = pyqtSignal(str)
    
    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        
    def run(self):
        try:
            self.progress_signal.emit("Открытие файла Excel...")
            
            # Read the Excel file
            self.progress_signal.emit("Чтение данных из листа 'Halfs Champs'...")
            excel_data = pd.read_excel(self.file_path, sheet_name='Halfs Champs')
            
            # Extract tournament names from column A
            self.progress_signal.emit("Извлечение названий турниров...")
            tournaments = excel_data.iloc[1:, 0].dropna().tolist()
            
            # Process the data structure
            self.progress_signal.emit("Анализ структуры данных...")
            
            # Create a dictionary to store all data
            data = {}
            
            # Get column headers to identify difference values
            headers = excel_data.iloc[0].tolist()
            
            # Process each tournament
            for i, tournament in enumerate(tournaments):
                if pd.isna(tournament) or not isinstance(tournament, str):
                    continue
                    
                self.progress_signal.emit(f"Обработка турнира {i+1}/{len(tournaments)}: {tournament}")
                
                try:
                    # Get the row for this tournament
                    row_idx = excel_data[excel_data.iloc[:, 0] == tournament].index[0]
                    row_data = excel_data.iloc[row_idx].tolist()
                    
                    # Store tournament data
                    data[tournament] = {}
                    
                    # Process columns
                    col_idx = 1  # Start from column B (index 1)
                    diff_value = 0.1
                    
                    while col_idx < len(row_data):
                        if col_idx + 8 < len(row_data):  # Ensure we have complete data for this difference
                            # Store data for this difference
                            data[tournament][diff_value] = {
                                'ОБЩЕЕ': {
                                    'кол-во': self._safe_value(row_data[col_idx]),
                                    'WIN': self._safe_value(row_data[col_idx + 1]),
                                    '%': self._safe_value(row_data[col_idx + 2])
                                },
                                'OVER': {
                                    'кол-во': self._safe_value(row_data[col_idx + 3]),
                                    'WIN': self._safe_value(row_data[col_idx + 4]),
                                    '%': self._safe_value(row_data[col_idx + 5])
                                },
                                'UNDER': {
                                    'кол-во': self._safe_value(row_data[col_idx + 6]),
                                    'WIN': self._safe_value(row_data[col_idx + 7]),
                                    '%': self._safe_value(row_data[col_idx + 8])
                                }
                            }
                            
                            # Move to next difference
                            col_idx += 9
                            if diff_value == 0.1:
                                diff_value = 0.5
                            else:
                                diff_value += 0.5
                        else:
                            break
                except Exception as e:
                    self.progress_signal.emit(f"Предупреждение: не удалось обработать турнир {tournament}: {str(e)}")
            
            self.progress_signal.emit("Обработка данных завершена успешно!")
            self.finished_signal.emit(data)
            
        except Exception as e:
            error_message = f"Ошибка при обработке файла: {str(e)}\n{traceback.format_exc()}"
            self.error_signal.emit(error_message)
            
    def _safe_value(self, value):
        """Convert value to a safe format"""
        if pd.isna(value):
            return 0
        try:
            if isinstance(value, (int, float)):
                return value
            return float(str(value).replace(',', '.'))
        except:
            return 0


class HighlightDelegate(QStyledItemDelegate):
    """Делегат для рисования рамки вокруг выделенной строки"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.highlighted_row = -1
        self.theme = "dark"
    
    def paint(self, painter, option, index):
        """Переопределенный метод для рисования ячеек с рамкой"""
        # Рисуем ячейку стандартным образом
        super().paint(painter, option, index)
        
        # Если это ячейка из выделенной строки, добавляем к ней рамку
        if index.row() == self.highlighted_row:
            painter.save()
            
            # Настройка толщины и цвета границы
            if self.theme == "dark":
                # Для темной темы - контрастная голубая граница
                pen = QPen(QColor(100, 180, 255), 3)  # Светло-голубой, толщина 3px
            else:
                # Для светлой темы - контрастная синяя граница
                pen = QPen(QColor(30, 100, 200), 3)   # Насыщенный синий, толщина 3px
            
            pen.setStyle(Qt.SolidLine)
            painter.setPen(pen)
            
            # Рисуем прямоугольник вокруг ячейки
            rect = option.rect
            # Уменьшаем прямоугольник чтобы границы не перекрывались
            adjusted_rect = rect.adjusted(1, 1, -2, -2)
            painter.drawRect(adjusted_rect)
            
            painter.restore()


class CustomTableWidget(QTableWidget):
    """Расширенная таблица с возможностью подсветки выбранных строк"""
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Отключаем возможность редактирования таблицы
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # Отключаем стандартную систему выделения Qt
        self.setSelectionMode(QTableWidget.NoSelection)  # Полностью отключаем стандартное выделение
        
        # Создаем делегат для рисования рамки
        self.highlight_delegate = HighlightDelegate(self)
        self.setItemDelegate(self.highlight_delegate)
        
        # Текущая выделенная строка
        self.highlighted_row = -1
        
        # Для рисования разделителей между группами столбцов
        self.vertical_separators = [3, 6]  # Индексы столбцов, после которых рисовать разделители

    def resizeEvent(self, event):
        """Обрабатывает изменение размера таблицы"""
        super().resizeEvent(event)
        
        # Получаем общую ширину таблицы
        total_width = self.width()
        
        # Ширина для колонки "Разница"
        raznica_width = 80
        
        # Оставшаяся ширина делится на 3 группы
        group_width = (total_width - raznica_width) // 3
        
        # Устанавливаем ширину колонок
        self.setColumnWidth(0, raznica_width)  # Разница
        
        # Распределяем ширину внутри каждой группы
        for group in range(3):
            base_idx = 1 + group * 3
            self.setColumnWidth(base_idx, int(group_width * 0.25))      # кол-во: 25%
            self.setColumnWidth(base_idx + 1, int(group_width * 0.45))  # WIN: 45%
            self.setColumnWidth(base_idx + 2, int(group_width * 0.30))  # %: 30%    
    
    def mousePressEvent(self, event):
        """Переопределяем обработку клика мыши для подсветки строки"""
        # Получаем индекс строки и передаем его в highlight_row
        clicked_index = self.indexAt(event.pos())
        if clicked_index.isValid():
            row = clicked_index.row()
            self.highlight_row(row)
        
        # Вызываем стандартный обработчик, чтобы сохранить другую функциональность
        super().mousePressEvent(event)
            
    def highlight_row(self, row):
        """Подсвечивает выбранную строку с помощью делегата"""
        # Если кликнули на той же строке, снимаем выделение
        if row == self.highlighted_row:
            self.highlighted_row = -1
            self.highlight_delegate.highlighted_row = -1
        else:
            # Устанавливаем новую выделенную строку
            self.highlighted_row = row
            self.highlight_delegate.highlighted_row = row
            
        # Обновляем отображение всей таблицы
        self.viewport().update()
    
    def clear_highlight(self):
        """Снимает подсветку с выделенной строки"""
        self.highlighted_row = -1
        self.highlight_delegate.highlighted_row = -1
        self.viewport().update()
    
    def update_theme(self, theme):
        """Обновляет тему делегата при смене темы"""
        self.highlight_delegate.theme = theme
    
    def paintEvent(self, event):
        """Переопределяем метод рисования для добавления вертикальных разделителей"""
        super().paintEvent(event)
        
        # Рисуем вертикальные разделители между группами колонок
        painter = QPainter(self.viewport())
        painter.setPen(QPen(QColor("#707070" if self.highlight_delegate.theme == "dark" else "#A0A0A0"), 2, Qt.SolidLine))
        
        header_height = self.horizontalHeader().height()
        
        for col_idx in self.vertical_separators:
            # Получаем позицию правой границы столбца
            x = 0
            for i in range(col_idx + 1):
                x += self.columnWidth(i)
            
            # Рисуем вертикальную линию
            painter.drawLine(x, header_height, x, self.height())


# Компонент боковой навигации
class SidebarNavigation(QListWidget):
    """Боковая панель навигации"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()
    
    def setup_ui(self):
        # Настройка стиля
        self. setStyleSheet("""
            QListWidget {
                border: none;
                background-color: #2a2a2a;
            }
            QListWidget:: item {
                padding: 15px;
                border-bottom:  1px solid #3a3a3a;
                color:  #cccccc;
            }
            QListWidget::item: hover {
                background-color:  #3a3a3a;
            }
            QListWidget::item:selected {
                background-color: #4A90E2;
                color: white;
            }
        """)
        
        
        # Добавляем пункты меню ТОЛЬКО ОДИН РАЗ
        # Оставляем только нужные разделы
        self.add_nav_item("Ройка", "Анализ статистики и отображение данных")
        # Раздел для просмотра базы данных половин
        self.add_nav_item("База половин", "База данных половин")
        # Раздел для статистики, коэффициентов, отклонений и побед/поражений
        self.add_nav_item(
            "Статистика из половин",
            "Статистика/коэффициенты, отклонения, победы/поражения и средние четверти"
        )
        # Раздел для сводной таблицы по турнирам
        self.add_nav_item(
            "Сводная таблица",
            "Сводная статистика по турнирам"
        )
        # Новый раздел – анализ данных первой половины
        self.add_nav_item("Анализ половин", "Анализ матчей по тоталам первой половины")
        # Разделы для кибер‑матчей
        self.add_nav_item("Cybers Bases", "База данных матчей для кибер‑анализа")
        self.add_nav_item("Cyber LIVE", "Прогнозы и темп по текущим матчам")

        # Раздел для сортировки половин (копирование данных между файлами)
        self.add_nav_item("Сортировка половин", "Перенос данных между файлами")
        
        # Устанавливаем первый элемент как активный по умолчанию
        self.setCurrentRow(0)
    
    def add_nav_item(self, title, description=""):
        item = QListWidgetItem(title)
        item.setToolTip(description)
        self.addItem(item)

class DateSortDelegate(QStyledItemDelegate):
    def __lt__(self, left, right):
        def parse_date(date_str):
            try:
                if date_str:
                    day, month, year = map(int, date_str.split('.'))
                    return datetime(year, month, day)
                return datetime.min
            except:
                return datetime.min

        # Преобразуем строки в объекты datetime для прямого сравнения
        left_date = parse_date(left.data(Qt.DisplayRole))
        right_date = parse_date(right.data(Qt.DisplayRole))
        
        # Сравниваем даты напрямую
        return left_date < right_date

class DatabaseViewDialog(QDialog):
    def __init__(self, db_path, parent=None):
        super().__init__(parent)
        self.db_path = db_path
        # Определяем schema по имени файла для PostgreSQL
        self._schema = 'royka'
        if 'halfs' in str(db_path):
            self._schema = 'halfs'
        elif 'cyber' in str(db_path):
            self._schema = 'cyber'
        self.setup_ui()
        self.current_sort_column = -1
        self.sort_order = Qt.AscendingOrder
        self.load_data()

    @contextmanager
    def _connect(self):
        from db_connection import db_connect
        with db_connect(schema=self._schema, sqlite_path=self.db_path) as conn:
            yield conn
        
    def setup_ui(self):
        self.setWindowTitle("Просмотр и удаление данных")
        self.setMinimumSize(800, 600)
        
        layout = QVBoxLayout(self)
        
        # Фильтры
        filter_layout = QHBoxLayout()
        
        # Фильтр по турниру
        self.tournament_combo = QComboBox()
        self.tournament_combo.setEditable(True)
        self.tournament_combo.addItem("Все турниры")
        filter_layout.addWidget(QLabel("Турнир:"))
        filter_layout.addWidget(self.tournament_combo)
        
        # Кнопка применения фильтра
        filter_btn = QPushButton("Применить фильтр")
        filter_btn.clicked.connect(self.apply_filter)
        filter_layout.addWidget(filter_btn)
        
        layout.addLayout(filter_layout)
        
        # Таблица
        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.MultiSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # Включаем сортировку
        self.table.setSortingEnabled(True)
        try:
            # Отдельная роль для сортировки, чтобы не конфликтовать с UserRole (id)
            self.table.model().setSortRole(Qt.UserRole + 1)
        except Exception:
            pass
        try:
            self.table.model().setSortRole(Qt.UserRole)
        except Exception:
            pass
        # Подключаем обработчик клика по заголовку
        self.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        
        layout.addWidget(self.table)
        
        # Кнопки
        button_layout = QHBoxLayout()
        
        delete_btn = QPushButton("Удалить выбранные")
        delete_btn.clicked.connect(self.delete_selected)
        button_layout.addWidget(delete_btn)
        
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)

    def on_header_clicked(self, logical_index):
        """Обработчик клика по заголовку колонки"""
        if self.current_sort_column == logical_index:
            # Меняем порядок сортировки если кликнули по той же колонке
            self.sort_order = Qt.DescendingOrder if self.sort_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            # Новая колонка - всегда начинаем с сортировки по возрастанию
            self.sort_order = Qt.AscendingOrder
            self.current_sort_column = logical_index
        
        # Отключаем сортировку на время обновления
        self.table.setSortingEnabled(False)
        # Применяем сортировку
        self.table.sortItems(logical_index, self.sort_order)
        # Включаем сортировку обратно
        self.table.setSortingEnabled(True)
    
    def load_data(self, tournament_filter=None):
        try:
            with self._connect() as conn:
                cursor = conn.cursor()
                
                # Получаем список турниров для комбобокса
                cursor.execute("SELECT DISTINCT tournament FROM matches ORDER BY tournament")
                tournaments = [row[0] for row in cursor.fetchall()]
                
                # Обновляем комбобокс, сохраняя текущий выбор
                current = self.tournament_combo.currentText()
                self.tournament_combo.clear()
                self.tournament_combo.addItem("Все турниры")
                self.tournament_combo.addItems(tournaments)
                if current in tournaments:
                    self.tournament_combo.setCurrentText(current)
                
                # Получаем данные с учетом фильтра
                query = """
                    SELECT id, date, tournament, team_home, team_away, 
                           t1h, t2h, tim, deviation, kickoff, predict, result
                    FROM matches
                """
                params = []
                
                if tournament_filter and tournament_filter != "Все турниры":
                    query += " WHERE tournament = ?"
                    params.append(tournament_filter)
                
                # Базовая сортировка по дате и турниру (корректная для dd.mm.yyyy)
                query += """
                    ORDER BY
                        CASE
                            WHEN instr(date, '-') > 0 THEN date
                            ELSE substr(date, 7, 4) || '-' || substr(date, 4, 2) || '-' || substr(date, 1, 2)
                        END ASC,
                        tournament
                """
                cursor.execute(query, params)
                data = cursor.fetchall()
                
                # Настраиваем таблицу
                headers = ["ID", "Дата", "Турнир", "Команда 1", "Команда 2", 
                          "T1H", "T2H", "TIM", "Deviation", "KickOff", "Predict", "Result"]
                
                self.table.setRowCount(len(data))
                self.table.setColumnCount(len(headers))
                self.table.setHorizontalHeaderLabels(headers)
                
                # Заполняем данные
                for row, record in enumerate(data):
                    for col, value in enumerate(record):
                        item = QTableWidgetItem()
                        
                        # Устанавливаем правильный тип данных для сортировки
                        if col == 0:  # ID
                            item.setData(Qt.DisplayRole, int(value) if value is not None else 0)
                        elif col == 1:  # Дата
                            item = QTableWidgetItem()
                            date_str = str(value) if value is not None else ""
                            item.setData(Qt.DisplayRole, date_str)
                            try:
                                if date_str:
                                    if "-" in date_str:
                                        d = datetime.strptime(date_str, "%Y-%m-%d").date()
                                    else:
                                        d = datetime.strptime(date_str, "%d.%m.%Y").date()
                                    item.setData(Qt.UserRole, QDate(d.year, d.month, d.day))
                            except Exception:
                                pass
                        elif col in [5, 6, 7, 8, 9, 11]:  # Числовые колонки
                            item.setData(Qt.DisplayRole, float(value) if value is not None else 0.0)
                        else:  # Текстовые колонки
                            item.setData(Qt.DisplayRole, str(value) if value is not None else "")
                            
                        self.table.setItem(row, col, item)
                
                # Устанавливаем делегат для колонки с датой
                date_delegate = DateSortDelegate()
                self.table.setItemDelegateForColumn(1, date_delegate)
                
                # Подгоняем размер колонок
                self.table.resizeColumnsToContents()
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки данных: {str(e)}")

    def apply_filter(self):
        """Применяет фильтр по турниру"""
        tournament = self.tournament_combo.currentText()
        self.load_data(tournament if tournament != "Все турниры" else None)

    def delete_selected(self):
        """Удаление выбранных записей"""
        selected_rows = self.table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "Предупреждение", "Выберите записи для удаления")
            return
        
        reply = QMessageBox.question(
            self,
            'Подтверждение',
            f'Вы действительно хотите удалить выбранные записи ({len(selected_rows)} шт.)?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                # Получаем уникальные ID выбранных записей
                selected_ids = set()
                for item in selected_rows:
                    row = item.row()
                    id_item = self.table.item(row, 0)
                    if id_item:
                        selected_ids.add(int(id_item.text()))
                
                # Удаляем записи
                with self._connect() as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        f"DELETE FROM matches WHERE id IN ({','.join(['?']*len(selected_ids))})",
                        list(selected_ids)
                    )
                    conn.commit()
                
                # Перезагружаем данные
                self.load_data(self.tournament_combo.currentText())
                
                QMessageBox.information(
                    self,
                    "Успех",
                    f"Удалено записей: {len(selected_ids)}"
                )
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении: {str(e)}")     

# Компонент страницы Ройка
class RoykaPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = RoykaDatabase()
        self.debug_log = None
        
        # Проверяем подключение к базе данных
        try:
            with self.db._connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM matches")
                count = cursor.fetchone()[0]
                print(f"Successfully connected to database. Found {count} records.")
        except Exception as e:
            print(f"Database connection error: {str(e)}")
            QMessageBox.critical(self, "Ошибка", 
                "Не удалось подключиться к базе данных. Проверьте файл базы данных.")
        
        self.setup_ui()  # Создаст self.table

    def setup_ui(self):
        """Настройка пользовательского интерфейса"""
        # Основной layout
        main_layout = QVBoxLayout(self)
        
        # Создаем вкладки
        self.tabs = QTabWidget()
        
        # Вкладка управления данными (первая)
        self.tab_data = QWidget()
        self.setup_data_management_tab()
        self.tabs.addTab(self.tab_data, "Управление данными")
        
        # Вкладка статистики (вторая)
        self.tab_stats = QWidget()
        self.setup_stats_tab()
        self.tabs.addTab(self.tab_stats, "Статистика")

        # Вкладка статистики по диапазонам (третья)
        # Здесь будем показывать ROI для диапазонов разницы, например 0.1-0.5, 0.5-1 и т.д.
        self.tab_range_stats = QWidget()
        self.setup_range_stats_tab()
        self.tabs.addTab(self.tab_range_stats, "Статистика по диапазонам")

        # Вкладка статистики T2H + Div >= 4.5 (четвертая) ← ДОБАВЛЯЕМ ЗДЕСЬ
        self.tab_half_stats = QWidget()
        self.setup_half_stats_tab()
        self.tabs.addTab(self.tab_half_stats, "Статистика 4.5+")

        # Вкладка статистики T2H + Div >= 4.5 CHANGE
        self.tab_half_stats_change = QWidget()
        self.setup_half_stats_change_tab()
        self.tabs.addTab(self.tab_half_stats_change, "Статистика 4.5+ CHANGE")
        
        # Добавляем вкладки в основной layout
        main_layout.addWidget(self.tabs)
        
        # Создаем область для логов
        self.debug_log = QPlainTextEdit()
        self.debug_log.setMaximumHeight(200)
        self.debug_log.setReadOnly(True)
        # Не задаём жёстко тёмный фон для области логов, чтобы она
        # наследовала текущую цветовую схему (тёмную или светлую).
        # По умолчанию QPlainTextEdit будет окрашен в цвета темы.
        main_layout.addWidget(self.debug_log)

        # Создаем labels для информации о базе данных
        stats_layout = QHBoxLayout()
        
        self.db_stats_label = QLabel()
        stats_layout.addWidget(self.db_stats_label)
        
        self.last_update_label = QLabel()
        stats_layout.addWidget(self.last_update_label)
        
        main_layout.addLayout(stats_layout)
        
        # Обновляем статистику базы данных
        self.update_database_stats()

    def find_duplicates(self):
        """Находит и отображает дубли матчей в базе"""
        try:
            # Получаем данные из базы данных
            with self.db._connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, date, tournament, team_home, team_away, 
                        t1h, t2h, tim, deviation, kickoff, predict, result
                    FROM matches
                """)
                data = cursor.fetchall()
                
                if not data:
                    QMessageBox.information(self, "Информация", "База данных пуста")
                    return
                
                # Создаем DataFrame из полученных данных
                columns = ["ID", "Дата", "Турнир", "Команда 1", "Команда 2", 
                        "T1H", "T2H", "TIM", "Deviation", "KickOff", "Predict", "Result"]
                df = pd.DataFrame(data, columns=columns)
                
                # Находим дубликаты по всем колонкам кроме ID
                duplicate_mask = df.duplicated(subset=df.columns[1:], keep='first')
                duplicates = df[duplicate_mask]
                
                if duplicates.empty:
                    QMessageBox.information(self, "Информация", "Дублирующиеся записи не найдены")
                    return
                
                # Создаем расширенный диалог для отображения результатов
                dialog = QDialog(self)
                dialog.setWindowTitle("Найденные дубли")
                dialog.setMinimumSize(800, 600)
                
                layout = QVBoxLayout(dialog)
                
                # Добавляем метку с количеством дублей
                count_label = QLabel(f"Найдено {len(duplicates)} дублирующихся записей:")
                layout.addWidget(count_label)
                
                # Текстовое поле для отображения результатов
                text_edit = QPlainTextEdit()
                text_edit.setReadOnly(True)
                
                # Формируем текст с группировкой по дубликатам
                msg = ""
                duplicate_ids = []  # Список ID для удаления
                
                # Группируем все данные по ключевым полям для выявления групп дубликатов
                key_columns = ['Дата', 'Турнир', 'Команда 1', 'Команда 2', 'TIM', 'Result']
                for _, group in df.groupby(key_columns):
                    if len(group) > 1:  # Если есть дубликаты
                        msg += "-" * 50 + "\n"
                        # Первая запись в группе - оригинал
                        original = group.iloc[0]
                        msg += "ОРИГИНАЛ:\n"
                        msg += (f"Дата: {original['Дата']}, Турнир: {original['Турнир']}\n"
                            f"Матч: {original['Команда 1']} vs {original['Команда 2']}\n"
                            f"TIM: {original['TIM']}, Result: {original['Result']}\n\n")
                        
                        msg += "ДУБЛИ (будут удалены):\n"
                        # Остальные записи - дубли
                        for idx, row in group.iloc[1:].iterrows():
                            duplicate_ids.append(row['ID'])
                            msg += (f"Дата: {row['Дата']}, Турнир: {row['Турнир']}\n"
                                f"Матч: {row['Команда 1']} vs {row['Команда 2']}\n"
                                f"TIM: {row['TIM']}, Result: {row['Result']}\n")
                        msg += "\n"
                
                text_edit.setPlainText(msg)
                layout.addWidget(text_edit)
                
                # Создаем горизонтальный layout для кнопок
                button_layout = QHBoxLayout()
                
                # Кнопка удаления дублей
                delete_btn = QPushButton(f"Удалить дубли ({len(duplicate_ids)} шт.)")
                delete_btn.clicked.connect(lambda: self.delete_duplicates(duplicate_ids, dialog))
                button_layout.addWidget(delete_btn)
                
                # Кнопка закрытия
                close_btn = QPushButton("Закрыть")
                close_btn.clicked.connect(dialog.accept)
                button_layout.addWidget(close_btn)
                
                layout.addLayout(button_layout)
                
                dialog.exec_()
                
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при поиске дублей: {str(e)}\n\n{traceback.format_exc()}")

    def delete_duplicates(self, duplicate_ids, dialog):
        """Удаляет дубли из базы данных"""
        try:
            if not duplicate_ids:
                return
                
            reply = QMessageBox.question(
                dialog,
                'Подтверждение',
                f'Вы действительно хотите удалить {len(duplicate_ids)} дублирующихся записей?\n'
                'Для каждой группы дублей будет оставлена одна оригинальная запись.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                with self.db._connect() as conn:
                    cursor = conn.cursor()
                    # Используем параметризованный запрос для безопасного удаления
                    cursor.execute(
                        f"DELETE FROM matches WHERE id IN ({','.join(['?']*len(duplicate_ids))})",
                        duplicate_ids
                    )
                    deleted_count = cursor.rowcount
                    conn.commit()
                
                QMessageBox.information(
                    dialog,
                    "Успех",
                    f"Удалено {deleted_count} дублирующихся записей.\n"
                    "Оригинальные записи сохранены."
                )
                
                # Обновляем статистику базы данных
                self.update_database_stats()
                
                # Закрываем диалог
                dialog.accept()
                
        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка", f"Ошибка при удалении дублей: {str(e)}")

    def add_delete_key_binding(self):
        """Добавляет обработку клавиши Delete"""
        self.table.keyPressEvent = self.handle_key_press  # Используем self.table вместо self.tree

    def handle_key_press(self, event):
        """Обработчик нажатий клавиш"""
        if event.key() == Qt.Key_Delete:
            self.delete_selected()
        else:
            # Вызываем стандартный обработчик для остальных клавиш
            QTableWidget.keyPressEvent(self.table, event)

    def setup_data_management_tab(self):
        """Настройка вкладки управления данными"""
        layout = QVBoxLayout()
        
        # Блок импорта данных
        import_group = QGroupBox("Импорт данных")
        import_layout = QVBoxLayout()
        
        # Кнопка для вставки из буфера
        paste_btn = QPushButton("Вставить данные из буфера")
        paste_btn.setMinimumHeight(40)
        paste_btn.clicked.connect(self.show_paste_dialog)
        
        # Кнопка для импорта из файла
        import_btn = QPushButton("Импорт из файла Excel")
        import_btn.setMinimumHeight(40)
        import_btn.clicked.connect(self.import_from_excel)
        
        import_layout.addWidget(paste_btn)
        import_layout.addWidget(import_btn)
        import_group.setLayout(import_layout)
        
        # Блок управления базой
        manage_group = QGroupBox("Управление базой")
        manage_layout = QVBoxLayout()
        
        # Кнопка поиска дублей
        self.find_duplicates_button = QPushButton("Найти дубли")
        self.find_duplicates_button.clicked.connect(self.find_duplicates)
        
        delete_btn = QPushButton("Редактировать базу данных")
        delete_btn.clicked.connect(self.show_delete_dialog)
        
        backup_btn = QPushButton("Создать резервную копию")
        backup_btn.clicked.connect(self.create_backup)
        
        restore_btn = QPushButton("Восстановить из копии")
        restore_btn.clicked.connect(self.restore_from_backup)
        
        manage_layout.addWidget(self.find_duplicates_button)
        manage_layout.addWidget(delete_btn)
        manage_layout.addWidget(backup_btn)
        manage_layout.addWidget(restore_btn)
        manage_group.setLayout(manage_layout)
        
        # Добавляем все блоки в layout вкладки
        layout.addWidget(import_group)
        layout.addWidget(manage_group)
        layout.addStretch()

        # Добавляем debug_log только на вкладку управления данными
        self.debug_log = QPlainTextEdit()
        self.debug_log.setMaximumHeight(200)
        self.debug_log.setReadOnly(True)
        self.debug_log.setStyleSheet("""
            QPlainTextEdit {
                background-color: #1a1a1f;
                color: #00ff00;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 12px;
                border: 1px solid #333339;
                padding: 5px;
            }
        """)
        layout.addWidget(self.debug_log)
        
        self.tab_data.setLayout(layout)

    def setup_stats_tab(self):
        """Настройка вкладки статистики"""
        layout = QVBoxLayout()
        
        # Блок поиска
        search_layout = QHBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите название турнира...")
        self.search_input.setMinimumWidth(400)
        self.search_input.returnPressed.connect(self.search_tournament)
        
        self.search_btn = QPushButton("Поиск")
        self.search_btn.clicked.connect(self.search_tournament)
        
        search_layout.addWidget(QLabel("Поиск турнира:"))
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_btn)
        
        layout.addLayout(search_layout)
        
        # Создаем контейнер для таблицы
        table_container = QWidget()
        self.table_layout = QVBoxLayout(table_container)
        self.table_layout.setContentsMargins(10, 0, 10, 10)
        
        # Создаем таблицу
        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.MultiSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_layout.addWidget(self.table)
        
        # Добавляем контейнер с таблицей в основной layout
        layout.addWidget(table_container)
        
        # Добавляем обработку клавиши Delete
        self.add_delete_key_binding()
        
        self.tab_stats.setLayout(layout)

    def setup_range_stats_tab(self):
        """
        Настройка вкладки "Статистика по диапазонам".

        Эта вкладка похожа на обычную статистику, но отображает ROI для
        диапазонов абсолютного значения исходной разницы: 0.1-0.5, 0.5-1.0,
        1.0-1.5 и так далее до 9.5-10.0. Исключаем значения, равные верхней
        границе (например, 0.5 не попадает в диапазон 0.1-0.5, а входит в
        следующий диапазон 0.5-1.0).
        """
        layout = QVBoxLayout()

        # Блок поиска турнира
        search_layout = QHBoxLayout()

        self.search_input_range = QLineEdit()
        self.search_input_range.setPlaceholderText("Введите название турнира...")
        self.search_input_range.setMinimumWidth(400)
        # При нажатии Enter выполняем поиск
        self.search_input_range.returnPressed.connect(self.search_tournament_range)

        # Используем тот же completer, что и для основной статистики
        try:
            self.search_input_range.setCompleter(self.tournament_completer)
        except Exception:
            pass

        self.search_btn_range = QPushButton("Поиск")
        self.search_btn_range.clicked.connect(self.search_tournament_range)

        search_layout.addWidget(QLabel("Поиск турнира:"))
        search_layout.addWidget(self.search_input_range)
        search_layout.addWidget(self.search_btn_range)
        layout.addLayout(search_layout)

        # Контейнер для таблицы
        table_container = QWidget()
        self.table_layout_range = QVBoxLayout(table_container)
        self.table_layout_range.setContentsMargins(10, 0, 10, 10)

        # Добавляем контейнер с таблицей в основной layout
        layout.addWidget(table_container)

        # Сохраняем ссылку на виджет вкладки
        self.tab_range_stats.setLayout(layout)

        # Переменные для хранения состояния
        self.current_table_range = None
        self.last_tournament_name_range = None
        self.last_stats_range = None

    def setup_half_stats_tab(self):
        """Настройка вкладки статистики по половине (T2H + Div >= 4. 5 / <= -4.5)"""
        layout = QVBoxLayout()
        
        # Блок поиска
        search_layout = QHBoxLayout()
        
        self.search_input_half = QLineEdit()
        self.search_input_half.setPlaceholderText("Введите название турнира...")
        self.search_input_half.setMinimumWidth(400)
        self.search_input_half.returnPressed.connect(self.search_tournament_half)
        
        self.search_btn_half = QPushButton("Поиск")
        self.search_btn_half.clicked.connect(self.search_tournament_half)
        
        search_layout.addWidget(QLabel("Поиск турнира:"))
        search_layout.addWidget(self.search_input_half)
        search_layout.addWidget(self. search_btn_half)
        
        layout.addLayout(search_layout)
        
        # Кнопка показа статистики по всем турнирам
        all_tournaments_btn = QPushButton("Статистика по всем турнирам")
        all_tournaments_btn.clicked.connect(self. show_all_tournaments_half_stats)
        layout.addWidget(all_tournaments_btn)
        
        # Контейнер для таблицы
        table_container = QWidget()
        self.table_layout_half = QVBoxLayout(table_container)
        self.table_layout_half.setContentsMargins(10, 0, 10, 10)
        
        layout.addWidget(table_container)
        
        # Добавляем debug_log
        self.debug_log_half = QPlainTextEdit()
        self.debug_log_half.setMaximumHeight(200)
        self.debug_log_half.setReadOnly(True)
        self.debug_log_half.setStyleSheet("""
            QPlainTextEdit {
                background-color: #1a1a1f;
                color: #00ff00;
                font-family:  Consolas, 'Courier New', monospace;
                font-size: 12px;
                border: 1px solid #333339;
                padding: 5px;
            }
        """)
        layout.addWidget(self.debug_log_half)
        
        self.tab_half_stats. setLayout(layout)

    def setup_half_stats_change_tab(self):
        """Настройка вкладки статистики по половине (CHANGE, 4.5+)"""
        layout = QVBoxLayout()

        search_layout = QHBoxLayout()
        self.search_input_half_change = QLineEdit()
        self.search_input_half_change.setPlaceholderText("Введите название турнира...")
        self.search_input_half_change.setMinimumWidth(400)
        self.search_input_half_change.returnPressed.connect(self.search_tournament_half_change)

        self.search_btn_half_change = QPushButton("Поиск")
        self.search_btn_half_change.clicked.connect(self.search_tournament_half_change)

        search_layout.addWidget(QLabel("Поиск турнира:"))
        search_layout.addWidget(self.search_input_half_change)
        search_layout.addWidget(self.search_btn_half_change)
        layout.addLayout(search_layout)

        all_tournaments_btn = QPushButton("Статистика по всем турнирам")
        all_tournaments_btn.clicked.connect(self.show_all_tournaments_half_stats_change)
        layout.addWidget(all_tournaments_btn)

        table_container = QWidget()
        self.table_layout_half_change = QVBoxLayout(table_container)
        self.table_layout_half_change.setContentsMargins(10, 0, 10, 10)
        layout.addWidget(table_container)

        self.debug_log_half_change = QPlainTextEdit()
        self.debug_log_half_change.setMaximumHeight(200)
        self.debug_log_half_change.setReadOnly(True)
        self.debug_log_half_change.setStyleSheet("""
            QPlainTextEdit {
                background-color: #1a1a1f;
                color: #00ff00;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 12px;
                border: 1px solid #333339;
                padding: 5px;
            }
        """)
        layout.addWidget(self.debug_log_half_change)

        self.tab_half_stats_change.setLayout(layout)
    
    def search_tournament_range(self):
        """
        Поиск турнира для вкладки "Статистика по диапазонам".

        Этот метод запрашивает у базы данных матчи указанного турнира,
        анализирует их по диапазонам разницы и отображает результат в таблице.
        """
        tournament_name = self.search_input_range.text().strip()
        if not tournament_name:
            QMessageBox.warning(self, "Предупреждение", "Введите название турнира")
            return

        # Очищаем лог и добавляем начальное сообщение
        if hasattr(self, 'debug_log') and self.debug_log is not None:
            try:
                self.debug_log.clear()
                self.add_debug_log(f"=== НАЧАЛО АНАЛИЗА ТУРНИРА (диапазоны): {tournament_name} ===")
            except Exception:
                pass

        try:
            # Проверяем наличие данных в базе
            stats = self.db.get_statistics()
            total_records = stats.get('total_records', 0)
            self.add_debug_log(f"Всего записей в базе: {total_records}")
            if total_records == 0:
                self.add_debug_log("ОШИБКА: База данных пуста")
                QMessageBox.warning(self, "Нет данных", "База данных пуста. Сначала импортируйте данные через вкладку 'Управление данными'")
                return

            # Получаем данные турнира
            with self.db._connect() as conn:
                cursor = conn.cursor()
                self.add_debug_log("Выполнение запроса к базе данных...")
                cursor.execute(
                    """
                        SELECT 
                            date, team_home, team_away, t1h, t2h, tim,
                            deviation, kickoff, predict, result
                        FROM matches 
                        WHERE tournament = ?
                        ORDER BY date
                    """,
                    (tournament_name,)
                )
                matches = cursor.fetchall()
                self.add_debug_log(f"Найдено матчей: {len(matches)}")
                if not matches:
                    self.add_debug_log("ОШИБКА: Турнир не найден в базе")
                    QMessageBox.warning(self, "Не найдено", f"Турнир '{tournament_name}' не найден в базе данных")
                    return

                # Преобразуем данные в структуру, аналогичную основной статистике
                tournament_data = {
                    'matches': [
                        {
                            'date': row[0],
                            'team1': row[1],
                            'team2': row[2],
                            't1h': row[3],
                            't2h': row[4],
                            'tim': row[5],
                            'deviation': row[6],
                            'kickoff': row[7],
                            'predict': row[8],
                            'result': row[9]
                        }
                        for row in matches
                    ]
                }

                # Отображаем статистику по диапазонам
                self.display_range_stats(tournament_name, tournament_data)

        except Exception as e:
            error_msg = f"Ошибка при поиске турнира: {str(e)}\n{traceback.format_exc()}"
            self.add_debug_log(f"ОШИБКА: {error_msg}")
            QMessageBox.critical(self, "Ошибка", str(e))

    def display_range_stats(self, tournament_name, tournament_data):
        """
        Отображает статистику турнира по диапазонам разницы.

        Таблица показывает значения ROI для каждого диапазона: общее количество ставок,
        суммарный выигрыш и процент доходности отдельно для OVER и UNDER, а также общий
        показатель.
        """
        try:
            # Очищаем предыдущий контент
            self.clear_layout(self.table_layout_range)

            # Заголовок
            title_label = QLabel(f"Статистика по диапазонам для турнира: {tournament_name}")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setAlignment(Qt.AlignCenter)
            self.table_layout_range.addWidget(title_label)

            # Подсказка про прокрутку
            scroll_hint = QLabel("* Используйте прокрутку, чтобы увидеть все диапазоны до 10")
            scroll_hint.setAlignment(Qt.AlignCenter)
            scroll_hint.setStyleSheet("color: #808080; font-size: 12px;")
            self.table_layout_range.addWidget(scroll_hint)

            # Анализируем данные по диапазонам
            stats = self.analyze_tournament_data_ranges(tournament_data)

            # Определение стилей
            header_bg = "#11244A"
            neutral_color = QColor(180, 180, 180)

            # Ширины колонок
            range_width = 100  # колонка диапазона
            kolvo_width = 80
            win_width = 80
            percent_width = 80

            # Создаем контейнер для заголовков
            header_container = QWidget()
            header_container.setFixedHeight(80)
            header_layout = QVBoxLayout(header_container)
            header_layout.setContentsMargins(0, 0, 0, 0)
            header_layout.setSpacing(0)

            # Верхний ряд заголовков (ОБЩЕЕ, OVER, UNDER)
            top_header = QWidget()
            top_header_layout = QHBoxLayout(top_header)
            top_header_layout.setContentsMargins(range_width, 0, 0, 0)
            top_header_layout.setSpacing(0)

            headers = [
                ("ОБЩЕЕ", "#CCCCCC"),
                ("OVER ↑", "#3A9B3A"),
                ("UNDER ↓", "#B33333")
            ]
            for text, color in headers:
                group_width = kolvo_width + win_width + percent_width
                label = QLabel(text)
                label.setFixedWidth(group_width)
                label.setAlignment(Qt.AlignCenter)
                label.setFont(QFont("Arial", 14, QFont.Bold))
                # Подгоняем положение заголовков относительно подзаголовков
                if text == "ОБЩЕЕ":
                    margin_left = -45
                elif text == "OVER ↑":
                    margin_left = -35
                elif text == "UNDER ↓":
                    margin_left = -25
                label.setStyleSheet(f"""
                    color: {color};
                    background-color: {header_bg};
                    border: none;
                    padding: 4px;
                    margin-left: {margin_left}px;
                """)
                top_header_layout.addWidget(label)

            # Нижний ряд заголовков (кол-во, WIN, %)
            bottom_header = QWidget()
            bottom_header_layout = QHBoxLayout(bottom_header)
            bottom_header_layout.setContentsMargins(range_width, 0, 0, 0)
            bottom_header_layout.setSpacing(0)

            subheaders = ["кол-во", "WIN", "%"]
            for _ in range(3):
                for subheader in subheaders:
                    width = kolvo_width if subheader == "кол-во" else (win_width if subheader == "WIN" else percent_width)
                    label = QLabel(subheader)
                    label.setFixedWidth(width)
                    label.setAlignment(Qt.AlignCenter)
                    label.setFont(QFont("Arial", 12))
                    label.setStyleSheet(f"""
                        color: #999999;
                        background-color: {header_bg};
                        border: none;
                        padding: 4px;
                    """)
                    bottom_header_layout.addWidget(label)

            # Добавляем оба ряда заголовков
            header_layout.addWidget(top_header)
            header_layout.addWidget(bottom_header)
            self.table_layout_range.addWidget(header_container)

            # Создаем основную таблицу
            table = CustomTableWidget()
            table.setFont(QFont("Arial", 14))
            table.setProperty("skipAutoResize", True)
            # Получаем список диапазонов в отсортированном порядке
            sorted_ranges = sorted(stats.keys(), key=lambda r: r[0])
            table.setRowCount(len(sorted_ranges))
            table.setColumnCount(10)

            # Настройка таблицы
            table.setStyleSheet(f"""
                QTableWidget {{
                    background-color: #0A192F;
                    gridline-color: #505050;
                    border: 2px solid #505050;
                }}
                QTableWidget::item {{
                    border: 1px solid #505050;
                    padding: 5px;
                    background-color: #0A192F;
                }}
                QTableWidget QHeaderView::section {{
                    background-color: #11244A;
                    color: #ffffff;
                    border: 2px solid #505050;
                    padding: 8px;
                    font-weight: bold;
                }}
                QTableWidget::item:selected {{
                    background-color: #2a2a35;
                    color: #ffffff;
                }}
            """)
            table.setShowGrid(True)
            table.setGridStyle(Qt.SolidLine)

            table.horizontalHeader().setDefaultSectionSize(120)
            table.verticalHeader().setDefaultSectionSize(40)
            table.horizontalHeader().setStyleSheet("""
                QHeaderView::section {
                    border: 2px solid #505050;
                    border-bottom: 2px solid #505050;
                    border-right: 2px solid #505050;
                }
            """)
            table.verticalHeader().setStyleSheet("""
                QHeaderView::section {
                    border: 2px solid #505050;
                    border-bottom: 2px solid #505050;
                    border-right: 2px solid #505050;
                }
            """)

            # Устанавливаем ширину колонок
            table.setColumnWidth(0, range_width)
            for group in range(3):
                base_idx = 1 + group * 3
                table.setColumnWidth(base_idx, kolvo_width)
                table.setColumnWidth(base_idx + 1, win_width)
                table.setColumnWidth(base_idx + 2, percent_width)

            # Заполняем таблицу данными
            for row, rng in enumerate(sorted_ranges):
                lower, upper = rng
                # Формат диапазона: показываем без .0, если число целое
                def fmt(x):
                    return str(int(x)) if float(x).is_integer() else str(x)
                range_text = f"{fmt(lower)}-{fmt(upper)}"
                range_item = QTableWidgetItem(range_text)
                range_item.setTextAlignment(Qt.AlignCenter)
                range_item.setBackground(QBrush(QColor("#0A192F")))
                range_item.setFont(QFont("Arial", 14))
                table.setItem(row, 0, range_item)

                categories = ['ОБЩЕЕ', 'OVER', 'UNDER']
                for cat_idx, category in enumerate(categories):
                    base_col = 1 + cat_idx * 3
                    cat_stats = stats[rng][category]

                    # кол-во
                    count_item = QTableWidgetItem(str(cat_stats['кол-во']))
                    count_item.setTextAlignment(Qt.AlignCenter)
                    count_item.setForeground(neutral_color)
                    count_item.setFont(QFont("Arial", 14))
                    table.setItem(row, base_col, count_item)

                    # WIN
                    win_value = cat_stats['WIN']
                    win_text = f"+{win_value}" if win_value > 0 else str(win_value)
                    win_item = QTableWidgetItem(win_text)
                    win_item.setTextAlignment(Qt.AlignCenter)
                    win_item.setFont(QFont("Arial", 14))
                    if win_value > 0:
                        win_item.setForeground(QColor("#3A9B3A"))
                    elif win_value < 0:
                        win_item.setForeground(QColor("#B33333"))
                    else:
                        win_item.setForeground(neutral_color)
                    table.setItem(row, base_col + 1, win_item)

                    # Процент
                    percent = cat_stats['%'] * 100
                    percent_text = f"+{percent:.1f}%" if percent > 0 else f"{percent:.1f}%"
                    percent_item = QTableWidgetItem(percent_text)
                    percent_item.setTextAlignment(Qt.AlignCenter)
                    percent_item.setFont(QFont("Arial", 14))
                    if percent > 0:
                        percent_item.setForeground(QColor("#3A9B3A"))
                    elif percent < 0:
                        percent_item.setForeground(QColor("#B33333"))
                    else:
                        percent_item.setForeground(neutral_color)
                    table.setItem(row, base_col + 2, percent_item)

            # Отключаем заголовки
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            table.verticalHeader().hide()
            table.horizontalHeader().hide()

            # Добавляем таблицу в layout
            self.table_layout_range.addWidget(table)
            self.current_table_range = table
            self.last_tournament_name_range = tournament_name
            self.last_stats_range = stats

        except Exception as e:
            error_msg = f"Ошибка при отображении статистики по диапазонам: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def setup_load_tab(self):
        layout = QVBoxLayout()
        
        # Title
        title_label = QLabel("Загрузка и обработка файла Excel")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addSpacing(20)
        layout.addWidget(title_label)
        layout.addSpacing(30)
        
        # File selection
        file_layout = QHBoxLayout()
        self.file_path_label = QLabel("Файл не выбран")
        self.file_path_label.setMinimumWidth(400)
        
        self.select_file_btn = QPushButton("Выбрать файл")
        self.select_file_btn.setMinimumWidth(150)
        self.select_file_btn.clicked.connect(self.select_file)
        
        file_layout.addWidget(QLabel("Путь к файлу:"))
        file_layout.addWidget(self.file_path_label)
        file_layout.addWidget(self.select_file_btn)
        layout.addLayout(file_layout)
        layout.addSpacing(20)
        
        # Process button
        self.process_btn = QPushButton("Обработать файл")
        self.process_btn.setMinimumWidth(200)
        self.process_btn.setMinimumHeight(50)
        self.process_btn.setFont(QFont("Arial", 12))
        self.process_btn.clicked.connect(self.process_file)
        self.process_btn.setEnabled(False)
        
        process_layout = QHBoxLayout()
        process_layout.addStretch()
        process_layout.addWidget(self.process_btn)
        process_layout.addStretch()
        layout.addLayout(process_layout)
        layout.addSpacing(30)
        
        # Progress bar and status
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        self.progress_bar.hide()
        
        self.status_label = QLabel("")
        self.status_label.setFont(QFont("Arial", 10))
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setWordWrap(True)
        
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.status_label)
        layout.addStretch()
        
        self.tab_load.setLayout(layout)
    
    def update_tournament_list(self):
        """Обновляет список турниров для автодополнения"""
        try:
            with self.db._connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT tournament FROM matches ORDER BY tournament")
                tournaments = [row[0] for row in cursor.fetchall()]
                
                # Обновляем модель автодополнения
                model = QStringListModel()
                model.setStringList(tournaments)
                self.tournament_completer.setModel(model)
                
                # Обновляем метку с доступными турнирами
                if tournaments:
                    self.available_label.setText(
                        f"Доступные турниры ({len(tournaments)}): {', '.join(tournaments[:5])}"
                        + ("..." if len(tournaments) > 5 else "")
                    )
                else:
                    self.available_label.setText("Доступные турниры: нет данных")
        except Exception as e:
            logging.error(f"Ошибка обновления списка турниров: {str(e)}")
    
    def select_file(self):
        try:
            log_step("Выбор файла")

            # Создаем настраиваемый фильтр для файлов - только файлы половин
            file_filter = "Файлы половин (Половины*.xlsx);; Все файлы Excel (*.xlsx *.xls)"
            
            # Получаем последнюю использованную директорию или используем домашнюю директорию
            last_dir = getattr(self, 'settings', QSettings("ExcelAnalyzer", "Preferences")).value("last_directory", os.path.expanduser("~"))

            # Запускаем диалог выбора файла с указанным фильтром
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Выберите файл половин", last_dir, file_filter
            )
            
            # Важно! QFileDialog возвращает пустую строку при нажатии Cancel
            if file_path:
                log_step(f"Выбран файл: {file_path}")
                # Сохраняем директорию для будущего использования
                getattr(self, 'settings', QSettings("ExcelAnalyzer", "Preferences")).setValue("last_directory", os.path.dirname(file_path))
                self.file_path_label.setText(file_path)
                self.process_btn.setEnabled(True)
                self.status_label.setText("Файл выбран. Нажмите 'Обработать файл' для продолжения.")
            else:
                # Пользователь нажал Cancel - просто логируем это событие и ничего не делаем
                log_step("Пользователь отменил выбор файла")
        except Exception as e:
            log_step(f"Ошибка при выборе файла: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось выбрать файл: {str(e)}")
    
    def process_file(self):
        try:
            log_step("Обработка файла")
            file_path = self.file_path_label.text()
            if file_path == "Файл не выбран":
                log_step("Файл не выбран")
                QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите файл Excel")
                return
            
            self.status_label.setText("Обработка файла...")
            self.progress_bar.show()
            self.process_btn.setEnabled(False)
            self.select_file_btn.setEnabled(False)
            
            # Start processing in a background thread
            log_step("Запуск потока обработки")
            self.processor_thread = RoykaProcessorThread(file_path)
            self.processor_thread.progress_signal.connect(self.update_progress)
            self.processor_thread.finished_signal.connect(self.processing_finished)
            self.processor_thread.error_signal.connect(self.processing_error)
            self.processor_thread.start()
        except Exception as e:
            log_step(f"Ошибка при запуске обработки: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось начать обработку файла: {str(e)}")
            self.process_btn.setEnabled(True)
            self.select_file_btn.setEnabled(True)
            self.progress_bar.hide()
    
    def update_progress(self, message):
        self.status_label.setText(message)
    
    def processing_finished(self, data):
        try:
            log_step("Завершение обработки файла")
            self.data = data
            self.progress_bar.hide()
            self.process_btn.setEnabled(True)
            self.select_file_btn.setEnabled(True)
            
            tournament_count = len(data) if data else 0
            log_step(f"Найдено турниров: {tournament_count}")
            self.status_label.setText(f"Файл успешно обработан. Найдено {tournament_count} турниров.")
            
            # Update available tournaments
            if data and tournament_count > 0:
                tournament_names = sorted(data.keys())
                self.available_label.setText(f"Доступные турниры: {', '.join(tournament_names[:5])}...")
                
                # Обновляем автозаполнение для поля поиска
                self.update_tournament_completer(tournament_names)
            else:
                self.available_label.setText("Доступные турниры: нет данных")
                
            QMessageBox.information(self, "Успешно", f"Файл успешно обработан. Найдено {tournament_count} турниров.")
            
            # Enable stats tab
            self.tabs.setTabEnabled(1, True)
            self.tabs.setCurrentIndex(1)  # Переключаемся на вкладку статистики
            
        except Exception as e:
            log_step(f"Ошибка при завершении обработки: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при завершении обработки: {str(e)}")
    
    def update_tournament_completer(self, tournament_names):
        """Обновляет список автодополнения для поиска турниров"""
        log_step("Обновление автодополнения для поиска")
        if not tournament_names:
            return
            
        # Создаем модель данных для автодополнения
        model = QStringListModel()
        model.setStringList(tournament_names)
        
        # Устанавливаем модель в автодополнение
        self.tournament_completer.setModel(model)
        log_step(f"Автодополнение обновлено: {len(tournament_names)} турниров")
    
    def processing_error(self, error_message):
        try:
            log_step(f"Ошибка обработки: {error_message}")
            self.progress_bar.hide()
            self.process_btn.setEnabled(True)
            self.select_file_btn.setEnabled(True)
            self.status_label.setText("Произошла ошибка при обработке файла.")
            
            error_dialog = QMessageBox()
            error_dialog.setIcon(QMessageBox.Critical)
            error_dialog.setWindowTitle("Ошибка")
            error_dialog.setText("Ошибка при обработке файла")
            error_dialog.setDetailedText(error_message)
            error_dialog.exec_()
        except Exception as e:
            log_step(f"Критическая ошибка: {str(e)}")
            QMessageBox.critical(self, "Критическая ошибка", f"Не удалось отобразить сообщение об ошибке: {str(e)}")
    
    def safe_search_tournament(self):
        try:
            log_step("Поиск турнира")
            self.search_tournament()
        except Exception as e:
            log_step(f"Ошибка при поиске: {str(e)}")
            error_msg = f"Ошибка при поиске турнира: {str(e)}\n{traceback.format_exc()}"
            QMessageBox.critical(self, "Ошибка поиска", error_msg)
    
    def search_tournament(self):
        """Поиск и отображение данных турнира"""
        tournament_name = self.search_input.text().strip()
        
        if not tournament_name:
            QMessageBox.warning(self, "Предупреждение", "Введите название турнира")
            return
        
        # Очищаем лог и добавляем начальное сообщение
        self.debug_log.clear()
        self.add_debug_log(f"=== НАЧАЛО АНАЛИЗА ТУРНИРА: {tournament_name} ===")
        
        try:
            # Проверяем есть ли данные в базе
            stats = self.db.get_statistics()
            self.add_debug_log(f"Всего записей в базе: {stats['total_records']}")
            
            if stats['total_records'] == 0:
                self.add_debug_log("ОШИБКА: База данных пуста")
                QMessageBox.warning(self, "Нет данных", 
                    "База данных пуста. Сначала импортируйте данные через вкладку 'Управление данными'")
                return
                
            # Получаем данные турнира
            with self.db._connect() as conn:
                cursor = conn.cursor()
                self.add_debug_log("Выполнение запроса к базе данных...")
                
                cursor.execute("""
                    SELECT 
                        date, team_home, team_away, t1h, t2h, tim,
                        deviation, kickoff, predict, result
                    FROM matches 
                    WHERE tournament = ?
                    ORDER BY date
                """, (tournament_name,))
                
                matches = cursor.fetchall()
                self.add_debug_log(f"Найдено матчей: {len(matches)}")
                
                if not matches:
                    self.add_debug_log("ОШИБКА: Турнир не найден в базе")
                    QMessageBox.warning(self, "Не найдено", 
                        f"Турнир '{tournament_name}' не найден в базе данных")
                    return
                
                # Преобразуем данные
                tournament_data = {
                    'matches': [
                        {
                            'date': row[0],
                            'team1': row[1],
                            'team2': row[2],
                            't1h': row[3],
                            't2h': row[4],
                            'tim': row[5],
                            'deviation': row[6],
                            'kickoff': row[7],
                            'predict': row[8],
                            'result': row[9]
                        }
                        for row in matches
                    ]
                }
                
                # Отображаем статистику
                self.display_tournament_stats(tournament_name, tournament_data)
                
        except Exception as e:
            error_msg = f"Ошибка при поиске турнира: {str(e)}\n{traceback.format_exc()}"
            self.add_debug_log(f"ОШИБКА: {error_msg}")
            QMessageBox.critical(self, "Ошибка", str(e))

    def add_debug_log(self, message):
        """Добавляет сообщение в лог отладки"""
        try:
            if hasattr(self, 'debug_log') and self.debug_log is not None:
                timestamp = datetime.now().strftime("%H:%M:%S")
                formatted_message = f"[{timestamp}] {message}"
                self.debug_log.appendPlainText(formatted_message)
                # Прокручиваем до конца
                self.debug_log.verticalScrollBar().setValue(
                    self.debug_log.verticalScrollBar().maximum()
                )
                # Принудительно обновляем отображение
                QApplication.processEvents()
        except Exception as e:
            print(f"Error in add_debug_log: {str(e)}")

    
    # Функция для форматирования процентов
    def format_percentage(self, value):
        try:
            if isinstance(value, (int, float)):
                # Умножаем на 100, чтобы получить проценты и округляем до одного знака
                formatted_value = round(value * 100, 1)
                
                # Добавляем знак плюса для положительных значений
                if formatted_value > 0:
                    return f"+{formatted_value}%"
                elif formatted_value < 0:
                    return f"{formatted_value}%"
                else:
                    return "0%"
            return str(value)
        except Exception:
            return str(value)
    
    # Функция для форматирования значений разницы
    def format_difference(self, value):
        try:
            if value == int(value):
                # Если число целое (например, 1.0), убираем десятичную часть
                return f"{int(value)}"
            else:
                # Если число не целое (например, 0.5), оставляем как есть
                return f"{value}"
        except Exception:
            return str(value)
    
    def display_tournament_stats(self, tournament_name, tournament_data):
        """Отображает статистику турнира"""
        try:
       
                
            # Очищаем предыдущий контент
            self.clear_layout(self.table_layout)

            # Title
            title_label = QLabel(f"Статистика для турнира: {tournament_name}")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setAlignment(Qt.AlignCenter)
            self.table_layout.addWidget(title_label)

          
            # Добавляем подсказку про прокрутку
            scroll_hint = QLabel("* Используйте прокрутку, чтобы увидеть все разницы до 10")
            scroll_hint.setAlignment(Qt.AlignCenter)
            scroll_hint.setStyleSheet("color: #808080; font-size: 12px;")
            self.table_layout.addWidget(scroll_hint)

            # Анализируем данные
            stats = self.analyze_tournament_data(tournament_data)
            
            # Очищаем предыдущий контент
            self.clear_layout(self.table_layout)

            # Title
            title_label = QLabel(f"Статистика для турнира: {tournament_name}")
            title_label.setFont(QFont("Arial", 18, QFont.Bold))
            title_label.setAlignment(Qt.AlignCenter)
            self.table_layout.addWidget(title_label)

            # Добавляем подсказку про прокрутку
            scroll_hint = QLabel("* Используйте прокрутку, чтобы увидеть все разницы до 10")
            scroll_hint.setAlignment(Qt.AlignCenter)
            scroll_hint.setStyleSheet("color: #808080; font-size: 12px;")
            self.table_layout.addWidget(scroll_hint)

            # Определение стилей и цветов
            header_bg = "#1E1E1E"  
            table_bg = "#1E1E1E"   
            neutral_color = QColor(180, 180, 180)
            grid_color = "#383838"

            # Ширины колонок
            # Ширины колонок
            raznica_width = 80  # Было 60
            kolvo_width = 80   # Не меняется
            win_width = 80     # Было 100
            percent_width = 80 # Не меняется

            # Создаем контейнер для заголовков
            header_container = QWidget()
            header_container.setFixedHeight(80)  # Увеличиваем высоту для двух строк
            header_layout = QVBoxLayout(header_container)  # Используем вертикальный layout
            header_layout.setContentsMargins(0, 0, 0, 0)
            header_layout.setSpacing(0)

            # Создаем верхний ряд заголовков (ОБЩЕЕ, OVER, UNDER)
            top_header = QWidget()
            top_header_layout = QHBoxLayout(top_header)
            top_header_layout.setContentsMargins(raznica_width, 0, 0, 0)  # Отступ слева для колонки "Разница"
            top_header_layout.setSpacing(0)

            # Заголовки групп с приглушенными цветами
            headers = [
                ("ОБЩЕЕ", "#CCCCCC"),  # Приглушенный белый
                ("OVER ↑", "#3A9B3A"),  # Приглушенный зеленый
                ("UNDER ↓", "#B33333")  # Приглушенный красный
            ]

            # Создаем группы основных заголовков
            for text, color in headers:
                group_width = kolvo_width + win_width + percent_width
                label = QLabel(text)
                label.setFixedWidth(group_width)
                label.setAlignment(Qt.AlignCenter)
                label.setFont(QFont("Arial", 14, QFont.Bold))

                # Добавляем отступы для центрирования над колонкой WIN
                # Добавляем отступы для центрирования над колонкой WIN
                if text == "ОБЩЕЕ":
                    margin_left = -45  # Сдвигаем ОБЩЕЕ сильнее влево
                elif text == "OVER ↑":
                    margin_left = -35  # Сдвигаем OVER влево
                elif text == "UNDER ↓":
                    margin_left = -25  # Сдвигаем UNDER чуть влево
                
                label.setStyleSheet(f"""
                    color: {color};
                    background-color: {header_bg};
                    border: none;
                    padding: 4px;
                    margin-left: {margin_left}px;
                """)
                top_header_layout.addWidget(label)

            # Создаем нижний ряд заголовков (кол-во, WIN, %)
            bottom_header = QWidget()
            bottom_header_layout = QHBoxLayout(bottom_header)
            bottom_header_layout.setContentsMargins(raznica_width, 0, 0, 0)
            bottom_header_layout.setSpacing(0)

            # Создаем подзаголовки для каждой группы
            subheaders = ["кол-во", "WIN", "%"]
            for _ in range(3):  # Для каждой группы (ОБЩЕЕ, OVER, UNDER)
                for subheader in subheaders:
                    width = kolvo_width if subheader == "кол-во" else (win_width if subheader == "WIN" else percent_width)
                    label = QLabel(subheader)
                    label.setFixedWidth(width)
                    label.setAlignment(Qt.AlignCenter)
                    label.setFont(QFont("Arial", 12))
                    label.setStyleSheet(f"""
                        color: #999999;
                        background-color: {header_bg};
                        border: none;
                        padding: 4px;
                    """)
                    bottom_header_layout.addWidget(label)

            # Добавляем оба ряда заголовков в контейнер
            header_layout.addWidget(top_header)
            header_layout.addWidget(bottom_header)

            # Добавляем контейнер с заголовками в основной layout
            self.table_layout.addWidget(header_container)

            # Создаем основную таблицу
            table = CustomTableWidget()
            table.setFont(QFont("Arial", 14))
            
            # Настройка таблицы
            differences = [0.1] + [x/2 for x in range(1, 21)]
            table.setRowCount(len(differences))
            table.setColumnCount(10)

            # Определяем стили таблицы
            table.setStyleSheet(f"""
                QTableWidget {{
                    background-color: #1a1a1f;  /* Более темный фон */
                    gridline-color: #505050;    /* Цвет линий сетки */
                    border: 2px solid #505050;  /* Жирная внешняя граница */
                }}
                
                QTableWidget::item {{
                    border: 1px solid #505050;  /* Границы для каждой ячейки */
                    padding: 5px;
                    background-color: #1a1a1f;  /* Фон ячеек */
                }}
                
                QTableWidget QHeaderView::section {{
                    background-color: #252529;   /* Фон заголовков */
                    color: #ffffff;              /* Цвет текста заголовков */
                    border: 2px solid #505050;   /* Жирные границы заголовков */
                    padding: 8px;
                    font-weight: bold;
                }}
                
                QTableWidget::item:selected {{
                    background-color: #2a2a35;   /* Цвет выделения */
                    color: #ffffff;              /* Цвет текста при выделении */
                }}
            """)
            # Включаем отображение сетки
            table.setShowGrid(True)
            table.setGridStyle(Qt.SolidLine)

            # Устанавливаем толщину линий сетки
            table.horizontalHeader().setDefaultSectionSize(120)
            table.verticalHeader().setDefaultSectionSize(40)

            # Устанавливаем стиль для вертикальных и горизонтальных заголовков
            table.horizontalHeader().setStyleSheet("""
                QHeaderView::section {
                    border: 2px solid #505050;
                    border-bottom: 2px solid #505050;
                    border-right: 2px solid #505050;
                }
            """)

            table.verticalHeader().setStyleSheet("""
                QHeaderView::section {
                    border: 2px solid #505050;
                    border-bottom: 2px solid #505050;
                    border-right: 2px solid #505050;
                }
            """)

            # Устанавливаем ширину колонок
            table.setColumnWidth(0, raznica_width)
            for group in range(3):
                base_idx = 1 + group * 3
                table.setColumnWidth(base_idx, kolvo_width)
                table.setColumnWidth(base_idx + 1, win_width)
                table.setColumnWidth(base_idx + 2, percent_width)

            # Заполняем данные
            for row, diff in enumerate(differences):
                # Разница
                diff_text = str(int(diff)) if diff.is_integer() else str(diff)
                diff_item = QTableWidgetItem(diff_text)
                diff_item.setTextAlignment(Qt.AlignCenter)
                diff_item.setBackground(QBrush(QColor("#1a1a2e")))
                diff_item.setFont(QFont("Arial", 14))
                table.setItem(row, 0, diff_item)

                # Заполняем данные для каждой категории
                categories = ['ОБЩЕЕ', 'OVER', 'UNDER']
                for cat_idx, category in enumerate(categories):
                    base_col = 1 + cat_idx * 3
                    cat_stats = stats[diff][category]

                    # кол-во
                    count_item = QTableWidgetItem(str(cat_stats['кол-во']))
                    count_item.setTextAlignment(Qt.AlignCenter)
                    count_item.setForeground(neutral_color)
                    count_item.setFont(QFont("Arial", 14))
                    table.setItem(row, base_col, count_item)

                    # WIN
                    win_value = cat_stats['WIN']
                    win_text = f"+{win_value}" if win_value > 0 else str(win_value)
                    win_item = QTableWidgetItem(win_text)
                    win_item.setTextAlignment(Qt.AlignCenter)
                    win_item.setFont(QFont("Arial", 14))
                    if win_value > 0:
                        win_item.setForeground(QColor("#3A9B3A"))  # Приглушенный зеленый
                    elif win_value < 0:
                        win_item.setForeground(QColor("#B33333"))  # Приглушенный красный
                    else:
                        win_item.setForeground(neutral_color)
                    table.setItem(row, base_col + 1, win_item)

                    # Процент
                    percent = cat_stats['%'] * 100
                    percent_text = f"+{percent:.1f}%" if percent > 0 else f"{percent:.1f}%"
                    percent_item = QTableWidgetItem(percent_text)
                    percent_item.setTextAlignment(Qt.AlignCenter)
                    percent_item.setFont(QFont("Arial", 14))
                    if percent > 0:
                        percent_item.setForeground(QColor("#3A9B3A"))  # Приглушенный зеленый
                    elif percent < 0:
                        percent_item.setForeground(QColor("#B33333"))  # Приглушенный красный
                    else:
                        percent_item.setForeground(neutral_color)
                    table.setItem(row, base_col + 2, percent_item)

            # Отключаем горизонтальный скролл и заголовки
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            table.verticalHeader().hide()
            table.horizontalHeader().hide()

            # Добавляем таблицу в layout
            self.table_layout.addWidget(table)
            
            # Сохраняем ссылки
            self.current_table = table
            self.last_tournament_name = tournament_name
            self.last_stats = stats

        except Exception as e:
            error_msg = f"Ошибка при отображении статистики: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def show_all_tournaments_half_stats_change(self):
        """Показывает статистику по всем турнирам для T2H + Div (CHANGE)"""
        try:
            self.debug_log_half_change.clear()
            self.add_debug_log_half_change("=== АНАЛИЗ ВСЕХ ТУРНИРОВ (T2H + Div CHANGE) ===")

            with self.db._connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT tournament FROM matches ORDER BY tournament")
                tournaments = [row[0] for row in cursor.fetchall()]

            if not tournaments:
                QMessageBox.warning(self, "Нет данных", "Нет турниров в базе данных")
                return

            self.add_debug_log_half_change(f"Найдено турниров: {len(tournaments)}")

            all_tournaments_stats = {}
            total_stats = {
                'OVER': {'кол-во': 0, 'WIN': 0, '%': 0},
                'UNDER': {'кол-во': 0, 'WIN': 0, '%': 0},
                'TOTAL': {'кол-во': 0, 'WIN': 0, '%': 0}
            }

            for tournament_name in tournaments:
                with self.db._connect() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT 
                            date, team_home, team_away, t1h, t2h, tim,
                            deviation, kickoff, predict, result
                        FROM matches 
                        WHERE tournament = ? 
                        ORDER BY date
                    """, (tournament_name,))
                    matches = cursor.fetchall()
                    tournament_data = {
                        'matches': [
                            {
                                'date': row[0],
                                'team1': row[1],
                                'team2': row[2],
                                't1h': row[3],
                                't2h': row[4],
                                'tim': row[5],
                                'deviation': row[6],
                                'kickoff': row[7],
                                'predict': row[8],
                                'result': row[9]
                            }
                            for row in matches
                        ]
                    }
                    tournament_stats = self.analyze_tournament_data_half_change(tournament_data, tournament_name)
                    all_tournaments_stats[tournament_name] = tournament_stats

                    for category in ['OVER', 'UNDER', 'TOTAL']:
                        total_stats[category]['кол-во'] += tournament_stats[category]['кол-во']
                        total_stats[category]['WIN'] += tournament_stats[category]['WIN']

                    self.add_debug_log_half_change(
                        f"  {tournament_name}: OVER={tournament_stats['OVER']['кол-во']}, "
                        f"UNDER={tournament_stats['UNDER']['кол-во']}, "
                        f"WIN={tournament_stats['TOTAL']['WIN']}"
                    )

            for category in ['OVER', 'UNDER', 'TOTAL']:
                if total_stats[category]['кол-во'] > 0:
                    win = total_stats[category]['WIN']
                    count = total_stats[category]['кол-во']
                    total_stats[category]['%'] = win / (count * 100)

            self.display_all_tournaments_half_stats_change(all_tournaments_stats, total_stats)
        except Exception as e:
            error_msg = f"Ошибка при анализе всех турниров: {str(e)}\n{traceback.format_exc()}"
            self.add_debug_log_half_change(f"ОШИБКА: {error_msg}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def display_all_tournaments_half_stats_change(self, all_tournaments_stats, total_stats):
        """Отображает статистику по всем турнирам для T2H + Div (CHANGE)"""
        try:
            self.clear_layout(self.table_layout_half_change)

            title_label = QLabel("Статистика T2H + Div CHANGE по всем турнирам")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setAlignment(Qt.AlignCenter)
            self.table_layout_half_change.addWidget(title_label)

            total_table = QTableWidget()
            total_table.setFont(QFont("Arial", 12, QFont.Bold))
            total_table.setRowCount(3)
            total_table.setColumnCount(4)
            total_table.setStyleSheet("""
                QTableWidget {
                    background-color: #1a1a1f;
                    gridline-color: #505050;
                    border: 2px solid #505050;
                }
                QTableWidget::item {
                    border: 1px solid #505050;
                    padding: 5px;
                    background-color: #1a1a1f;
                }
            """)
            total_table.setShowGrid(True)
            total_table.setGridStyle(Qt.SolidLine)
            total_table.setColumnWidth(0, 150)
            total_table.setColumnWidth(1, 100)
            total_table.setColumnWidth(2, 100)
            total_table.setColumnWidth(3, 100)

            neutral_color = QColor(180, 180, 180)
            category_colors = {
                'OVER': QColor(50, 100, 50),
                'UNDER': QColor(100, 50, 50),
                'TOTAL': QColor(50, 50, 100)
            }

            for row, category in enumerate(['OVER', 'UNDER', 'TOTAL']):
                label_item = QTableWidgetItem(category)
                label_item.setTextAlignment(Qt.AlignCenter)
                label_item.setBackground(QBrush(category_colors[category]))
                label_item.setFont(QFont("Arial", 12, QFont.Bold))
                total_table.setItem(row, 0, label_item)

                count_item = QTableWidgetItem(str(total_stats[category]['кол-во']))
                count_item.setTextAlignment(Qt.AlignCenter)
                count_item.setForeground(neutral_color)
                total_table.setItem(row, 1, count_item)

                win_value = total_stats[category]['WIN']
                win_text = f"+{win_value}" if win_value > 0 else str(win_value)
                win_item = QTableWidgetItem(win_text)
                win_item.setTextAlignment(Qt.AlignCenter)
                if win_value > 0:
                    win_item.setForeground(QColor("#3A9B3A"))
                elif win_value < 0:
                    win_item.setForeground(QColor("#B33333"))
                else:
                    win_item.setForeground(neutral_color)
                total_table.setItem(row, 2, win_item)

                count = total_stats[category]['кол-во']
                if count > 0:
                    total_bet = count * 100
                    roi_percent = int((win_value / total_bet) * 100)
                else:
                    roi_percent = 0
                roi_item = QTableWidgetItem(f"{roi_percent}%")
                roi_item.setTextAlignment(Qt.AlignCenter)
                if roi_percent > 0:
                    roi_item.setForeground(QColor("#3A9B3A"))
                elif roi_percent < 0:
                    roi_item.setForeground(QColor("#B33333"))
                else:
                    roi_item.setForeground(neutral_color)
                total_table.setItem(row, 3, roi_item)

            self.table_layout_half_change.addWidget(total_table)
            self.table_layout_half_change.addSpacing(20)

            tournaments_table = QTableWidget()
            tournaments_table.setFont(QFont("Arial", 10))
            tournaments_table.setColumnCount(5)
            tournaments_table.setHorizontalHeaderLabels(["Турнир", "OVER WIN", "UNDER WIN", "TOTAL WIN", "ROI %"])
            tournaments_table.setRowCount(len(all_tournaments_stats))
            tournaments_table.setStyleSheet("""
                QTableWidget {
                    background-color: #1a1a1f;
                    gridline-color: #505050;
                    border: 2px solid #505050;
                }
            """)

            for row, (tournament_name, stats) in enumerate(sorted(all_tournaments_stats.items())):
                name_item = QTableWidgetItem(tournament_name)
                tournaments_table.setItem(row, 0, name_item)

                over_win = stats['OVER']['WIN']
                over_item = QTableWidgetItem(f"+{over_win}" if over_win > 0 else str(over_win))
                over_item.setTextAlignment(Qt.AlignCenter)
                if over_win > 0:
                    over_item.setForeground(QColor("#3A9B3A"))
                elif over_win < 0:
                    over_item.setForeground(QColor("#B33333"))
                tournaments_table.setItem(row, 1, over_item)

                under_win = stats['UNDER']['WIN']
                under_item = QTableWidgetItem(f"+{under_win}" if under_win > 0 else str(under_win))
                under_item.setTextAlignment(Qt.AlignCenter)
                if under_win > 0:
                    under_item.setForeground(QColor("#3A9B3A"))
                elif under_win < 0:
                    under_item.setForeground(QColor("#B33333"))
                tournaments_table.setItem(row, 2, under_item)

                total_win = stats['TOTAL']['WIN']
                total_item = QTableWidgetItem(f"+{total_win}" if total_win > 0 else str(total_win))
                total_item.setTextAlignment(Qt.AlignCenter)
                if total_win > 0:
                    total_item.setForeground(QColor("#3A9B3A"))
                elif total_win < 0:
                    total_item.setForeground(QColor("#B33333"))
                tournaments_table.setItem(row, 3, total_item)

                total_count = stats['TOTAL']['кол-во']
                if total_count > 0:
                    total_bet = total_count * 100
                    roi_percent = int((total_win / total_bet) * 100)
                else:
                    roi_percent = 0
                roi_item = QTableWidgetItem(f"{roi_percent}%")
                roi_item.setTextAlignment(Qt.AlignCenter)
                if roi_percent > 0:
                    roi_item.setForeground(QColor("#3A9B3A"))
                elif roi_percent < 0:
                    roi_item.setForeground(QColor("#B33333"))
                tournaments_table.setItem(row, 4, roi_item)

            tournaments_table.resizeColumnsToContents()
            self.table_layout_half_change.addWidget(tournaments_table)
        except Exception as e:
            error_msg = f"Ошибка при отображении статистики: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)
    def show_debug_logs(self):
        """Показывает окно с логами"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Отладочная информация")
        dialog.setMinimumSize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        # Создаем поле для логов
        log_text = QPlainTextEdit()
        log_text.setReadOnly(True)
        log_text.setStyleSheet("""
            QPlainTextEdit {
                background-color: #1a1a1f;
                color: #00ff00;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 12px;
                padding: 5px;
            }
        """)
        layout.addWidget(log_text)
        
        # Кнопка закрытия
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        self.debug_dialog = dialog
        self.debug_log = log_text
        
        dialog.show()

    def add_debug_log(self, message):
        """Добавляет сообщение в лог отладки"""
        if hasattr(self, 'debug_log'):
            self.debug_log.appendPlainText(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")        

    def calculate_prediction(self, match_data, difference):
        """Расчет прогноза"""
        try:
            # Получаем значения
            t1h = float(match_data.get('t1h', 0) or 0)
            t2h = float(match_data.get('t2h', 0) or 0)
            tim = float(match_data.get('tim', 0) or 0)
            dev = float(match_data.get('deviation', 0) or 0)
            kickoff = float(match_data.get('kickoff', 0) or 0)
            predict = match_data.get('predict', '')
            result = float(match_data.get('result', 0) or 0)

            # Этап 1
            initial_diff = t1h + t2h - tim
            if initial_diff >= difference:
                stage1 = "OVER"
            elif initial_diff <= -difference:
                stage1 = "UNDER"
            else:
                return ("No bet", 1, result)

            # Этап 2
            modified_diff = t1h + t2h + dev - tim
            if (stage1 == "OVER" and modified_diff >= difference) or \
            (stage1 == "UNDER" and modified_diff <= -difference):
                stage2 = stage1
            else:
                return ("No bet", 2, result)

            # Этап 3
            try:
                predict_value = float(str(predict).replace(',', '.')) if predict else 0
                kickoff_diff = predict_value - kickoff if kickoff != 0 else 0
            except (ValueError, TypeError):
                return (stage2, 2, result)

            # Проверяем predict = 0
            if predict_value == 0:
                return (stage2, 2, result)

            # Проверяем диапазон (-3; 3)
            if -3 < kickoff_diff < 3:
                return (stage2, 2, result)

            # Новая логика для UNDER и OVER
            if stage2 == "UNDER":
                if kickoff_diff < 0:  # Для UNDER если разница отрицательная
                    return (stage2, 2, result)
            else:  # OVER
                if kickoff_diff > 0:  # Для OVER если разница положительная
                    return (stage2, 2, result)

            # Расчет для этапа 3 и 4
            if kickoff != 0:
                modified_value = t1h + t2h * (1 + (kickoff_diff/kickoff)) - tim
                if (stage2 == "OVER" and kickoff_diff <= -3 and modified_value >= difference) or \
                (stage2 == "UNDER" and kickoff_diff >= 3 and modified_value <= -difference):
                    stage3 = stage2
                else:
                    return ("No bet", 3, result)
            else:
                return (stage2, 2, result)

            # Этап 4
            final_value = t1h + t2h * (1 + (kickoff_diff/kickoff)) + dev - tim
            if (stage3 == "OVER" and kickoff_diff <= -3 and final_value >= difference) or \
            (stage3 == "UNDER" and kickoff_diff >= 3 and final_value <= -difference):
                return (stage3, 4, result)

            return ("No bet", 4, result)

        except Exception as e:
            return ("Error", 0, None)
        
    def calculate_prediction_half(self, match_data, half_threshold=4.5):
        """
        Расчет прогноза для T2H + Div >= 4.5 (OVER) и T2H + Div <= -4.5 (UNDER)
        """
        try:
            # Получаем значения
            t1h = float(match_data.get('t1h', 0) or 0)
            t2h = float(match_data. get('t2h', 0) or 0)
            tim = float(match_data.get('tim', 0) or 0)
            dev = float(match_data.get('deviation', 0) or 0)
            kickoff = float(match_data. get('kickoff', 0) or 0)
            predict = match_data.get('predict', '')
            result = float(match_data.get('result', 0) or 0)

            # Этап 1: Начальная разница (направление)
            initial_diff = t1h + t2h - tim
            
            if abs(initial_diff) < 0.1: 
                return ("No bet", 1, result)

            # Запоминаем исходное направление для первого этапа
            initial_direction = "OVER" if initial_diff >= 0.1 else "UNDER"

            # Этап 2: Добавляем deviation
            stage2_value = t1h + t2h + dev - tim
            stage2_direction = "OVER" if stage2_value >= 0.1 else "UNDER"

            # Если направление изменилось — No bet
            if stage2_direction != initial_direction:
                return ("No bet", 2, result)

            # Проверяем порог 4.5 на этапе 2
            if abs(stage2_value) < half_threshold:
                return ("No bet", 2, result)

            # Проверяем predict и kickoff
            try:
                predict_value = float(str(predict).replace(',', '.')) if predict else 0
                kickoff_diff = predict_value - kickoff if kickoff != 0 else 0
            except (ValueError, TypeError):
                return (stage2_direction, 2, result)

            # Если predict = 0, используем этап 2
            if predict_value == 0:
                return (stage2_direction, 2, result)

            # Если диапазон (-3; 3), используем этап 2
            if -3 < kickoff_diff < 3:
                return (stage2_direction, 2, result)

            # Этап 3-4: Работаем с predict/kickoff
            # Рассчет для этапа 3 и 4
            if kickoff != 0:
                ratio = kickoff_diff / kickoff
                stage3_value = t1h + t2h * (1 + ratio) - tim
                
                # Этап 4: Добавляем deviation
                final_value = t1h + t2h * (1 + ratio) + dev - tim
                
                # Для OVER: проверяем условие этапа 3 (kickoff_diff <= -3)
                if stage2_direction == "OVER" and kickoff_diff <= -3 and stage3_value >= 0.1:
                    # Переходим на этап 4
                    if final_value >= half_threshold:
                        return ("OVER", 4, result)
                    elif final_value <= -half_threshold:
                        return ("UNDER", 4, result)
                    else:
                        return (stage2_direction, 2, result)
                
                # Для UNDER: проверяем условие этапа 3 (kickoff_diff >= 3)
                if stage2_direction == "UNDER" and kickoff_diff >= 3 and stage3_value <= -0.1:
                    # Переходим на этап 4
                    if final_value >= half_threshold:
                        return ("OVER", 4, result)
                    elif final_value <= -half_threshold:
                        return ("UNDER", 4, result)
                    else:
                        return (stage2_direction, 2, result)
                
                # Если не прошли условие этапа 3, используем stage2_value
                return (stage2_direction, 2, result)
            else:
                # Если kickoff = 0, используем stage2_value
                return (stage2_direction, 2, result)

        except Exception as e:
            logging.error(f"Ошибка в calculate_prediction_half: {str(e)}")
            return ("Error", 0, None)

    def calculate_prediction_half_change(self, match_data, half_threshold=4.5):
        """
        Расчет прогноза для T2H + Div >= 4.5 (OVER) и T2H + Div <= -4.5 (UNDER)
        Вариант CHANGE: игнорируем смену направления на этапе 2.
        """
        try:
            t1h = float(match_data.get('t1h', 0) or 0)
            t2h = float(match_data.get('t2h', 0) or 0)
            tim = float(match_data.get('tim', 0) or 0)
            dev = float(match_data.get('deviation', 0) or 0)
            kickoff = float(match_data.get('kickoff', 0) or 0)
            predict = match_data.get('predict', '')
            result = float(match_data.get('result', 0) or 0)

            initial_diff = t1h + t2h - tim
            if abs(initial_diff) < 0.1:
                return ("No bet", 1, result)

            stage2_value = t1h + t2h + dev - tim
            if abs(stage2_value) < half_threshold:
                return ("No bet", 2, result)

            stage2_direction = "OVER" if stage2_value >= 0.1 else "UNDER"

            try:
                predict_value = float(str(predict).replace(',', '.')) if predict else 0
                kickoff_diff = predict_value - kickoff if kickoff != 0 else 0
            except (ValueError, TypeError):
                return (stage2_direction, 2, result)

            if predict_value == 0:
                return (stage2_direction, 2, result)

            if -3 < kickoff_diff < 3:
                return (stage2_direction, 2, result)

            if kickoff != 0:
                ratio = kickoff_diff / kickoff
                stage3_value = t1h + t2h * (1 + ratio) - tim
                final_value = t1h + t2h * (1 + ratio) + dev - tim

                if stage2_direction == "OVER" and kickoff_diff <= -3 and stage3_value >= 0.1:
                    if final_value >= half_threshold:
                        return ("OVER", 4, result)
                    elif final_value <= -half_threshold:
                        return ("UNDER", 4, result)
                    return (stage2_direction, 2, result)

                if stage2_direction == "UNDER" and kickoff_diff >= 3 and stage3_value <= -0.1:
                    if final_value >= half_threshold:
                        return ("OVER", 4, result)
                    elif final_value <= -half_threshold:
                        return ("UNDER", 4, result)
                    return (stage2_direction, 2, result)

                return (stage2_direction, 2, result)

            return (stage2_direction, 2, result)

        except Exception as e:
            logging.error(f"Ошибка в calculate_prediction_half_change: {str(e)}")
            return ("Error", 0, None)

    def calculate_prediction_half_ncaa(self, match_data, half_threshold=4.5):
        """
        Расчет прогноза для NCAA D1 с особой логикой T2H + Div >= 4.5 / <= -4.5
        
        Для NCAA D1: игнорируем смену направления, суммируем значения в любом случае
        
        Returns:
            tuple: (prediction, stage, result)
        """
        try:
            # Получаем значения
            t1h = float(match_data.get('t1h', 0) or 0)
            t2h = float(match_data.get('t2h', 0) or 0)
            tim = float(match_data.get('tim', 0) or 0)
            dev = float(match_data.get('deviation', 0) or 0)
            kickoff = float(match_data. get('kickoff', 0) or 0)
            predict = match_data.get('predict', '')
            result = float(match_data.get('result', 0) or 0)

            # Этап 1: Проверяем начальную разницу первой половины
            initial_diff = t1h + t2h - tim
            
            # Для NCAA D1 определяем направление, но не отклоняем если смена
            if initial_diff >= 0.1:  
                stage1 = "OVER"
            elif initial_diff <= -0.1:
                stage1 = "UNDER"
            else:
                return ("No bet", 1, result)

            # Этап 2: Добавляем deviation
            # ДЛЯ NCAA D1: НЕ ПРОВЕРЯЕМ СМЕНУ НАПРАВЛЕНИЯ, просто суммируем
            stage2_value = t1h + t2h + dev - tim
            
            # Для NCAA D1 продолжаем в любом случае
            stage2 = stage1  # Сохраняем исходное направление
            
            # Если значение очень близко к 0, может быть "No bet"
            if abs(stage2_value) < 0.1:
                return ("No bet", 2, result)

            # Этап 3: Проверяем predict и kickoff
            try:
                predict_value = float(str(predict).replace(',', '.')) if predict else 0
                kickoff_diff = predict_value - kickoff if kickoff != 0 else 0
            except (ValueError, TypeError):
                # Если predict=0 или ошибка, проверяем критерий 4.5
                if (stage2_value >= half_threshold):  # OVER
                    return ("OVER", 2, result)
                elif (stage2_value <= -half_threshold):  # UNDER
                    return ("UNDER", 2, result)
                else:
                    return ("No bet", 2, result)

            # Если predict = 0, проверяем критерий 4.5
            if predict_value == 0:
                if stage2_value >= half_threshold: 
                    return ("OVER", 2, result)
                elif stage2_value <= -half_threshold:
                    return ("UNDER", 2, result)
                else:
                    return ("No bet", 2, result)

            # Если диапазон (-3; 3), проверяем критерий 4.5
            if -3 < kickoff_diff < 3:
                if stage2_value >= half_threshold:
                    return ("OVER", 2, result)
                elif stage2_value <= -half_threshold:
                    return ("UNDER", 2, result)
                else:
                    return ("No bet", 2, result)

            # Проверяем условия для этапа 4
            # Для NCAA D1: проверяем по итоговому значению, не по исходному направлению
            if stage2_value > 0:  # Потенциальная OVER
                if kickoff_diff > 0:
                    # Просто проверяем критерий 4.5
                    if stage2_value >= half_threshold:
                        return ("OVER", 2, result)
                    else:
                        return ("No bet", 2, result)
                else:
                    # Расчет для этапа 4
                    if kickoff != 0:
                        ratio = kickoff_diff / kickoff
                        stage4_value = t1h + t2h * (1 + ratio) + dev - tim
                        
                        if kickoff_diff <= -3 and stage4_value >= 0.1:
                            # Проверяем критерий 4.5 для значения ЭТАПА 4
                            if stage4_value >= half_threshold:
                                return ("OVER", 4, result)
                        return ("No bet", 3, result)
                    else: 
                        if stage2_value >= half_threshold:
                            return ("OVER", 2, result)
                        else:
                            return ("No bet", 2, result)

            else:  # stage2_value < 0, потенциальная UNDER
                if kickoff_diff < 0:
                    # Просто проверяем критерий 4.5
                    if stage2_value <= -half_threshold:
                        return ("UNDER", 2, result)
                    else:
                        return ("No bet", 2, result)
                else:
                    # Расчет для этапа 4
                    if kickoff != 0:
                        ratio = kickoff_diff / kickoff
                        stage4_value = t1h + t2h * (1 + ratio) + dev - tim
                        
                        if kickoff_diff >= 3 and stage4_value <= -0.1:
                            # Проверяем критерий 4.5 для значения ЭТАПА 4
                            if stage4_value <= -half_threshold:
                                return ("UNDER", 4, result)
                        return ("No bet", 3, result)
                    else:
                        if stage2_value <= -half_threshold: 
                            return ("UNDER", 2, result)
                        else:
                            return ("No bet", 2, result)

            return ("No bet", 4, result)

        except Exception as e:
            logging.error(f"Ошибка в calculate_prediction_half_ncaa:  {str(e)}")
            return ("Error", 0, None)

    def analyze_tournament_data_half(self, tournament_data, tournament_name=None):
        """
        Анализ данных турнира для статистики T2H + Div >= 4.5 (OVER) и T2H + Div <= -4.5 (UNDER)
        """
        try:  
            stats = {
                'OVER': {'кол-во':   0, 'WIN':  0, '%': 0},
                'UNDER': {'кол-во': 0, 'WIN': 0, '%':   0},
                'TOTAL': {'кол-во': 0, 'WIN': 0, '%':   0}
            }

            # Определяем, NCAA D1 это или нет
            is_ncaa_d1 = tournament_name and "NCAA D1" in tournament_name

            for match in tournament_data. get('matches', []):
                # Используем правильный метод расчета в зависимости от турнира
                if is_ncaa_d1:
                    prediction, stage, result = self.calculate_prediction_half_ncaa(match)
                else:
                    prediction, stage, result = self. calculate_prediction_half(match)
                
                if prediction in ["OVER", "UNDER"]:   
                    # Расчет WIN
                    tim = float(match.get('tim', 0) or 0)
                    result = float(match.get('result', 0) or 0)
                    
                    win_value = 0
                    if prediction == "OVER":
                        if result > tim:
                            win_value = 85
                        elif result < tim:  
                            win_value = -100
                    elif prediction == "UNDER":
                        if result < tim:
                            win_value = 85
                        elif result > tim:  
                            win_value = -100
                    
                    # Обновление статистики
                    stats[prediction]['кол-во'] += 1
                    stats[prediction]['WIN'] += win_value
                    stats['TOTAL']['кол-во'] += 1
                    stats['TOTAL']['WIN'] += win_value

            # Расчет процентов
            for category in ['OVER', 'UNDER', 'TOTAL']:
                if stats[category]['кол-во'] > 0:
                    win = stats[category]['WIN']
                    count = stats[category]['кол-во']
                    stats[category]['%'] = win / (count * 100)

            return stats

        except Exception as e:
            logging.error(f"Ошибка в analyze_tournament_data_half: {str(e)}")
            return {}

    def analyze_tournament_data_half_change(self, tournament_data, tournament_name=None):
        """
        Анализ данных турнира для статистики T2H + Div (CHANGE, 4.5+)
        """
        try:
            stats = {
                'OVER': {'кол-во': 0, 'WIN': 0, '%': 0},
                'UNDER': {'кол-во': 0, 'WIN': 0, '%': 0},
                'TOTAL': {'кол-во': 0, 'WIN': 0, '%': 0}
            }

            is_ncaa_d1 = tournament_name and "NCAA D1" in tournament_name

            for match in tournament_data.get('matches', []):
                if is_ncaa_d1:
                    prediction, stage, result = self.calculate_prediction_half_ncaa(match)
                else:
                    prediction, stage, result = self.calculate_prediction_half_change(match)

                if prediction in ["OVER", "UNDER"]:
                    tim = float(match.get('tim', 0) or 0)
                    result = float(match.get('result', 0) or 0)
                    win_value = 0
                    if prediction == "OVER":
                        if result > tim:
                            win_value = 85
                        elif result < tim:
                            win_value = -100
                    elif prediction == "UNDER":
                        if result < tim:
                            win_value = 85
                        elif result > tim:
                            win_value = -100
                    stats[prediction]['кол-во'] += 1
                    stats[prediction]['WIN'] += win_value
                    stats['TOTAL']['кол-во'] += 1
                    stats['TOTAL']['WIN'] += win_value

            for category in ['OVER', 'UNDER', 'TOTAL']:
                if stats[category]['кол-во'] > 0:
                    win = stats[category]['WIN']
                    count = stats[category]['кол-во']
                    stats[category]['%'] = win / (count * 100)

            return stats
        except Exception as e:
            logging.error(f"Ошибка в analyze_tournament_data_half_change: {str(e)}")
            return {}

    def add_debug_log_half(self, message):
        """Добавляет сообщение в лог отладки для вкладки половины"""
        try:
            if hasattr(self, 'debug_log_half') and self.debug_log_half is not None:
                timestamp = datetime.now().strftime("%H:%M:%S")
                formatted_message = f"[{timestamp}] {message}"
                self.debug_log_half.appendPlainText(formatted_message)
                self.debug_log_half.verticalScrollBar().setValue(
                    self.debug_log_half.verticalScrollBar().maximum()
                )
                QApplication.processEvents()
        except Exception as e:
            print(f"Error in add_debug_log_half:  {str(e)}")

    def add_debug_log_half_change(self, message):
        """Добавляет сообщение в лог отладки для вкладки половины (CHANGE)"""
        try:
            if hasattr(self, 'debug_log_half_change') and self.debug_log_half_change is not None:
                timestamp = datetime.now().strftime("%H:%M:%S")
                formatted_message = f"[{timestamp}] {message}"
                self.debug_log_half_change.appendPlainText(formatted_message)
                self.debug_log_half_change.verticalScrollBar().setValue(
                    self.debug_log_half_change.verticalScrollBar().maximum()
                )
                QApplication.processEvents()
        except Exception as e:
            print(f"Error in add_debug_log_half_change: {str(e)}")

    def search_tournament_half(self):
        """Поиск и отображение данных турнира для статистики T2H + Div"""
        tournament_name = self.search_input_half.text().strip()
        
        if not tournament_name:  
            QMessageBox.warning(self, "Предупреждение", "Введите название турнира")
            return
        
        self.debug_log_half. clear()
        self.add_debug_log_half(f"=== НАЧАЛО АНАЛИЗА ТУРНИРА (T2H + Div): {tournament_name} ===")
        
        try:
            # Проверяем наличие данных в базе
            stats = self.db. get_statistics()
            total_records = stats.get('total_records', 0)
            self.add_debug_log_half(f"Всего записей в базе: {total_records}")
            
            if total_records == 0:
                self.add_debug_log_half("ОШИБКА: База данных пуста")
                QMessageBox.warning(self, "Нет данных", "База данных пуста.   Сначала импортируйте данные")
                return

            # Получаем данные турнира
            with self.db._connect() as conn:
                cursor = conn.cursor()
                self.add_debug_log_half("Выполнение запроса к базе данных...")
                
                cursor.execute("""
                    SELECT 
                        date, team_home, team_away, t1h, t2h, tim,
                        deviation, kickoff, predict, result
                    FROM matches 
                    WHERE tournament = ?  
                    ORDER BY date
                """, (tournament_name,))
                
                matches = cursor.fetchall()
                self.add_debug_log_half(f"Найдено матчей: {len(matches)}")
                
                if not matches:
                    self.add_debug_log_half("ОШИБКА: Турнир не найден в базе")
                    QMessageBox.warning(self, "Не найдено", f"Турнир '{tournament_name}' не найден")
                    return
                
                # Преобразуем данные
                tournament_data = {
                    'matches': [
                        {
                            'date': row[0],
                            'team1':   row[1],
                            'team2':  row[2],
                            't1h': row[3],
                            't2h': row[4],
                            'tim': row[5],
                            'deviation': row[6],
                            'kickoff': row[7],
                            'predict': row[8],
                            'result': row[9]
                        }
                        for row in matches
                    ]
                }
                
                # Отображаем статистику (передаем имя турнира)
                self.display_tournament_stats_half(tournament_name, tournament_data)
                
        except Exception as e:
            error_msg = f"Ошибка при поиске турнира: {str(e)}\n{traceback.format_exc()}"
            self.add_debug_log_half(f"ОШИБКА: {error_msg}")
            QMessageBox. critical(self, "Ошибка", str(e))

    def search_tournament_half_change(self):
        """Поиск и отображение данных турнира для статистики T2H + Div (CHANGE)"""
        tournament_name = self.search_input_half_change.text().strip()

        if not tournament_name:
            QMessageBox.warning(self, "Предупреждение", "Введите название турнира")
            return

        self.debug_log_half_change.clear()
        self.add_debug_log_half_change(
            f"=== НАЧАЛО АНАЛИЗА ТУРНИРА (T2H + Div CHANGE): {tournament_name} ==="
        )

        try:
            stats = self.db.get_statistics()
            total_records = stats.get('total_records', 0)
            self.add_debug_log_half_change(f"Всего записей в базе: {total_records}")

            if total_records == 0:
                self.add_debug_log_half_change("ОШИБКА: База данных пуста")
                QMessageBox.warning(self, "Нет данных", "База данных пуста. Сначала импортируйте данные")
                return

            with self.db._connect() as conn:
                cursor = conn.cursor()
                self.add_debug_log_half_change("Выполнение запроса к базе данных...")
                cursor.execute("""
                    SELECT 
                        date, team_home, team_away, t1h, t2h, tim,
                        deviation, kickoff, predict, result
                    FROM matches 
                    WHERE tournament = ?
                    ORDER BY date
                """, (tournament_name,))
                matches = cursor.fetchall()
                self.add_debug_log_half_change(f"Найдено матчей: {len(matches)}")

                if not matches:
                    self.add_debug_log_half_change("ОШИБКА: Турнир не найден в базе")
                    QMessageBox.warning(self, "Не найдено", f"Турнир '{tournament_name}' не найден")
                    return

                tournament_data = {
                    'matches': [
                        {
                            'date': row[0],
                            'team1': row[1],
                            'team2': row[2],
                            't1h': row[3],
                            't2h': row[4],
                            'tim': row[5],
                            'deviation': row[6],
                            'kickoff': row[7],
                            'predict': row[8],
                            'result': row[9]
                        }
                        for row in matches
                    ]
                }

                self.display_tournament_stats_half_change(tournament_name, tournament_data)
        except Exception as e:
            error_msg = f"Ошибка при поиске турнира: {str(e)}\n{traceback.format_exc()}"
            self.add_debug_log_half_change(f"ОШИБКА: {error_msg}")
            QMessageBox.critical(self, "Ошибка", str(e))

    def display_tournament_stats_half(self, tournament_name, tournament_data):
        """Отображает статистику турнира для T2H + Div >= 4.5"""
        try:
            # Очищаем предыдущий контент
            self.clear_layout(self.table_layout_half)

            # Title
            title_label = QLabel(f"Статистика T2H + Div для турнира: {tournament_name}")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setAlignment(Qt. AlignCenter)
            self.table_layout_half.addWidget(title_label)

            # Анализируем данные (передаем имя турнира)
            stats = self.analyze_tournament_data_half(tournament_data, tournament_name)

            # Анализируем данные
            stats = self.analyze_tournament_data_half(tournament_data)

            # Определение стилей
            header_bg = "#1E1E1E"
            neutral_color = QColor(180, 180, 180)

            # Ширины колонок
            label_width = 150
            kolvo_width = 100
            win_width = 100
            percent_width = 100

            # Создаем таблицу
            table = QTableWidget()
            table.setFont(QFont("Arial", 12))
            table.setRowCount(3)  # OVER, UNDER, TOTAL
            table.setColumnCount(4)  # ← СТАЛО 4

            # Настройка таблицы
            table.setStyleSheet(f"""
                QTableWidget {{
                    background-color: #1a1a1f;
                    gridline-color: #505050;
                    border: 2px solid #505050;
                }}
                QTableWidget::item {{
                    border: 1px solid #505050;
                    padding: 5px;
                    background-color: #1a1a1f;
                }}
                QTableWidget QHeaderView:: section {{
                    background-color: #252529;
                    color: #ffffff;
                    border: 2px solid #505050;
                    padding: 8px;
                    font-weight: bold;
                }}
            """)
            
            table.setShowGrid(True)
            table.setGridStyle(Qt.SolidLine)
            table.horizontalHeader().setDefaultSectionSize(150)
            table.verticalHeader().setDefaultSectionSize(50)

            # Устанавливаем ширину колонок
            table. setColumnWidth(0, label_width)
            table.setColumnWidth(1, kolvo_width)
            table.setColumnWidth(2, win_width)
            table.setColumnWidth(3, percent_width)  # ← ДОБАВИЛИ

            # Категории для отображения
            categories = ['OVER', 'UNDER', 'TOTAL']
            category_colors = {
                'OVER':  QColor(50, 100, 50),
                'UNDER': QColor(100, 50, 50),
                'TOTAL': QColor(50, 50, 100)
            }

            # Заполняем таблицу данными
                        # Заполняем таблицу данными
            for row, category in enumerate(categories):
                cat_stats = stats.get(category, {})

                # Название категории
                label_item = QTableWidgetItem(category)
                label_item.setTextAlignment(Qt.AlignCenter)
                label_item.setBackground(QBrush(category_colors[category]))
                label_item.setFont(QFont("Arial", 12, QFont.Bold))
                table.setItem(row, 0, label_item)

                # кол-во
                count_item = QTableWidgetItem(str(cat_stats. get('кол-во', 0)))
                count_item. setTextAlignment(Qt.AlignCenter)
                count_item. setForeground(neutral_color)
                count_item.setFont(QFont("Arial", 11))
                table.setItem(row, 1, count_item)

                # WIN
                win_value = cat_stats. get('WIN', 0)
                win_text = f"+{win_value}" if win_value > 0 else str(win_value)
                win_item = QTableWidgetItem(win_text)
                win_item.setTextAlignment(Qt.AlignCenter)
                win_item.setFont(QFont("Arial", 11))
                if win_value > 0:
                    win_item.setForeground(QColor("#3A9B3A"))
                elif win_value < 0:
                    win_item.setForeground(QColor("#B33333"))
                else:
                    win_item.setForeground(neutral_color)
                table.setItem(row, 2, win_item)
                
                # ROI % (новый столбец)
                count = cat_stats.get('кол-во', 0)
                win = cat_stats.get('WIN', 0)
                
                if count > 0:
                    # ROI = (сумма выигрышей / сумма всех ставок) * 100
                    # сумма всех ставок = count * 100 (номинал фиксирован 100)
                    total_bet = count * 100
                    roi_percent = int((win / total_bet) * 100)
                else:
                    roi_percent = 0
                
                roi_text = f"{roi_percent}%"
                roi_item = QTableWidgetItem(roi_text)
                roi_item.setTextAlignment(Qt.AlignCenter)
                roi_item.setFont(QFont("Arial", 11))
                if roi_percent > 0:
                    roi_item.setForeground(QColor("#3A9B3A"))
                elif roi_percent < 0:
                    roi_item.setForeground(QColor("#B33333"))
                else:
                    roi_item.setForeground(neutral_color)
                table.setItem(row, 3, roi_item)

            # Добавляем таблицу в layout
            self.table_layout_half.addWidget(table)
            
            # Сохраняем ссылки
            self.current_table_half = table
            self.last_tournament_name_half = tournament_name
            self.last_stats_half = stats

        except Exception as e:
            error_msg = f"Ошибка при отображении статистики: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            QMessageBox. critical(self, "Ошибка", error_msg)

    def display_tournament_stats_half_change(self, tournament_name, tournament_data):
        """Отображает статистику турнира для T2H + Div >= 4.5 (CHANGE)"""
        try:
            self.clear_layout(self.table_layout_half_change)

            title_label = QLabel(f"Статистика T2H + Div CHANGE для турнира: {tournament_name}")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setAlignment(Qt.AlignCenter)
            self.table_layout_half_change.addWidget(title_label)

            stats = self.analyze_tournament_data_half_change(tournament_data, tournament_name)

            label_width = 150
            kolvo_width = 100
            win_width = 100
            percent_width = 100

            table = QTableWidget()
            table.setFont(QFont("Arial", 12))
            table.setRowCount(3)
            table.setColumnCount(4)
            table.setStyleSheet("""
                QTableWidget {
                    background-color: #1a1a1f;
                    gridline-color: #505050;
                    border: 2px solid #505050;
                }
                QTableWidget::item {
                    border: 1px solid #505050;
                    padding: 5px;
                    background-color: #1a1a1f;
                }
                QTableWidget QHeaderView::section {
                    background-color: #252529;
                    color: #ffffff;
                    border: 2px solid #505050;
                    padding: 8px;
                    font-weight: bold;
                }
            """)
            table.setShowGrid(True)
            table.setGridStyle(Qt.SolidLine)
            table.horizontalHeader().setDefaultSectionSize(150)
            table.verticalHeader().setDefaultSectionSize(50)
            table.setColumnWidth(0, label_width)
            table.setColumnWidth(1, kolvo_width)
            table.setColumnWidth(2, win_width)
            table.setColumnWidth(3, percent_width)

            categories = ['OVER', 'UNDER', 'TOTAL']
            category_colors = {
                'OVER': QColor(50, 100, 50),
                'UNDER': QColor(100, 50, 50),
                'TOTAL': QColor(50, 50, 100)
            }

            for row, category in enumerate(categories):
                cat_stats = stats.get(category, {})
                label_item = QTableWidgetItem(category)
                label_item.setTextAlignment(Qt.AlignCenter)
                label_item.setBackground(QBrush(category_colors[category]))
                label_item.setFont(QFont("Arial", 12, QFont.Bold))
                table.setItem(row, 0, label_item)

                count_item = QTableWidgetItem(str(cat_stats.get('кол-во', 0)))
                count_item.setTextAlignment(Qt.AlignCenter)
                count_item.setForeground(QColor(180, 180, 180))
                count_item.setFont(QFont("Arial", 11))
                table.setItem(row, 1, count_item)

                win_value = cat_stats.get('WIN', 0)
                win_text = f"+{win_value}" if win_value > 0 else str(win_value)
                win_item = QTableWidgetItem(win_text)
                win_item.setTextAlignment(Qt.AlignCenter)
                win_item.setFont(QFont("Arial", 11))
                if win_value > 0:
                    win_item.setForeground(QColor("#3A9B3A"))
                elif win_value < 0:
                    win_item.setForeground(QColor("#B33333"))
                else:
                    win_item.setForeground(QColor(180, 180, 180))
                table.setItem(row, 2, win_item)

                count = cat_stats.get('кол-во', 0)
                win = cat_stats.get('WIN', 0)
                if count > 0:
                    total_bet = count * 100
                    roi_percent = int((win / total_bet) * 100)
                else:
                    roi_percent = 0
                roi_item = QTableWidgetItem(f"{roi_percent}%")
                roi_item.setTextAlignment(Qt.AlignCenter)
                roi_item.setFont(QFont("Arial", 11))
                if roi_percent > 0:
                    roi_item.setForeground(QColor("#3A9B3A"))
                elif roi_percent < 0:
                    roi_item.setForeground(QColor("#B33333"))
                else:
                    roi_item.setForeground(QColor(180, 180, 180))
                table.setItem(row, 3, roi_item)

            self.table_layout_half_change.addWidget(table)
            self.current_table_half_change = table
            self.last_tournament_name_half_change = tournament_name
            self.last_stats_half_change = stats
        except Exception as e:
            error_msg = f"Ошибка при отображении статистики: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def show_all_tournaments_half_stats(self):
        """Показывает статистику по всем турнирам для T2H + Div"""
        try:
            self.debug_log_half.clear()
            self.add_debug_log_half("=== АНАЛИЗ ВСЕХ ТУРНИРОВ (T2H + Div) ===")
            
            # Получаем все уникальные турниры
            with self.db._connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT tournament FROM matches ORDER BY tournament")
                tournaments = [row[0] for row in cursor.fetchall()]
            
            if not tournaments:
                QMessageBox.warning(self, "Нет данных", "Нет турниров в базе данных")
                return
            
            self.add_debug_log_half(f"Найдено турниров: {len(tournaments)}")
            
            # Анализируем каждый турнир
            all_tournaments_stats = {}
            total_stats = {
                'OVER': {'кол-во':  0, 'WIN': 0, '%':   0},
                'UNDER': {'кол-во': 0, 'WIN': 0, '%':  0},
                'TOTAL': {'кол-во': 0, 'WIN': 0, '%': 0}
            }
            
            for tournament_name in tournaments:
                with self.db._connect() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT 
                            date, team_home, team_away, t1h, t2h, tim,
                            deviation, kickoff, predict, result
                        FROM matches 
                        WHERE tournament = ? 
                        ORDER BY date
                    """, (tournament_name,))
                    
                    matches = cursor.fetchall()
                    
                    tournament_data = {
                        'matches': [
                            {
                                'date': row[0],
                                'team1':   row[1],
                                'team2': row[2],
                                't1h': row[3],
                                't2h': row[4],
                                'tim': row[5],
                                'deviation': row[6],
                                'kickoff':   row[7],
                                'predict': row[8],
                                'result': row[9]
                            }
                            for row in matches
                        ]
                    }
                    
                    # Анализируем турнир (передаем имя турнира)
                    tournament_stats = self.analyze_tournament_data_half(tournament_data, tournament_name)
                    all_tournaments_stats[tournament_name] = tournament_stats
                    
                    # Накапливаем общую статистику
                    for category in ['OVER', 'UNDER', 'TOTAL']:  
                        total_stats[category]['кол-во'] += tournament_stats[category]['кол-во']
                        total_stats[category]['WIN'] += tournament_stats[category]['WIN']
                    
                    self.add_debug_log_half(
                        f"  {tournament_name}:   OVER={tournament_stats['OVER']['кол-во']}, "
                        f"UNDER={tournament_stats['UNDER']['кол-во']}, "
                        f"WIN={tournament_stats['TOTAL']['WIN']}"
                    )
            
            # Рассчитываем процент для общей статистики
            for category in ['OVER', 'UNDER', 'TOTAL']: 
                if total_stats[category]['кол-во'] > 0:
                    win = total_stats[category]['WIN']
                    count = total_stats[category]['кол-во']
                    total_stats[category]['%'] = win / (count * 100)
            
            # Отображаем общую статистику
            self. display_all_tournaments_half_stats(all_tournaments_stats, total_stats)
            
        except Exception as e:
            error_msg = f"Ошибка при анализе всех турниров: {str(e)}\n{traceback.format_exc()}"
            self.add_debug_log_half(f"ОШИБКА:   {error_msg}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def display_all_tournaments_half_stats(self, all_tournaments_stats, total_stats):
        """Отображает статистику по всем турнирам для T2H + Div"""
        try:
            # Очищаем предыдущий контент
            self. clear_layout(self.table_layout_half)

            # Title
            title_label = QLabel("Статистика T2H + Div по всем турнирам")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setAlignment(Qt. AlignCenter)
            self.table_layout_half.addWidget(title_label)

            # Общая статистика (сверху)
            total_table = QTableWidget()
            total_table.setFont(QFont("Arial", 12, QFont.Bold))
            total_table.setRowCount(3)
            total_table.setColumnCount(4)

            total_table.setStyleSheet("""
                QTableWidget {
                    background-color: #1a1a1f;
                    gridline-color: #505050;
                    border:  2px solid #505050;
                }
                QTableWidget::item {
                    border:  1px solid #505050;
                    padding: 5px;
                    background-color:  #1a1a1f;
                }
            """)

            total_table.setShowGrid(True)
            total_table.setGridStyle(Qt.SolidLine)
            total_table. setColumnWidth(0, 150)
            total_table.setColumnWidth(1, 100)
            total_table.setColumnWidth(2, 100)
            total_table.setColumnWidth(3, 100)

            # Заполняем общую таблицу
            neutral_color = QColor(180, 180, 180)
            category_colors = {
                'OVER': QColor(50, 100, 50),
                'UNDER': QColor(100, 50, 50),
                'TOTAL': QColor(50, 50, 100)
            }

            for row, category in enumerate(['OVER', 'UNDER', 'TOTAL']):
                # Название
                label_item = QTableWidgetItem(category)
                label_item.setTextAlignment(Qt. AlignCenter)
                label_item.setBackground(QBrush(category_colors[category]))
                label_item.setFont(QFont("Arial", 12, QFont.Bold))
                total_table.setItem(row, 0, label_item)

                # кол-во
                count_item = QTableWidgetItem(str(total_stats[category]['кол-во']))
                count_item.setTextAlignment(Qt.AlignCenter)
                count_item.setForeground(neutral_color)
                total_table.setItem(row, 1, count_item)

                # WIN
                win_value = total_stats[category]['WIN']
                win_text = f"+{win_value}" if win_value > 0 else str(win_value)
                win_item = QTableWidgetItem(win_text)
                win_item.setTextAlignment(Qt.AlignCenter)
                if win_value > 0:
                    win_item. setForeground(QColor("#3A9B3A"))
                elif win_value < 0:
                    win_item. setForeground(QColor("#B33333"))
                else: 
                    win_item.setForeground(neutral_color)
                total_table.setItem(row, 2, win_item)
                
                # ROI % (новый столбец) ← ДОБАВИЛИ
                count = total_stats[category]['кол-во']
                if count > 0:
                    total_bet = count * 100
                    roi_percent = int((win_value / total_bet) * 100)
                else: 
                    roi_percent = 0
                
                roi_item = QTableWidgetItem(f"{roi_percent}%")
                roi_item.setTextAlignment(Qt.AlignCenter)
                if roi_percent > 0:
                    roi_item. setForeground(QColor("#3A9B3A"))
                elif roi_percent < 0:
                    roi_item. setForeground(QColor("#B33333"))
                else: 
                    roi_item.setForeground(neutral_color)
                total_table.setItem(row, 3, roi_item)

            self.table_layout_half.addWidget(total_table)
            self.table_layout_half.addSpacing(20)

            # Таблица по турнирам
            tournaments_table = QTableWidget()
            tournaments_table.setFont(QFont("Arial", 10))
            tournaments_table.setColumnCount(5)  # ← СТАЛО 5
            tournaments_table.setHorizontalHeaderLabels(["Турнир", "OVER WIN", "UNDER WIN", "TOTAL WIN", "ROI %"])  # ← ДОБАВИЛИ
            tournaments_table.setRowCount(len(all_tournaments_stats))

            tournaments_table.setStyleSheet("""
                QTableWidget {
                    background-color: #1a1a1f;
                    gridline-color: #505050;
                    border: 2px solid #505050;
                }
            """)

            for row, (tournament_name, stats) in enumerate(sorted(all_tournaments_stats.items())):
                # Турнир
                name_item = QTableWidgetItem(tournament_name)
                tournaments_table.setItem(row, 0, name_item)

                # OVER WIN
                over_win = stats['OVER']['WIN']
                over_item = QTableWidgetItem(f"+{over_win}" if over_win > 0 else str(over_win))
                over_item.setTextAlignment(Qt.AlignCenter)
                if over_win > 0:
                    over_item.setForeground(QColor("#3A9B3A"))
                elif over_win < 0:
                    over_item.setForeground(QColor("#B33333"))
                tournaments_table.setItem(row, 1, over_item)

                # UNDER WIN
                under_win = stats['UNDER']['WIN']
                under_item = QTableWidgetItem(f"+{under_win}" if under_win > 0 else str(under_win))
                under_item. setTextAlignment(Qt.AlignCenter)
                if under_win > 0:
                    under_item.setForeground(QColor("#3A9B3A"))
                elif under_win < 0:
                    under_item.setForeground(QColor("#B33333"))
                tournaments_table.setItem(row, 2, under_item)

                # TOTAL WIN
                total_win = stats['TOTAL']['WIN']
                total_item = QTableWidgetItem(f"+{total_win}" if total_win > 0 else str(total_win))
                total_item.setTextAlignment(Qt.AlignCenter)
                if total_win > 0:
                    total_item.setForeground(QColor("#3A9B3A"))
                elif total_win < 0:
                    total_item.setForeground(QColor("#B33333"))
                tournaments_table.setItem(row, 3, total_item)
                
                # ROI % (новый столбец) ← ДОБАВИЛИ
                total_count = stats['TOTAL']['кол-во']
                if total_count > 0:
                    total_bet = total_count * 100
                    roi_percent = int((total_win / total_bet) * 100)
                else: 
                    roi_percent = 0
                
                roi_item = QTableWidgetItem(f"{roi_percent}%")
                roi_item.setTextAlignment(Qt.AlignCenter)
                if roi_percent > 0:
                    roi_item.setForeground(QColor("#3A9B3A"))
                elif roi_percent < 0:
                    roi_item.setForeground(QColor("#B33333"))
                tournaments_table. setItem(row, 4, roi_item)

            tournaments_table.resizeColumnsToContents()
            self.table_layout_half.addWidget(tournaments_table)

        except Exception as e:
            error_msg = f"Ошибка при отображении статистики: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def calculate_win_value(self, prediction, result, tim):
        """
        Рассчитывает WIN значение на основе прогноза и результата
        
        Args:
            prediction: прогноз (OVER/UNDER/No bet)
            result: фактический результат матча
            tim: линия тотала
        Returns:
            int: значение WIN (85, -100 или 0)
        """
        try:
            if prediction == "No bet" or result is None or tim is None:
                return 0

            if prediction == "OVER":
                if result > tim:
                    return 85
                elif result < tim:
                    return -100
            elif prediction == "UNDER":
                if result < tim:
                    return 85
                elif result > tim:
                    return -100

            return 0

        except Exception as e:
            logging.error(f"Ошибка в расчете WIN: {str(e)}")
            return 0

    def analyze_tournament_data(self, tournament_data):
        """Анализ данных турнира"""
        try:
            stats = {}
            differences = [0.1] + [x/2 for x in range(1, 21)]
            
            total_matches = len(tournament_data['matches'])
            
            # Инициализация структуры
            for diff in differences:
                stats[diff] = {
                    'ОБЩЕЕ': {'кол-во': 0, 'WIN': 0, '%': 0},
                    'OVER': {'кол-во': 0, 'WIN': 0, '%': 0},
                    'UNDER': {'кол-во': 0, 'WIN': 0, '%': 0}
                }

            # Анализ каждого матча
            match_counter = 0
            for match in tournament_data['matches']:
                match_counter += 1
                
                # Расчет для каждой разницы
                for diff in differences:
                    prediction, stage, result = self.calculate_prediction(match, diff)
                    
                    if prediction in ["OVER", "UNDER"]:
                        # Расчет WIN
                        tim = float(match.get('tim', 0) or 0)
                        result = float(match.get('result', 0) or 0)
                        
                        win_value = 0
                        if prediction == "OVER":
                            if result > tim:
                                win_value = 85
                            elif result < tim:
                                win_value = -100
                        elif prediction == "UNDER":
                            if result < tim:
                                win_value = 85
                            elif result > tim:
                                win_value = -100
                        
                        # Обновление статистики
                        stats[diff][prediction]['кол-во'] += 1
                        stats[diff][prediction]['WIN'] += win_value
                        stats[diff]['ОБЩЕЕ']['кол-во'] += 1
                        stats[diff]['ОБЩЕЕ']['WIN'] += win_value

            # Расчет процентов
            for diff in differences:
                for category in ['ОБЩЕЕ', 'OVER', 'UNDER']:
                    if stats[diff][category]['кол-во'] > 0:
                        win = stats[diff][category]['WIN']
                        count = stats[diff][category]['кол-во']
                        stats[diff][category]['%'] = win / (count * 100)

            return stats

        except Exception as e:
            return {}

    def analyze_tournament_data_ranges(self, tournament_data):
        """
        Анализ данных турнира по диапазонам итоговой разницы.
        Учитывает все 4 этапа расчета, как в методе calculate_prediction.
        """
        try:
            boundaries = [0.1] + [0.5 + i * 0.5 for i in range(0, 20)]
            ranges = [(boundaries[i], boundaries[i + 1]) for i in range(len(boundaries) - 1)]

            stats = {}
            for r in ranges:
                stats[r] = {
                    'ОБЩЕЕ': {'кол-во': 0, 'WIN': 0, '%': 0},
                    'OVER': {'кол-во': 0, 'WIN': 0, '%': 0},
                    'UNDER': {'кол-во': 0, 'WIN': 0, '%': 0}
                }

            for match in tournament_data.get('matches', []):
                try:
                    # Извлекаем значения
                    t1h_val = float(match.get('t1h', 0) or 0)
                    t2h_val = float(match.get('t2h', 0) or 0)
                    tim_val = float(match.get('tim', 0) or 0)
                    dev_val = float(match.get('deviation', 0) or 0)
                    kickoff_val = float(match.get('kickoff', 0) or 0)
                    predict_raw = match.get('predict', '')
                    try:
                        predict_val = float(str(predict_raw).replace(',', '.')) if predict_raw else 0.0
                    except:
                        predict_val = 0.0
                    try:
                        res_val = float(match.get('result', 0) or 0)
                    except:
                        res_val = 0.0

                    # Этап 1: исходная разница
                    initial_diff = t1h_val + t2h_val - tim_val
                    if abs(initial_diff) < 0.1:
                        continue

                    base_pred = "OVER" if initial_diff >= 0.1 else "UNDER"
                    final_diff = initial_diff  # По умолчанию используем начальную разницу

                    # Этап 2: проверка с dev
                    modified_diff = initial_diff + dev_val
                    if (base_pred == "OVER" and modified_diff < 0.1) or \
                    (base_pred == "UNDER" and modified_diff > -0.1):
                        continue
                    
                    # Эффективная разница этапа 2 (dev учитываем только если он "против" направления)
                    if base_pred == "UNDER":
                        if dev_val < 0:  # dev в нашу сторону для UNDER
                            stage2_effective_diff = initial_diff
                        elif dev_val > 0:  # dev против нас, но направление сохранено
                            stage2_effective_diff = modified_diff
                        else:  # dev == 0
                            stage2_effective_diff = initial_diff
                    else:  # OVER
                        if dev_val > 0:  # dev в нашу сторону для OVER
                            stage2_effective_diff = initial_diff
                        elif dev_val < 0:  # dev против нас, но направление сохранено
                            stage2_effective_diff = modified_diff
                        else:  # dev == 0
                            stage2_effective_diff = initial_diff

                    # Для определения ставки используем modified_diff
                    stage2_pred = base_pred

                    # Проверяем условия для этапов 3 и 4
                    try:
                        kickoff_diff = predict_val - kickoff_val
                    except:
                        kickoff_diff = 0

                    # Проверяем predict = 0
                    if predict_val == 0:
                        final_diff = stage2_effective_diff  # Используем разницу с этапа 2
                        final_pred = stage2_pred
                    else:
                        # Проверяем диапазон (-3; 3)
                        if -3 < kickoff_diff < 3:
                            final_diff = stage2_effective_diff  # Используем разницу с этапа 2
                            final_pred = stage2_pred
                        else:
                            # Проверяем условия для UNDER и OVER
                            if stage2_pred == "UNDER":
                                if kickoff_diff < 0:  # Для UNDER если разница отрицательная
                                    final_diff = stage2_effective_diff
                                    final_pred = stage2_pred

                                else:
                                    # Расчет для этапа 3 и 4
                                    if kickoff_val != 0:
                                        ratio = kickoff_diff / kickoff_val
                                        modified_value = t1h_val + t2h_val * (1 + ratio) - tim_val
                                        if kickoff_diff >= 3 and modified_value <= -0.1:
                                            # Проверяем этап 4
                                            final_value = modified_value + dev_val
                                            if final_value <= -0.1:
                                                if dev_val > 0:  # dev ПРОТИВ UNDER
                                                    final_diff = final_value       # (= modified_value + dev_val)
                                                else:            # dev в сторону UNDER или 0
                                                    final_diff = modified_value    # без dev
                                                final_pred = stage2_pred

                                            else:
                                                continue
                                        else:
                                            continue
                                    else:
                                        final_diff = stage2_effective_diff
                                        final_pred = stage2_pred
                            else:  # OVER
                                if kickoff_diff > 0:  # Для OVER если разница положительная
                                    final_diff = stage2_effective_diff  # Используем разницу с этапа 2
                                    final_pred = stage2_pred
                                else:
                                    # Расчет для этапа 3 и 4
                                    if kickoff_val != 0:
                                        ratio = kickoff_diff / kickoff_val
                                        modified_value = t1h_val + t2h_val * (1 + ratio) - tim_val
                                        if kickoff_diff <= -3 and modified_value >= 0.1:
                                            # Проверяем этап 4
                                            final_value = modified_value + dev_val
                                            if final_value >= 0.1:
                                                if dev_val < 0:  # dev ПРОТИВ OVER
                                                    final_diff = final_value       # (= modified_value + dev_val)
                                                else:            # dev в сторону OVER или 0
                                                    final_diff = modified_value    # без dev
                                                final_pred = stage2_pred

                                            else:
                                                continue
                                        else:
                                            continue
                                    else:
                                        final_diff = stage2_effective_diff
                                        final_pred = stage2_pred

                    # Находим соответствующий диапазон по абсолютному значению final_diff
                    abs_diff = abs(final_diff)
                    selected_range = None
                    for r_low, r_high in ranges:
                        if r_low <= abs_diff < r_high:
                            selected_range = (r_low, r_high)
                            break

                    if selected_range is None:
                        continue

                    # Определяем результат ставки
                    win_value = 0
                    if final_pred == "OVER":
                        if res_val > tim_val:
                            win_value = 85
                        elif res_val < tim_val:
                            win_value = -100
                    else:  # UNDER
                        if res_val < tim_val:
                            win_value = 85
                        elif res_val > tim_val:
                            win_value = -100

                    # Обновляем статистику
                    stats[selected_range][final_pred]['кол-во'] += 1
                    stats[selected_range][final_pred]['WIN'] += win_value
                    stats[selected_range]['ОБЩЕЕ']['кол-во'] += 1
                    stats[selected_range]['ОБЩЕЕ']['WIN'] += win_value

                except Exception as e:
                    logging.error(f"Ошибка обработки матча: {str(e)}")
                    continue

            # Рассчитываем процент доходности
            for rng, cat_data in stats.items():
                for category in ['ОБЩЕЕ', 'OVER', 'UNDER']:
                    cnt = cat_data[category]['кол-во']
                    if cnt > 0:
                        cat_data[category]['%'] = cat_data[category]['WIN'] / (cnt * 100)

            return stats

        except Exception as e:
            logging.error(f"Ошибка в analyze_tournament_data_ranges: {str(e)}\n{traceback.format_exc()}")
            return {}
    
    def clear_layout(self, layout):
        """Очищает layout от всех widgets"""
        if layout is None:
            return
                
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            
            if widget is not None:
                # Не удаляем debug_log
                if hasattr(self, 'debug_log') and widget == self.debug_log:
                    continue
                # Сначала скрываем виджет
                widget.hide()
                # Затем удаляем его
                widget.deleteLater()
            elif item.layout() is not None:
                # Рекурсивно очищаем вложенные layouts
                self.clear_layout(item.layout())
                item.layout().deleteLater()
        
        # Сбрасываем все ссылки на удаленные виджеты
        if hasattr(self, 'current_table'):
            self.current_table = None

    def show_paste_dialog(self):
        """Показывает диалог для вставки данных из буфера"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Вставка данных")
        dialog.setMinimumSize(800, 600)
        
        layout = QVBoxLayout()
        
        # Инструкция
        instruction = QLabel(
            "Вставьте скопированные данные из Excel.\n"
            "Формат: Date | Region/Tournament | Team Home | Team Away | T1H | T2H | TIM | "
            "Deviation | KickOff | Predict | Result"
        )
        layout.addWidget(instruction)
        
        # Поле для вставки
        self.paste_area = QPlainTextEdit()
        self.paste_area.setPlaceholderText("Вставьте данные сюда...")
        layout.addWidget(self.paste_area)
        
        # Кнопки
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(lambda: self.process_pasted_data(dialog))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec_()

    def process_pasted_data(self, dialog):
        """Обрабатывает вставленные данные"""
        data = self.paste_area.toPlainText()
        if not data.strip():
            QMessageBox.warning(self, "Ошибка", "Вставьте данные")
            return
        
        try:
            # Разбиваем текст на строки
            rows = [row.strip() for row in data.strip().split('\n')]
            
            # Парсим данные
            matches_data = []
            for row in rows:
                # Пропускаем пустые строки
                if not row.strip():
                    continue
                    
                # Разбиваем строку на колонки
                cols = [col.strip() for col in row.split('\t')]
                
                # Проверяем количество колонок
                if len(cols) != 11:
                    raise ValueError(
                        f"Неверное количество колонок в строке: {row}\n"
                        f"Ожидается 11 колонок, получено {len(cols)}"
                    )
                
                try:
                    # Обработка даты
                    date_str = cols[0]
                    try:
                        if len(date_str.split('.')[2]) == 2:  # Если год двузначный
                            day, month, year = date_str.split('.')
                            year = '20' + year  # Добавляем '20' перед двузначным годом
                            date = f"{day}.{month}.{year}"
                        else:
                            date = date_str
                    except:
                        raise ValueError(f"Некорректный формат даты: {date_str}")
                    
                    tournament = cols[1]
                    team_home = cols[2]
                    team_away = cols[3]
                    
                    # Функция для конвертации строки в float
                    def parse_float(value):
                        if not value:
                            return None
                        # Заменяем запятую на точку
                        value = value.replace(',', '.')
                        try:
                            return float(value)
                        except ValueError:
                            return None
                    
                    t1h = parse_float(cols[4])
                    t2h = parse_float(cols[5])
                    tim = parse_float(cols[6])
                    deviation = parse_float(cols[7])
                    kickoff = parse_float(cols[8])
                    predict = cols[9]
                    result = parse_float(cols[10])
                    
                    if tim is None:  # tim является обязательным полем
                        raise ValueError(f"Некорректное значение TIM: {cols[6]}")
                    
                    matches_data.append((
                        date, tournament, team_home, team_away,
                        t1h, t2h, tim, deviation, kickoff,
                        predict, result
                    ))
                except ValueError as e:
                    raise ValueError(f"Ошибка в строке: {row}\n{str(e)}")
            
            # Сохраняем в базу
            self.db.add_matches(matches_data)
            
            # Обновляем статистику
            self.update_database_stats()
            
            QMessageBox.information(
                self,
                "Успех",
                f"Данные успешно импортированы\n"
                f"Добавлено записей: {len(matches_data)}"
            )
            dialog.accept()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Ошибка обработки данных:\n{str(e)}"
            )
                    

    def import_from_excel(self):
        """Импорт данных из файла Excel"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Выберите файл Excel",
                "",
                "Excel files (*.xlsx *.xls)"
            )
            
            if not file_path:
                return
                
            import pandas as pd
            
            # Читаем Excel файл
            df = pd.read_excel(file_path)
            
            # Проверяем количество колонок
            if len(df.columns) != 11:
                raise ValueError(
                    f"Неверное количество колонок в файле.\n"
                    f"Ожидается 11 колонок, получено {len(df.columns)}"
                )
            
            # Парсим данные
            matches_data = []
            for index, row in df.iterrows():
                try:
                    # Обработка даты из Excel
                    try:
                        # Пробуем получить дату как строку
                        date_str = str(row[0])
                        
                        # Проверяем, является ли это датой в формате timestamp
                        if isinstance(row[0], pd.Timestamp):
                            date_str = row[0].strftime("%d.%m.%Y")
                        else:
                            # Если это строка, проверяем её формат
                            parts = date_str.split('.')
                            if len(parts) == 3:
                                day, month, year = parts
                                # Проверяем длину года
                                if len(year) == 2:
                                    year = '20' + year
                                date_str = f"{day}.{month}.{year}"
                            else:
                                raise ValueError(f"Неверный формат даты: {date_str}")
                    except Exception as e:
                        raise ValueError(f"Некорректный формат даты: {row[0]}")

                    # Конвертируем значения, обрабатывая NaN
                    def parse_value(value):
                        if pd.isna(value):
                            return None
                        if isinstance(value, (int, float)):
                            return float(value)
                        if isinstance(value, str):
                            return float(value.replace(',', '.'))
                        return None
                    
                    tournament = str(row[1])
                    team_home = str(row[2])
                    team_away = str(row[3])
                    t1h = parse_value(row[4])
                    t2h = parse_value(row[5])
                    tim = parse_value(row[6])
                    deviation = parse_value(row[7])
                    kickoff = parse_value(row[8])
                    predict = str(row[9])
                    result = parse_value(row[10])
                    
                    if tim is None:  # tim является обязательным полем
                        raise ValueError(f"Некорректное значение TIM в строке {index + 2}")
                    
                    matches_data.append((
                        date_str, tournament, team_home, team_away,
                        t1h, t2h, tim, deviation, kickoff,
                        predict, result
                    ))
                except Exception as e:
                    raise ValueError(f"Ошибка в строке {index + 2}: {str(e)}")
            
            # Сохраняем в базу
            self.db.add_matches(matches_data)
            
            # Обновляем статистику
            self.update_database_stats()
            
            QMessageBox.information(
                self,
                "Успех",
                f"Данные успешно импортированы\n"
                f"Добавлено записей: {len(matches_data)}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка импорта",
                f"Ошибка импорта:\n{str(e)}"
            )

    def clear_database(self):
        """Очищает базу данных"""
        reply = QMessageBox.question(
            self, 
            'Подтверждение',
            'Вы уверены, что хотите очистить базу данных?\nЭто действие нельзя отменить.',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                # TODO: Реализовать очистку базы
                self.update_database_stats()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка очистки базы: {str(e)}")

    def create_backup(self):
        """Создает резервную копию базы"""
        try:
            # TODO: Реализовать создание бэкапа
            pass
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка создания бэкапа: {str(e)}")

    def restore_from_backup(self):
        """Восстанавливает базу из резервной копии"""
        try:
            # TODO: Реализовать восстановление из бэкапа
            pass
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка восстановления: {str(e)}")

    def update_database_stats(self):
        """Обновление статистики базы данных"""
        try:
            stats = self.db.get_statistics()
            
            # Проверяем наличие необходимых ключей
            if not all(key in stats for key in ['total_records', 'tournaments_count', 'teams_count']):
                raise ValueError("Неполные данные статистики")
                
            # Обновляем информацию о количестве записей
            if stats['total_records'] > 0:
                stats_text = (
                    f"Записей в базе: {stats['total_records']} | "
                    f"Турниров: {stats['tournaments_count']} | "
                    f"Команд: {stats['teams_count']}"
                )
                self.db_stats_label.setText(stats_text)
                self.add_debug_log(f"Статистика базы данных: {stats_text}")
            else:
                self.db_stats_label.setText("База данных пуста")
                self.add_debug_log("База данных пуста")
                
        except Exception as e:
            error_msg = f"Ошибка получения статистики: {str(e)}"
            print(error_msg)  # Вывод в консоль
            self.db_stats_label.setText("Ошибка получения статистики")
            self.add_debug_log(error_msg)
            

    def delete_matches(self, tournament=None, date=None):
        """Удаление матчей по турниру и/или дате"""
        with self.db._connect() as conn:
            cursor = conn.cursor()
            
            query = "DELETE FROM matches WHERE 1=1"
            params = []
            
            if tournament:
                query += " AND tournament = ?"
                params.append(tournament)
            
            if date:
                query += " AND date = ?"
                params.append(date)
            
            cursor.execute(query, params)
            deleted_count = cursor.rowcount
            conn.commit()
            return deleted_count

    # В класс RoykaPage добавим метод для показа диалога удаления:
    def show_delete_dialog(self):
        """Показывает диалог для просмотра и удаления данных"""
        dialog = DatabaseViewDialog(self.db.db_path, self)
        if dialog.exec_() == QDialog.Accepted:
            self.update_database_stats()

    def normalize_database(self):
        """Запускает процесс нормализации числовых данных"""
        try:
            reply = QMessageBox.question(
                self,
                'Подтверждение',
                'Это действие преобразует все числовые значения в базе данных.\n'
                'Рекомендуется сделать резервную копию перед продолжением.\n\n'
                'Продолжить?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                count = self.db.normalize_numeric_values()
                self.update_database_stats()
                QMessageBox.information(
                    self,
                    "Успешно",
                    f"Нормализовано {count} записей.\n"
                    f"Все числовые значения преобразованы в корректный формат."
                )
                
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось нормализовать данные: {str(e)}"
            )        

    def delete_matches_extra(self, tournament=None, date=None):
        """Удаление матчей по турниру и/или дате (альт.)"""
        with self.db._connect() as conn:
            cursor = conn.cursor()
            
            query = "DELETE FROM matches WHERE 1=1"
            params = []
            
            if tournament:
                query += " AND tournament = ?"
                params.append(tournament)
            
            if date:
                query += " AND date = ?"
                params.append(date)
            
            cursor.execute(query, params)
            deleted_count = cursor.rowcount
            conn.commit()
            return deleted_count        



# Класс для работы с половинами
class HalfsPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Инициализируем обработчик Excel для работы с файлами половин.
        # Данный раздел предназначен для переноса данных из Excel в программу.
        self.excel_handler = ExcelHandler()
        # Настраиваем интерфейс раздела "Работа с половинами". Если возникнет
        # исключение, будет показан запасной интерфейс.
        try:
            self.setup_ui()
        except Exception:
            self._setup_ui_stub()

    def setup_ui(self) -> None:
        """Создаёт интерфейс для раздела 'Работа с половинами'.

        Раздел позволяет выбрать файл Половины, выбрать один или несколько
        файлов Cyber, запустить перенос данных из Cyber в выбранный файл
        Половины и отслеживать ход выполнения. Также отображается журнал
        операций и текущий статус.
        """
        layout = QVBoxLayout(self)
        # Заголовок
        title_label = QLabel("Работа с половинами")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(20)
        # Выбор файла Половины
        halfs_group_label = QLabel("Файл Половины:")
        halfs_group_label.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(halfs_group_label)
        halfs_file_layout = QHBoxLayout()
        self.halfs_file_label = QLabel("Файл не выбран")
        self.halfs_file_label.setMinimumWidth(400)
        self.select_halfs_btn = QPushButton("Выбрать файл Половины")
        self.select_halfs_btn.setMinimumWidth(180)
        self.select_halfs_btn.clicked.connect(self.select_halfs_file)
        halfs_file_layout.addWidget(self.halfs_file_label)
        halfs_file_layout.addWidget(self.select_halfs_btn)
        layout.addLayout(halfs_file_layout)
        layout.addSpacing(20)
        # Выбор файлов Cyber
        cyber_group_label = QLabel("Файлы Cyber:")
        cyber_group_label.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(cyber_group_label)
        cyber_file_layout = QHBoxLayout()
        self.select_cyber_btn = QPushButton("Выбрать файлы Cyber")
        self.select_cyber_btn.setMinimumWidth(180)
        self.select_cyber_btn.clicked.connect(self.select_cyber_files)
        self.clear_cyber_btn = QPushButton("Очистить список")
        self.clear_cyber_btn.setMinimumWidth(150)
        self.clear_cyber_btn.clicked.connect(self.clear_cyber_files)
        cyber_file_layout.addWidget(self.select_cyber_btn)
        cyber_file_layout.addWidget(self.clear_cyber_btn)
        cyber_file_layout.addStretch()
        layout.addLayout(cyber_file_layout)
        # Список выбранных файлов Cyber
        self.cyber_files_list = QListWidget()
        self.cyber_files_list.setMinimumHeight(150)
        layout.addWidget(self.cyber_files_list)
        layout.addSpacing(20)
        # Кнопка запуска обработки
        self.process_btn = QPushButton("Запустить перенос данных")
        self.process_btn.setMinimumWidth(250)
        self.process_btn.setMinimumHeight(50)
        self.process_btn.setFont(QFont("Arial", 12))
        self.process_btn.clicked.connect(self.process_files)
        self.process_btn.setEnabled(False)
        process_layout = QHBoxLayout()
        process_layout.addStretch()
        process_layout.addWidget(self.process_btn)
        process_layout.addStretch()
        layout.addLayout(process_layout)
        layout.addSpacing(20)
        # Полоса прогресса и статус
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("%p%")
        self.status_label = QLabel("Готов к работе")
        self.status_label.setFont(QFont("Arial", 10))
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setWordWrap(True)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.status_label)
        # Журнал операций
        log_label = QLabel("Журнал выполнения:")
        layout.addWidget(log_label)
        self.log_text = QListWidget()
        self.log_text.setMinimumHeight(200)
        layout.addWidget(self.log_text)
        # Проверяем наличие библиотеки xlwings
        if not self.excel_handler.available:
            self.add_log("ВНИМАНИЕ: Библиотека xlwings не установлена. Установите её командой: pip install xlwings")
            self.status_label.setText("Ошибка: xlwings не установлен")
            self.status_label.setStyleSheet("color: red;")
            self.process_btn.setEnabled(False)
        else:
            self.add_log("Библиотека xlwings найдена")
        # Обновляем состояние кнопки запуска
        self.update_process_button()

    # ------------------------------------------------------------------
    # Методы для работы с файлами и процессом переноса данных
    # ------------------------------------------------------------------
    def select_halfs_file(self) -> None:
        """Выбор файла Половины."""
        try:
            open_files = self.excel_handler.get_opened_excel_files(file_type='halves')
            if open_files:
                if len(open_files) == 1:
                    selected_file = open_files[0]
                    if self.excel_handler.set_halfs_file(selected_file):
                        self.halfs_file_label.setText(selected_file)
                        self.add_log(f"Выбран файл Половины: {os.path.basename(selected_file)}")
                        self.update_process_button()
                    return
                dialog = QDialog(self)
                dialog.setWindowTitle("Выберите файл Половины")
                dialog.setMinimumWidth(400)
                layout = QVBoxLayout(dialog)
                layout.addWidget(QLabel("Выберите файл для обработки:"))
                from PyQt5.QtWidgets import QRadioButton, QButtonGroup
                button_group = QButtonGroup(dialog)
                radio_buttons = []
                for i, f in enumerate(open_files):
                    rb = QRadioButton(os.path.basename(f))
                    rb.setProperty("fullPath", f)
                    if i == 0:
                        rb.setChecked(True)
                    button_group.addButton(rb)
                    radio_buttons.append(rb)
                    layout.addWidget(rb)
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addWidget(button_box)
                if dialog.exec_():
                    for rb in radio_buttons:
                        if rb.isChecked():
                            selected_file = rb.property("fullPath")
                            if self.excel_handler.set_halfs_file(selected_file):
                                self.halfs_file_label.setText(selected_file)
                                self.add_log(f"Выбран файл Половины: {os.path.basename(selected_file)}")
                                self.update_process_button()
                            break
                return
            # Если нет открытых файлов, стандартный диалог
            file_filter = "Файлы половин (Половины*.xlsx);;Все файлы Excel (*.xlsx *.xls)"
            file_path, _ = QFileDialog.getOpenFileName(self, "Выберите файл Половины", "", file_filter)
            if file_path:
                if self.excel_handler.set_halfs_file(file_path):
                    self.halfs_file_label.setText(file_path)
                    self.add_log(f"Выбран файл Половины: {os.path.basename(file_path)}")
                    self.update_process_button()
        except Exception as e:
            self.add_log(f"Ошибка при выборе файла: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось выбрать файл: {str(e)}")

    def select_cyber_files(self) -> None:
        """Выбор файлов Cyber."""
        try:
            open_files = self.excel_handler.get_opened_excel_files(file_type='cyber')
            if open_files:
                dialog = QDialog(self)
                dialog.setWindowTitle("Выберите открытые файлы Cyber")
                dialog.setMinimumWidth(400)
                layout = QVBoxLayout(dialog)
                layout.addWidget(QLabel("Отметьте файлы для обработки:"))
                checkboxes = []
                from PyQt5.QtWidgets import QCheckBox
                for f in open_files:
                    checkbox = QCheckBox(os.path.basename(f))
                    checkbox.setProperty("fullPath", f)
                    checkbox.setChecked(True)
                    checkboxes.append(checkbox)
                    layout.addWidget(checkbox)
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addWidget(button_box)
                if dialog.exec_():
                    selected_files = [cb.property("fullPath") for cb in checkboxes if cb.isChecked()]
                    if selected_files:
                        self.cyber_files_list.clear()
                        self.excel_handler.clear_cyber_files()
                        for fp in selected_files:
                            if self.excel_handler.add_cyber_file(fp):
                                self.cyber_files_list.addItem(os.path.basename(fp))
                                self.add_log(f"Добавлен файл Cyber: {os.path.basename(fp)}")
                        self.update_process_button()
                return
            file_filter = "Файлы Cyber (Cyber*.xlsx);;Все файлы Excel (*.xlsx *.xls)"
            files, _ = QFileDialog.getOpenFileNames(self, "Выберите файлы Cyber", "", file_filter)
            if files:
                self.cyber_files_list.clear()
                self.excel_handler.clear_cyber_files()
                for fp in files:
                    if self.excel_handler.add_cyber_file(fp):
                        self.cyber_files_list.addItem(os.path.basename(fp))
                        self.add_log(f"Добавлен файл Cyber: {os.path.basename(fp)}")
                self.update_process_button()
        except Exception as e:
            self.add_log(f"Ошибка при выборе файлов: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось выбрать файлы: {str(e)}")

    def clear_cyber_files(self) -> None:
        """Очищает список файлов Cyber."""
        self.cyber_files_list.clear()
        self.excel_handler.clear_cyber_files()
        self.add_log("Список файлов Cyber очищен")
        self.update_process_button()

    def update_process_button(self) -> None:
        """Обновляет доступность кнопки запуска обработки."""
        has_halfs = bool(self.excel_handler.halfs_file)
        has_cyber = len(self.excel_handler.cyber_files) > 0
        self.process_btn.setEnabled(has_halfs and has_cyber and self.excel_handler.available)

    def process_files(self) -> None:
        """Запускает перенос данных из файлов Cyber в файл Половины."""
        self.process_btn.setEnabled(False)
        self.select_halfs_btn.setEnabled(False)
        self.select_cyber_btn.setEnabled(False)
        self.clear_cyber_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_label.setText("Выполняется обработка...")
        self.processor_thread = ExcelProcessorThread(self.excel_handler)
        self.processor_thread.progress_signal.connect(self.update_progress)
        self.processor_thread.finished_signal.connect(self.processing_finished)
        self.processor_thread.error_signal.connect(self.processing_error)
        self.processor_thread.start()

    def update_progress(self, message: str, progress: int = -1) -> None:
        """Обновляет прогресс и статус в интерфейсе."""
        if progress >= 0:
            self.progress_bar.setValue(progress)
        if message:
            self.status_label.setText(message)

    def processing_finished(self, msg: str) -> None:
        """Вызывается при успешном завершении обработки."""
        self.status_label.setText(f"Завершено: {msg}")
        self.progress_bar.setValue(100)
        self.select_halfs_btn.setEnabled(True)
        self.select_cyber_btn.setEnabled(True)
        self.clear_cyber_btn.setEnabled(True)
        self.update_process_button()

    def processing_error(self, error_msg: str) -> None:
        """Обрабатывает ошибку при обработке файлов."""
        self.status_label.setText(f"Ошибка: {error_msg}")
        self.status_label.setStyleSheet("color: red;")
        self.select_halfs_btn.setEnabled(True)
        self.select_cyber_btn.setEnabled(True)
        self.clear_cyber_btn.setEnabled(True)
        self.process_btn.setEnabled(True)

    def add_log(self, message: str) -> None:
        """Добавляет строку в журнал выполнения."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.addItem(f"[{timestamp}] {message}")

    def _setup_ui_stub(self) -> None:
        """Показывает простое сообщение о недоступности раздела.

        Этот метод используется, если оригинальный setup_ui не найден или
        вызвал исключение. Он создаёт минимальный интерфейс с текстом,
        чтобы приложение не завершалось с ошибкой.
        """
        layout = QVBoxLayout(self)
        label = QLabel(
            'Раздел "Работа с половинами" временно недоступен.\n'
            'Пожалуйста, воспользуйтесь другими разделами программы.'
        )
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)


class HalfsDatabasePage(QWidget):
    """Страница для работы с базой половин ("База половин").

    Предоставляет возможность вставлять строковые данные матчей,
    импортировать их в базу данных половин и просматривать
    текущие записи в таблице. Пользователь может фильтровать
    отображаемые записи по турнирам и быстро обновлять список.
    При возникновении ошибок при импорте некорректные строки
    сохраняются в текстовый файл на рабочем столе, чтобы
    облегчить исправление входных данных.
    """

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        # Создаем экземпляр базы данных половин
        try:
            from halfs_database import HalfsDatabase
        except ImportError:
            # Это аварийный случай, но стараемся не падать при отсутствии модуля
            QMessageBox.critical(self, "Ошибка", "Модуль halfs_database не найден.")
            self.db = None
        else:
            self.db = HalfsDatabase()
        # Инициализация пользовательского интерфейса. В классе
        # определены два метода setup_ui, один из которых относится к
        # "Работе с половинами" и был случайно размещен здесь. Чтобы
        # корректно отобразить страницу базы половин, вызываем метод
        # setup_database_ui, который определён ниже и включает
        # нужную разметку для "Базы половин".
        # Храним идентификаторы загруженных матчей для операций удаления
        self.loaded_match_ids: List[int] = []
        # Флаг, указывающий, что таблица сейчас заполняется.
        # Используется, чтобы не реагировать на программные изменения
        # элементов (itemChanged) и не вызывать обновление базы данных.
        self._updating_table: bool = False
        self._loaded_once: bool = False
        self.setup_database_ui()

    def setup_database_ui(self) -> None:
        """Настройка пользовательского интерфейса для раздела 'База половин'.

        Этот метод создаёт виджеты, необходимые для ввода строк матчей,
        отображения списка матчей, их фильтрации и удаления. Он вызывается
        в конструкторе и повторно не используется в других местах.
        """
        layout = QVBoxLayout(self)
        # Заголовок раздела
        title_label = QLabel("База половин")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(10)
        # Поле для ввода строк матчей
        input_label = QLabel("Вставьте строки матчей (каждая строка — отдельная игра):")
        layout.addWidget(input_label)
        self.input_text = QPlainTextEdit()
        self.input_text.setPlaceholderText(
            "Например:\n21.01.2026 China-2 Changsha Jiangxi 23 35 26 18 24 23 23 32\n..."
        )
        self.input_text.setMinimumHeight(100)
        layout.addWidget(self.input_text)
        # Кнопки для импорта и очистки поля
        btn_layout = QHBoxLayout()
        self.import_btn = QPushButton("Импортировать матчи")
        self.import_btn.setMinimumHeight(40)
        self.import_btn.clicked.connect(self.import_matches)
        btn_layout.addWidget(self.import_btn)
        self.clear_input_btn = QPushButton("Очистить")
        self.clear_input_btn.setMinimumHeight(40)
        self.clear_input_btn.clicked.connect(lambda: self.input_text.clear())
        btn_layout.addWidget(self.clear_input_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        # Фильтр по турниру и команде
        filter_layout = QHBoxLayout()
        # Турнир
        filter_layout.addWidget(QLabel("Турнир:"))
        self.tournament_combo = QComboBox()
        self.tournament_combo.addItem("Все турниры")
        self.tournament_combo.currentIndexChanged.connect(self.on_tournament_changed)
        filter_layout.addWidget(self.tournament_combo)
        # Команда
        filter_layout.addWidget(QLabel("Команда:"))
        self.team_combo = QComboBox()
        self.team_combo.addItem("Все команды")
        self.team_combo.setEnabled(False)
        self.team_combo.currentIndexChanged.connect(self.load_matches)
        filter_layout.addWidget(self.team_combo)
        # Кнопка обновления списка
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.setMinimumHeight(30)
        self.refresh_btn.clicked.connect(self.load_matches)
        filter_layout.addWidget(self.refresh_btn)
        # Кнопка переименования турнира
        self.rename_tournament_btn = QPushButton("Переименовать турнир")
        self.rename_tournament_btn.setMinimumHeight(30)
        self.rename_tournament_btn.clicked.connect(self.rename_tournament_dialog)
        filter_layout.addWidget(self.rename_tournament_btn)
        # Кнопка проверки количества игр
        self.check_games_btn = QPushButton("Проверить игры")
        self.check_games_btn.setMinimumHeight(30)
        self.check_games_btn.clicked.connect(self.show_games_summary)
        filter_layout.addWidget(self.check_games_btn)
        # Кнопка поиска и замены
        self.replace_btn = QPushButton("Заменить")
        self.replace_btn.setMinimumHeight(30)
        self.replace_btn.clicked.connect(self.replace_values_dialog)
        filter_layout.addWidget(self.replace_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        layout.addSpacing(10)
        # Таблица для отображения матчей
        self.table = QTableWidget()
        # Разрешаем выделение по строкам и множественный выбор
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.MultiSelection)
        # Разрешаем редактирование всех ячеек. Обновление БД произойдет
        # в обработчике itemChanged, определённом ниже.
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed | QAbstractItemView.AnyKeyPressed)
        # По умолчанию сортировка выключена во время загрузки для повышения скорости
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table, 1)

        # Подключаем обработчик изменения ячеек, чтобы сохранять правки
        self.table.itemChanged.connect(self.on_table_item_changed)
        # Поле для вывода статуса/ошибок
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)
        # Кнопки для удаления записей
        delete_layout = QHBoxLayout()
        self.delete_selected_btn = QPushButton("Удалить выбранные")
        self.delete_selected_btn.setMinimumHeight(30)
        self.delete_selected_btn.clicked.connect(self.delete_selected_matches)
        delete_layout.addWidget(self.delete_selected_btn)
        self.delete_all_btn = QPushButton("Удалить все")
        self.delete_all_btn.setMinimumHeight(30)
        self.delete_all_btn.clicked.connect(self.delete_all_matches)
        delete_layout.addWidget(self.delete_all_btn)
        delete_layout.addStretch()
        layout.addLayout(delete_layout)
        # Загрузка данных отложена до первого показа страницы

    def showEvent(self, event):
        super().showEvent(event)
        if not self._loaded_once:
            self._loaded_once = True
            QTimer.singleShot(0, self.load_matches)

    def import_matches(self) -> None:
        """Импортирует матчи из текстового поля в базу данных половин."""
        if not self.db:
            QMessageBox.critical(self, "Ошибка", "База данных недоступна.")
            return
        raw_text = self.input_text.toPlainText().strip()
        if not raw_text:
            QMessageBox.warning(self, "Внимание", "Пожалуйста, вставьте строки для импорта.")
            return
        # Разбиваем ввод на непустые строки
        lines = [ln for ln in raw_text.splitlines() if ln.strip()]
        if not lines:
            QMessageBox.warning(self, "Внимание", "Нет корректных строк для импорта.")
            return
        # Обработка импортируемых строк. При копировании из Excel
        # ячейки разделяются символом табуляции. Если такие
        # разделители присутствуют, нужно корректно собрать значения
        # каждого столбца, чтобы названия турниров и команд с
        # несколькими словами не разбивались на части. Для этого
        # заменяем пробелы внутри ячеек на символ подчёркивания ("_")
        # исключительно для строковых ячеек. Числовые значения и даты
        # оставляем без изменений. Такой подход позволяет
        # обеспечить корректную длину строки при последующем разборе
        # в HalfsDatabase.import_lines. В дальнейшем, если импортёр
        # заменяет подчёркивания на пробелы, данные будут приведены
        # к исходному виду.
        processed_lines: List[str] = []
        for line in lines:
            # Используем символ табуляции для разделения ячеек, если он присутствует.
            if "\t" in line:
                cells = [c.strip() for c in line.split("\t")]
                new_cells: List[str] = []
                for c in cells:
                    has_alpha = any(ch.isalpha() for ch in c)
                    if has_alpha:
                        # Внутри каждой строковой ячейки заменяем пробелы и
                        # подчёркивания на специальный маркер '~'. Это позволяет
                        # сохранить многословные названия и избежать
                        # разбиения на части при разборе строки по пробелам в
                        # HalfsDatabase.import_lines. Символ '~' выбран
                        # потому, что в исходных данных он практически не
                        # встречается. Дефисы оставляем, так как они входят в
                        # обозначение турниров (например, "Hungary-2").
                        # Сначала заменяем подчёркивания на пробелы, затем
                        # объединяем слова маркером.
                        tmp = c.replace("_", " ").split()
                        new_cells.append("~".join(tmp))
                    else:
                        new_cells.append(c)
                processed_lines.append(" ".join(new_cells))
            else:
                # В случае отсутствия табуляции предполагаем, что строка
                # уже содержит корректный формат с одним пробелом между
                # значениями. Для единообразия удаляем лишние пробелы.
                processed_lines.append(" ".join(line.split()))
        # Далее выполняем импорт, используя файл для записи ошибок
        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        os.makedirs(desktop_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        error_file = os.path.join(desktop_dir, f"halfs_import_errors_{timestamp}.txt")
        inserted, errors = self.db.import_lines(processed_lines, error_file_path=error_file)
        # После импорта восстанавливаем пробелы вместо маркера '~' в названиях
        # турниров и команд. Загружаем все записи и заменяем '~' на пробел в
        # соответствующих полях. Этот шаг нужен, чтобы пользователю
        # отображались корректные названия без временных маркеров.
        try:
            df_all = self.db._load_matches()
            # Перебираем строки, заменяем в нужных столбцах
            for _, r in df_all.iterrows():
                match_id = r.get("id")
                if match_id is None:
                    continue
                for fld in ("tournament", "team_home", "team_away"):
                    val = r.get(fld)
                    if isinstance(val, str) and "~" in val:
                        new_val = val.replace("~", " ")
                        try:
                            self.db.update_match_field(match_id, fld, new_val)
                        except Exception:
                            pass
        except Exception:
            # Игнорируем любые ошибки на этапе исправления маркеров
            pass
        # Перезагружаем таблицу
        self.load_matches()
        # Примечание: ранее после импорта вызывалось self.show_games_summary(),
        # что отображало таблицу «Нарушения нормы игр в турнире». По требованиям
        # функционала сортировки половин этот вызов удалён, чтобы
        # не отображать таблицу нарушений после импорта.
        msg_parts = [f"Успешно импортировано матчей: {inserted}"]
        if errors:
            msg_parts.append(
                f"Строк с ошибками: {len(errors)}. Файл со списком ошибок сохранён на рабочем столе:\n{error_file}"
            )
        QMessageBox.information(self, "Импорт завершён", "\n".join(msg_parts))

    def on_tournament_changed(self) -> None:
        """Обрабатывает изменение выбранного турнира.

        Заполняет список команд для выбранного турнира и перезагружает таблицу.
        Если выбран пункт "Все турниры", список команд очищается и отключается.
        """
        current = self.tournament_combo.currentText()
        if not self.db:
            return
        if current and current != "Все турниры":
            try:
                df = self.db._load_matches(tournament=current)
            except Exception:
                df = pd.DataFrame()
            teams = set(df["team_home"]).union(set(df["team_away"]))
            teams = sorted(teams)
            self.team_combo.blockSignals(True)
            self.team_combo.clear()
            self.team_combo.addItem("Все команды")
            self.team_combo.addItems(teams)
            self.team_combo.setEnabled(True)
            # Optionally keep previous team selection if still valid
            self.team_combo.setCurrentIndex(0)
            self.team_combo.blockSignals(False)
        else:
            # Disable team filter when no specific tournament selected
            self.team_combo.blockSignals(True)
            self.team_combo.clear()
            self.team_combo.addItem("Все команды")
            self.team_combo.setEnabled(False)
            self.team_combo.blockSignals(False)
        # Reload matches with new filters
        self.load_matches()

    def load_matches(self) -> None:
        """Загружает матчи из базы данных в таблицу с учётом выбранного турнира."""
        if not self.db:
            return
        tournament_filter = self.tournament_combo.currentText()
        if tournament_filter == "Все турниры":
            df = self.db._load_matches()
        else:
            df = self.db._load_matches(tournament=tournament_filter)
        # Применяем фильтр по команде, если включен
        if hasattr(self, 'team_combo') and self.team_combo.isEnabled():
            team_filter = self.team_combo.currentText()
            if team_filter and team_filter != "Все команды":
                try:
                    df = df[(df['team_home'] == team_filter) | (df['team_away'] == team_filter)]
                except Exception:
                    pass
        # Обновляем список турниров в комбобоксе
        try:
            all_df = self.db._load_matches()
            tournaments = sorted(set(all_df["tournament"]))
            current = self.tournament_combo.currentText()
            self.tournament_combo.blockSignals(True)
            self.tournament_combo.clear()
            self.tournament_combo.addItem("Все турниры")
            self.tournament_combo.addItems(tournaments)
            if current and current in tournaments:
                idx = self.tournament_combo.findText(current)
                if idx >= 0:
                    self.tournament_combo.setCurrentIndex(idx)
            self.tournament_combo.blockSignals(False)
        except Exception:
            pass
        headers = [
            "Дата", "Турнир", "Команда 1", "Команда 2",
            "Q1 (дом)", "Q1 (гость)", "Q2 (дом)", "Q2 (гость)",
            "Q3 (дом)", "Q3 (гость)", "Q4 (дом)", "Q4 (гость)",
            "ОТ (дом)", "ОТ (гость)",
            "Итог (дом)", "-", "Итог (гость)"
        ]
        # ВАЖНО: блокируем сигналы и сортировку ДО любых изменений таблицы,
        # чтобы Qt не переставлял строки во время заполнения и не вызывал
        # itemChanged, из-за чего данные «рассыпались».
        self._updating_table = True
        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(0)
        self.loaded_match_ids = []
        rows: List[List] = []
        for _, row in df.iterrows():
            # Сохраняем идентификатор матча, чтобы корректно обновлять/удалять записи
            try:
                match_id = int(row.get("id"))
            except Exception:
                match_id = None
            self.loaded_match_ids.append(match_id)
            # Форматируем дату в человеческий формат
            date_val = row.get("date")
            try:
                d = datetime.strptime(date_val, "%Y-%m-%d").strftime("%d.%m.%Y") if date_val else ""
            except Exception:
                d = str(date_val) if date_val is not None else ""
            # Считаем суммарные очки хозяев и гостей по всем четвертям и овертайму
            home_total = 0
            away_total = 0
            for q in ("q1", "q2", "q3", "q4"):
                hv = row.get(f"{q}_home")
                av = row.get(f"{q}_away")
                if hv is not None and pd.notna(hv):
                    try:
                        home_total += int(hv)
                    except Exception:
                        pass
                if av is not None and pd.notna(av):
                    try:
                        away_total += int(av)
                    except Exception:
                        pass
            # Учитываем очки в овертайме, если они есть
            hv = row.get("ot_home")
            av = row.get("ot_away")
            if hv is not None and pd.notna(hv):
                try:
                    home_total += int(hv)
                except Exception:
                    pass
            if av is not None and pd.notna(av):
                try:
                    away_total += int(av)
                except Exception:
                    pass
            # Формируем строку таблицы. Вместо одной колонки "Итог"
            # добавляем значение для хозяев, дефис и значение для гостей.
            rows.append([
                d,
                row.get("tournament", ""),
                row.get("team_home", ""),
                row.get("team_away", ""),
                row.get("q1_home"), row.get("q1_away"),
                row.get("q2_home"), row.get("q2_away"),
                row.get("q3_home"), row.get("q3_away"),
                row.get("q4_home"), row.get("q4_away"),
                row.get("ot_home"), row.get("ot_away"),
                home_total, "-", away_total
            ])
        self.table.setRowCount(len(rows))
        for row_idx, row_data in enumerate(rows):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem()
                # Показываем пустую ячейку для None или NaN
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    item.setData(Qt.DisplayRole, "")
                else:
                    # numeric columns: indices >=4 and <=13 hold integers
                    # Колонки Q1–Q4 и OT содержат числовые значения
                    if 4 <= col_idx <= 13 or col_idx in (14, 16):
                        try:
                            item.setData(Qt.DisplayRole, int(value))
                        except (ValueError, TypeError):
                            # В колонке 15 находится дефис, который должен оставаться строкой
                            item.setData(Qt.DisplayRole, value)
                    else:
                        item.setData(Qt.DisplayRole, value)
                if col_idx == 0:
                    # store match_id in UserRole to use in editing
                    item.setData(Qt.UserRole, self.loaded_match_ids[row_idx])
                # Подсветка победителя: если очки хозяев больше, подсвечиваем
                # ячейку итогового счёта хозяев, если гостей — ячейку гостей.
                try:
                    home_total = row_data[14]
                    away_total = row_data[16]
                except Exception:
                    home_total = None
                    away_total = None
                if home_total is not None and away_total is not None:
                    try:
                        ht = int(home_total)
                        at = int(away_total)
                        if ht != at:
                            winner_color = QColor(60, 179, 113, 80)  # soft green with alpha
                            if ht > at and col_idx == 14:
                                item.setBackground(winner_color)
                            elif at > ht and col_idx == 16:
                                item.setBackground(winner_color)
                    except Exception:
                        pass
                self.table.setItem(row_idx, col_idx, item)
        # Включаем сортировку и сигналы обратно ПОСЛЕ заполнения всех ячеек
        self.table.setSortingEnabled(True)
        self.table.blockSignals(False)
        self._updating_table = False
        # Автоматически подгоняем ширину столбцов под содержимое только для небольших таблиц
        header = self.table.horizontalHeader()
        if header is not None:
            if len(rows) > 10000:
                header.setSectionResizeMode(QHeaderView.Interactive)
            else:
                header.setSectionResizeMode(QHeaderView.ResizeToContents)
        try:
            date_delegate = DateSortDelegate()
            self.table.setItemDelegateForColumn(0, date_delegate)
        except Exception:
            pass

    def delete_selected_matches(self) -> None:
        """Удаляет выбранные строки из базы данных после подтверждения."""
        if not self.db:
            QMessageBox.critical(self, "Ошибка", "База данных недоступна.")
            return
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.information(self, "Удаление", "Выберите хотя бы одну строку для удаления.")
            return
        match_ids = []
        for index in selected:
            row_idx = index.row()
            match_id = self.loaded_match_ids[row_idx] if 0 <= row_idx < len(self.loaded_match_ids) else None
            if match_id is not None:
                match_ids.append(match_id)
        if not match_ids:
            QMessageBox.warning(self, "Удаление", "Невозможно определить идентификаторы выбранных матчей.")
            return
        reply = QMessageBox.question(
            self,
            'Подтверждение',
            f'Вы действительно хотите удалить выбранные записи ({len(match_ids)} шт.)?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                deleted = self.db.delete_matches(match_ids)
                self.load_matches()
                QMessageBox.information(
                    self,
                    "Удаление",
                    f"Удалено записей: {deleted}"
                )
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении: {str(e)}")

    def delete_all_matches(self) -> None:
        """Удаляет все записи из базы или записи выбранного турнира после подтверждения."""
        if not self.db:
            QMessageBox.critical(self, "Ошибка", "База данных недоступна.")
            return
        tournament_filter = self.tournament_combo.currentText()
        if tournament_filter and tournament_filter != "Все турниры":
            msg = f"Вы действительно хотите удалить все матчи из турнира '{tournament_filter}'?"
        else:
            msg = "Вы действительно хотите удалить ВСЕ матчи из базы данных?"
        reply = QMessageBox.question(
            self,
            'Подтверждение',
            msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                deleted = self.db.delete_all_matches(
                    tournament_filter if tournament_filter and tournament_filter != "Все турниры" else None
                )
                self.load_matches()
                QMessageBox.information(
                    self,
                    "Удаление",
                    f"Удалено записей: {deleted}"
                )
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении: {str(e)}")

    def rename_tournament_dialog(self) -> None:
        """Переименовывает выбранный турнир через диалог."""
        if not self.db:
            QMessageBox.critical(self, "Ошибка", "База данных недоступна.")
            return
        # Загрузить все уникальные названия турниров
        try:
            all_df = self.db._load_matches()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось получить список турниров: {str(e)}")
            return
        tournaments = sorted(set(all_df["tournament"]))
        if not tournaments:
            QMessageBox.information(self, "Переименование", "В базе нет ни одного турнира.")
            return
        # Создать диалог для выбора турнира и ввода нового названия
        dialog = QDialog(self)
        dialog.setWindowTitle("Переименование турнира")
        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.addWidget(QLabel("Выберите турнир для переименования:"))
        combo = QComboBox(dialog)
        combo.addItems(tournaments)
        dlg_layout.addWidget(combo)
        dlg_layout.addWidget(QLabel("Введите новое название:"))
        name_edit = QLineEdit(dialog)
        dlg_layout.addWidget(name_edit)
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Сохранить")
        cancel_btn = QPushButton("Отмена")
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        dlg_layout.addLayout(btn_layout)

        def on_save() -> None:
            old_name = combo.currentText()
            new_name = name_edit.text().strip()
            if not new_name:
                QMessageBox.warning(dialog, "Внимание", "Новое название не может быть пустым.")
                return
            if old_name == new_name:
                QMessageBox.information(dialog, "Переименование", "Новое название совпадает с текущим.")
                return
            try:
                updated = self.db.rename_tournament(old_name, new_name)
                if updated > 0:
                    QMessageBox.information(dialog, "Готово", f"Обновлено записей: {updated}")
                else:
                    QMessageBox.information(dialog, "Внимание", "Изменений не было произведено.")
                # Перезагрузим таблицу и обновим списки турниров
                self.load_matches()
                dialog.accept()
            except Exception as ex:
                QMessageBox.critical(dialog, "Ошибка", f"Ошибка при переименовании: {str(ex)}")

        save_btn.clicked.connect(on_save)
        cancel_btn.clicked.connect(dialog.reject)
        dialog.exec_()

    def show_games_summary(self) -> None:
        """Показывает сводку по количеству игр в турнирах и сравнивает с нормой.

        Вычисляет количество команд и нормативное количество игр (n_teams//2) для
        каждого турнира. Выводит турнир, фактическое число игр и норму.
        """
        if not self.db:
            QMessageBox.warning(self, "Ошибка", "База данных недоступна.")
            return
        summary_df = self.db.get_games_summary()
        if summary_df.empty:
            QMessageBox.information(self, "Сводка игр", "Нет данных для отображения.")
            return
        # Фильтруем те турниры, где фактическое число игр отличается от нормы
        mismatches = summary_df[summary_df['actual_games'] != summary_df['normative_games']]
        if mismatches.empty:
            QMessageBox.information(self, "Сводка игр", "Все турниры соответствуют норме игр.")
            return
        # Создаем диалог с таблицей
        dialog = QDialog(self)
        dialog.setWindowTitle("Нарушения нормы игр в турнирах")
        dlg_layout = QVBoxLayout(dialog)
        info_label = QLabel("Турниры, где количество игр отличается от нормы:")
        dlg_layout.addWidget(info_label)
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Турнир", "Игры", "Норма"])
        table.setRowCount(len(mismatches))
        for row_idx, (tournament, row) in enumerate(mismatches.iterrows()):
            # Турнир
            t_item = QTableWidgetItem(str(tournament))
            t_item.setFont(QFont("Arial", 10, QFont.Bold))
            table.setItem(row_idx, 0, t_item)
            # Фактическое
            a_item = QTableWidgetItem(str(int(row['actual_games'])))
            a_item.setFont(QFont("Arial", 10, QFont.Bold))
            table.setItem(row_idx, 1, a_item)
            # Норма
            n_item = QTableWidgetItem(str(int(row['normative_games'])))
            n_item.setFont(QFont("Arial", 10, QFont.Bold))
            table.setItem(row_idx, 2, n_item)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        table.horizontalHeader().setStretchLastSection(False)
        dlg_layout.addWidget(table)
        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(dialog.reject)
        dlg_layout.addWidget(btn_box)
        dialog.exec_()

    def replace_values_dialog(self) -> None:
        """Открывает диалог для поиска и замены строк в таблице.

        Пользователь может выбрать, какие значения заменить и на что. Замена
        производится в выбранных ячейках или во всей таблице, если нет
        выделения. Все изменения записываются в базу данных.
        """
        if not self.db:
            QMessageBox.warning(self, "Ошибка", "База данных недоступна.")
            return
        # Создаем диалог
        dialog = QDialog(self)
        dialog.setWindowTitle("Поиск и замена")
        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.addWidget(QLabel("Что заменить:"))
        old_edit = QLineEdit()
        dlg_layout.addWidget(old_edit)
        dlg_layout.addWidget(QLabel("На что заменить:"))
        new_edit = QLineEdit()
        dlg_layout.addWidget(new_edit)
        # Checkbox: только выделенные ячейки
        only_selected_cb = QCheckBox("Только в выделенных ячейках")
        dlg_layout.addWidget(only_selected_cb)
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        dlg_layout.addWidget(btn_box)

        def on_replace() -> None:
            old_val = old_edit.text()
            new_val = new_edit.text()
            if not old_val:
                QMessageBox.warning(dialog, "Внимание", "Введите текст для замены.")
                return
            # Собираем список ячеек
            indexes = []
            if only_selected_cb.isChecked():
                indexes = self.table.selectedIndexes()
            # Если не выделено, заменяем во всей таблице
            if not indexes:
                # Соберем все индексы
                for r in range(self.table.rowCount()):
                    for c in range(self.table.columnCount()):
                        indexes.append(self.table.model().index(r, c))
            # Применяем замену
            replaced = 0
            # Temporarily block signals to avoid recursion
            self._updating_table = True
            for idx in indexes:
                r = idx.row()
                c = idx.column()
                item = self.table.item(r, c)
                if item is None:
                    continue
                current_text = item.text()
                if old_val not in current_text:
                    continue
                new_text = current_text.replace(old_val, new_val)
                if new_text == current_text:
                    continue
                # Обновляем отображение
                item.setText(new_text)
                replaced += 1
                # Обновляем базу данных, если колонка редактируемая
                # Получаем match id
                id_item = self.table.item(r, 0)
                match_id = id_item.data(Qt.UserRole) if id_item else None
                if match_id is None:
                    continue
                # Определяем имя поля в базе
                field_map = {
                    0: 'date',
                    1: 'tournament',
                    2: 'team_home',
                    3: 'team_away',
                    4: 'q1_home', 5: 'q1_away',
                    6: 'q2_home', 7: 'q2_away',
                    8: 'q3_home', 9: 'q3_away',
                    10: 'q4_home', 11: 'q4_away',
                    12: 'ot_home', 13: 'ot_away',
                    14: None  # Итоговая колонка не хранится в БД
                }
                field_name = field_map.get(c)
                if not field_name:
                    continue
                # Преобразуем текст в соответствующий тип
                if field_name == 'date':
                    val = None
                    if new_text:
                        try:
                            val = datetime.strptime(new_text, '%d.%m.%Y').strftime('%Y-%m-%d')
                        except Exception:
                            val = new_text
                elif field_name in {'tournament', 'team_home', 'team_away'}:
                    val = new_text
                else:
                    if not new_text:
                        val = None
                    else:
                        try:
                            val = int(new_text)
                        except Exception:
                            val = None
                try:
                    self.db.update_match_field(match_id, field_name, val)
                except Exception:
                    pass
            self._updating_table = False
            # Перезагрузить таблицу для корректного отображения итогов
            self.load_matches()
            QMessageBox.information(dialog, "Завершено", f"Заменено {replaced} ячеек.")
            dialog.accept()

        btn_box.accepted.connect(on_replace)
        btn_box.rejected.connect(dialog.reject)
        dialog.exec_()

    def _recalc_row_totals(self, row: int) -> None:
        """Пересчитывает итоговые колонки (14 и 16) для указанной строки таблицы."""
        self._updating_table = True
        self.table.blockSignals(True)
        try:
            home_total = 0
            away_total = 0
            # Колонки Q1–Q4: 4,6,8,10 — home; 5,7,9,11 — away
            # Колонки OT: 12 — home; 13 — away
            for c in (4, 6, 8, 10, 12):
                it = self.table.item(row, c)
                if it and it.text().strip():
                    try:
                        home_total += int(it.text().strip())
                    except ValueError:
                        pass
            for c in (5, 7, 9, 11, 13):
                it = self.table.item(row, c)
                if it and it.text().strip():
                    try:
                        away_total += int(it.text().strip())
                    except ValueError:
                        pass
            # Обновляем ячейки итогов
            home_item = self.table.item(row, 14)
            if not home_item:
                home_item = QTableWidgetItem()
                self.table.setItem(row, 14, home_item)
            home_item.setData(Qt.DisplayRole, home_total)

            away_item = self.table.item(row, 16)
            if not away_item:
                away_item = QTableWidgetItem()
                self.table.setItem(row, 16, away_item)
            away_item.setData(Qt.DisplayRole, away_total)

            # Подсветка победителя
            winner_color = QColor(60, 179, 113, 80)
            no_color = QColor(0, 0, 0, 0)
            if home_total > away_total:
                home_item.setBackground(winner_color)
                away_item.setBackground(no_color)
            elif away_total > home_total:
                home_item.setBackground(no_color)
                away_item.setBackground(winner_color)
            else:
                home_item.setBackground(no_color)
                away_item.setBackground(no_color)
        finally:
            self.table.blockSignals(False)
            self._updating_table = False

    def on_table_item_changed(self, item: QTableWidgetItem) -> None:
        """Обработчик изменения ячейки таблицы.

        Сохраняет внесённые пользователем правки в базу данных.
        Игнорирует вызовы во время программной загрузки (см. self._updating_table).
        """
        # Не обрабатывать изменения, возникшие при загрузке данных
        if getattr(self, '_updating_table', False):
            return
        if not self.db:
            return
        row = item.row()
        col = item.column()
        # Получаем соответствующий id матча
        match_id = None
        # id хранится в UserRole первого столбца для каждой строки
        id_item = self.table.item(row, 0)
        if id_item:
            match_id = id_item.data(Qt.UserRole)
        if match_id is None:
            return
        # Определяем, какое поле обновлять
        field_map = {
            0: 'date',
            1: 'tournament',
            2: 'team_home',
            3: 'team_away',
            4: 'q1_home', 5: 'q1_away',
            6: 'q2_home', 7: 'q2_away',
            8: 'q3_home', 9: 'q3_away',
            10: 'q4_home', 11: 'q4_away',
            12: 'ot_home', 13: 'ot_away'
        }
        field_name = field_map.get(col)
        if not field_name:
            return
        # Получаем новое значение
        new_text = item.text().strip()
        # Преобразуем в нужный тип
        value: Optional[any]
        if field_name == 'date':
            # допустимы форматы dd.mm.yyyy или yyyy-mm-dd или пустая строка
            if not new_text:
                value = None
            else:
                try:
                    # попытаться распознать dd.mm.yyyy
                    value = datetime.strptime(new_text, '%d.%m.%Y').strftime('%Y-%m-%d')
                except Exception:
                    try:
                        # может уже ISO
                        datetime.strptime(new_text, '%Y-%m-%d')
                        value = new_text
                    except Exception:
                        # неверный формат, покажем предупреждение и откатим
                        QMessageBox.warning(self, "Неверный формат даты", "Введите дату в формате ДД.ММ.ГГГГ или YYYY-MM-DD.")
                        # восстанавливаем старое значение из БД
                        self.load_matches()
                        return
        elif field_name in {'tournament', 'team_home', 'team_away'}:
            value = new_text if new_text else None
        else:
            # столбцы очков: должны быть целым числом или пустыми
            if not new_text:
                value = None
            else:
                try:
                    value = int(new_text)
                except ValueError:
                    QMessageBox.warning(self, "Неверный формат", "Значение должно быть числом или пустым.")
                    self.load_matches()
                    return
        try:
            self.db.update_match_field(match_id, field_name, value)
        except Exception as exc:
            QMessageBox.critical(self, "Ошибка", f"Не удалось обновить запись: {exc}")
            self.load_matches()
            return
        # Пересчитываем итоги строки «на месте» без полной перезагрузки,
        # чтобы данные не пропадали из-за пересортировки таблицы.
        self._recalc_row_totals(row)


# Новая страница статистики для раздела "Половины"
class HalfsStatisticsPage(QWidget):
    """Страница для отображения статистики, коэффициентов, отклонений и распределения четвертей.

    Этот виджет предоставляет три вкладки:

    1. Статистика/коэффициенты: таблица со средними очками, пропущенными очками и
       дополнительным калькулятором коэффициентов "больше/меньше" для пары команд.
    2. Отклонения: таблица с разницей между второй и первой половиной для каждой команды.
    3. Средние четверти: калькулятор распределения вводимого тотала по четвертям для пары команд.

    Все вычисления выполняются на основе базы данных половин.
    """

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        try:
            from halfs_database import HalfsDatabase
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "Модуль halfs_database не найден.")
            self.db = None
        else:
            self.db = HalfsDatabase()
        # Используем отдельный метод для инициализации пользовательского интерфейса статистики
        self._stats_loaded = False
        self._deviation_loaded = False
        self._quarter_loaded = False
        self._wins_loaded = False
        self.setup_statistics_ui()

    def setup_statistics_ui(self) -> None:
        layout = QVBoxLayout(self)
        # Создаем вкладки
        self.tabs = QTabWidget()
        # Страница статистики и коэффициентов
        self.stats_tab = QWidget()
        self.setup_stats_tab()
        self.tabs.addTab(self.stats_tab, "Статистика/коэффициенты")
        # Страница отклонений
        self.deviation_tab = QWidget()
        self.setup_deviation_tab()
        self.tabs.addTab(self.deviation_tab, "Отклонения")
        # Страница распределения четвертей
        self.quarter_tab = QWidget()
        self.setup_quarter_tab()
        self.tabs.addTab(self.quarter_tab, "Средние четверти")

        # Вкладка побед/поражений
        # Эта вкладка отображает количество побед и поражений для каждой команды в выбранном турнире.
        self.wins_losses_tab = QWidget()
        self.setup_wins_losses_tab()
        self.tabs.addTab(self.wins_losses_tab, "Победы/поражения")
        layout.addWidget(self.tabs)
        self.tabs.currentChanged.connect(self.on_tab_changed)

    # ------------------------------------------------------------------
    # Статистика/коэффициенты
    # ------------------------------------------------------------------
    def setup_stats_tab(self) -> None:
        layout = QVBoxLayout(self.stats_tab)
        # Фильтр по турниру
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Турнир:"))
        self.stats_tournament_combo = QComboBox()
        self.stats_tournament_combo.addItem("Все турниры")
        self.stats_tournament_combo.currentIndexChanged.connect(self.load_stats_table)
        filter_layout.addWidget(self.stats_tournament_combo)
        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_stats_table)
        filter_layout.addWidget(refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        # Две таблицы статистики: забитые и пропущенные
        # Таблица забитых очков
        scored_group = QGroupBox("Забитые очки")
        scored_layout = QVBoxLayout(scored_group)
        self.stats_scored_table = QTableWidget()
        self.stats_scored_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.stats_scored_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.stats_scored_table.setSortingEnabled(True)
        scored_layout.addWidget(self.stats_scored_table)
        # Таблица пропущенных очков
        conceded_group = QGroupBox("Пропущенные очки")
        conceded_layout = QVBoxLayout(conceded_group)
        self.stats_conceded_table = QTableWidget()
        self.stats_conceded_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.stats_conceded_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.stats_conceded_table.setSortingEnabled(True)
        conceded_layout.addWidget(self.stats_conceded_table)
        # Размещаем две таблицы рядом
        tables_layout = QHBoxLayout()
        tables_layout.addWidget(scored_group)
        tables_layout.addWidget(conceded_group)
        layout.addLayout(tables_layout)
        # Раздел коэффициентов
        coeff_group = QGroupBox("Коэффициенты (Over/Under)")
        coeff_layout = QGridLayout(coeff_group)
        # Выбор команд
        coeff_layout.addWidget(QLabel("Команда 1:"), 0, 0)
        self.coeff_team1_combo = QComboBox()
        coeff_layout.addWidget(self.coeff_team1_combo, 0, 1)
        coeff_layout.addWidget(QLabel("Команда 2:"), 0, 2)
        self.coeff_team2_combo = QComboBox()
        coeff_layout.addWidget(self.coeff_team2_combo, 0, 3)
        # Порог четверти
        coeff_layout.addWidget(QLabel("Порог четверти:"), 1, 0)
        self.coeff_q_thresh_edit = QLineEdit()
        self.coeff_q_thresh_edit.setPlaceholderText("Напр.: 40")
        coeff_layout.addWidget(self.coeff_q_thresh_edit, 1, 1)
        # Порог половины
        coeff_layout.addWidget(QLabel("Порог половины:"), 1, 2)
        self.coeff_h_thresh_edit = QLineEdit()
        self.coeff_h_thresh_edit.setPlaceholderText("Напр.: 80")
        coeff_layout.addWidget(self.coeff_h_thresh_edit, 1, 3)
        # Порог матча
        coeff_layout.addWidget(QLabel("Порог матча:"), 2, 0)
        self.coeff_m_thresh_edit = QLineEdit()
        self.coeff_m_thresh_edit.setPlaceholderText("Напр.: 160")
        coeff_layout.addWidget(self.coeff_m_thresh_edit, 2, 1)
        # Кнопка расчета
        self.calc_coeff_btn = QPushButton("Рассчитать")
        self.calc_coeff_btn.clicked.connect(self.calculate_coefficients)
        coeff_layout.addWidget(self.calc_coeff_btn, 2, 3)
        # Таблица для результатов коэффициентов (будет транспонирована)
        self.coeff_table = QTableWidget()
        # Сразу создаём таблицу с 2 строками (Over/Under) и 7 столбцами (1Q..M)
        self.coeff_table.setRowCount(2)
        self.coeff_table.setColumnCount(7)
        self.coeff_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.coeff_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.coeff_table.setSortingEnabled(False)
        # Устанавливаем вертикальные подписи (Больше/Меньше)
        self.coeff_table.setVerticalHeaderLabels(["Больше", "Меньше"])
        # Изначально задаём заголовки по умолчанию; они будут обновлены при расчёте
        self.coeff_table.setHorizontalHeaderLabels(["1Q", "2Q", "3Q", "4Q", "1H", "2H", "M"])
        # Добавляем таблицу в coeff_layout (новая строка, растягиваем на все колонки)
        coeff_layout.addWidget(self.coeff_table, 3, 0, 1, 4)
        # Добавляем коэффициенты блок на основную вкладку
        layout.addWidget(coeff_group)
        # Обновляем исходные данные при первом показе вкладки

    # ------------------------------------------------------------------
    # Победы/поражения
    # ------------------------------------------------------------------
    def setup_wins_losses_tab(self) -> None:
        """Настройка вкладки, отображающей количество побед и поражений."""
        layout = QVBoxLayout(self.wins_losses_tab)
        # Фильтр по турниру
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Турнир:"))
        self.wl_tournament_combo = QComboBox()
        self.wl_tournament_combo.addItem("Все турниры")
        self.wl_tournament_combo.currentIndexChanged.connect(self.load_wins_losses_table)
        filter_layout.addWidget(self.wl_tournament_combo)
        wl_refresh_btn = QPushButton("Обновить")
        wl_refresh_btn.clicked.connect(self.load_wins_losses_table)
        filter_layout.addWidget(wl_refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        # Таблица побед/поражений
        self.wins_losses_table = QTableWidget()
        self.wins_losses_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.wins_losses_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.wins_losses_table.setSortingEnabled(True)
        layout.addWidget(self.wins_losses_table)
        # Загрузка данных при первом показе вкладки

    def showEvent(self, event):
        super().showEvent(event)
        self.on_tab_changed(self.tabs.currentIndex())

    def on_tab_changed(self, index: int) -> None:
        # 0: stats, 1: deviation, 2: quarter, 3: wins/losses
        if index == 0 and not self._stats_loaded:
            self._stats_loaded = True
            QTimer.singleShot(0, self.load_stats_table)
        elif index == 1 and not self._deviation_loaded:
            self._deviation_loaded = True
            QTimer.singleShot(0, self.load_deviation_table)
        elif index == 2 and not self._quarter_loaded:
            self._quarter_loaded = True
            QTimer.singleShot(0, self.load_quarter_teams)
        elif index == 3 and not self._wins_loaded:
            self._wins_loaded = True
            QTimer.singleShot(0, self.load_wins_losses_table)

    def load_wins_losses_table(self) -> None:
        """Загружает таблицу побед и поражений для выбранного турнира."""
        if not self.db:
            return
        # Обновляем список турниров для выпадающего списка
        try:
            all_df = self.db._load_matches()
            tournaments = sorted(set(all_df["tournament"]))
            current = self.wl_tournament_combo.currentText()
            self.wl_tournament_combo.blockSignals(True)
            self.wl_tournament_combo.clear()
            self.wl_tournament_combo.addItem("Все турниры")
            self.wl_tournament_combo.addItems(tournaments)
            if current and current in tournaments:
                idx = self.wl_tournament_combo.findText(current)
                if idx >= 0:
                    self.wl_tournament_combo.setCurrentIndex(idx)
            self.wl_tournament_combo.blockSignals(False)
        except Exception:
            pass
        # Определяем выбранный турнир
        t_filter = self.wl_tournament_combo.currentText()
        tournament = None if not t_filter or t_filter == "Все турниры" else t_filter
        # Получаем данные
        wl_df = self.db.get_wins_losses(tournament)
        if wl_df is None or wl_df.empty:
            self.wins_losses_table.clear()
            self.wins_losses_table.setRowCount(0)
            self.wins_losses_table.setColumnCount(0)
            return
        # Настройка таблицы
        headers = ["Команда", "Победы", "Поражения"]
        self.wins_losses_table.setColumnCount(len(headers))
        self.wins_losses_table.setHorizontalHeaderLabels(headers)
        self.wins_losses_table.setRowCount(len(wl_df))
        # Зелёный для побед, красный для поражений
        green_brush = QBrush(QColor(0, 128, 0))
        red_brush = QBrush(QColor(178, 34, 34))
        for row_idx, (team, row) in enumerate(wl_df.iterrows()):
            values = [team, row["wins"], row["losses"]]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem()
                if isinstance(val, (int, float)):
                    item.setData(Qt.DisplayRole, val)
                else:
                    item.setData(Qt.DisplayRole, str(val))
                # Применяем цвет
                font = item.font()
                font.setBold(True)
                item.setFont(font)
                if col_idx == 1:
                    item.setForeground(green_brush)
                elif col_idx == 2:
                    item.setForeground(red_brush)
                self.wins_losses_table.setItem(row_idx, col_idx, item)
        # Подгоняем ширину
        self.wins_losses_table.resizeColumnsToContents()
        header = self.wins_losses_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
            header.setStretchLastSection(False)

    def load_stats_table(self) -> None:
        """Загружает статистику команд и обновляет списки команд для расчета коэффициентов."""
        if not self.db:
            return
        # Определяем выбранный турнир
        t_filter = self.stats_tournament_combo.currentText() if hasattr(self, 'stats_tournament_combo') else "Все турниры"
        tournament = None if not t_filter or t_filter == "Все турниры" else t_filter
        # Загружаем все турниры для выпадающего списка, если список пуст или обновляем
        try:
            all_df = self.db._load_matches()
            tournaments = sorted(set(all_df["tournament"]))
            current = self.stats_tournament_combo.currentText()
            self.stats_tournament_combo.blockSignals(True)
            self.stats_tournament_combo.clear()
            self.stats_tournament_combo.addItem("Все турниры")
            self.stats_tournament_combo.addItems(tournaments)
            # Восстанавливаем выбор
            if current and current in tournaments:
                idx = self.stats_tournament_combo.findText(current)
                if idx >= 0:
                    # Устанавливаем индекс без смещения
                    self.stats_tournament_combo.setCurrentIndex(idx)
            self.stats_tournament_combo.blockSignals(False)
        except Exception:
            pass
        # Если выбран фильтр "Все турниры" или не указан, не загружаем статистику.
        # Пользователь должен выбрать турнир для просмотра статистики.
        if tournament is None:
            # Очистка таблиц
            self.stats_scored_table.clear()
            self.stats_scored_table.setRowCount(0)
            self.stats_scored_table.setColumnCount(0)
            self.stats_conceded_table.clear()
            self.stats_conceded_table.setRowCount(0)
            self.stats_conceded_table.setColumnCount(0)
            # Очищаем списки команд
            self.coeff_team1_combo.clear()
            self.coeff_team2_combo.clear()
            # Никаких дальнейших действий, пока не выбран конкретный турнир
            return
        # Получаем статистику только для выбранного турнира
        stats_df = self.db.get_team_statistics(tournament)
        if stats_df is None or stats_df.empty:
            # Очистка таблиц и списков, если нет данных по выбранному турниру
            self.stats_scored_table.clear()
            self.stats_scored_table.setRowCount(0)
            self.stats_scored_table.setColumnCount(0)
            self.stats_conceded_table.clear()
            self.stats_conceded_table.setRowCount(0)
            self.stats_conceded_table.setColumnCount(0)
            self.coeff_team1_combo.clear()
            self.coeff_team2_combo.clear()
            return
        # Для NCAA D1 и NCAA D1 (W): скрываем команды с играми ниже среднего
        tournament_norm = " ".join(str(tournament).replace("~", " ").split()).upper()
        if tournament_norm in ("NCAA D1", "NCAA D1 (W)"):
            try:
                avg_games = float(stats_df["games"].mean())
                stats_df = stats_df[stats_df["games"] >= avg_games]
            except Exception:
                pass
        if stats_df.empty:
            self.stats_scored_table.clear()
            self.stats_scored_table.setRowCount(0)
            self.stats_scored_table.setColumnCount(0)
            self.stats_conceded_table.clear()
            self.stats_conceded_table.setRowCount(0)
            self.stats_conceded_table.setColumnCount(0)
            self.coeff_team1_combo.clear()
            self.coeff_team2_combo.clear()
            return

        # Заполняем таблицу забитых
        scored_headers = [
            "Команда", "Игры",
            "Q1 (заб.)", "Q2 (заб.)", "Q3 (заб.)", "Q4 (заб.)",
            "1-я пол. (заб.)", "2-я пол. (заб.)", "Всего (заб.)"
        ]
        conceded_headers = [
            "Команда",
            "Q1 (проп.)", "Q2 (проп.)", "Q3 (проп.)", "Q4 (проп.)",
            "1-я пол. (проп.)", "2-я пол. (проп.)", "Всего (проп.)"
        ]
        # Настройка столбцов
        self.stats_scored_table.setColumnCount(len(scored_headers))
        self.stats_scored_table.setHorizontalHeaderLabels(scored_headers)
        self.stats_scored_table.setRowCount(len(stats_df))
        self.stats_conceded_table.setColumnCount(len(conceded_headers))
        self.stats_conceded_table.setHorizontalHeaderLabels(conceded_headers)
        self.stats_conceded_table.setRowCount(len(stats_df))
        # Обновляем списки команд для коэффициентов
        teams = list(stats_df.index)
        self.coeff_team1_combo.blockSignals(True)
        self.coeff_team2_combo.blockSignals(True)
        self.coeff_team1_combo.clear()
        self.coeff_team2_combo.clear()
        self.coeff_team1_combo.addItems(teams)
        self.coeff_team2_combo.addItems(teams)
        self.coeff_team1_combo.blockSignals(False)
        self.coeff_team2_combo.blockSignals(False)
        # Заполняем таблицы
        # Заполняем строки таблиц
        # Выключаем сортировку и обновления для ускорения
        self.stats_scored_table.setSortingEnabled(False)
        self.stats_conceded_table.setSortingEnabled(False)
        self.stats_scored_table.setUpdatesEnabled(False)
        self.stats_conceded_table.setUpdatesEnabled(False)
        for row_idx, (team, row) in enumerate(stats_df.iterrows()):
            # Формируем список значений для таблицы забитых очков
            scored_values = [
                team,
                row["games"],
                row["avg_scored_q1"], row["avg_scored_q2"], row["avg_scored_q3"], row["avg_scored_q4"],
                row["first_half_scored"], row["second_half_scored"], row["total_scored"]
            ]
            for col_idx, val in enumerate(scored_values):
                item = QTableWidgetItem()
                # Пропущенные или NaN значения заменяем нулями
                if isinstance(val, (int, float)):
                    val = 0.0 if pd.isna(val) else float(val)
                    # Округляем до 1 знака после запятой
                    item.setData(Qt.DisplayRole, round(val, 1))
                else:
                    item.setData(Qt.DisplayRole, str(val))
                # Устанавливаем элемент
                self.stats_scored_table.setItem(row_idx, col_idx, item)
            # Формируем список значений для таблицы пропущенных очков
            conceded_values = [
                team,
                row["avg_conceded_q1"], row["avg_conceded_q2"], row["avg_conceded_q3"], row["avg_conceded_q4"],
                row["first_half_conceded"], row["second_half_conceded"], row["total_conceded"]
            ]
            for col_idx, val in enumerate(conceded_values):
                item = QTableWidgetItem()
                if isinstance(val, (int, float)):
                    val = 0.0 if pd.isna(val) else float(val)
                    item.setData(Qt.DisplayRole, round(val, 1))
                else:
                    item.setData(Qt.DisplayRole, str(val))
                self.stats_conceded_table.setItem(row_idx, col_idx, item)
        # Включаем обновления и сортировку обратно
        self.stats_scored_table.setUpdatesEnabled(True)
        self.stats_conceded_table.setUpdatesEnabled(True)
        self.stats_scored_table.setSortingEnabled(True)
        self.stats_conceded_table.setSortingEnabled(True)
        # Подсветка столбцов половин: одинаковая палитра для обеих таблиц
        first_half_color = QColor(50, 130, 184, 60)  # полупрозрачный синий
        second_half_color = QColor(92, 184, 92, 60)  # полупрозрачный зелёный
        # Колонки половин в таблицах: забитые: 6 (1-я), 7 (2-я); пропущенные: 5 (1-я), 6 (2-я)
        for row_idx in range(self.stats_scored_table.rowCount()):
            # Таблица забитых очков
            h1_item = self.stats_scored_table.item(row_idx, 6)
            h2_item = self.stats_scored_table.item(row_idx, 7)
            if h1_item:
                h1_item.setBackground(first_half_color)
            if h2_item:
                h2_item.setBackground(second_half_color)
        for row_idx in range(self.stats_conceded_table.rowCount()):
            h1_item = self.stats_conceded_table.item(row_idx, 5)
            h2_item = self.stats_conceded_table.item(row_idx, 6)
            if h1_item:
                h1_item.setBackground(first_half_color)
            if h2_item:
                h2_item.setBackground(second_half_color)
        # Подгоняем ширину столбцов. Для больших таблиц
        # использовать интерактивный режим, чтобы не тормозить при вычислениях.
        header_scored = self.stats_scored_table.horizontalHeader()
        header_conceded = self.stats_conceded_table.horizontalHeader()
        row_count = self.stats_scored_table.rowCount()
        if header_scored is not None:
            if row_count > 10000:
                # при больших данных не выполняем resizeColumnsToContents
                header_scored.setSectionResizeMode(QHeaderView.Interactive)
                header_scored.setStretchLastSection(False)
            else:
                self.stats_scored_table.resizeColumnsToContents()
                header_scored.setSectionResizeMode(QHeaderView.ResizeToContents)
                header_scored.setStretchLastSection(False)
        if header_conceded is not None:
            if row_count > 10000:
                header_conceded.setSectionResizeMode(QHeaderView.Interactive)
                header_conceded.setStretchLastSection(False)
            else:
                self.stats_conceded_table.resizeColumnsToContents()
                header_conceded.setSectionResizeMode(QHeaderView.ResizeToContents)
                header_conceded.setStretchLastSection(False)

    def calculate_coefficients(self) -> None:
        """Рассчитывает коэффициенты Over/Under для выбранной пары команд и порогов."""
        if not self.db:
            return
        team1 = self.coeff_team1_combo.currentText()
        team2 = self.coeff_team2_combo.currentText()
        if not team1 or not team2:
            QMessageBox.warning(self, "Внимание", "Выберите обе команды для расчёта коэффициентов.")
            return
        try:
            q_text = self.coeff_q_thresh_edit.text().strip().replace(",", ".")
            h_text = self.coeff_h_thresh_edit.text().strip().replace(",", ".")
            m_text = self.coeff_m_thresh_edit.text().strip().replace(",", ".")
            q_thr = float(q_text) if q_text else 0.0
            h_thr = float(h_text) if h_text else 0.0
            m_thr = float(m_text) if m_text else 0.0
        except ValueError:
            QMessageBox.warning(self, "Внимание", "Введите числовые значения порогов.")
            return
        # Определяем турнир
        t_filter = self.stats_tournament_combo.currentText()
        tournament = None if not t_filter or t_filter == "Все турниры" else t_filter
        tournament_norm = " ".join(str(tournament).replace("~", " ").split()).upper() if tournament else ""
        is_ncaa_d1 = tournament_norm == "NCAA D1"
        ncaa_half_threshold = None
        if is_ncaa_d1 and h_thr > 0.0:
            # NCAA D1: half thresholds are based on Q1/Q2 values
            ncaa_half_threshold = h_thr
            h_thr = 0.0
            q_thr = ncaa_half_threshold
        result = self.db.get_tot_coefficients(team1, team2, q_thr, h_thr, m_thr, tournament)
        if result is None:
            QMessageBox.warning(self, "Внимание", "Не удалось вычислить коэффициенты для выбранных команд.")
            return
        # Формируем таблицу коэффициентов с динамическими столбцами
        selected_periods: List[str] = []
        selected_labels: List[str] = []
        # NCAA D1: 1H/2H берутся из Q1/Q2
        if is_ncaa_d1 and ncaa_half_threshold and ncaa_half_threshold > 0.0:
            selected_periods.extend(["q1", "q2"])
            selected_labels.extend(["1H", "2H"])
        # Добавляем четверти, если указан порог четверти
        elif q_thr > 0.0:
            selected_periods.extend(["q1", "q2", "q3", "q4"])
            selected_labels.extend(["1Q", "2Q", "3Q", "4Q"])
        # Добавляем половины, если указан порог половины
        if h_thr > 0.0:
            selected_periods.extend(["h1", "h2"])
            selected_labels.extend(["1H", "2H"])
        # Добавляем матч, если указан порог матча
        if m_thr > 0.0:
            selected_periods.append("match")
            selected_labels.append("M")
        # Если пользователь не ввёл ни одного порога, очищаем таблицу и прекращаем
        if not selected_periods:
            self.coeff_table.clear()
            self.coeff_table.setRowCount(0)
            self.coeff_table.setColumnCount(0)
            return
        # Настраиваем таблицу: 2 строки (Больше/Меньше) и нужное количество столбцов
        self.coeff_table.setRowCount(2)
        self.coeff_table.setColumnCount(len(selected_periods))
        self.coeff_table.setHorizontalHeaderLabels(selected_labels)
        self.coeff_table.setVerticalHeaderLabels(["Больше", "Меньше"])
        # Цвета для коэффициентов
        over_color = QColor("green")
        under_color = QColor("red")
        # Заполняем таблицу по выбранным периодам
        for col_idx, period in enumerate(selected_periods):
            over_coeff = result["over"].get(period, 0.0)
            under_coeff = result["under"].get(period, 0.0)
            # Округляем значения до сотых
            over_val = round(float(over_coeff), 2) if isinstance(over_coeff, (int, float)) else over_coeff
            under_val = round(float(under_coeff), 2) if isinstance(under_coeff, (int, float)) else under_coeff
            # 'Больше'
            over_item = QTableWidgetItem()
            over_item.setData(Qt.DisplayRole, over_val)
            over_item.setForeground(QBrush(over_color))
            font_over = over_item.font()
            # Всегда выделяем жирным, если колонка отображается
            font_over.setBold(True)
            over_item.setFont(font_over)
            self.coeff_table.setItem(0, col_idx, over_item)
            # 'Меньше'
            under_item = QTableWidgetItem()
            under_item.setData(Qt.DisplayRole, under_val)
            under_item.setForeground(QBrush(under_color))
            font_under = under_item.font()
            font_under.setBold(True)
            under_item.setFont(font_under)
            self.coeff_table.setItem(1, col_idx, under_item)
        # Подгоняем ширину столбцов
        self.coeff_table.resizeColumnsToContents()
        header = self.coeff_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
            # Не растягиваем последний столбец, чтобы все колонки были равномерны
            header.setStretchLastSection(False)

    # ------------------------------------------------------------------
    # Отклонения
    # ------------------------------------------------------------------
    def setup_deviation_tab(self) -> None:
        layout = QVBoxLayout(self.deviation_tab)
        # Фильтр по турниру
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Турнир:"))
        self.dev_tournament_combo = QComboBox()
        self.dev_tournament_combo.addItem("Все турниры")
        self.dev_tournament_combo.currentIndexChanged.connect(self.load_deviation_table)
        filter_layout.addWidget(self.dev_tournament_combo)
        dev_refresh_btn = QPushButton("Обновить")
        dev_refresh_btn.clicked.connect(self.load_deviation_table)
        filter_layout.addWidget(dev_refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        # Таблица отклонений
        self.deviation_table = QTableWidget()
        self.deviation_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.deviation_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.deviation_table.setSortingEnabled(True)
        layout.addWidget(self.deviation_table)
        # Сводные показатели: средний тотал и среднее отклонение по турниру, размещаем под таблицей
        summary_layout = QHBoxLayout()
        # Средний тотал
        self.avg_total_label = QLabel("Средний тотал: —")
        bold_font = QFont()
        bold_font.setBold(True)
        self.avg_total_label.setFont(bold_font)
        summary_layout.addWidget(self.avg_total_label)
        # Среднее отклонение
        self.avg_deviation_label = QLabel("Среднее отклонение: —")
        self.avg_deviation_label.setFont(bold_font)
        summary_layout.addWidget(self.avg_deviation_label)
        summary_layout.addStretch()
        layout.addLayout(summary_layout)

        # Блок для расчета отклонений пары команд
        pair_group = QGroupBox("Пара команд — комбинированное отклонение")
        pair_layout = QGridLayout(pair_group)
        pair_layout.addWidget(QLabel("Команда 1:"), 0, 0)
        self.pair_team1_combo = QComboBox()
        pair_layout.addWidget(self.pair_team1_combo, 0, 1)
        pair_layout.addWidget(QLabel("Команда 2:"), 0, 2)
        self.pair_team2_combo = QComboBox()
        pair_layout.addWidget(self.pair_team2_combo, 0, 3)
        self.pair_calc_btn = QPushButton("Рассчитать")
        self.pair_calc_btn.clicked.connect(self.calculate_pair_deviation)
        pair_layout.addWidget(self.pair_calc_btn, 1, 3)
        pair_layout.addWidget(QLabel("Результат ( (Dev1 + Dev2) / 4 ):"), 1, 0, 1, 2)
        self.pair_result_label = QLabel("—")
        pair_layout.addWidget(self.pair_result_label, 1, 2)

        # Метка для отображения среднего тотала выбранных команд
        pair_layout.addWidget(QLabel("Средний тотал двух команд:"), 2, 0, 1, 2)
        self.pair_total_label = QLabel("—")
        pair_layout.addWidget(self.pair_total_label, 2, 2)
        # Добавляем группу для пары команд
        layout.addWidget(pair_group)
        # Загрузка данных
        self.load_deviation_table()

    def load_deviation_table(self) -> None:
        """Загружает таблицу отклонений для выбранного турнира."""
        if not self.db:
            return
        # Обновляем список турниров
        try:
            all_df = self.db._load_matches()
            tournaments = sorted(set(all_df["tournament"]))
            current = self.dev_tournament_combo.currentText()
            self.dev_tournament_combo.blockSignals(True)
            self.dev_tournament_combo.clear()
            self.dev_tournament_combo.addItem("Все турниры")
            self.dev_tournament_combo.addItems(tournaments)
            if current and current in tournaments:
                idx = self.dev_tournament_combo.findText(current)
                if idx >= 0:
                    # Выбираем индекс напрямую, без смещения
                    self.dev_tournament_combo.setCurrentIndex(idx)
            self.dev_tournament_combo.blockSignals(False)
        except Exception:
            pass
        # Определяем фильтр
        t_filter = self.dev_tournament_combo.currentText()
        tournament = None if not t_filter or t_filter == "Все турниры" else t_filter
        dev_df = self.db.get_team_deviations(tournament)
        if dev_df is None or dev_df.empty:
            self.deviation_table.clear()
            self.deviation_table.setRowCount(0)
            self.deviation_table.setColumnCount(0)
            return
        headers = ["Команда", "Отклонение", "Средний тотал"]
        self.deviation_table.setColumnCount(len(headers))
        self.deviation_table.setHorizontalHeaderLabels(headers)
        self.deviation_table.setRowCount(len(dev_df))
        # Цвет текста для отклонения: отрицательное – красный, положительное – зелёный, ноль – оранжевый
        for row_idx, (team, row) in enumerate(dev_df.iterrows()):
            dev_val = row["deviation"]
            avg_val = row["average_total"]
            dev_rounded = round(float(dev_val), 1) if pd.notna(dev_val) else 0.0
            avg_rounded = round(float(avg_val), 1) if pd.notna(avg_val) else 0.0
            values = [team, dev_rounded, avg_rounded]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem()
                if isinstance(val, (int, float)):
                    item.setData(Qt.DisplayRole, val)
                else:
                    item.setData(Qt.DisplayRole, str(val))
                # Применяем цвет шрифта и жирное начертание для колонки отклонения (index 1)
                if col_idx == 1:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    if dev_rounded < 0:
                        item.setForeground(QBrush(QColor("red")))
                    elif dev_rounded > 0:
                        item.setForeground(QBrush(QColor("green")))
                    else:
                        item.setForeground(QBrush(QColor("orange")))
                self.deviation_table.setItem(row_idx, col_idx, item)
        self.deviation_table.resizeColumnsToContents()
        header = self.deviation_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
            header.setStretchLastSection(False)

        # Вычисляем средние значения для отображения внизу
        try:
            avg_total = dev_df["average_total"].mean()
            avg_dev = dev_df["deviation"].mean()
        except Exception:
            avg_total = 0.0
            avg_dev = 0.0
        # Обновляем подписи
        if hasattr(self, 'avg_total_label'):
            self.avg_total_label.setText(f"Средний тотал: {avg_total:.1f}")
        if hasattr(self, 'avg_deviation_label'):
            self.avg_deviation_label.setText(f"Среднее отклонение: {avg_dev:.1f}")
            # Цвет текста в зависимости от знака
            if avg_dev < 0:
                self.avg_deviation_label.setStyleSheet("color: red; font-weight: bold;")
            elif avg_dev > 0:
                self.avg_deviation_label.setStyleSheet("color: green; font-weight: bold;")
            else:
                self.avg_deviation_label.setStyleSheet("color: orange; font-weight: bold;")

        # Обновление списков команд для блока пары отклонений
        teams = list(dev_df.index)
        current1 = self.pair_team1_combo.currentText() if hasattr(self, 'pair_team1_combo') else None
        current2 = self.pair_team2_combo.currentText() if hasattr(self, 'pair_team2_combo') else None
        if hasattr(self, 'pair_team1_combo'):
            self.pair_team1_combo.blockSignals(True)
            self.pair_team2_combo.blockSignals(True)
            self.pair_team1_combo.clear()
            self.pair_team2_combo.clear()
            self.pair_team1_combo.addItems(teams)
            self.pair_team2_combo.addItems(teams)
            # Restore previous selections if possible
            if current1 in teams:
                idx1 = self.pair_team1_combo.findText(current1)
                if idx1 >= 0:
                    self.pair_team1_combo.setCurrentIndex(idx1)
            if current2 in teams:
                idx2 = self.pair_team2_combo.findText(current2)
                if idx2 >= 0:
                    self.pair_team2_combo.setCurrentIndex(idx2)
            self.pair_team1_combo.blockSignals(False)
            self.pair_team2_combo.blockSignals(False)

    def calculate_pair_deviation(self) -> None:
        """Вычисляет комбинированное отклонение для выбранной пары команд."""
        if not self.db:
            return
        team1 = self.pair_team1_combo.currentText() if hasattr(self, 'pair_team1_combo') else None
        team2 = self.pair_team2_combo.currentText() if hasattr(self, 'pair_team2_combo') else None
        if not team1 or not team2:
            QMessageBox.warning(self, "Внимание", "Выберите обе команды для расчёта отклонения.")
            return
        # Фильтр по турниру
        t_filter = self.dev_tournament_combo.currentText() if hasattr(self, 'dev_tournament_combo') else None
        tournament = None if not t_filter or t_filter == "Все турниры" else t_filter
        deviation = self.db.get_pair_deviation(team1, team2, tournament)
        # Получаем данные для средних тоталов
        dev_df = self.db.get_team_deviations(tournament)
        avg_total = None
        if dev_df is not None and not dev_df.empty:
            try:
                total1 = dev_df.loc[team1, "average_total"] if team1 in dev_df.index else None
                total2 = dev_df.loc[team2, "average_total"] if team2 in dev_df.index else None
                if total1 is not None and total2 is not None:
                    avg_total = (float(total1) + float(total2)) / 2.0
            except Exception:
                avg_total = None
        if deviation is None:
            self.pair_result_label.setText("—")
            self.pair_total_label.setText("—")
            QMessageBox.warning(self, "Внимание", "Не удалось вычислить комбинированное отклонение для выбранных команд.")
        else:
            # Отображаем число с одним знаком после запятой, как в Excel
            self.pair_result_label.setText(f"{deviation:.1f}")
            if avg_total is not None:
                self.pair_total_label.setText(f"{avg_total:.1f}")
            else:
                self.pair_total_label.setText("—")

    # ------------------------------------------------------------------
    # Средние четверти
    # ------------------------------------------------------------------
    def setup_quarter_tab(self) -> None:
        layout = QVBoxLayout(self.quarter_tab)
        # Фильтр по турниру (опционально)
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Турнир:"))
        self.q_tournament_combo = QComboBox()
        self.q_tournament_combo.addItem("Все турниры")
        self.q_tournament_combo.currentIndexChanged.connect(self.load_quarter_teams)
        filter_layout.addWidget(self.q_tournament_combo)
        q_refresh_btn = QPushButton("Обновить")
        q_refresh_btn.clicked.connect(self.load_quarter_teams)
        filter_layout.addWidget(q_refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        # Выбор команд и тотала
        form_layout = QGridLayout()
        form_layout.addWidget(QLabel("Команда 1:"), 0, 0)
        self.q_team1_combo = QComboBox()
        form_layout.addWidget(self.q_team1_combo, 0, 1)
        form_layout.addWidget(QLabel("Команда 2:"), 0, 2)
        self.q_team2_combo = QComboBox()
        form_layout.addWidget(self.q_team2_combo, 0, 3)
        form_layout.addWidget(QLabel("Тотал на матч:"), 1, 0)
        self.match_total_edit = QLineEdit()
        self.match_total_edit.setPlaceholderText("Напр.: 160")
        form_layout.addWidget(self.match_total_edit, 1, 1)
        self.calc_quarter_btn = QPushButton("Рассчитать распределение")
        self.calc_quarter_btn.clicked.connect(self.calculate_distribution)
        form_layout.addWidget(self.calc_quarter_btn, 1, 3)
        # Равномерно распределяем ширину колонок с комбобоксами команд
        form_layout.setColumnStretch(1, 1)
        form_layout.setColumnStretch(3, 1)
        layout.addLayout(form_layout)
        # Таблица для отображения распределения (транспонированный вид)
        self.quarter_table = QTableWidget()
        # Одна строка и четыре колонки (1Q, 2Q, 3Q, 4Q)
        self.quarter_table.setRowCount(1)
        self.quarter_table.setColumnCount(4)
        self.quarter_table.setHorizontalHeaderLabels(["1Q", "2Q", "3Q", "4Q"])
        # Скрываем вертикальный заголовок, так как таблица имеет единственную строку
        self.quarter_table.verticalHeader().setVisible(False)
        self.quarter_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # Выбор по ячейке — удобнее для одной строки
        self.quarter_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.quarter_table.setSortingEnabled(False)
        layout.addWidget(self.quarter_table)
        # Загрузка начальных данных
        self.load_quarter_teams()

    def load_quarter_teams(self) -> None:
        """Обновляет список команд для распределения четвертей."""
        if not self.db:
            return
        # Обновляем список турниров
        try:
            all_df = self.db._load_matches()
            tournaments = sorted(set(all_df["tournament"]))
            current = self.q_tournament_combo.currentText()
            self.q_tournament_combo.blockSignals(True)
            self.q_tournament_combo.clear()
            self.q_tournament_combo.addItem("Все турниры")
            self.q_tournament_combo.addItems(tournaments)
            if current and current in tournaments:
                idx = self.q_tournament_combo.findText(current)
                if idx >= 0:
                    # Устанавливаем индекс напрямую без смещения
                    self.q_tournament_combo.setCurrentIndex(idx)
            self.q_tournament_combo.blockSignals(False)
        except Exception:
            pass
        # Определяем турнир
        t_filter = self.q_tournament_combo.currentText()
        tournament = None if not t_filter or t_filter == "Все турниры" else t_filter
        stats_df = self.db.get_team_statistics(tournament)
        if stats_df is None or stats_df.empty:
            self.q_team1_combo.clear()
            self.q_team2_combo.clear()
            return
        teams = list(stats_df.index)
        self.q_team1_combo.blockSignals(True)
        self.q_team2_combo.blockSignals(True)
        self.q_team1_combo.clear()
        self.q_team2_combo.clear()
        self.q_team1_combo.addItems(teams)
        self.q_team2_combo.addItems(teams)
        self.q_team1_combo.blockSignals(False)
        self.q_team2_combo.blockSignals(False)

    def calculate_distribution(self) -> None:
        """Рассчитывает распределение тотала по четвертям для выбранной пары команд."""
        if not self.db:
            return
        team1 = self.q_team1_combo.currentText()
        team2 = self.q_team2_combo.currentText()
        if not team1 or not team2:
            QMessageBox.warning(self, "Внимание", "Выберите обе команды для расчёта.")
            return
        try:
            match_total = float(self.match_total_edit.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Внимание", "Введите числовое значение тотала.")
            return
        t_filter = self.q_tournament_combo.currentText()
        tournament = None if not t_filter or t_filter == "Все турниры" else t_filter
        distribution = self.db.distribute_total(team1, team2, match_total, tournament)
        if distribution is None:
            QMessageBox.warning(self, "Внимание", "Не удалось вычислить распределение для выбранных команд.")
            return
        # Заполняем транспонированную таблицу: одна строка и четыре колонки
        period_order = ["q1", "q2", "q3", "q4"]
        # Устанавливаем размер таблицы
        self.quarter_table.setRowCount(1)
        self.quarter_table.setColumnCount(4)
        self.quarter_table.setHorizontalHeaderLabels(["1Q", "2Q", "3Q", "4Q"])
        for col_idx, period in enumerate(period_order):
            val = distribution.get(period, 0.0)
            # Округляем значение до десятых
            try:
                val = round(float(val), 1)
            except Exception:
                pass
            item = QTableWidgetItem()
            item.setData(Qt.DisplayRole, val)
            self.quarter_table.setItem(0, col_idx, item)
        # Подгоняем ширину столбцов
        self.quarter_table.resizeColumnsToContents()
        header = self.quarter_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
            header.setStretchLastSection(False)


class TournamentSummaryPage(QWidget):
    """Страница для отображения сводной статистики по турнирам.

    Эта страница показывает для каждого турнира усреднённые значения по
    четвертям, половинам, матчу, количество игр и команд, а также среднее
    отклонение между второй и первой половиной. Цветовая подсветка
    применяется к колонке отклонений: зелёный — положительное значение,
    красный — отрицательное, оранжевый — ноль.
    """

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        try:
            from halfs_database import HalfsDatabase
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "Модуль halfs_database не найден.")
            self.db = None
        else:
            self.db = HalfsDatabase()
        # Храним выбранные турниры и данные сводной таблицы
        # множество выбранных турниров
        self.selected_tournaments = set()
        # актуальная таблица сводных данных (DataFrame)
        self.current_summary_df = None
        self._loaded_once = False
        # Инициализируем интерфейс сводной таблицы
        self._setup_baza_polovins_ui()

    def _setup_baza_polovins_ui(self) -> None:
        layout = QVBoxLayout(self)
        # Заголовок
        title_label = QLabel("Сводная таблица по турнирам")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(10)
        # Фильтр по турнирам
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Турниры:"))
        self.tour_filter_btn = QPushButton("Все турниры")
        self.tour_filter_menu = QMenu(self.tour_filter_btn)
        # При выборе из меню обновляем отображение
        self.tour_filter_btn.setMenu(self.tour_filter_menu)
        filter_layout.addWidget(self.tour_filter_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        # Таблица для сводных данных
        self.summary_table = QTableWidget()
        self.summary_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.summary_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.summary_table.setSortingEnabled(True)
        layout.addWidget(self.summary_table)
        # Кнопка обновления
        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_summary)
        layout.addWidget(refresh_btn)
        # Загрузка данных при первом показе страницы

    def showEvent(self, event):
        super().showEvent(event)
        if not self._loaded_once:
            self._loaded_once = True
            QTimer.singleShot(0, self.load_summary)

    def load_summary(self) -> None:
        """Загружает и отображает сводную статистику по турнирам."""
        if not self.db:
            return
        summary_df = self.db.get_tournament_summary()
        # Храним полную таблицу
        if summary_df is None or summary_df.empty:
            self.current_summary_df = None
            self.summary_table.clear()
            self.summary_table.setRowCount(0)
            self.summary_table.setColumnCount(0)
            # Также очищаем меню турниров
            self.tour_filter_menu.clear()
            return
        self.current_summary_df = summary_df
        # Обновляем меню выбора турниров
        self.populate_tour_menu(list(summary_df.index))
        # Обновляем отображение таблицы на основе выбранных турниров
        self.refresh_summary_table()

    def populate_tour_menu(self, tournaments: List[str]) -> None:
        """Создаёт выпадающее меню со списком турниров и чекбоксами."""
        self.tour_filter_menu.clear()
        # Пункт для выбора всех
        all_action = QAction("Все турниры", self.tour_filter_menu)
        all_action.setCheckable(True)
        all_action.setData("all")
        # Помечаем как выбранный, если в множестве нет конкретных турниров
        all_action.setChecked(len(self.selected_tournaments) == 0)
        all_action.toggled.connect(self.on_tour_filter_changed)
        self.tour_filter_menu.addAction(all_action)
        self.tour_filter_menu.addSeparator()
        for t in tournaments:
            act = QAction(t, self.tour_filter_menu)
            act.setCheckable(True)
            act.setData(t)
            act.setChecked(t in self.selected_tournaments)
            act.toggled.connect(self.on_tour_filter_changed)
            self.tour_filter_menu.addAction(act)

    def on_tour_filter_changed(self, checked: bool) -> None:
        """Обработчик изменения фильтра турниров."""
        action = self.sender()
        if not isinstance(action, QAction):
            return
        key = action.data()
        # Обработка выбора "Все турниры"
        if key == "all":
            if checked:
                # Очистить набор выбранных и снять отметки с других пунктов
                self.selected_tournaments.clear()
                for act in self.tour_filter_menu.actions()[2:]:  # пропускаем "Все турниры" и разделитель
                    act.blockSignals(True)
                    act.setChecked(False)
                    act.blockSignals(False)
            else:
                # Если "Все" снят, но ничего не выбрано, оставляем его выбранным
                if not self.selected_tournaments:
                    action.blockSignals(True)
                    action.setChecked(True)
                    action.blockSignals(False)
                    return
        else:
            if checked:
                self.selected_tournaments.add(key)
                # Снять отметку с "Все"
                first_action = self.tour_filter_menu.actions()[0]
                first_action.blockSignals(True)
                first_action.setChecked(False)
                first_action.blockSignals(False)
            else:
                if key in self.selected_tournaments:
                    self.selected_tournaments.remove(key)
                # Если ни одного турнира не выбрано, выбрать "Все"
                if not self.selected_tournaments:
                    first_action = self.tour_filter_menu.actions()[0]
                    first_action.blockSignals(True)
                    first_action.setChecked(True)
                    first_action.blockSignals(False)
        # Обновляем название кнопки
        if not self.selected_tournaments:
            self.tour_filter_btn.setText("Все турниры")
        else:
            self.tour_filter_btn.setText(f"Выбрано: {len(self.selected_tournaments)}")
        # Перерисовываем таблицу
        self.refresh_summary_table()

    def refresh_summary_table(self) -> None:
        """Применяет текущий фильтр турниров и обновляет отображение таблицы."""
        df = self.current_summary_df
        if df is None or df.empty:
            self.summary_table.clear()
            self.summary_table.setRowCount(0)
            self.summary_table.setColumnCount(0)
            return
        # Фильтруем по выбранным турнирам
        if self.selected_tournaments:
            filtered_df = df[df.index.isin(self.selected_tournaments)]
        else:
            filtered_df = df
        # Настройка заголовков с разделителями для половин
        # Добавляем пустые колонки после H1 и H2 для создания вертикальных разделителей
        headers = [
            "Отклонение", "Турнир", "Q1", "Q2", "H1", "", "Q3", "Q4", "H2", "", "Матч", "Игры", "Команды"
        ]
        self.summary_table.setColumnCount(len(headers))
        self.summary_table.setHorizontalHeaderLabels(headers)
        self.summary_table.setRowCount(len(filtered_df))
        # Цвета для текста отклонения
        color_neg = QColor(220, 0, 0)     # тёмно‑красный для отрицательных
        color_pos = QColor(0, 128, 0)     # тёмно‑зелёный для положительных
        color_zero = QColor(255, 140, 0)  # оранжевый для нуля
        # Цвета текста для четвертей, половин и матча (приглушённые и более спокойные)
        col_text_colors = {
            2: QColor(100, 149, 237),  # Q1 – приглушённый голубой
            3: QColor(60, 179, 113),  # Q2 – приглушённый зелёный
            4: QColor(210, 180, 140),  # H1 – тёплый бежевый
            6: QColor(147, 112, 219),  # Q3 – приглушённый фиолетовый
            7: QColor(32, 178, 170),  # Q4 – приглушённый бирюзовый
            8: QColor(218, 165, 32),  # H2 – золотистый
            10: QColor(219, 112, 147)   # Match – тёплый розовый
        }
        # Перебор строк и заполнение таблицы
        for row_idx, (tournament, row) in enumerate(filtered_df.iterrows()):
            deviation = row["deviation"]
            # Подготовка значений (без разделителей)
            row_values = [
                round(float(deviation), 1) if pd.notna(deviation) else 0.0,
                tournament,
                round(float(row["q1_avg"]), 1) if pd.notna(row["q1_avg"]) else 0.0,
                round(float(row["q2_avg"]), 1) if pd.notna(row["q2_avg"]) else 0.0,
                round(float(row["h1_avg"]), 1) if pd.notna(row["h1_avg"]) else 0.0,
                # Разделитель после H1 (пустой столбец)
                None,
                round(float(row["q3_avg"]), 1) if pd.notna(row["q3_avg"]) else 0.0,
                round(float(row["q4_avg"]), 1) if pd.notna(row["q4_avg"]) else 0.0,
                round(float(row["h2_avg"]), 1) if pd.notna(row["h2_avg"]) else 0.0,
                # Разделитель после H2 (пустой столбец)
                None,
                round(float(row["match_avg"]), 1) if pd.notna(row["match_avg"]) else 0.0,
                int(row["games_count"]),
                int(row["teams_count"])
            ]
            for col_idx, val in enumerate(row_values):
                item = QTableWidgetItem()
                # Разделители: задаём пустой текст и тёмный фон, отключаем выбор
                if val is None:
                    item.setText("")
                    item.setFlags(Qt.NoItemFlags)
                    item.setBackground(QBrush(QColor(50, 50, 50)))
                    self.summary_table.setItem(row_idx, col_idx, item)
                    # Установим минимальную ширину для разделителя
                    self.summary_table.setColumnWidth(col_idx, 4)
                    continue
                # Устанавливаем значение отображения
                item.setData(Qt.DisplayRole, val)
                font = item.font()
                # Колонка отклонения: цвет текста и жирный
                if col_idx == 0:
                    font.setBold(True)
                    if val < 0:
                        item.setForeground(QBrush(color_neg))
                    elif val > 0:
                        item.setForeground(QBrush(color_pos))
                    else:
                        item.setForeground(QBrush(color_zero))
                    item.setFont(font)
                # Колонка турнира: жирный текст
                elif col_idx == 1:
                    font.setBold(True)
                    item.setFont(font)
                # Четверти, половины и матч: устанавливаем цвет текста и жирный
                elif col_idx in col_text_colors:
                    font.setBold(True)
                    item.setFont(font)
                    item.setForeground(QBrush(col_text_colors[col_idx]))
                # Игры и команды: жирный
                elif col_idx in (11, 12):
                    font.setBold(True)
                    item.setFont(font)
                self.summary_table.setItem(row_idx, col_idx, item)
        # Подгоняем ширину столбцов (кроме разделителей)
        self.summary_table.resizeColumnsToContents()
        header = self.summary_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
            header.setStretchLastSection(False)

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        # Заголовок
        title_label = QLabel("База половин")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(10)

        # Блок ввода строк матчей
        input_label = QLabel("Вставьте строки матчей (каждая строка — отдельная игра):")
        layout.addWidget(input_label)
        self.input_text = QPlainTextEdit()
        self.input_text.setPlaceholderText(
            "Например:\n21.01.2026 China-2 Changsha Jiangxi 23 35 26 18 24 23 23 32\n..."
        )
        self.input_text.setMinimumHeight(100)
        layout.addWidget(self.input_text)

        # Кнопки для импорта и очистки
        btn_layout = QHBoxLayout()
        self.import_btn = QPushButton("Импортировать матчи")
        self.import_btn.setMinimumHeight(40)
        self.import_btn.clicked.connect(self.import_matches)
        btn_layout.addWidget(self.import_btn)

        self.clear_input_btn = QPushButton("Очистить")
        self.clear_input_btn.setMinimumHeight(40)
        self.clear_input_btn.clicked.connect(lambda: self.input_text.clear())
        btn_layout.addWidget(self.clear_input_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Фильтр по турниру
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Турнир:"))
        self.tournament_combo = QComboBox()
        self.tournament_combo.addItem("Все турниры")
        self.tournament_combo.currentIndexChanged.connect(self.load_matches)
        filter_layout.addWidget(self.tournament_combo)
        # Кнопка обновления
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.setMinimumHeight(30)
        self.refresh_btn.clicked.connect(self.load_matches)
        filter_layout.addWidget(self.refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        layout.addSpacing(10)

        # Таблица для отображения матчей
        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table, 1)

        # Сообщение об ошибках
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)

        # Загрузка данных при инициализации
        self.load_matches()

    def setup_database_ui(self) -> None:
        """Настройка пользовательского интерфейса для раздела "База половин".

        Внутренний метод дублирует первоначальную реализацию
        setup_ui для HalfsDatabasePage. Он используется для
        корректной инициализации интерфейса базы половин,
        оставляя метод setup_ui свободным для других назначений.
        """
        layout = QVBoxLayout(self)
        # Заголовок
        title_label = QLabel("База половин")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(10)
        # Блок ввода строк матчей
        input_label = QLabel("Вставьте строки матчей (каждая строка — отдельная игра):")
        layout.addWidget(input_label)
        self.input_text = QPlainTextEdit()
        self.input_text.setPlaceholderText(
            "Например:\n21.01.2026 China-2 Changsha Jiangxi 23 35 26 18 24 23 23 32\n..."
        )
        self.input_text.setMinimumHeight(100)
        layout.addWidget(self.input_text)
        # Кнопки для импорта и очистки
        btn_layout = QHBoxLayout()
        self.import_btn = QPushButton("Импортировать матчи")
        self.import_btn.setMinimumHeight(40)
        self.import_btn.clicked.connect(self.import_matches)
        btn_layout.addWidget(self.import_btn)
        self.clear_input_btn = QPushButton("Очистить")
        self.clear_input_btn.setMinimumHeight(40)
        self.clear_input_btn.clicked.connect(lambda: self.input_text.clear())
        btn_layout.addWidget(self.clear_input_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        # Фильтр по турниру
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Турнир:"))
        self.tournament_combo = QComboBox()
        self.tournament_combo.addItem("Все турниры")
        self.tournament_combo.currentIndexChanged.connect(self.load_matches)
        filter_layout.addWidget(self.tournament_combo)
        # Кнопка обновления
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.setMinimumHeight(30)
        self.refresh_btn.clicked.connect(self.load_matches)
        filter_layout.addWidget(self.refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        layout.addSpacing(10)
        # Таблица для отображения матчей
        self.table = QTableWidget()
        # Разрешаем выделять строки полностью и множественный выбор
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.MultiSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table, 1)
        # Сообщение об ошибках
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)
        # Кнопки для удаления выбранных и всех записей
        delete_layout = QHBoxLayout()
        self.delete_selected_btn = QPushButton("Удалить выбранные")
        self.delete_selected_btn.setMinimumHeight(30)
        self.delete_selected_btn.clicked.connect(self.delete_selected_matches)
        delete_layout.addWidget(self.delete_selected_btn)
        self.delete_all_btn = QPushButton("Удалить все")
        self.delete_all_btn.setMinimumHeight(30)
        self.delete_all_btn.clicked.connect(self.delete_all_matches)
        delete_layout.addWidget(self.delete_all_btn)
        delete_layout.addStretch()
        layout.addLayout(delete_layout)
        # Загрузка данных при инициализации
        self.load_matches()

    def import_matches(self) -> None:
        """Импортирует матчи из текстового поля в базу данных."""
        if not self.db:
            QMessageBox.critical(self, "Ошибка", "База данных недоступна.")
            return
        raw_text = self.input_text.toPlainText().strip()
        if not raw_text:
            QMessageBox.warning(self, "Внимание", "Пожалуйста, вставьте строки для импорта.")
            return
        lines = [ln for ln in raw_text.splitlines() if ln.strip()]
        if not lines:
            QMessageBox.warning(self, "Внимание", "Нет корректных строк для импорта.")
            return
        # Обработка импортируемых строк, аналогичная HalfsDatabasePage.import_matches.
        # Если данные скопированы из Excel, значения в строке разделены табуляцией.
        # Чтобы не потерять целостность названий турниров и команд, заменяем
        # пробелы внутри ячеек на "_" для строковых ячеек.
        processed_lines: List[str] = []
        for line in lines:
            if "\t" in line:
                cells = [c.strip() for c in line.split("\t")]
                new_cells: List[str] = []
                for c in cells:
                    has_alpha = any(ch.isalpha() for ch in c)
                    if has_alpha:
                        tmp = c.replace("_", " ").split()
                        new_cells.append("~".join(tmp))
                    else:
                        new_cells.append(c)
                processed_lines.append(" ".join(new_cells))
            else:
                processed_lines.append(" ".join(line.split()))
        # Формируем путь для файла ошибок на рабочем столе
        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        os.makedirs(desktop_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        error_file = os.path.join(desktop_dir, f"halfs_import_errors_{timestamp}.txt")
        inserted, errors = self.db.import_lines(processed_lines, error_file_path=error_file)
        # После импорта восстанавливаем пробелы вместо маркера '~'
        try:
            df_all = self.db._load_matches()
            for _, r in df_all.iterrows():
                match_id = r.get("id")
                if match_id is None:
                    continue
                for fld in ("tournament", "team_home", "team_away"):
                    val = r.get(fld)
                    if isinstance(val, str) and "~" in val:
                        new_val = val.replace("~", " ")
                        try:
                            self.db.update_match_field(match_id, fld, new_val)
                        except Exception:
                            pass
        except Exception:
            pass
        # Обновляем таблицу после импорта
        self.load_matches()
        msg_parts = [f"Успешно импортировано матчей: {inserted}"]
        if errors:
            msg_parts.append(
                f"Строк с ошибками: {len(errors)}. Файл со списком ошибок сохранён на рабочем столе:\n{error_file}"
            )
        QMessageBox.information(self, "Импорт завершён", "\n".join(msg_parts))

    def load_matches(self) -> None:
        """Загружает список матчей в таблицу с учётом фильтра по турниру."""
        if not self.db:
            return
        tournament_filter = self.tournament_combo.currentText()
        if tournament_filter == "Все турниры":
            df = self.db._load_matches()
        else:
            df = self.db._load_matches(tournament=tournament_filter)
        # Обновляем выпадающий список турниров (если фильтр не был выбран)
        try:
            all_df = self.db._load_matches()
            tournaments = sorted(set(all_df["tournament"]))
            current = self.tournament_combo.currentText()
            self.tournament_combo.blockSignals(True)
            self.tournament_combo.clear()
            self.tournament_combo.addItem("Все турниры")
            self.tournament_combo.addItems(tournaments)
            # Восстанавливаем выбор, если он есть в списке
            if current and current in tournaments:
                index = self.tournament_combo.findText(current)
                if index >= 0:
                    # Устанавливаем индекс напрямую, без смещения
                    self.tournament_combo.setCurrentIndex(index)
            self.tournament_combo.blockSignals(False)
        except Exception:
            pass
        # Настраиваем таблицу
        headers = [
            "Дата", "Турнир", "Команда 1", "Команда 2",
            "Q1 (дом)", "Q1 (гость)", "Q2 (дом)", "Q2 (гость)",
            "Q3 (дом)", "Q3 (гость)", "Q4 (дом)", "Q4 (гость)",
            "ОТ (дом)", "ОТ (гость)"
        ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        # Convert date to display format
        # Подготавливаем строки и список идентификаторов
        self.loaded_match_ids = []
        rows: List[List] = []
        for _, row in df.iterrows():
            # Сохраняем идентификатор матча для операций удаления
            try:
                match_id = int(row.get("id"))
            except Exception:
                match_id = None
            self.loaded_match_ids.append(match_id)
            # Форматируем дату
            date_val = row.get("date")
            try:
                # База хранит дату в формате YYYY-MM-DD
                d = datetime.strptime(date_val, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                d = str(date_val)
            rows.append([
                d,
                row.get("tournament", ""),
                row.get("team_home", ""),
                row.get("team_away", ""),
                row.get("q1_home"), row.get("q1_away"),
                row.get("q2_home"), row.get("q2_away"),
                row.get("q3_home"), row.get("q3_away"),
                row.get("q4_home"), row.get("q4_away"),
                row.get("ot_home"), row.get("ot_away")
            ])
        self.table.setRowCount(len(rows))
        # Заполняем таблицу
        for row_idx, row_data in enumerate(rows):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem()
                # Format None values as empty strings
                if value is None:
                    item.setData(Qt.DisplayRole, "")
                else:
                    # Пытаемся установить числовые значения как числа для корректной сортировки
                    if col_idx >= 4:  # Счета
                        try:
                            item.setData(Qt.DisplayRole, int(value))
                        except (ValueError, TypeError):
                            item.setData(Qt.DisplayRole, value)
                    else:
                        item.setData(Qt.DisplayRole, value)
                # Дополнительно сохраняем ID в пользовательских данных первого столбца
                if col_idx == 0:
                    item.setData(Qt.UserRole, self.loaded_match_ids[row_idx])
                self.table.setItem(row_idx, col_idx, item)
        # Подгоняем ширины
        self.table.resizeColumnsToContents()
        if self.table.horizontalHeader().count() > 0:
            self.table.horizontalHeader().setStretchLastSection(True)
        # Настраиваем делегат для колонки с датой, чтобы сортировка по дате была корректной
        try:
            date_delegate = DateSortDelegate()
            self.table.setItemDelegateForColumn(0, date_delegate)
        except Exception:
            pass

    def delete_selected_matches(self) -> None:
        """Удаляет выбранные строки из базы данных после подтверждения пользователя."""
        if not self.db:
            QMessageBox.critical(self, "Ошибка", "База данных недоступна.")
            return
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.information(self, "Удаление", "Выберите хотя бы одну строку для удаления.")
            return
        # Собираем идентификаторы выбранных матчей
        match_ids = []
        for index in selected:
            row_idx = index.row()
            match_id = self.loaded_match_ids[row_idx] if 0 <= row_idx < len(self.loaded_match_ids) else None
            if match_id is not None:
                match_ids.append(match_id)
        if not match_ids:
            QMessageBox.warning(self, "Удаление", "Невозможно определить идентификаторы выбранных матчей.")
            return
        # Запрос подтверждения
        reply = QMessageBox.question(
            self,
            'Подтверждение',
            f'Вы действительно хотите удалить выбранные записи ({len(match_ids)} шт.)?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                deleted = self.db.delete_matches(match_ids)
                self.load_matches()
                QMessageBox.information(
                    self,
                    "Удаление",
                    f"Удалено записей: {deleted}"
                )
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении: {str(e)}")

    def delete_all_matches(self) -> None:
        """Удаляет все матчи из базы данных (или из выбранного турнира) после подтверждения."""
        if not self.db:
            QMessageBox.critical(self, "Ошибка", "База данных недоступна.")
            return
        # Определяем фильтр по турниру
        tournament_filter = self.tournament_combo.currentText()
        # Запрос подтверждения
        if tournament_filter and tournament_filter != "Все турниры":
            msg = f"Вы действительно хотите удалить все матчи из турнира '{tournament_filter}'?"
        else:
            msg = "Вы действительно хотите удалить ВСЕ матчи из базы данных?"
        reply = QMessageBox.question(
            self,
            'Подтверждение',
            msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                deleted = self.db.delete_all_matches(
                    tournament_filter if tournament_filter and tournament_filter != "Все турниры" else None
                )
                self.load_matches()
                QMessageBox.information(
                    self,
                    "Удаление",
                    f"Удалено записей: {deleted}"
                )
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении: {str(e)}")
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Title
        title_label = QLabel("Работа с половинами")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(20)
        
        # File selection section - Halfs file
        halfs_group_label = QLabel("Файл Половины:")
        halfs_group_label.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(halfs_group_label)
        
        halfs_file_layout = QHBoxLayout()
        self.halfs_file_label = QLabel("Файл не выбран")
        self.halfs_file_label.setMinimumWidth(400)
        
        self.select_halfs_btn = QPushButton("Выбрать файл Половины")
        self.select_halfs_btn.setMinimumWidth(180)
        self.select_halfs_btn.clicked.connect(self.select_halfs_file)  # Правильное имя метода
        
        halfs_file_layout.addWidget(self.halfs_file_label)
        halfs_file_layout.addWidget(self.select_halfs_btn)
        layout.addLayout(halfs_file_layout)
        layout.addSpacing(20)
        
        # File selection section - Cyber files
        cyber_group_label = QLabel("Файлы Cyber:")
        cyber_group_label.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(cyber_group_label)
        
        cyber_file_layout = QHBoxLayout()
        self.select_cyber_btn = QPushButton("Выбрать файлы Cyber")
        self.select_cyber_btn.setMinimumWidth(180)
        self.select_cyber_btn.clicked.connect(self.select_cyber_files)  # Правильное имя метода
        
        self.clear_cyber_btn = QPushButton("Очистить список")
        self.clear_cyber_btn.setMinimumWidth(150)
        self.clear_cyber_btn.clicked.connect(self.clear_cyber_files)
        
        cyber_file_layout.addWidget(self.select_cyber_btn)
        cyber_file_layout.addWidget(self.clear_cyber_btn)
        cyber_file_layout.addStretch()
        layout.addLayout(cyber_file_layout)
        
        # Selected Cyber files list
        self.cyber_files_list = QListWidget()
        self.cyber_files_list.setMinimumHeight(150)
        layout.addWidget(self.cyber_files_list)
        layout.addSpacing(20)
        
        # Process button
        self.process_btn = QPushButton("Запустить перенос данных")
        self.process_btn.setMinimumWidth(250)
        self.process_btn.setMinimumHeight(50)
        self.process_btn.setFont(QFont("Arial", 12))
        self.process_btn.clicked.connect(self.process_files)
        self.process_btn.setEnabled(False)
        
        process_layout = QHBoxLayout()
        process_layout.addStretch()
        process_layout.addWidget(self.process_btn)
        process_layout.addStretch()
        layout.addLayout(process_layout)
        layout.addSpacing(20)
        
        # Progress bar and status
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("%p%")
        
        self.status_label = QLabel("Готов к работе")
        self.status_label.setFont(QFont("Arial", 10))
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setWordWrap(True)
        
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.status_label)
        
        # Log area
        log_label = QLabel("Журнал выполнения:")
        layout.addWidget(log_label)
        
        self.log_text = QListWidget()
        self.log_text.setMinimumHeight(200)
        layout.addWidget(self.log_text)
        
        # Check xlwings availability
        if not self.excel_handler.available:
            self.add_log("ВНИМАНИЕ: Библиотека xlwings не установлена. Установите её командой: pip install xlwings")
            self.status_label.setText("Ошибка: xlwings не установлен")
            self.status_label.setStyleSheet("color: red;")
            self.process_btn.setEnabled(False)
        else:
            self.add_log("Библиотека xlwings найдена")
    
    def select_halfs_file(self):
        """Выбор файла Половины"""
        try:
            # Проверяем открытые файлы Excel с фильтром по типу 'halves'
            open_files = self.excel_handler.get_opened_excel_files(file_type='halves')
            
            if open_files:
                # Если только один файл - сразу его выбираем
                if len(open_files) == 1:
                    selected_file = open_files[0]
                    if self.excel_handler.set_halfs_file(selected_file):
                        self.halfs_file_label.setText(selected_file)
                        self.add_log(f"Выбран файл Половины: {os.path.basename(selected_file)}")
                        self.update_process_button()
                    return
                
                # Если несколько файлов - создаем диалог с радиокнопками (только один выбор)
                dialog = QDialog(self)
                dialog.setWindowTitle("Выберите файл Половины")
                dialog.setMinimumWidth(400)
                
                layout = QVBoxLayout(dialog)
                layout.addWidget(QLabel("Выберите файл для обработки:"))
                
                # Используем радиокнопки вместо чекбоксов для единичного выбора
                from PyQt5.QtWidgets import QRadioButton, QButtonGroup
                
                button_group = QButtonGroup(dialog)
                radio_buttons = []
                
                for i, f in enumerate(open_files):
                    radio_button = QRadioButton(os.path.basename(f))
                    radio_button.setProperty("fullPath", f)
                    if i == 0:  # Первый файл выбран по умолчанию
                        radio_button.setChecked(True)
                    button_group.addButton(radio_button)
                    radio_buttons.append(radio_button)
                    layout.addWidget(radio_button)
                
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addWidget(button_box)
                
                if dialog.exec_():
                    # Получаем выбранный файл
                    for radio_button in radio_buttons:
                        if radio_button.isChecked():
                            selected_file = radio_button.property("fullPath")
                            if self.excel_handler.set_halfs_file(selected_file):
                                self.halfs_file_label.setText(selected_file)
                                self.add_log(f"Выбран файл Половины: {os.path.basename(selected_file)}")
                                self.update_process_button()
                            break
                return
            
            # Если нет открытых файлов Половины - стандартный диалог
            file_filter = "Файлы половин (Половины*.xlsx);;Все файлы Excel (*.xlsx *.xls)"
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Выберите файл Половины", "", file_filter
            )
            
            if file_path:
                if self.excel_handler.set_halfs_file(file_path):
                    self.halfs_file_label.setText(file_path)
                    self.add_log(f"Выбран файл Половины: {os.path.basename(file_path)}")
                    self.update_process_button()
                    
        except Exception as e:
            self.add_log(f"Ошибка при выборе файла: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось выбрать файл: {str(e)}")
    
    def select_cyber_files(self):
        """Выбор файлов Cyber"""
        try:
            # Проверяем открытые файлы Excel с фильтром по типу 'cyber'
            open_files = self.excel_handler.get_opened_excel_files(file_type='cyber')
            
            if open_files:
                # Создаем диалоговое окно с чекбоксами
                dialog = QDialog(self)
                dialog.setWindowTitle("Выберите открытые файлы Cyber")
                dialog.setMinimumWidth(400)
                
                layout = QVBoxLayout(dialog)
                layout.addWidget(QLabel("Отметьте файлы для обработки:"))
                
                checkboxes = []
                for f in open_files:
                    checkbox = QCheckBox(os.path.basename(f))
                    checkbox.setProperty("fullPath", f)
                    checkbox.setChecked(True)  # По умолчанию выбраны все
                    checkboxes.append(checkbox)
                    layout.addWidget(checkbox)
                
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addWidget(button_box)
                
                if dialog.exec_():
                    # Получаем выбранные файлы
                    selected_files = []
                    for checkbox in checkboxes:
                        if checkbox.isChecked():
                            selected_files.append(checkbox.property("fullPath"))
                    
                    if selected_files:
                        # Очищаем старый список
                        self.cyber_files_list.clear()
                        self.excel_handler.clear_cyber_files()
                        
                        # Добавляем новые файлы
                        for file_path in selected_files:
                            if self.excel_handler.add_cyber_file(file_path):
                                self.cyber_files_list.addItem(os.path.basename(file_path))
                                self.add_log(f"Добавлен файл Cyber: {os.path.basename(file_path)}")
                        self.update_process_button()
                return
            
            # Если нет открытых файлов Cyber - стандартный диалог
            file_filter = "Файлы Cyber (Cyber*.xlsx);;Все файлы Excel (*.xlsx *.xls)"
            files, _ = QFileDialog.getOpenFileNames(
                self, "Выберите файлы Cyber", "", file_filter
            )
            
            if files:
                # Очищаем старый список
                self.cyber_files_list.clear()
                self.excel_handler.clear_cyber_files()
                
                for file_path in files:
                    if self.excel_handler.add_cyber_file(file_path):
                        self.cyber_files_list.addItem(os.path.basename(file_path))
                        self.add_log(f"Добавлен файл Cyber: {os.path.basename(file_path)}")
                self.update_process_button()
                
        except Exception as e:
            self.add_log(f"Ошибка при выборе файлов: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось выбрать файлы: {str(e)}")

    def clear_cyber_files(self):
        """Очистка списка файлов Cyber"""
        self.cyber_files_list.clear()
        self.excel_handler.clear_cyber_files()
        self.add_log("Список файлов Cyber очищен")
        self.update_process_button()
    
    def update_process_button(self):
        """Обновляет состояние кнопки запуска процесса"""
        has_halfs = self.excel_handler.halfs_file is not None
        has_cyber = len(self.excel_handler.cyber_files) > 0
        self.process_btn.setEnabled(has_halfs and has_cyber and self.excel_handler.available)
    
    def process_files(self):
        """Запуск процесса обработки файлов"""
        self.process_btn.setEnabled(False)
        self.select_halfs_btn.setEnabled(False)
        self.select_cyber_btn.setEnabled(False)
        self.clear_cyber_btn.setEnabled(False)
        
        self.progress_bar.setValue(0)
        self.status_label.setText("Выполняется обработка...")
        
        # Создаем и запускаем поток для обработки файлов
        self.processor_thread = ExcelProcessorThread(self.excel_handler)
        self.processor_thread.progress_signal.connect(self.update_progress)
        self.processor_thread.finished_signal.connect(self.processing_finished)
        self.processor_thread.error_signal.connect(self.processing_error)
        self.processor_thread.start()
    
    def update_progress(self, message, progress=-1):
        """Обновление прогресса обработки"""
        self.add_log(message)
        if progress >= 0:
            self.progress_bar.setValue(progress)
        self.status_label.setText(message)
    
    def processing_finished(self, message):
        """Завершение процесса обработки"""
        self.progress_bar.setValue(100)
        self.status_label.setText(message)
        self.add_log(message)
        
        self.process_btn.setEnabled(True)
        self.select_halfs_btn.setEnabled(True)
        self.select_cyber_btn.setEnabled(True)
        self.clear_cyber_btn.setEnabled(True)
        
        QMessageBox.information(self, "Обработка завершена", message)
    
    def processing_error(self, error_message):
        """Обработка ошибки при обработке файлов"""
        self.progress_bar.setValue(0)
        self.status_label.setText("Произошла ошибка!")
        self.add_log(f"ОШИБКА: {error_message}")
        
        self.process_btn.setEnabled(True)
        self.select_halfs_btn.setEnabled(True)
        self.select_cyber_btn.setEnabled(True)
        self.clear_cyber_btn.setEnabled(True)
        
        QMessageBox.critical(self, "Ошибка обработки", error_message)
    
    def add_log(self, message):
        """Добавляет сообщение в лог"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.addItem(f"[{timestamp}] {message}")
        self.log_text.scrollToBottom()

# Сначала добавьте класс потока для обработки данных (добавить ПЕРЕД классом QuartersPage)

class QuartersProcessorThread(QThread):
    """Поток для загрузки данных четвертей в фоне"""
    progress_signal = pyqtSignal(str, int)  # Сообщение и процент
    finished_signal = pyqtSignal(dict, int)  # Данные и количество турниров
    error_signal = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.data_cache = {}
        self.excel_app = None

    def safe_float(self, value):
        """Безопасное преобразование в float"""
        try:
            if value is None:
                return 0
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                return float(value.replace(',', '.'))
            return 0
        except:
            return 0

    def get_excel_app(self):
        """Безопасное получение приложения Excel с повторными попытками"""
        max_attempts = 5
        for attempt in range(max_attempts):
            try:
                if xw.apps:
                    apps_count = len(xw.apps)
                    self.progress_signal.emit(f"Найдено процессов Excel: {apps_count}", 5)
                    
                    # Проверяем каждый процесс Excel
                    for i, app in enumerate(xw.apps):
                        try:
                            pid = app.pid
                            self.progress_signal.emit(f"Проверка процесса Excel {i+1}/{apps_count} (PID: {pid})", 6)
                            
                            # Получаем список книг
                            try:
                                books = list(app.books)
                                self.progress_signal.emit(f"Книг в процессе {pid}: {len(books)}", 7)
                                
                                # Проверяем каждую книгу
                                for book in books:
                                    try:
                                        book_name = book.name.lower()
                                        full_path = book.fullname.lower()
                                        
                                        self.progress_signal.emit(f"Проверка книги: {book_name}", 8)
                                        
                                        # Проверяем имя файла разными способами
                                        if ("половин" in book_name or 
                                            "половины" in book_name or 
                                            "половин" in full_path or
                                            "половины" in full_path):
                                            
                                            self.progress_signal.emit(f"Найден файл Половины: {book_name}", 9)
                                            return app
                                            
                                    except Exception as e:
                                        self.progress_signal.emit(f"Ошибка при проверке книги: {str(e)}", 8)
                                        continue
                                        
                            except Exception as e:
                                self.progress_signal.emit(f"Ошибка при получении списка книг: {str(e)}", 7)
                                continue
                                
                        except Exception as e:
                            self.progress_signal.emit(f"Ошибка при проверке процесса Excel: {str(e)}", 6)
                            continue
                    
                    # Если прошли все процессы и не нашли файл
                    self.progress_signal.emit("Файл Половины не найден в открытых книгах", 10)
                    
                else:
                    self.progress_signal.emit("Не найдено запущенных процессов Excel", 5)
                
                time.sleep(1)
                
            except Exception as e:
                self.progress_signal.emit(f"Ошибка при поиске Excel: {str(e)}", 5)
                time.sleep(1)
        
        return None

    def run(self):
        try:
            if not xw:
                self.error_signal.emit("Библиотека xlwings не установлена")
                return

            self.progress_signal.emit("Подключение к Excel...", 5)
            
            # Получаем Excel с подробным логированием
            self.excel_app = self.get_excel_app()
            if not self.excel_app:
                self.error_signal.emit("Не удалось найти файл 'Половины' среди открытых файлов Excel")
                return

            # Пробуем получить список файлов Половины
            halfs_files = []
            try:
                for book in self.excel_app.books:
                    try:
                        book_name = book.name.lower()
                        if "половин" in book_name:
                            halfs_files.append(book)
                            self.progress_signal.emit(f"Добавлен файл: {book.name}", 15)
                    except Exception as e:
                        self.progress_signal.emit(f"Ошибка при проверке книги: {str(e)}", 15)
                        continue
            except Exception as e:
                self.error_signal.emit(f"Ошибка при получении списка книг: {str(e)}")
                return

            if not halfs_files:
                self.error_signal.emit("Файлы 'Половины' не найдены среди открытых файлов Excel")
                return

            # Продолжаем выполнение с найденными файлами...

            self.progress_signal.emit(f"Найдено файлов: {len(halfs_files)}", 15)

            quarters_data = {}
            total_tournaments = 0

            # Обработка каждого файла
            for book_idx, book in enumerate(halfs_files):
                try:
                    book_name = os.path.basename(book.fullname)
                    self.progress_signal.emit(f"Обработка файла: {book_name}", 20)

                    # Получаем список листов
                    sheets = [sheet for sheet in book.sheets 
                             if sheet.name not in ["Оглавление", "Шаблон", "Halfs Champs"]]

                    for sheet in sheets:
                        try:
                            # Читаем данные
                            data_range = sheet.range("AD1:AJ200").value

                            # Ищем строку с "Team 1"
                            team1_row = None
                            for i, row in enumerate(data_range):
                                if row and row[0] == "Team 1":
                                    team1_row = i
                                    break

                            if team1_row is None:
                                continue

                            # Собираем матчи
                            matches = []
                            row = team1_row + 1
                            match_number = 1

                            while row < len(data_range):
                                row_data = data_range[row]
                                if not row_data or not row_data[0] or not row_data[1]:
                                    break

                                match_data = {
                                    "row": match_number,
                                    "team1": str(row_data[0] or ""),
                                    "team2": str(row_data[1] or ""),
                                    "total": self.safe_float(row_data[2]),
                                    "q1": self.safe_float(row_data[3]),
                                    "q2": self.safe_float(row_data[4]),
                                    "q3": self.safe_float(row_data[5]),
                                    "q4": self.safe_float(row_data[6])
                                }

                                # ручные поля для сохранения введённых значений
                                match_data["total_manual"] = None
                                match_data["half2_manual"] = None   # "2 Half" (ручной ввод)

                                matches.append(match_data)
                                row += 1
                                match_number += 1

                            if matches:
                                quarters_data[sheet.name] = {
                                    "matches": matches,
                                    "file": book_name,
                                    "header_row": team1_row + 1
                                }
                                total_tournaments += 1

                                self.progress_signal.emit(
                                    f"Обработан турнир: {sheet.name} ({len(matches)} матчей)", 
                                    20 + (70 * total_tournaments // len(sheets))
                                )

                        except Exception as e:
                            logging.error(f"Ошибка при обработке листа {sheet.name}: {str(e)}")
                            continue

                except Exception as e:
                    logging.error(f"Ошибка при обработке файла {book_name}: {str(e)}")
                    continue

            if total_tournaments > 0:
                self.progress_signal.emit(
                    f"Обработка завершена. Найдено турниров: {total_tournaments}", 
                    100
                )
                self.finished_signal.emit(quarters_data, total_tournaments)
            else:
                self.error_signal.emit("Не найдено данных о четвертях в открытых файлах")

        except Exception as e:
            error_msg = f"Критическая ошибка: {str(e)}"
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            self.error_signal.emit(error_msg)

        finally:
            # Очищаем ссылки на объекты Excel
            self.excel_app = None

class QuartersPage(QWidget):
    """Страница для работы с четвертями"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.quarters_data = {}  # Словарь для хранения данных по турнирам
        self.current_table = None
        self.current_tournament = None
        self.processor_thread = None
        self.betsapi = BetsAPIHandler()
        self.bets_results_data = {}  # Для хранения результатов анализа линий

        # Словарь для хранения рассчитанных ставок по всем турнирам
        # Формат: {tournament_name: {"half": [...], "q1": [...], "missing": [...]}}
        self.bets_results_data = {}
        
        # Путь к файлу для сохранения данных
        self.data_file_path = os.path.join(
            os.path.expanduser("~"), 
            "AppData", 
            "Local", 
            "ExcelAnalyzer",
            "quarters_data.json"
        )
        
        # Создаем директорию если её нет
        os.makedirs(os.path.dirname(self.data_file_path), exist_ok=True)
        
        # ВАЖНО: сначала создаем UI
        self.setup_ui()
        
        # ПОТОМ загружаем сохраненные данные
        self.load_saved_data()

    def keyPressEvent(self, event):
        """Обработка нажатий клавиш"""
        if event.key() == Qt.Key_F9:
            self.recalculate_second_half()
        super().keyPressEvent(event)

    def recalculate_second_half(self):
        """Пересчет значений 3й и 4й четверти на основе нового тотала второй половины"""
        if not self.current_table:
            return
            
        for row in range(self.current_table.rowCount()):
            # Получаем текущие значения 3й и 4й четверти
            # 3я и 4я четверти находятся в колонках 7 и 8 соответственно (с учетом колонки 1 HALF)
            q3_item = self.current_table.item(row, 7)  # 3я четверть
            q4_item = self.current_table.item(row, 8)  # 4я четверть
            
            # Получаем значение нового тотала для второй половины
            total_2h_item = self.current_table.item(row, 3)
            
            if not all([q3_item, q4_item, total_2h_item]):
                continue
                
            try:
                # Получаем текущие значения
                q3_old = float(q3_item.text().split('\n')[0])
                q4_old = float(q4_item.text().split('\n')[0])
                new_total = float(total_2h_item.text())
                
                # Считаем сумму второй половины
                old_total = q3_old + q4_old
                
                # Считаем процентные доли
                q3_percent = q3_old / old_total if old_total else 0
                q4_percent = q4_old / old_total if old_total else 0
                
                # Рассчитываем новые значения
                q3_new = round(new_total * q3_percent, 1)
                q4_new = round(new_total * q4_percent, 1)
                
                # Обновляем значения в таблице
                theme = QApplication.instance().property("theme") or "dark"
                text_color = QColor(200, 200, 200) if theme == "dark" else QColor(0, 0, 0)
                
                # Цвета для четвертей
                q3_color = QColor(60, 40, 40) if theme == "dark" else QColor(240, 220, 220)
                q4_color = QColor(40, 40, 60) if theme == "dark" else QColor(220, 220, 240)
                
                # Обновляем 3ю четверть
                q3_text = f"{q3_new:.1f}\n({q3_percent*100:.1f}%)"
                q3_new_item = QTableWidgetItem(q3_text)
                q3_new_item.setTextAlignment(Qt.AlignCenter)
                q3_new_item.setFont(QFont("Arial", 10))
                q3_new_item.setBackground(QBrush(q3_color))
                q3_new_item.setForeground(text_color)
                self.current_table.setItem(row, 7, q3_new_item)
                
                # Обновляем 4ю четверть
                q4_text = f"{q4_new:.1f}\n({q4_percent*100:.1f}%)"
                q4_new_item = QTableWidgetItem(q4_text)
                q4_new_item.setTextAlignment(Qt.AlignCenter)
                q4_new_item.setFont(QFont("Arial", 10))
                q4_new_item.setBackground(QBrush(q4_color))
                q4_new_item.setForeground(text_color)
                self.current_table.setItem(row, 8, q4_new_item)
                
                # Обновляем данные в self.quarters_data
                if self.current_tournament:
                    match_data = self.quarters_data[self.current_tournament]["matches"][row]
                    match_data["q3"] = q3_new
                    match_data["q4"] = q4_new
                
            except (ValueError, ZeroDivisionError, AttributeError) as e:
                self.add_log(f"Ошибка при пересчете строки {row+1}: {str(e)}")
                continue        
        
    def setup_ui(self):
        """Настройка пользовательского интерфейса"""
        # Основной layout
        main_layout = QVBoxLayout(self)
        
        # Заголовок
        title_label = QLabel("Анализ четвертей")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        main_layout.addSpacing(20)
        
        # Кнопки управления
        control_layout = QHBoxLayout()
        
        self.load_btn = QPushButton("Загрузить данные из открытых файлов")
        self.load_btn.setMinimumHeight(40)
        self.load_btn.clicked.connect(self.load_quarters_data)
        
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.setMinimumHeight(40)
        self.refresh_btn.clicked.connect(self.load_quarters_data)
        
        # НОВАЯ КНОПКА для обновления тоталов через API
        self.update_totals_btn = QPushButton("🔄 Обновить тоталы (BetsAPI)")
        self.update_totals_btn.setMinimumHeight(40)
        self.update_totals_btn.clicked.connect(self.update_totals_from_api)
        self.update_totals_btn.setEnabled(False)
        
        self.clear_cache_btn = QPushButton("Очистить сохраненные данные")
        self.clear_cache_btn.setMinimumHeight(40)
        self.clear_cache_btn.clicked.connect(self.clear_saved_data)
        
        control_layout.addWidget(self.load_btn)
        control_layout.addWidget(self.refresh_btn)
        control_layout.addWidget(self.update_totals_btn)  # ДОБАВИЛИ НОВУЮ КНОПКУ
        control_layout.addWidget(self.clear_cache_btn)

        # Кнопка для загрузки файла линий (ставки на половины/четверти)
        # При нажатии вызывается метод загрузки и анализа линий
        main_layout.addLayout(control_layout)
        main_layout.addSpacing(10)
        
        # Статус с информацией о последнем обновлении
        self.status_label = QLabel("Откройте файл 'Половины' в Excel и нажмите 'Загрузить данные'")
        self.status_label.setWordWrap(True)
        main_layout.addWidget(self.status_label)
        
        # Метка с датой последнего обновления
        self.last_update_label = QLabel("")
        self.last_update_label.setStyleSheet("color: gray; font-size: 10px;")
        main_layout.addWidget(self.last_update_label)
        main_layout.addSpacing(10)
        
        # Поиск турнира
        search_layout = QHBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите название турнира...")
        self.search_input.setMinimumWidth(400)
        self.search_input.returnPressed.connect(self.search_tournament)
        
        # Автодополнение для поиска
        self.tournament_completer = QCompleter()
        self.tournament_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.tournament_completer.setFilterMode(Qt.MatchContains)
        self.tournament_completer.setCompletionMode(QCompleter.PopupCompletion)
        self.search_input.setCompleter(self.tournament_completer)
        
        self.search_btn = QPushButton("Поиск")
        self.search_btn.clicked.connect(self.search_tournament)
        
        search_layout.addWidget(QLabel("Поиск турнира:"))
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_btn)
        search_layout.addStretch()
        
        main_layout.addLayout(search_layout)
        main_layout.addSpacing(10)
        
        # Список доступных турниров
        self.available_label = QLabel("Доступные турниры: нет данных")
        self.available_label.setFont(QFont("Arial", 10))
        main_layout.addWidget(self.available_label)
        main_layout.addSpacing(10)
        
        # Область для отображения таблицы (верхняя часть)
        self.table_container = QWidget()
        self.table_layout = QVBoxLayout(self.table_container)
        self.table_layout.setContentsMargins(0, 0, 0, 0)

        # Скролл для таблицы
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(self.table_container)
        # Добавляем scroll_area непосредственно в основной layout
        main_layout.addWidget(scroll_area, 3)

        # Журнал выполнения – скрыт по умолчанию
        self.log_label = QLabel("Журнал выполнения:")
        self.log_label.setFont(QFont("Arial", 10, QFont.Bold))
        main_layout.addWidget(self.log_label)
        
        self.log_text = QListWidget()
        self.log_text.setStyleSheet("""
            QListWidget {
                background-color: #1a1a1f;
                border: 1px solid #333339;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 10px;
            }
            QListWidget::item {
                padding: 2px;
                border-bottom: 1px solid #2a2a2f;
            }
        """)
        main_layout.addWidget(self.log_text, 1)
        
        # Скрываем лог по умолчанию
        self.log_label.hide()
        self.log_text.hide()

        # Добавляем кнопку для показа/скрытия лога
        self.toggle_log_btn = QPushButton("Показать лог")
        self.toggle_log_btn.setCheckable(True)
        self.toggle_log_btn.setChecked(False)
        self.toggle_log_btn.toggled.connect(self.toggle_log_visibility)
        main_layout.addWidget(self.toggle_log_btn)

        # Добавляем начальное сообщение в лог
        self.add_log("Система готова к работе")

    def on_cell_changed(self, item):
        """Обработка изменения значения в ячейке"""
        try:
            if not self.current_table:
                return
                
            column = item.column()
            row = item.row()
            
            # Получаем значение из ячейки
            new_value_str = item.text().strip().replace(',', '.')
            
            # Если ячейка пустая, игнорируем изменение
            if not new_value_str:
                return
                
            try:
                new_value = float(new_value_str)
            except ValueError:
                self.add_log("Ошибка: Введите числовое значение")
                return

            # Обработка колонки Total (индекс 3)
            if column == 3:
                row = item.row()
                new_total = new_value  # из парсинга выше

                # сохраняем ручное значение Total
                if self.current_tournament:
                    match_data = self.quarters_data[self.current_tournament]["matches"][row]
                    match_data["total_manual"] = new_total
                    match_data["total"] = new_total          # чтобы расчёты брали его
                    self.save_data_to_file()

                # колонки четвертей: 1Q..4Q (с учетом колонки 1 HALF в позиции 6)
                # 1Q -> 4, 2Q -> 5, 3Q -> 7, 4Q -> 8
                quarter_cols = [4, 5, 7, 8]

                def parse_percent(txt: str, val_fallback: float, sum_fallback: float) -> float:
                    # пробуем вытащить "(xx.x%)" из подписи ячейки
                    import re
                    m = re.search(r"\(([\d\.,]+)%\)", txt or "")
                    if m:
                        return float(m.group(1).replace(",", ".")) / 100.0
                    # иначе считаем из текущих значений
                    return (val_fallback / sum_fallback) if sum_fallback > 0 else 0.0

                # читаем текущие значения и проценты
                q_vals, q_pcts = [], []
                sum_old = 0.0
                for c in quarter_cols:
                    cell = self.current_table.item(row, c)
                    v_txt = (cell.text().split("\n")[0] if cell and cell.text() else "")
                    v = float(v_txt.replace(",", ".")) if v_txt else 0.0
                    q_vals.append(v)
                    sum_old += v

                for c, v in zip(quarter_cols, q_vals):
                    cell = self.current_table.item(row, c)
                    txt = cell.text() if cell else ""
                    q_pcts.append(parse_percent(txt, v, sum_old))

                # пересчёт всех четвертей по новым Total и уже посчитанным %
                new_q = [round(new_total * p, 1) for p in q_pcts]

                # чтобы не ловить рекурсивный itemChanged
                self.current_table.blockSignals(True)
                try:
                    # Обновляем значения четвертей в таблице
                    for c, val, pct in zip(quarter_cols, new_q, q_pcts):
                        cell = self.current_table.item(row, c)
                        if cell is None:
                            cell = QTableWidgetItem()
                            self.current_table.setItem(row, c, cell)
                        cell.setText(f"{val:.1f}\n({pct*100:.1f}%)")

                    # После обновления четвертей пересчитываем и обновляем ячейку 1 HALF (колонка 6)
                    half1_val = new_q[0] + new_q[1]
                    half1_item = QTableWidgetItem(f"{half1_val:.1f}" if half1_val else "")
                    half1_item.setTextAlignment(Qt.AlignCenter)
                    # Устанавливаем цвет аналогичный определению в display_tournament_quarters
                    theme = QApplication.instance().property("theme") or "dark"
                    half1_color = QColor(50, 80, 50) if theme == "dark" else QColor(220, 230, 220)
                    half1_item.setBackground(half1_color)
                    text_color_local = QColor(200, 200, 200) if theme == "dark" else QColor(0, 0, 0)
                    half1_item.setForeground(text_color_local)
                    self.current_table.setItem(row, 6, half1_item)

                    # Обновляем total в данных и сами значения четвертей (НЕ трогаем 2 Half, Q3 NEW, Q4 NEW)
                    if self.current_tournament:
                        match_data = self.quarters_data[self.current_tournament]["matches"][row]
                        match_data["total"] = new_total
                        match_data["q1"], match_data["q2"], match_data["q3"], match_data["q4"] = new_q
                        # обновляем сохранённое значение первой половины
                        match_data["half1"] = half1_val
                        self.save_data_to_file()
                finally:
                    self.current_table.blockSignals(False)

            
            # Обработка колонки 2 Half (колонка 9 после вставки 1 HALF)
            elif column == 9:
                try:
                    # сохранить ручной ввод "2 Half"
                    if self.current_tournament:
                        md = self.quarters_data[self.current_tournament]["matches"][row]
                        md["half2_manual"] = new_value
                        md["half2"] = new_value      # если используешь это поле в формулах
                        self.save_data_to_file()
                    
                    # Получаем значения Q3 и Q4 для пересчета
                    q3_item = self.current_table.item(row, 7)  # 3Q (колонка 7)
                    q4_item = self.current_table.item(row, 8)  # 4Q (колонка 8)
                    
                    if q3_item and q4_item:
                        # Получаем текущие значения
                        q3_text = q3_item.text().split('\n')[0]
                        q4_text = q4_item.text().split('\n')[0]
                        
                        q3_val = float(q3_text) if q3_text else 0
                        q4_val = float(q4_text) if q4_text else 0
                        
                        old_total = q3_val + q4_val
                        if old_total > 0:
                            # Рассчитываем пропорции
                            q3_percent = q3_val / old_total
                            q4_percent = q4_val / old_total
                            
                            # Рассчитываем новые значения
                            new_q3 = round(new_value * q3_percent, 1)
                            new_q4 = round(new_value * q4_percent, 1)
                            
                            # Обновляем значения в новых колонках Q3 NEW и Q4 NEW
                            q3_new_item = QTableWidgetItem(f"{new_q3:.1f}")
                            q4_new_item = QTableWidgetItem(f"{new_q4:.1f}")
                            
                            q3_new_item.setTextAlignment(Qt.AlignCenter)
                            q4_new_item.setTextAlignment(Qt.AlignCenter)
                            
                            self.current_table.setItem(row, 10, q3_new_item)  # Q3 NEW (колонка 10)
                            self.current_table.setItem(row, 11, q4_new_item)  # Q4 NEW (колонка 11)

                            # Сохраняем ручной 2 Half и пересчитанные NEW
                            if self.current_tournament:
                                match_data = self.quarters_data[self.current_tournament]["matches"][row]

                                # half2
                                match_data["half2_manual"] = new_value
                                match_data["half2"] = new_value

                                # q3_new / q4_new — достанем числа (если у тебя есть уже переменные new_q3/new_q4 — используй их)
                                def _num_from_item(it):
                                    txt = it.text() if it else ""
                                    txt = txt.split("\n", 1)[0] if txt else ""
                                    return float(txt.replace(",", ".")) if txt else 0.0

                                match_data["q3_new"] = _num_from_item(q3_new_item)
                                match_data["q4_new"] = _num_from_item(q4_new_item)

                                self.save_data_to_file()

                            
                except Exception as e:
                    self.add_log(f"Ошибка при расчете новых значений: {str(e)}")
                    
        except Exception as e:
            self.add_log(f"Ошибка в on_cell_changed: {str(e)}")

    
    def update_quarter_value(self, row, col, value, percentage):
        """Обновляет значение четверти в таблице"""
        if not self.current_table:
            return
            
        theme = QApplication.instance().property("theme") or "dark"
        text_color = QColor(200, 200, 200) if theme == "dark" else QColor(0, 0, 0)
        
        # Определяем цвет фона для четверти (с учетом смещения колонки 1 HALF)
        # 3Q располагается в колонке 7, 4Q — в колонке 8
        if col == 7:  # 3Q
            bg_color = QColor(60, 40, 40) if theme == "dark" else QColor(240, 220, 220)
        else:  # 4Q
            bg_color = QColor(40, 40, 60) if theme == "dark" else QColor(220, 220, 240)
        
        # Создаем новый item
        display_text = f"{value:.1f}\n({percentage:.1f}%)"
        item = QTableWidgetItem(display_text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("Arial", 12))
        item.setBackground(QBrush(bg_color))
        item.setForeground(text_color)
        
        self.current_table.setItem(row, col, item)

    def toggle_log_visibility(self, checked):
        """
        Показывает или скрывает журнал выполнения в зависимости от состояния кнопки.
        Если checked=True, лог отображается и кнопка меняет текст на «Скрыть лог».
        Если checked=False, лог скрывается и кнопка меняет текст на «Показать лог».
        """
        try:
            if checked:
                self.log_label.show()
                self.log_text.show()
                self.toggle_log_btn.setText("Скрыть лог")
            else:
                self.log_label.hide()
                self.log_text.hide()
                self.toggle_log_btn.setText("Показать лог")
        except Exception:
            pass
    
    def save_data_to_file(self):
        """Сохраняет данные в JSON файл"""
        try:
            # Проверяем что есть данные для сохранения
            if not self.quarters_data:
                self.add_log("Нет данных для сохранения")
                return False
            
            # Подготавливаем данные для сохранения
            save_data = {
                "last_update": datetime.now().isoformat(),
                "tournaments": self.quarters_data,
                "version": "1.0"  # Добавим версию для совместимости
            }
            
            # Создаем временный файл сначала
            temp_file = self.data_file_path + ".tmp"
            
            # Сохраняем во временный файл
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
            
            # Проверяем что файл создан и не пустой
            if os.path.exists(temp_file) and os.path.getsize(temp_file) > 0:
                # Заменяем старый файл новым
                if os.path.exists(self.data_file_path):
                    os.remove(self.data_file_path)
                os.rename(temp_file, self.data_file_path)
                
                file_size = os.path.getsize(self.data_file_path)
                self.add_log(f"✓ Данные сохранены в кэш: {len(self.quarters_data)} турниров, размер файла: {file_size} байт")
                self.add_log(f"Путь к файлу: {self.data_file_path}")
                return True
            else:
                self.add_log("Ошибка: временный файл не создан или пустой")
                return False
            
        except Exception as e:
            self.add_log(f"✗ Ошибка при сохранении данных: {str(e)}")
            logging.error(f"Ошибка при сохранении данных в файл: {str(e)}\n{traceback.format_exc()}")
            
            # Удаляем временный файл если он остался
            temp_file = self.data_file_path + ".tmp"
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
            
            return False
    
    def load_saved_data(self):
        """Загружает сохраненные данные из JSON файла"""
        try:
            # Проверяем существование файла
            if not os.path.exists(self.data_file_path):
                self.add_log(f"Файл с данными не найден: {self.data_file_path}")
                return False
            
            # Проверяем размер файла
            file_size = os.path.getsize(self.data_file_path)
            self.add_log(f"Найден файл с данными, размер: {file_size} байт")
            
            # Читаем файл
            with open(self.data_file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                if not content:
                    self.add_log("Файл с данными пуст")
                    return False
                
                save_data = json.loads(content)
            
            # Восстанавливаем данные
            self.quarters_data = save_data.get("tournaments", {})
            last_update = save_data.get("last_update", "")
            
            if self.quarters_data:
                # Обновляем интерфейс
                tournament_names = sorted(self.quarters_data.keys())
                total_tournaments = len(tournament_names)
                
                self.available_label.setText(
                    f"Доступные турниры ({total_tournaments}): {', '.join(tournament_names[:5])}"
                    + ("..." if len(tournament_names) > 5 else "")
                )
                
                # Обновляем автодополнение
                model = QStringListModel()
                model.setStringList(tournament_names)
                self.tournament_completer.setModel(model)
                
                # Показываем дату последнего обновления
                if last_update:
                    try:
                        update_time = datetime.fromisoformat(last_update)
                        formatted_time = update_time.strftime("%d.%m.%Y %H:%M:%S")
                        self.last_update_label.setText(f"Последнее обновление: {formatted_time}")
                    except Exception as e:
                        self.add_log(f"Ошибка при парсинге даты: {str(e)}")
                
                self.status_label.setText(f"Загружено {total_tournaments} турниров из кэша")
                self.add_log(f"✓ Успешно загружены сохраненные данные: {total_tournaments} турниров")
                
                # Показываем список турниров в логе
                self.add_log("Доступные турниры из кэша:")
                for i, name in enumerate(tournament_names[:10], 1):  # Показываем первые 10
                    matches_count = len(self.quarters_data[name]["matches"])
                    self.add_log(f"  {i}. {name} ({matches_count} матчей)")
                
                if len(tournament_names) > 10:
                    self.add_log(f"  ... и еще {len(tournament_names) - 10} турниров")

                # Сообщаем другим разделам о том, что данные загружены из кэша
                try:
                    cb = getattr(self, "on_data_updated", None)
                    if callable(cb):
                        cb(self.quarters_data)
                except Exception:
                    pass
                
                return True
            else:
                self.add_log("В файле нет данных о турнирах")
                return False
                
        except json.JSONDecodeError as e:
            self.add_log(f"Ошибка при разборе JSON: {str(e)}")
            self.add_log(f"Файл поврежден, удаляем его")
            # Удаляем поврежденный файл
            try:
                os.remove(self.data_file_path)
            except:
                pass
            return False
            
        except Exception as e:
            self.add_log(f"Ошибка при загрузке сохраненных данных: {str(e)}")
            logging.error(f"Ошибка при загрузке данных из файла: {str(e)}\n{traceback.format_exc()}")
            return False
    
    def clear_saved_data(self):
        """Очищает сохраненные данные"""
        try:
            reply = QMessageBox.question(
                self, 
                "Подтверждение", 
                "Вы уверены, что хотите удалить все сохраненные данные?\n"
                "Это действие нельзя отменить.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # --- СБРОС РУЧНЫХ ЗНАЧЕНИЙ (добавлено) ---
                for tname, tdata in (self.quarters_data or {}).items():
                    for md in tdata.get("matches", []):
                        md["total_manual"] = None
                        md["half2_manual"] = None
                        md["q3_new"] = None        # добавить
                        md["q4_new"] = None        # добавить
                # --- конец добавленного блока ---

                # Удаляем файл с данными
                if os.path.exists(self.data_file_path):
                    os.remove(self.data_file_path)
                
                # Очищаем текущие данные
                self.quarters_data = {}
                
                # Обновляем интерфейс
                self.available_label.setText("Доступные турниры: нет данных")
                self.last_update_label.setText("")
                self.status_label.setText("Сохраненные данные удалены")
                
                # Очищаем автодополнение
                model = QStringListModel()
                model.setStringList([])
                self.tournament_completer.setModel(model)
                
                # Очищаем таблицу если отображена
                self.clear_layout(self.table_layout)
                
                self.add_log("Сохраненные данные успешно удалены")
                QMessageBox.information(self, "Успешно", "Сохраненные данные удалены")
                
        except Exception as e:
            error_msg = f"Ошибка при удалении данных: {str(e)}"
            self.add_log(f"ОШИБКА: {error_msg}")
            logging.error(error_msg)
            QMessageBox.critical(self, "Ошибка", error_msg)

    
    def add_log(self, message):
        """Добавляет сообщение в журнал выполнения"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.addItem(f"[{timestamp}] {message}")
        self.log_text.scrollToBottom()
        
        # Ограничиваем количество строк в логе (например, 100)
        while self.log_text.count() > 100:
            self.log_text.takeItem(0)
    
    def clear_log(self):
        """Очищает журнал выполнения"""
        self.log_text.clear()
        self.add_log("Журнал очищен")
        
    def load_quarters_data(self):
        """Запускает загрузку данных в отдельном потоке"""
        # Спрашиваем подтверждение если есть сохраненные данные
        if self.quarters_data:
            reply = QMessageBox.question(
                self,
                "Подтверждение",
                "Текущие данные будут заменены новыми.\nПродолжить?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.No:
                return
        
        # Очищаем лог и добавляем стартовое сообщение
        self.clear_log()
        self.add_log("Начало загрузки данных из открытых файлов Excel")
        
        # Блокируем кнопки
        self.load_btn.setEnabled(False)
        self.refresh_btn.setEnabled(False)
        self.search_btn.setEnabled(False)
        self.clear_cache_btn.setEnabled(False)
        
        # Создаем и запускаем поток
        self.processor_thread = QuartersProcessorThread()
        self.processor_thread.progress_signal.connect(self.update_progress)
        self.processor_thread.finished_signal.connect(self.processing_finished)
        self.processor_thread.error_signal.connect(self.processing_error)
        self.processor_thread.start()
    
    def update_progress(self, message, progress):
        """Обновляет прогресс и добавляет сообщение в лог"""
        # Добавляем сообщение в лог
        self.add_log(f"{message} [{progress}%]")
        
        # Обновляем основной статус
        self.status_label.setText(f"Загрузка: {message} ({progress}%)")
    
    def processing_finished(self, data, total_tournaments):
        """Обработка завершения загрузки"""
        self.quarters_data = data

        # Сообщаем другим разделам о том, что данные обновились (например, "Половины и четверти")
        try:
            cb = getattr(self, "on_data_updated", None)
            if callable(cb):
                cb(self.quarters_data)
        except Exception:
            pass
        
        # Сохраняем данные в файл
        if self.save_data_to_file():
            self.add_log("Данные автоматически сохранены")
        
        # Обновляем дату последнего обновления
        current_time = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        self.last_update_label.setText(f"Последнее обновление: {current_time}")
        
        # Разблокируем кнопки
        self.load_btn.setEnabled(True)
        self.refresh_btn.setEnabled(True)
        self.search_btn.setEnabled(True)
        self.clear_cache_btn.setEnabled(True)
        
        # Добавляем в лог результат
        self.add_log(f"Загрузка завершена успешно!")
        self.add_log(f"Найдено турниров: {total_tournaments}")
        
        # Обновляем интерфейс
        if total_tournaments > 0:
            tournament_names = sorted(self.quarters_data.keys())
            
            # Добавляем список турниров в лог
            self.add_log("Доступные турниры:")
            for i, name in enumerate(tournament_names, 1):
                matches_count = len(self.quarters_data[name]["matches"])
                self.add_log(f"  {i}. {name} ({matches_count} матчей)")
            
            self.available_label.setText(
                f"Доступные турниры ({total_tournaments}): {', '.join(tournament_names[:5])}"
                + ("..." if len(tournament_names) > 5 else "")
            )
            
            # Обновляем автодополнение
            model = QStringListModel()
            model.setStringList(tournament_names)
            self.tournament_completer.setModel(model)
            
            self.status_label.setText(f"Загружено {total_tournaments} турниров")
            
            # Показываем уведомление
            QMessageBox.information(self, "Успешно", 
                f"Загружено {total_tournaments} турниров\n"
                f"Данные сохранены и будут доступны при следующем запуске программы")
        else:
            self.status_label.setText("Не найдено данных о четвертях в открытых файлах")
            self.available_label.setText("Доступные турниры: нет данных")
            self.add_log("ВНИМАНИЕ: Не найдено данных о четвертях")
            self.add_log("Убедитесь, что в файлах есть листы с таблицами четвертей")
    
    def processing_error(self, error_message):
        """Обработка ошибки загрузки"""
        # Разблокируем кнопки
        self.load_btn.setEnabled(True)
        self.refresh_btn.setEnabled(True)
        self.search_btn.setEnabled(True)
        self.clear_cache_btn.setEnabled(True)
        
        # Добавляем ошибку в лог
        self.add_log(f"ОШИБКА: {error_message}")
        
        self.status_label.setText(f"Ошибка при загрузке данных")
        QMessageBox.critical(self, "Ошибка", error_message)
    
    def search_tournament(self):
        """Поиск и отображение данных турнира"""
        tournament_name = self.search_input.text().strip()
        logging.info(f"=== Начало поиска турнира: {tournament_name} ===")
        
        try:
            if not tournament_name:
                logging.warning("Пустое название турнира")
                QMessageBox.warning(self, "Предупреждение", "Введите название турнира")
                return
            
            logging.info(f"Текущие загруженные данные: {bool(self.quarters_data)}")
            if not self.quarters_data:
                logging.error("self.quarters_data пуст")
                QMessageBox.warning(self, "Предупреждение", "Нет загруженных данных. Сначала загрузите данные из Excel")
                return
            
            # Выводим список доступных турниров
            logging.info(f"Доступные турниры: {list(self.quarters_data.keys())}")
            
            if tournament_name not in self.quarters_data:
                logging.warning(f"Турнир '{tournament_name}' не найден в списке")
                similar = [name for name in self.quarters_data.keys() 
                        if tournament_name.lower() in name.lower()]
                logging.info(f"Похожие турниры: {similar}")
                
                if similar:
                    msg = f"Турнир '{tournament_name}' не найден.\n\nПохожие турниры:\n"
                    msg += "\n".join(similar[:5])
                    QMessageBox.information(self, "Турнир не найден", msg)
                else:
                    QMessageBox.warning(self, "Не найдено", f"Турнир '{tournament_name}' не найден")
                return
            
            # Получаем данные турнира
            logging.info(f"Турнир '{tournament_name}' найден, получаем данные...")
            tournament_data = self.quarters_data[tournament_name]
            
            # Проверяем структуру данных
            logging.info(f"Проверка структуры данных турнира:")
            logging.info(f"Тип данных: {type(tournament_data)}")
            logging.info(f"Ключи в данных: {tournament_data.keys() if isinstance(tournament_data, dict) else 'не словарь'}")
            
            if not isinstance(tournament_data, dict):
                raise ValueError(f"Неверный формат данных турнира: {type(tournament_data)}")
            
            if 'matches' not in tournament_data:
                raise ValueError("В данных турнира отсутствует ключ 'matches'")
            
            logging.info(f"Количество матчей: {len(tournament_data['matches'])}")
            if tournament_data['matches']:
                logging.info(f"Пример первого матча: {tournament_data['matches'][0]}")
            
            # Отображаем данные
            logging.info(f"Вызываем display_tournament_quarters...")
            self.display_tournament_quarters(tournament_name, tournament_data)
            
        except Exception as e:
            error_msg = f"Ошибка при поиске турнира: {str(e)}\n{traceback.format_exc()}"
            logging.error(error_msg)
            QMessageBox.critical(self, "Ошибка", str(e))
    
    def display_tournament_quarters(self, tournament_name, tournament_data):
        """Отображает таблицу четвертей для турнира"""
        try:
            logging.info("=== Начало display_tournament_quarters ===")
            
            # Проверяем наличие table_layout
            if not hasattr(self, 'table_layout'):
                raise RuntimeError("table_layout не инициализирован")
            if self.table_layout is None:
                raise RuntimeError("table_layout is None")

            logging.info("Шаг 1: Очистка предыдущей таблицы")
            if self.table_layout.count() > 0:
                logging.info(f"В layout {self.table_layout.count()} элементов")
                self.clear_layout(self.table_layout)

            # Обновляем текущий турнир.  
            # Не скрываем результаты ставок сразу – вместо этого в конце
            # отобразим ставки для нового выбранного турнира, если они есть.
            self.current_tournament = tournament_name
            logging.info("Шаг 1 завершен успешно")

            # Создаем контейнер
            logging.info("Создание контейнера для таблицы")
            table_container = QWidget()
            container_layout = QVBoxLayout(table_container)
            container_layout.setContentsMargins(10, 10, 10, 10)
            container_layout.setSpacing(10)

            # Заголовок и информация
            title_label = QLabel(f"Четверти турнира: {tournament_name}")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(title_label)

            info_label = QLabel(f"Файл: {tournament_data.get('file', 'Неизвестно')}")
            info_label.setFont(QFont("Arial", 10))
            info_label.setAlignment(Qt.AlignCenter)
            container_layout.addWidget(info_label)

            # Проверяем наличие матчей
            matches = tournament_data.get("matches", [])
            if not matches:
                logging.warning("Нет данных о матчах")
                no_data_label = QLabel("Нет данных о матчах")
                no_data_label.setAlignment(Qt.AlignCenter)
                container_layout.addWidget(no_data_label)
                self.table_layout.addWidget(table_container)
                return

            logging.info(f"Найдено матчей: {len(matches)}")

            # Создаем таблицу
            table = QTableWidget()
            self.current_table = table
            
            # Настраиваем таблицу
            table.setRowCount(len(matches))
            # Колонки: добавляем "1 HALF" после 2Q, всего 12 колонок
            table.setColumnCount(12)

            headers = [
                "№", "Team 1", "Team 2", "Total", "1Q", "2Q", "1 HALF", "3Q", "4Q", "2 Half", "Q3 NEW", "Q4 NEW"
            ]
            table.setHorizontalHeaderLabels(headers)

            # Настройка цветов
            theme = QApplication.instance().property("theme") or "dark"
            text_color = QColor(200, 200, 200) if theme == "dark" else QColor(0, 0, 0)
            
            # Цвета для четвертей
            q_colors = [
                QColor(40, 60, 40) if theme == "dark" else QColor(220, 240, 220),  # Q1
                QColor(60, 60, 40) if theme == "dark" else QColor(240, 240, 220),  # Q2
                QColor(60, 40, 40) if theme == "dark" else QColor(240, 220, 220),  # Q3
                QColor(40, 40, 60) if theme == "dark" else QColor(220, 220, 240)   # Q4
            ]

            logging.info("Заполнение таблицы данными")
            for row, match in enumerate(matches):
                try:
                    # Номер матча
                    num_item = QTableWidgetItem(str(match.get("row", row + 1)))
                    num_item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row, 0, num_item)

                    # Команды
                    table.setItem(row, 1, QTableWidgetItem(str(match.get("team1", ""))))
                    table.setItem(row, 2, QTableWidgetItem(str(match.get("team2", ""))))

                    # Тотал
                    total = float((match.get("total_manual")
                                if match.get("total_manual") is not None
                                else match.get("total", 0)) or 0)
                    total_item = QTableWidgetItem(f"{total:.1f}" if total else "")
                    total_item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row, 3, total_item)
                    total_item = QTableWidgetItem(f"{total:.1f}" if total else "")
                    total_item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row, 3, total_item)

                    # Четверти и 1-я половина
                    # Будем заполнять 1Q и 2Q (колонки 4 и 5), затем 1 HALF (колонка 6), затем 3Q и 4Q (7 и 8)
                    quarters = ["q1", "q2", "q3", "q4"]
                    q_values = []
                    # Заполняем 1Q и 2Q
                    for i, (quarter, color) in enumerate(zip(quarters[:2], q_colors[:2])):
                        value = float(match.get(quarter, 0) or 0)
                        q_values.append(value)
                        if total > 0:
                            percent = (value / total) * 100
                            text = f"{value:.1f}\n({percent:.1f}%)"
                        else:
                            text = f"{value:.1f}" if value else ""
                        item = QTableWidgetItem(text)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setBackground(color)
                        item.setForeground(text_color)
                        table.setItem(row, 4 + i, item)
                    # Рассчитываем и устанавливаем 1 HALF (сумма 1Q + 2Q)
                    half1_val = sum(q_values)
                    half1_item = QTableWidgetItem(f"{half1_val:.1f}" if half1_val else "")
                    half1_item.setTextAlignment(Qt.AlignCenter)
                    # Цвет для 1 Half: берем усредненный цвет первых двух четвертей или оставляем стандартный
                    # Здесь используем чуть отличающийся зеленоватый оттенок
                    half1_color = QColor(50, 80, 50) if theme == "dark" else QColor(220, 230, 220)
                    half1_item.setBackground(half1_color)
                    half1_item.setForeground(text_color)
                    table.setItem(row, 6, half1_item)
                    # сохраняем значение первой половины в match для дальнейшего анализа
                    match["half1"] = half1_val
                    # Заполняем 3Q и 4Q (i=2,3) в колонки 7 и 8
                    for i2, (quarter, color) in enumerate(zip(quarters[2:], q_colors[2:])):
                        idx = 4 + 2 + 1 + i2  # 4 (start) + 2 quarters + 1 half1 + index within second half
                        value = float(match.get(quarter, 0) or 0)
                        if total > 0:
                            percent = (value / total) * 100
                            text = f"{value:.1f}\n({percent:.1f}%)"
                        else:
                            text = f"{value:.1f}" if value else ""
                        item = QTableWidgetItem(text)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setBackground(color)
                        item.setForeground(text_color)
                        table.setItem(row, idx, item)

                    # Очищаем только Q3 NEW (колонка 10) и Q4 NEW (колонка 11)
                    for col in (10, 11):
                        item = QTableWidgetItem("")
                        item.setTextAlignment(Qt.AlignCenter)
                        table.setItem(row, col, item)

                    # 2 Half (ручное → иначе исходное). Колонка 9
                    half2_val = match.get("half2_manual")
                    if half2_val is None:
                        half2_val = match.get("half2")
                    item_half2 = QTableWidgetItem("" if half2_val in (None, "") else f"{float(half2_val):.1f}")
                    item_half2.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row, 9, item_half2)

                    # Q3 NEW / Q4 NEW — подставляем сохранённые. Колонки 10 и 11
                    q3n = match.get("q3_new")
                    q4n = match.get("q4_new")

                    if q3n not in (None, ""):
                        it = QTableWidgetItem(f"{float(q3n):.1f}")
                        it.setTextAlignment(Qt.AlignCenter)
                        table.setItem(row, 10, it)

                    if q4n not in (None, ""):
                        it = QTableWidgetItem(f"{float(q4n):.1f}")
                        it.setTextAlignment(Qt.AlignCenter)
                        table.setItem(row, 11, it)


                except Exception as e:
                    logging.error(f"Ошибка при заполнении строки {row}: {str(e)}")
                    continue

            # Настройка размеров и стилей
            table.resizeColumnsToContents()
            table.setColumnWidth(0, 40)
            table.setColumnWidth(1, 150)
            table.setColumnWidth(2, 150)

            # Установка высоты строк
            for row in range(table.rowCount()):
                table.setRowHeight(row, 45)

            # Подключение обработчика изменений
            table.itemChanged.connect(self.on_cell_changed)

            # Добавление таблицы в layout
            container_layout.addWidget(table)
            self.table_layout.addWidget(table_container)
            
            # Активация кнопки обновления тоталов
            self.update_totals_btn.setEnabled(True)

            logging.info("Таблица успешно создана и отображена")

            # При отображении турнира не обновляем таблицу ставок, 
            # ставки выводятся агрегированно после загрузки линий.

        except Exception as e:
            error_msg = f"Ошибка в display_tournament_quarters: {str(e)}\n{traceback.format_exc()}"
            logging.error(error_msg)
            self.add_log(f"ОШИБКА: {error_msg}")
            QMessageBox.critical(self, "Ошибка", str(e))
        

    def clear_layout(self, layout):
        """Очищает layout от всех виджетов"""
        try:
            logging.info("Начало очистки layout")
            if layout is None:
                logging.warning("Layout is None")
                return
                
            while layout.count():
                logging.info(f"Удаление элемента {layout.count()}")
                item = layout.takeAt(0)
                if item is None:
                    continue
                    
                widget = item.widget()
                if widget is not None:
                    widget.setParent(None)
                    widget.deleteLater()
                else:
                    # Если это вложенный layout
                    self.clear_layout(item.layout())
                    
            logging.info("Layout очищен успешно")
        except Exception as e:
            logging.error(f"Ошибка при очистке layout: {str(e)}")
            raise

    def update_totals_from_api(self):
        """Обновление тоталов через BetsAPI с выбором даты"""
        try:
            # Сначала проверяем доступ к API
            success, message = self.betsapi.test_api_access()
            if not success:
                QMessageBox.critical(self, "Ошибка API", 
                    f"Проблема с доступом к BetsAPI:\n{message}\n\n"
                    f"Проверьте токен и доступ к API")
                return
            if not self.quarters_data:
                QMessageBox.warning(self, "Предупреждение", "Нет данных для обновления")
                return
            
            # Показываем диалог выбора даты
            date_dialog = DateSelectionDialog(self)
            if date_dialog.exec_() != QDialog.Accepted:
                return
            
            selected_date = date_dialog.get_selected_date()
            self.add_log(f"Выбрана дата: {selected_date}")
            
            # Инициализируем API handler
            if not hasattr(self, 'betsapi') or not self.betsapi:
                self.betsapi = BetsAPIHandler()
            
            # Получаем список всех матчей на выбранную дату
            self.add_log("Загрузка матчей из BetsAPI...")
            all_matches = self.betsapi.get_basketball_matches(selected_date)
            
            if not all_matches:
                QMessageBox.warning(self, "Нет матчей", f"Не найдено матчей на {selected_date}")
                return
            
            self.add_log(f"Найдено матчей в API: {len(all_matches)}")
            
            # Создаем словарь для быстрого поиска матчей
            matches_dict = {}
            for match in all_matches:
                home = match.get("home", {}).get("name", "").lower()
                away = match.get("away", {}).get("name", "").lower()
                matches_dict[f"{home} vs {away}"] = match
                matches_dict[f"{away} vs {home}"] = match
            
            # Получаем общее количество матчей для прогресс-бара
            total_matches = 0
            for tournament_data in self.quarters_data.values():
                total_matches += len(tournament_data["matches"])
            
            # Создаем и показываем прогресс-диалог
            progress = QProgressDialog("Обновление тоталов...", "Отмена", 0, total_matches, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            current_match = 0
            updates_count = 0
            
            # Обрабатываем каждый турнир
            for tournament_name, tournament_data in self.quarters_data.items():
                self.add_log(f"\nОбработка турнира: {tournament_name}")
                
                for match in tournament_data["matches"]:
                    if progress.wasCanceled():
                        break
                    
                    current_match += 1
                    progress.setValue(current_match)
                    progress.setLabelText(
                        f"Турнир: {tournament_name}\n"
                        f"Матч {current_match}/{total_matches}:\n"
                        f"{match['team1']} vs {match['team2']}"
                    )
                    
                    # Ищем матч в API
                    match_key = f"{match['team1'].lower()} vs {match['team2'].lower()}"
                    api_match = matches_dict.get(match_key)
                    
                    if not api_match:
                        match_key = f"{match['team2'].lower()} vs {match['team1'].lower()}"
                        api_match = matches_dict.get(match_key)
                    
                    if api_match:
                        self.add_log(f"\nНайден матч в API: {api_match['id']}")
                        self.add_log(f"Статус матча: {api_match['match_type']}")
                        
                        # Получаем тотал
                        total_data = self.betsapi.get_match_total(
                            api_match['id'], 
                            api_match['match_type']
                        )
                        
                        if total_data:
                            new_total = total_data['value']
                            old_total = match['total']
                            
                            # Обновляем значение в данных
                            match['total'] = new_total
                            updates_count += 1
                            
                            self.add_log(
                                f"✓ Тотал обновлен: {old_total:.1f} → {new_total:.1f} "
                                f"({total_data['bookmaker']} {total_data['type']})"
                            )
                            
                            # Если это текущий отображаемый турнир, обновляем таблицу
                            if (self.current_tournament == tournament_name and 
                                self.current_table):
                                self.update_total_in_table(
                                    match['row'] - 1,  # row в match начинается с 1
                                    new_total
                                )
                        else:
                            self.add_log(f"✗ Тотал не найден для матча")
                    else:
                        self.add_log(f"✗ Матч не найден в API")
                    
                    QApplication.processEvents()
            
            progress.setValue(total_matches)
            
            # Сохраняем обновленные данные
            if updates_count > 0:
                if self.save_data_to_file():
                    self.add_log("Обновленные данные сохранены")
                
                # Обновляем отображение текущего турнира
                if self.current_tournament and self.current_table:
                    self.display_tournament_quarters(
                        self.current_tournament,
                        self.quarters_data[self.current_tournament]
                    )
            
            QMessageBox.information(
                self,
                "Обновление завершено",
                f"Всего обновлено: {updates_count} матчей\n"
                f"Не удалось обновить: {total_matches - updates_count} матчей"
            )
            
        except Exception as e:
            error_msg = f"Ошибка при обновлении тоталов: {str(e)}"
            self.add_log(f"ОШИБКА: {error_msg}")
            logging.error(f"{error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)

    def update_total_in_table(self, row, new_total):
        """Обновляет значение тотала в таблице и пересчитывает четверти"""
        try:
            if self.current_table:
                # Получаем текущее значение для лога
                current_item = self.current_table.item(row, 3)
                current_total = float(current_item.text()) if current_item else 0
                
                # Создаем новый item с новым значением
                total_item = QTableWidgetItem(f"{new_total:.1f}")
                total_item.setTextAlignment(Qt.AlignCenter)
                total_item.setFont(QFont("Arial", 11, QFont.Bold))
                
                # Устанавливаем новое значение
                self.current_table.setItem(row, 3, total_item)
                
                self.add_log(f"Обновлен тотал в строке {row + 1}: {current_total:.1f} → {new_total:.1f}")
                
                # Получаем текущие значения четвертей
                quarter_values = []
                for i in range(4):
                    item = self.current_table.item(row, 4 + i)
                    if item and item.text():
                        try:
                            # Извлекаем процент из текста (например, "41.9\n(24.9%)")
                            text = item.text()
                            if "\n" in text and "%" in text:
                                percent = float(text.split("(")[1].split("%")[0])
                                quarter_values.append(percent / 100)  # переводим процент в долю
                        except:
                            quarter_values.append(0)
                
                # Если есть значения четвертей, пересчитываем их
                if len(quarter_values) == 4:
                    theme = QApplication.instance().property("theme") or "dark"
                    
                    # Определяем цвета в зависимости от темы
                    if theme == "dark":
                        q_colors = [
                            QColor(40, 60, 40),   # Q1 - темно-зеленый
                            QColor(60, 60, 40),   # Q2 - темно-желтый
                            QColor(60, 40, 40),   # Q3 - темно-красный
                            QColor(40, 40, 60)    # Q4 - темно-синий
                        ]
                        text_color = QColor(200, 200, 200)
                    else:
                        q_colors = [
                            QColor(220, 240, 220),  # Q1 - светло-зеленый
                            QColor(240, 240, 220),  # Q2 - светло-желтый
                            QColor(240, 220, 220),  # Q3 - светло-красный
                            QColor(220, 220, 240)   # Q4 - светло-синий
                        ]
                        text_color = QColor(0, 0, 0)
                    
                    # Обновляем каждую четверть
                    for i, percent in enumerate(quarter_values):
                        # Рассчитываем новое значение четверти
                        new_quarter = new_total * percent
                        
                        # Создаем текст с значением и процентом
                        display_text = f"{new_quarter:.1f}\n({percent*100:.1f}%)"
                        
                        # Создаем и настраиваем item
                        q_item = QTableWidgetItem(display_text)
                        q_item.setTextAlignment(Qt.AlignCenter)
                        q_item.setFont(QFont("Arial", 12))
                        q_item.setBackground(QBrush(q_colors[i]))
                        q_item.setForeground(text_color)
                        
                        # Устанавливаем item в таблицу
                        self.current_table.setItem(row, 4 + i, q_item)
                        
                        self.add_log(f"Обновлена четверть {i+1}: {new_quarter:.1f} ({percent*100:.1f}%)")
                
                # Обновляем отображение
                self.current_table.viewport().update()
                
        except Exception as e:
            self.add_log(f"Ошибка при обновлении тотала в таблице: {str(e)}")
            logging.error(f"Ошибка обновления тотала: {str(e)}\n{traceback.format_exc()}")
        
    def process_batch(self, batch, updates):
        """Обработка пакета матчей"""
        for match, api_match, tournament_name in batch:
            try:
                self.add_log(f"\n{'='*50}")
                self.add_log(f"Обработка матча: {match['team1']} vs {match['team2']}")
                self.add_log(f"ID матча в API: {api_match['id']}")
                self.add_log(f"Тип матча: {api_match['match_type']}")
                
                # Логируем параметры запроса
                self.add_log("Отправка запроса к BetsAPI...")
                
                total_data = self.betsapi.get_match_total(
                    api_match['id'], 
                    api_match['match_type']
                )
                
                if total_data:
                    new_total = total_data['value']
                    old_total = match['total']
                    
                    updates.append((match, new_total))
                    
                    self.add_log(
                        f"✓ Тотал обновлен: {old_total:.1f} → {new_total:.1f} "
                        f"({total_data['bookmaker']} {total_data['type']})"
                    )
                else:
                    self.add_log(f"✗ Тотал не найден для матча: {match['team1']} vs {match['team2']}")
                    self.add_log("Полные данные матча из API:")
                    self.add_log(str(api_match))
                    
                # Небольшая задержка между запросами
                time.sleep(0.1)
                
            except Exception as e:
                self.add_log(f"Ошибка при обработке матча: {str(e)}")
                self.add_log("Traceback:")
                import traceback
                self.add_log(traceback.format_exc())
                continue

    def recalculate_quarters(self, row, match, new_total):
        """Пересчитывает значения четвертей на основе процентов"""
        
        theme = QApplication.instance().property("theme") or "dark"
        
        # Цвета для четвертей
        if theme == "dark":
            q_colors = [QColor(40, 60, 40), QColor(60, 60, 40), 
                    QColor(60, 40, 40), QColor(40, 40, 60)]
            text_color = QColor(200, 200, 200)
        else:
            q_colors = [QColor(220, 240, 220), QColor(240, 240, 220),
                    QColor(240, 220, 220), QColor(220, 220, 240)]
            text_color = QColor(0, 0, 0)
        
        quarter_keys = ['q1', 'q2', 'q3', 'q4']
        
        for i, q_key in enumerate(quarter_keys):
            # Получаем текущее значение и процент из ячейки
            item = self.current_table.item(row, 4 + i)
            if item:
                text = item.text()
                if "\n" in text and "%" in text:
                    # Извлекаем процент
                    percent_str = text.split("(")[1].split("%")[0]
                    try:
                        percentage = float(percent_str)
                        
                        # Пересчитываем значение четверти
                        new_quarter_value = new_total * (percentage / 100)
                        
                        # Обновляем в данных
                        match[q_key] = new_quarter_value
                        
                        # Создаем новый текст для ячейки
                        display_text = f"{new_quarter_value:.1f}\n({percentage:.1f}%)"
                        
                        # Создаем новый item
                        q_item = QTableWidgetItem(display_text)
                        q_item.setTextAlignment(Qt.AlignCenter)
                        q_item.setFont(QFont("Arial", 10))
                        q_item.setBackground(QBrush(q_colors[i]))
                        q_item.setForeground(text_color)
                        
                        self.current_table.setItem(row, 4 + i, q_item)
                        
                    except ValueError:
                        self.add_log(f"Ошибка при пересчете четверти {i+1}")

# Добавьте этот код перед классом MainWindow

from PyQt5.QtCore import QSettings

class ThemeManager:
    def __init__(self, parent=None):
        self.parent = parent
        
        self.dark_theme = """
        /* Основная тёмная тема c чёрно‑синими и зелёными оттенками */
        QMainWindow {
            background-color: #0A192F;
        }

        QWidget {
            background-color: #0A192F;
            color: #E0E5EC;
        }

        QWidget#header {
            background-color: #11244A;
            border-bottom: 2px solid #0F3460;
        }

        QPushButton {
            background-color: #11244A;
            border: 1px solid #1C3A6A;
            padding: 8px;
            border-radius: 6px;
            min-height: 30px;
            color: #E0E5EC;
        }

        QPushButton:hover {
            background-color: #173766;
            border: 1px solid #2A4B82;
        }

        QPushButton:pressed {
            background-color: #0F3460;
        }

        QLineEdit {
            background-color: #11244A;
            border: 1px solid #1C3A6A;
            padding: 8px;
            border-radius: 6px;
            color: #E0E5EC;
        }

        QLineEdit:focus {
            border: 2px solid #2EC4B6;
        }

        QTableWidget {
            background-color: #11244A;
            gridline-color: #1C3A6A;
            border: 1px solid #1C3A6A;
            color: #E0E5EC;
        }

        QTableWidget::item {
            padding: 5px;
        }

        QTableWidget::item:selected {
            background-color: #2EC4B6;
            color: #0A192F;
        }

        QTableWidget QLineEdit, QTableView QLineEdit {
            background-color: #11244A;
            color: #E0E5EC;
            selection-background-color: #2EC4B6;
            selection-color: #0A192F;
        }

        QHeaderView::section {
            background-color: #142C54;
            padding: 8px;
            border: 1px solid #1C3A6A;
            color: #E0E5EC;
        }

        QListWidget {
            background-color: #11244A;
            border: 1px solid #1C3A6A;
            color: #E0E5EC;
        }

        QListWidget::item {
            padding: 10px;
            border-bottom: 1px solid #1C3A6A;
        }

        QListWidget::item:hover {
            background-color: #173766;
        }

        QListWidget::item:selected {
            background-color: #2EC4B6;
            color: #0A192F;
        }

        QLabel {
            background-color: transparent;
        }

        QScrollBar:vertical {
            background-color: #11244A;
            width: 12px;
            border-radius: 6px;
        }

        QScrollBar::handle:vertical {
            background-color: #1C3A6A;
            border-radius: 6px;
            min-height: 20px;
        }

        QScrollBar::handle:vertical:hover {
            background-color: #2A4B82;
        }

        QComboBox {
            background-color: #11244A;
            border: 1px solid #1C3A6A;
            padding: 5px;
            border-radius: 6px;
            min-height: 30px;
            color: #E0E5EC;
        }

        QComboBox:hover {
            border: 1px solid #2A4B82;
        }

        QComboBox::drop-down {
            border: none;
        }

        QComboBox::down-arrow {
            image: none;
            border-left: 5px solid transparent;
            border-right: 5px solid transparent;
            border-top: 5px solid #E0E5EC;
            margin-right: 5px;
        }

        QMessageBox {
            background-color: #11244A;
        }

        QMessageBox QLabel {
            color: #E0E5EC;
        }

        QMessageBox QPushButton {
            min-width: 80px;
        }

        QScrollArea {
            background-color: transparent;
            border: none;
        }

        QSplitter::handle {
            background-color: #142C54;
        }
        """

        # Специальные стили для бокового меню, соответствующие выбранной теме. Эти
        # значения применяются в методе apply_theme, чтобы боковое меню
        # гармонировало с основной цветовой схемой.
        self.sidebar_dark = """
        /* Боковое меню для тёмной темы. Цвет фона совпадает с цветом основного окна,
           чтобы панель выглядела частью единого интерфейса. */
        QListWidget {
            border: none;
            background-color: #0A192F;
        }
        QListWidget::item {
            padding: 15px;
            border-bottom: 1px solid #1C3A6A;
            color: #E0E5EC;
        }
        QListWidget::item:hover {
            background-color: #173766;
        }
        QListWidget::item:selected {
            background-color: #2EC4B6;
            color: #0A192F;
        }
        """

        self.sidebar_light = """
        /* Боковое меню для светлой темы. Цвет фона совпадает с основным
           фоном (#F5FAF7) для единообразия. */
        QListWidget {
            border: none;
            background-color: #F5FAF7;
        }
        QListWidget::item {
            padding: 15px;
            border-bottom: 1px solid #E5EEF3;
            color: #0A192F;
        }
        QListWidget::item:hover {
            background-color: #EFF6F0;
        }
        QListWidget::item:selected {
            background-color: #2EC4B6;
            color: #FFFFFF;
        }
        """
        
        self.light_theme = """
        /* Обновлённая светлая тема с акцентом на зелёный и читаемым текстом */
        QMainWindow {
            background-color: #F5FAF7;
        }

        QWidget {
            background-color: #F5FAF7;
            color: #0A192F;
        }

        QWidget#header {
            background-color: #FFFFFF;
            border-bottom: 2px solid #E5EEF3;
        }

        QPushButton {
            background-color: #FFFFFF;
            border: 1px solid #C8D3E5;
            padding: 8px;
            border-radius: 6px;
            min-height: 30px;
            color: #0A192F;
        }

        QPushButton:hover {
            background-color: #EFF6F0;
            border: 1px solid #AFC9DA;
        }

        QPushButton:pressed {
            background-color: #DCE6F2;
        }

        QLineEdit {
            background-color: #FFFFFF;
            border: 1px solid #C8D3E5;
            padding: 8px;
            border-radius: 6px;
            color: #0A192F;
        }

        QLineEdit:focus {
            border: 2px solid #2EC4B6;
        }

        QTableWidget {
            background-color: #FFFFFF;
            gridline-color: #E5EEF3;
            border: 1px solid #C8D3E5;
            color: #0A192F;
        }

        QTableWidget::item {
            padding: 5px;
        }

        QTableWidget::item:selected {
            background-color: #2EC4B6;
            color: #FFFFFF;
        }

        QTableWidget QLineEdit, QTableView QLineEdit {
            background-color: #F5FAF7;
            color: #0A192F;
            selection-background-color: #2EC4B6;
            selection-color: #0A192F;
        }

        QHeaderView::section {
            background-color: #EEF5F9;
            padding: 8px;
            border: 1px solid #C8D3E5;
            color: #0A192F;
        }

        QListWidget {
            background-color: #FFFFFF;
            border: 1px solid #C8D3E5;
            color: #0A192F;
        }

        QListWidget::item {
            padding: 10px;
            border-bottom: 1px solid #E5EEF3;
        }

        QListWidget::item:hover {
            background-color: #EFF6F0;
        }

        QListWidget::item:selected {
            background-color: #2EC4B6;
            color: #FFFFFF;
        }

        QLabel {
            background-color: transparent;
            color: #0A192F;
        }

        QScrollBar:vertical {
            background-color: #EFF6F0;
            width: 12px;
            border-radius: 6px;
        }

        QScrollBar::handle:vertical {
            background-color: #C8D3E5;
            border-radius: 6px;
            min-height: 20px;
        }

        QScrollBar::handle:vertical:hover {
            background-color: #AFC9DA;
        }

        QComboBox {
            background-color: #FFFFFF;
            border: 1px solid #C8D3E5;
            padding: 5px;
            border-radius: 6px;
            min-height: 30px;
            color: #0A192F;
        }

        QComboBox:hover {
            border: 1px solid #AFC9DA;
        }

        QComboBox::drop-down {
            border: none;
        }

        QComboBox::down-arrow {
            image: none;
            border-left: 5px solid transparent;
            border-right: 5px solid transparent;
            border-top: 5px solid #0A192F;
            margin-right: 5px;
        }

        QMessageBox {
            background-color: #FFFFFF;
        }

        QMessageBox QLabel {
            color: #0A192F;
        }

        QMessageBox QPushButton {
            min-width: 80px;
        }

        QScrollArea {
            background-color: transparent;
            border: none;
        }

        QSplitter::handle {
            background-color: #E5EEF3;
        }
        """
    
        # Создание страницы ставок
        self.bets_page = BetsPage()

    def apply_theme(self, theme_name):
        """Применяет выбранную тему"""
        if self.parent:
            if theme_name == "dark":
                self.parent.setStyleSheet(self.dark_theme)
                # Применяем специальный стиль к боковой панели, чтобы она
                # соответствовала основной тёмной теме
                try:
                    self.parent.sidebar.setStyleSheet(self.sidebar_dark)
                except Exception:
                    pass
            else:
                self.parent.setStyleSheet(self.light_theme)
                # Применяем специальный стиль к боковой панели для светлой темы
                try:
                    self.parent.sidebar.setStyleSheet(self.sidebar_light)
                except Exception:
                    pass
        
        # Сохраняем выбранную тему
        self.save_theme_preference(theme_name)
        # После применения темы обновляем стили логов, если у родителя есть такой метод
        try:
            if self.parent and hasattr(self.parent, 'update_debug_log_styles'):
                self.parent.update_debug_log_styles()
        except Exception:
            pass
    
    def save_theme_preference(self, theme_name):
        """Сохраняет предпочтение темы"""
        settings = QSettings("ExcelAnalyzer", "ThemePreferences")
        settings.setValue("theme", theme_name)
    
    def load_theme_preference(self):
        """Загружает сохраненное предпочтение темы"""
        settings = QSettings("ExcelAnalyzer", "ThemePreferences")
        return settings.value("theme", "dark")  # По умолчанию темная тема
    
class BetsPage(QWidget):
    """Страница для управления ставками на четверти и половины"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.quarters_page = None  # Ссылка на страницу четвертей для получения данных
        self.bets_data = {
            "half":  [],
            "q1": []
        }
        self.notifier = None  # Инициализируется после загрузки данных
        self.setup_ui()
    
    def setup_ui(self):
        """Настройка интерфейса"""
        main_layout = QVBoxLayout(self)
        
        # Заголовок
        title_label = QLabel("Управление ставками")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt. AlignCenter)
        main_layout.addWidget(title_label)
        main_layout.addSpacing(10)
        
        # Инструкции
        info_label = QLabel(
            "Эта страница отображает сгруппированные ставки на первую половину и первую четверть.\n"
            "Ставки загружаются из раздела 'Половины и четверти' при загрузке файла 'Линии...'."
        )
        info_label.setWordWrap(True)
        info_label.setFont(QFont("Arial", 10))
        main_layout.addWidget(info_label)
        main_layout. addSpacing(10)
        
        # Кнопки управления
        button_layout = QHBoxLayout()
        
        self.refresh_btn = QPushButton("Обновить ставки")
        self.refresh_btn.clicked.connect(self.refresh_bets)
        
        self.clear_history_btn = QPushButton("Очистить историю уведомлений")
        self.clear_history_btn.clicked.connect(self.clear_notification_history)
        
        self.monitor_checkbox = QCheckBox("Включить мониторинг уведомлений")
        self.monitor_checkbox.setChecked(False)
        self.monitor_checkbox.stateChanged.connect(self.toggle_monitoring)
        
        button_layout.addWidget(self.refresh_btn)
        button_layout.addWidget(self.clear_history_btn)
        button_layout.addWidget(self.monitor_checkbox)
        button_layout.addStretch()
        
        main_layout.addLayout(button_layout)
        main_layout.addSpacing(10)
        
        # Контейнер для таблиц
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setContentsMargins(0, 0, 0, 0)
        self.scroll_layout.setSpacing(20)
        
        scroll_area.setWidget(self.scroll_content)
        main_layout.addWidget(scroll_area)
        
        # Статус мониторинга
        self. monitor_status_label = QLabel("Мониторинг выключен")
        self.monitor_status_label.setFont(QFont("Arial", 10))
        self.monitor_status_label.setStyleSheet("color: #ff6b6b;")
        main_layout.addWidget(self.monitor_status_label)
        
        # Лог
        self.log_label = QLabel("Журнал:")
        self.log_label. setFont(QFont("Arial", 10, QFont.Bold))
        main_layout.addWidget(self. log_label)
        
        self.log_text = QListWidget()
        self.log_text.setMaximumHeight(150)
        main_layout. addWidget(self.log_text)
    
    def set_quarters_page(self, quarters_page):
        """Устанавливает ссылку на страницу четвертей"""
        self.quarters_page = quarters_page
    
    def update_bets(self, bets_half:  List, bets_q1: List):
        """
        Обновляет ставки с данными
        
        Args:
            bets_half: Список ставок на половину
            bets_q1: Список ставок на первую четверть
        """
        self.bets_data["half"] = bets_half
        self.bets_data["q1"] = bets_q1
        
        self.display_bets()
        self.add_log(f"Ставки обновлены: {len(bets_half)} на половину, {len(bets_q1)} на четверть")
        
        # Инициализируем уведомителя если нужно
        if not self.notifier:
            from bets_notifier import BetsNotifier
            self.notifier = BetsNotifier(
                telegram_token=TELEGRAM_BOT_TOKEN,
                telegram_chat_id=TELEGRAM_CHAT_ID
            )
        
        # Запускаем мониторинг если он включен
        if self.monitor_checkbox.isChecked():
            self.notifier.start_monitoring(self. bets_data)
    
    def display_bets(self):
        """Отображает ставки в таблицах"""
        try:
            # Очищаем предыдущий контент
            while self.scroll_layout.count():
                item = self.scroll_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
            
            # Таблица ставок на половину
            if self.bets_data["half"]:
                self.display_bets_table(
                    "Ставки на первую половину",
                    self.bets_data["half"],
                    "#2d5c8f"  # Синий цвет
                )
            
            # Таблица ставок на четверть
            if self.bets_data["q1"]: 
                self.display_bets_table(
                    "Ставки на первую четверть",
                    self.bets_data["q1"],
                    "#5c4a2d"  # Коричневый цвет
                )
            
            # Если нет ставок
            if not self.bets_data["half"] and not self. bets_data["q1"]:
                no_bets_label = QLabel("Ставок не найдено.  Загрузите файл 'Линии...' в разделе 'Четверти'.")
                no_bets_label.setAlignment(Qt.AlignCenter)
                no_bets_label.setFont(QFont("Arial", 12))
                self.scroll_layout.addWidget(no_bets_label)
            
            self.scroll_layout.addStretch()
            
        except Exception as e:
            self.add_log(f"ОШИБКА при отображении ставок: {str(e)}")
    
    def display_bets_table(self, title: str, bets_list: List, color: str):
        """Отображает таблицу ставок"""
        try:
            # Заголовок
            title_label = QLabel(title)
            title_label.setFont(QFont("Arial", 14, QFont.Bold))
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setStyleSheet(f"color: {color};")
            self.scroll_layout.addWidget(title_label)
            
            # Таблица
            table = QTableWidget()
            table.setColumnCount(7)
            table.setHorizontalHeaderLabels([
                "Время", "Турнир", "Команда 1", "Команда 2", "Тип", "Линия", "Разница"
            ])
            
            table.setRowCount(len(bets_list))
            
            # Устанавливаем размеры колонок
            header = table.horizontalHeader()
            header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Время
            header.setSectionResizeMode(1, QHeaderView. Stretch)  # Турнир
            header.setSectionResizeMode(2, QHeaderView.Stretch)  # Команда 1
            header. setSectionResizeMode(3, QHeaderView.Stretch)  # Команда 2
            header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Тип
            header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Линия
            header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Разница
            
            table.setSelectionBehavior(QAbstractItemView.SelectRows)
            table.setSelectionMode(QAbstractItemView.SingleSelection)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            
            # Заполняем таблицу
            for row, bet_data in enumerate(bets_list):
                try:
                    # Распаковываем данные
                    tournament, team1, team2, bet_type_line, line, diff = bet_data[: 6]
                    match_time = bet_data[6] if len(bet_data) > 6 else "-"
                    
                    # Время
                    time_item = QTableWidgetItem(str(match_time))
                    time_item.setTextAlignment(Qt.AlignCenter)
                    
                    # Турнир
                    tournament_item = QTableWidgetItem(str(tournament))
                    
                    # Команды
                    team1_item = QTableWidgetItem(str(team1))
                    team2_item = QTableWidgetItem(str(team2))
                    
                    # Тип ставки
                    bet_type_item = QTableWidgetItem(str(bet_type_line))
                    bet_type_item.setTextAlignment(Qt.AlignCenter)
                    
                    # Линия
                    line_item = QTableWidgetItem(f"{float(line):.1f}")
                    line_item.setTextAlignment(Qt.AlignCenter)
                    
                    # Разница
                    diff_item = QTableWidgetItem(f"{float(diff):.1f}")
                    diff_item. setTextAlignment(Qt.AlignCenter)
                    
                    # Определяем цвет строки в зависимости от типа
                    if "OVER" in str(bet_type_line):
                        row_color = QColor(50, 100, 50) if QApplication.instance().property("theme") == "dark" else QColor(220, 240, 220)
                    else:
                        row_color = QColor(100, 50, 50) if QApplication.instance().property("theme") == "dark" else QColor(240, 220, 220)
                    
                    # Применяем цвет к строке
                    for item in [time_item, tournament_item, team1_item, team2_item, bet_type_item, line_item, diff_item]:
                        item. setBackground(QBrush(row_color))
                    
                    # Добавляем в таблицу
                    table.setItem(row, 0, time_item)
                    table. setItem(row, 1, tournament_item)
                    table.setItem(row, 2, team1_item)
                    table.setItem(row, 3, team2_item)
                    table.setItem(row, 4, bet_type_item)
                    table. setItem(row, 5, line_item)
                    table.setItem(row, 6, diff_item)
                    
                except Exception as e: 
                    self.add_log(f"Ошибка при обработке ставки {row}: {str(e)}")
                    continue
            
            # Настройка высоты строк
            table.resizeRowsToContents()
            
            # Добавляем таблицу в layout
            self.scroll_layout. addWidget(table)
            
        except Exception as e:
            self.add_log(f"ОШИБКА при отображении таблицы ставок: {str(e)}")
    
    def refresh_bets(self):
        """Обновляет ставки"""
        if self.quarters_page and hasattr(self.quarters_page, 'bets_results_data'):
            # Преобразуем данные из страницы четвертей
            all_half = []
            all_q1 = []
            
            for tournament_name, bets_for_tournament in self.quarters_page.bets_results_data. items():
                all_half.extend(bets_for_tournament. get("half", []))
                all_q1.extend(bets_for_tournament.get("q1", []))
            
            self.update_bets(all_half, all_q1)
        else:
            self.add_log("Данные о четвертях не загружены")
    
    def toggle_monitoring(self, state):
        """Переключает мониторинг уведомлений"""
        if state == Qt.Checked:
            if not self.notifier:
                from bets_notifier import BetsNotifier
                self.notifier = BetsNotifier(
                    telegram_token=TELEGRAM_BOT_TOKEN,
                    telegram_chat_id=TELEGRAM_CHAT_ID
                )
            
            self.notifier.start_monitoring(self.bets_data)
            self.monitor_status_label.setText("Мониторинг включен ✓")
            self.monitor_status_label.setStyleSheet("color: #51cf66;")
            self.add_log("Мониторинг уведомлений включен")
        else:
            if self.notifier:
                self.notifier.stop_monitoring()
            self.monitor_status_label.setText("Мониторинг выключен")
            self.monitor_status_label.setStyleSheet("color: #ff6b6b;")
            self.add_log("Мониторинг уведомлений выключен")
    
    def clear_notification_history(self):
        """Очищает историю отправленных уведомлений"""
        if self.notifier:
            self. notifier.clear_history()
            self.add_log("История уведомлений очищена")
            QMessageBox.information(self, "Успех", "История уведомлений очищена")
    
    def add_log(self, message: str):
        """Добавляет сообщение в журнал"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.addItem(f"[{timestamp}] {message}")
        self.log_text.scrollToBottom()
        
        # Ограничиваем размер лога
        while self.log_text.count() > 50:
            self.log_text.takeItem(0)

class HalfsQuartersPage(QWidget):
    """Страница для анализа ставок на ��оловину и четверть"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.quarters_data = {}  # Ссылка на данные четвертей из QuartersPage
        self.bets_results_data = {}
        self.setup_ui()
    
    def setup_ui(self):
        """Настройка интерфейса"""
        layout = QVBoxLayout(self)
        
        # Заголовок
        title = QLabel("Ставки на первую половину и четверть")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(20)
        
        # Кнопка загрузки линий
        self.load_lines_btn = QPushButton("Загрузить файл 'Линии...'")
        self.load_lines_btn.setMinimumHeight(40)
        self.load_lines_btn.clicked.connect(self.load_lines_file)
        layout.addWidget(self.load_lines_btn)
        layout.addSpacing(10)
        
        # Статус
        self.status_label = QLabel("Дождитесь загрузки данных четвертей")
        self.status_label. setWordWrap(True)
        layout.addWidget(self.status_label)
        layout.addSpacing(10)
        
        # Контейнер для таблиц ставок
        self.bets_container = QWidget()
        self.bets_layout = QVBoxLayout(self.bets_container)
        layout.addWidget(self.bets_container, 1)
        
        # Лог
        self.log_text = QPlainTextEdit()
        self.log_text.setMaximumHeight(150)
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)
    
    def set_quarters_data(self, quarters_data):
        """Устанавливает ссылку на данные четвертей"""
        self.quarters_data = quarters_data
        if quarters_data:
            self. status_label.setText("✓ Данные четвертей загружены.  Готово к анализу линий")
            self.load_lines_btn. setEnabled(True)
        else:
            self.load_lines_btn.setEnabled(False)
    
    def add_log(self, message):
        """Добавляет сообщение в лог"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.appendPlainText(f"[{timestamp}] {message}")

    # ------------------------------
    # Helpers
    # ------------------------------
    @staticmethod
    def _excel_col_to_idx(col: str) -> int:
        """Excel column letters (A, B, ..., AA, AB...) -> 0-based index."""
        col = (col or "").strip().upper()
        if not col:
            raise ValueError("Empty excel column")
        idx = 0
        for ch in col:
            if not ('A' <= ch <= 'Z'):
                raise ValueError(f"Invalid excel column: {col}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    @staticmethod
    def _parse_float(value) -> float | None:
        """Безопасный парсинг числа (поддержка запятой)."""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        s = str(value).strip()
        if not s or s.lower() in {"nan", "none", "-"}:
            return None
        s = s.replace(" ", "").replace(",", ".")
        # убираем лишние символы вроде '+'
        if s.startswith("+"):
            s = s[1:]
        try:
            return float(s)
        except Exception:
            return None

    @staticmethod
    def _parse_time(value) -> str | None:
        """Парсинг времени из Excel.

        Поддержка:
        - строка "HH:MM" (или "H:MM")
        - datetime/time
        - excel-serial float (например 0.5 -> 12:00)
        """
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        # datetime
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        # pandas Timestamp
        try:
            if hasattr(value, "to_pydatetime"):
                dt = value.to_pydatetime()
                if isinstance(dt, datetime):
                    return dt.strftime("%H:%M")
        except Exception:
            pass

        # excel float time (fraction of day)
        if isinstance(value, (int, float)):
            try:
                v = float(value)
                # ограничим только "время", не дату
                if 0 <= v < 1.0:
                    total_minutes = int(round(v * 24 * 60))
                    hh = (total_minutes // 60) % 24
                    mm = total_minutes % 60
                    return f"{hh:02d}:{mm:02d}"
            except Exception:
                pass

        s = str(value).strip()
        if not s or s.lower() in {"nan", "none", "-"}:
            return None
        # Иногда Excel отдаёт "11:00:00" – режем до HH:MM
        if ":" in s:
            parts = s.split(":")
            if len(parts) >= 2:
                hh = parts[0].zfill(2)
                mm = parts[1].zfill(2)
                if hh.isdigit() and mm.isdigit():
                    return f"{int(hh):02d}:{int(mm):02d}"
        return None

    @staticmethod
    def _norm_text(s: str) -> str:
        return " ".join((s or "").strip().lower().split())

    @staticmethod
    def _ceil_to_half(x: float) -> float:
        return math.ceil(x * 2.0) / 2.0

    @staticmethod
    def _floor_to_int(x: float) -> int:
        return int(math.floor(x))

    def _compute_bet(self, market_line: float, program_value: float, threshold: float = 3.5):
        """
        Возвращает (bet_type, rounded_line, diff_after_rounding) или None.

        Логика:
        1) Сначала проверяем, что |market_line - program_value| > threshold.
        2) Затем считаем предельную линию: program_value ± threshold.
        3) Округляем "в сторону, обратную ставке":
           - OVER -> вниз до целого
           - UNDER -> вверх до шага 0.5
        4) Проверяем, что после округления условие всё ещё выполняется
           и что market_line не "хуже" округлённой линии:
           - OVER: market_line <= rounded_line
           - UNDER: market_line >= rounded_line
        """
        if market_line is None or program_value is None:
            return None
        diff = market_line - program_value
        if abs(diff) <= threshold:
            return None

        if diff < 0:  # market ниже программы -> OVER
            bet_type = "OVER"
            candidate = program_value - threshold  # максимум, который ещё держит порог
            rounded = float(self._floor_to_int(candidate))
            # проверяем, что порог сохранился
            if rounded - program_value > -threshold:
                return None
            # проверяем, что рыночная линия не хуже (нам нужна <=)
            if market_line > rounded:
                return None
            return bet_type, rounded, rounded - program_value
        else:  # market выше программы -> UNDER
            bet_type = "UNDER"
            candidate = program_value + threshold  # минимум, который ещё держит порог
            rounded = float(self._ceil_to_half(candidate))
            if rounded - program_value < threshold:
                return None
            if market_line < rounded:
                return None
            return bet_type, rounded, rounded - program_value
    
    def load_lines_file(self):
        """Загружает файл "Линии..." и считает ставки на 1 HALF и 1Q.

        В файле матчи лежат в нескольких "кибер-блоках".
        Турнир и линия 1 HALF находятся в парах колонок:
        B-G, J-O, R-W, Z-AE, AH-AM, AP-AU, AX-BC, BF-BK.
        Время — колонка слева от турнира.
        Team1/Team2 — две колонки справа от турнира.
        """
        try:
            if not self.quarters_data:
                QMessageBox.warning(self, "Ошибка",
                                    "Сначала загрузите данные четвертей в разделе 'Четверти'")
                return

            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Выберите файл 'Линии...'",
                "",
                "Excel files (*.xlsx *.xls)"
            )
            if not file_path:
                return

            self.add_log("Начало анализа файла линий...")

            # Читаем без заголовков, потому что файл — "табличная простыня".
            df = pd.read_excel(file_path, header=None)

            # mapping: TournamentCol -> HalfCol
            tournament_columns = {
                'B': 'G',
                'J': 'O',
                'R': 'W',
                'Z': 'AE',
                'AH': 'AM',
                'AP': 'AU',
                'AX': 'BC',
                'BF': 'BK'
            }

            # Сохраняем матчи из Lines в формате:
            # lines_matches[tournament][(team1_norm, team2_norm)] = {time, half, team1, team2}
            lines_matches: Dict[str, Dict[Tuple[str, str], dict]] = {}
            # Дополнительно — индекс по нормализованному названию турнира (на случай пробелов)
            norm_tournament_map: Dict[str, str] = {}

            rows, cols = df.shape

            for t_col, h_col in tournament_columns.items():
                t_idx = self._excel_col_to_idx(t_col)
                h_idx = self._excel_col_to_idx(h_col)
                time_idx = t_idx - 1
                team1_idx = t_idx + 1
                team2_idx = t_idx + 2

                # безопасность по границам
                if max(h_idx, team2_idx) >= cols or time_idx < 0:
                    continue

                for r in range(rows):
                    tournament_raw = df.iat[r, t_idx] if t_idx < cols else None
                    if tournament_raw is None or (isinstance(tournament_raw, float) and pd.isna(tournament_raw)):
                        continue
                    tournament = str(tournament_raw).strip()
                    if not tournament:
                        continue

                    team1_raw = df.iat[r, team1_idx] if team1_idx < cols else None
                    team2_raw = df.iat[r, team2_idx] if team2_idx < cols else None
                    if team1_raw is None or team2_raw is None:
                        continue
                    team1 = str(team1_raw).strip()
                    team2 = str(team2_raw).strip()
                    if not team1 or not team2:
                        continue

                    half_val = self._parse_float(df.iat[r, h_idx] if h_idx < cols else None)
                    if half_val is None:
                        continue

                    time_val = self._parse_time(df.iat[r, time_idx] if time_idx < cols else None)

                    t_norm = self._norm_text(tournament)
                    norm_tournament_map.setdefault(t_norm, tournament)

                    lines_matches.setdefault(tournament, {})
                    key = (self._norm_text(team1), self._norm_text(team2))
                    # если дубль — перезаписываем (обычно последние данные актуальнее)
                    lines_matches[tournament][key] = {
                        "time": time_val or "",
                        "half": float(half_val),
                        "team1": team1,
                        "team2": team2,
                    }

            self.add_log(f"✓ Найдено турниров в файле: {len(lines_matches)}")

            # Анализируем ставки
            self.analyze_bets(lines_matches, norm_tournament_map)

        except Exception as e:
            error_msg = f"Ошибка: {str(e)}"
            self.add_log(f"ОШИБКА: {error_msg}")
            QMessageBox.critical(self, "Ошибка", error_msg)
    
    def analyze_bets(self, lines_matches: Dict[str, Dict[Tuple[str, str], dict]], norm_tournament_map: Dict[str, str]):
        """Сравнивает данные из 'Четверти' с файлом 'Линии...' и формирует 2 таблицы ставок."""
        try:
            self.add_log("Начало анализа ставок...")

            bets_half: List[Tuple[str, str, str, str, str, str, str]] = []  # time, tournament, t1, t2, bet, line, diff
            bets_q1: List[Tuple[str, str, str, str, str, str, str]] = []
            missing: List[str] = []

            for tournament_name, qdata in self.quarters_data.items():
                # Строгое совпадение турнира, но делаем небольшой fallback на нормализацию пробелов
                t_lines = lines_matches.get(tournament_name)
                if t_lines is None:
                    t_norm = self._norm_text(tournament_name)
                    t_real = norm_tournament_map.get(t_norm)
                    if t_real:
                        t_lines = lines_matches.get(t_real)

                if not t_lines:
                    continue

                matches = qdata.get("matches", [])
                for match in matches:
                    team1_raw = str(match.get("team1", "") or "").strip()
                    team2_raw = str(match.get("team2", "") or "").strip()
                    if not team1_raw or not team2_raw:
                        continue

                    key1 = (self._norm_text(team1_raw), self._norm_text(team2_raw))
                    key2 = (key1[1], key1[0])
                    line_row = t_lines.get(key1) or t_lines.get(key2)
                    if not line_row:
                        missing.append(f"{tournament_name}: {team1_raw} vs {team2_raw}")
                        continue

                    market_half = float(line_row["half"])
                    time_str = str(line_row.get("time", "") or "")

                    program_half = self._parse_float(match.get("half1"))
                    program_q1 = self._parse_float(match.get("q1"))

                    # 1 HALF bet
                    if program_half is not None:
                        res_half = self._compute_bet(market_half, float(program_half), threshold=3.5)
                        if res_half:
                            bet_type, rounded_line, diff = res_half
                            # формат линии: целое без .0, иначе 0.5
                            line_txt = f"{rounded_line:.1f}".rstrip('0').rstrip('.')
                            bets_half.append((
                                time_str,
                                tournament_name,
                                team1_raw,
                                team2_raw,
                                f"{bet_type} {line_txt}",
                                line_txt,
                                f"{diff:.1f}"
                            ))

                    # 1Q bet: market_half / 2
                    if program_q1 is not None:
                        market_q1 = market_half / 2.0
                        res_q1 = self._compute_bet(market_q1, float(program_q1), threshold=3.5)
                        if res_q1:
                            bet_type, rounded_line, diff = res_q1
                            line_txt = f"{rounded_line:.1f}".rstrip('0').rstrip('.')
                            bets_q1.append((
                                time_str,
                                tournament_name,
                                team1_raw,
                                team2_raw,
                                f"{bet_type} {line_txt}",
                                line_txt,
                                f"{diff:.1f}"
                            ))

            # Сортировка по времени (пустые внизу)
            def _time_key(t: str):
                if not t:
                    return (1, 99, 99)
                try:
                    hh, mm = t.split(":")[:2]
                    return (0, int(hh), int(mm))
                except Exception:
                    return (1, 99, 99)

            bets_half.sort(key=lambda r: (_time_key(r[0]), r[1], r[2], r[3]))
            bets_q1.sort(key=lambda r: (_time_key(r[0]), r[1], r[2], r[3]))

            self.display_bets(bets_half, bets_q1, missing)

        except Exception as e:
            self.add_log(f"ОШИБКА в analyze_bets: {str(e)}")
            logging.error(traceback.format_exc())
    
    def display_bets(self, bets_half, bets_q1, missing):
        """Отображает ставки в таблицах"""
        try:
            # Очищаем контейнер
            while self.bets_layout.count():
                item = self.bets_layout.takeAt(0)
                if item. widget():
                    item.widget().deleteLater()
            
            # Таблица ставок на половину
            label_half = QLabel("Ставки на первую половину")
            label_half.setFont(QFont("Arial", 12, QFont.Bold))
            self.bets_layout.addWidget(label_half)

            if bets_half:
                table_half = QTableWidget()
                table_half.setColumnCount(7)
                table_half.setHorizontalHeaderLabels(
                    ["Время", "Турнир", "Команда 1", "Команда 2", "Ставка", "Линия", "Разница"]
                )
                table_half.setRowCount(len(bets_half))

                for row, bet in enumerate(bets_half):
                    for col, value in enumerate(bet):
                        item = QTableWidgetItem(str(value))
                        item.setTextAlignment(Qt.AlignCenter)
                        table_half.setItem(row, col, item)

                table_half.resizeColumnsToContents()
                table_half.horizontalHeader().setStretchLastSection(True)
                self.bets_layout.addWidget(table_half)
            else:
                self.bets_layout.addWidget(QLabel("Нет ставок по критерию 3.5 после округления."))
            
            # Таблица ставок на четверть
            label_q1 = QLabel("Ставки на первую четверть")
            label_q1.setFont(QFont("Arial", 12, QFont.Bold))
            self.bets_layout.addWidget(label_q1)

            if bets_q1:
                table_q1 = QTableWidget()
                table_q1.setColumnCount(7)
                table_q1.setHorizontalHeaderLabels(
                    ["Время", "Турнир", "Команда 1", "Команда 2", "Ставка", "Линия", "Разница"]
                )
                table_q1.setRowCount(len(bets_q1))

                for row, bet in enumerate(bets_q1):
                    for col, value in enumerate(bet):
                        item = QTableWidgetItem(str(value))
                        item.setTextAlignment(Qt.AlignCenter)
                        table_q1.setItem(row, col, item)

                table_q1.resizeColumnsToContents()
                table_q1.horizontalHeader().setStretchLastSection(True)
                self.bets_layout.addWidget(table_q1)
            else:
                self.bets_layout.addWidget(QLabel("Нет ставок по критерию 3.5 после округления."))
            
            self.bets_layout.addStretch()
            
            # Логируем результаты
            self.add_log(f"✓ Найдено ставок на половину: {len(bets_half)}")
            self.add_log(f"✓ Найдено ставок на четверть: {len(bets_q1)}")
            if missing:
                self.add_log(f"⚠ Пропущенных матчей: {len(missing)}")
            
        except Exception as e:
            self. add_log(f"ОШИБКА в display_bets: {str(e)}")

# =====================================================================

# =====================================================================
# Разделы Cybers: база и live

class CybersDatabase:
    """Хранит базу матчей Cybers в SQLite / PostgreSQL."""

    _SCHEMA = 'cyber'

    def __init__(self) -> None:
        self.columns = [
            "date", "tournament", "team", "home_away",
            "two_pt_made", "two_pt_attempt",
            "three_pt_made", "three_pt_attempt",
            "fta_made", "fta_attempt",
            "off_rebound", "turnovers",
            "controls", "points",
            "opponent", "attak_kef", "status",
        ]
        self.excluded_tournaments: Set[str] = set()
        self._enriched_cache: Optional[pd.DataFrame] = None
        self._aggregate_cache: Dict[str, Dict[str, Tuple[float, float]]] = {}
        self._tournament_avg: Dict[str, Tuple[float, float]] = {}
        self._tournament_predict_cache: Dict[str, Dict[str, Dict[str, Tuple[float, float]]]] = {}
        self.db_path = os.path.join(get_data_dir(), "cyber_bases.db")
        self.init_database()

    @contextmanager
    def _connect(self):
        from db_connection import db_connect
        with db_connect(schema=self._SCHEMA, sqlite_path=self.db_path) as conn:
            yield conn

    def init_database(self) -> None:
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS cyber_matches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT,
                    tournament TEXT,
                    team TEXT,
                    home_away TEXT,
                    two_pt_made REAL,
                    two_pt_attempt REAL,
                    three_pt_made REAL,
                    three_pt_attempt REAL,
                    fta_made REAL,
                    fta_attempt REAL,
                    off_rebound REAL,
                    turnovers REAL,
                    controls REAL,
                    points REAL,
                    opponent TEXT,
                    attak_kef REAL,
                    status TEXT
                )
                """
            )
            cur.execute("CREATE INDEX IF NOT EXISTS idx_cyber_tournament ON cyber_matches(tournament)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_cyber_team ON cyber_matches(team)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_cyber_opponent ON cyber_matches(opponent)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_cyber_date ON cyber_matches(date)")
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS cyber_live_matches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tournament TEXT,
                    team1 TEXT,
                    team2 TEXT,
                    total REAL,
                    calc_temp REAL
                )
                """
            )
            conn.commit()

    def add_rows(self, rows: List[dict]) -> int:
        if not rows:
            return 0
        values = [
            (
                r.get("date"),
                r.get("tournament"),
                r.get("team"),
                r.get("home_away"),
                r.get("two_pt_made"),
                r.get("two_pt_attempt"),
                r.get("three_pt_made"),
                r.get("three_pt_attempt"),
                r.get("fta_made"),
                r.get("fta_attempt"),
                r.get("off_rebound"),
                r.get("turnovers"),
                r.get("controls"),
                r.get("points"),
                r.get("opponent"),
                r.get("attak_kef"),
                r.get("status"),
            )
            for r in rows
        ]
        with self._connect() as conn:
            cur = conn.cursor()
            cur.executemany(
                """
                INSERT INTO cyber_matches (
                    date, tournament, team, home_away,
                    two_pt_made, two_pt_attempt, three_pt_made, three_pt_attempt,
                    fta_made, fta_attempt, off_rebound, turnovers,
                    controls, points, opponent, attak_kef, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                values
            )
            conn.commit()
        self.invalidate_cache()
        return len(rows)

    def clear(self) -> None:
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM cyber_matches")
            conn.commit()
        self.invalidate_cache()

    def get_dataframe(self) -> pd.DataFrame:
        with self._connect() as conn:
            df = pd.read_sql_query(
                """
                SELECT id, date, tournament, team, home_away,
                       two_pt_made, two_pt_attempt, three_pt_made, three_pt_attempt,
                       fta_made, fta_attempt, off_rebound, turnovers,
                       controls, points, opponent, attak_kef, status
                FROM cyber_matches
                ORDER BY id ASC
                """,
                conn
            )
        return df

    def get_dataframe_for_tournament(self, tournament: str) -> pd.DataFrame:
        with self._connect() as conn:
            df = pd.read_sql_query(
                """
                SELECT id, date, tournament, team, home_away,
                       two_pt_made, two_pt_attempt, three_pt_made, three_pt_attempt,
                       fta_made, fta_attempt, off_rebound, turnovers,
                       controls, points, opponent, attak_kef, status
                FROM cyber_matches
                WHERE tournament = ?
                ORDER BY id ASC
                """,
                conn,
                params=(tournament,)
            )
        return df

    def load_live_matches(self) -> List[Tuple[str, str, str, float, float]]:
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT tournament, team1, team2, total, calc_temp FROM cyber_live_matches ORDER BY id ASC"
            )
            rows = cur.fetchall()
        return [(r[0], r[1], r[2], r[3] if r[3] is not None else "", r[4] if r[4] is not None else 0.0) for r in rows]

    def save_live_matches(self, rows: List[Tuple[str, str, str, float, float]]) -> None:
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM cyber_live_matches")
            if rows:
                cur.executemany(
                    "INSERT INTO cyber_live_matches (tournament, team1, team2, total, calc_temp) VALUES (?, ?, ?, ?, ?)",
                    rows
                )
            conn.commit()

    def clear_live_matches(self) -> None:
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM cyber_live_matches")
            conn.commit()

    def find_duplicate_pairs(self) -> List[List[int]]:
        df = self.get_dataframe()
        if df.empty:
            return []
        pairs = []
        for i in range(0, len(df), 2):
            if i + 1 >= len(df):
                break
            pair = df.iloc[i:i + 2]
            if len(pair) < 2:
                continue
            row1 = pair.iloc[0]
            row2 = pair.iloc[1]
            signature = (
                tuple(row1.get(col) for col in self.columns),
                tuple(row2.get(col) for col in self.columns),
            )
            pairs.append((signature, [int(row1.get("id")), int(row2.get("id"))]))
        seen = {}
        duplicates = []
        for sig, ids in pairs:
            if sig in seen:
                duplicates.append(ids)
            else:
                seen[sig] = ids
        return duplicates

    def get_tournaments(self) -> List[str]:
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT tournament FROM cyber_matches WHERE tournament IS NOT NULL AND tournament <> ''")
            rows = cur.fetchall()
        return sorted([r[0] for r in rows])

    def delete_rows(self, ids: List[int]) -> int:
        if not ids:
            return 0
        with self._connect() as conn:
            cur = conn.cursor()
            cur.executemany("DELETE FROM cyber_matches WHERE id = ?", [(i,) for i in ids])
            conn.commit()
            deleted = cur.rowcount
        self.invalidate_cache()
        return deleted

    def delete_tournament(self, tournament: str) -> int:
        if not tournament:
            return 0
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM cyber_matches WHERE tournament = ?", (tournament,))
            conn.commit()
            deleted = cur.rowcount
        self.invalidate_cache()
        return deleted

    def update_match_field(self, row_id: int, field: str, value) -> None:
        if not row_id or not field:
            return
        if field not in self.columns:
            return
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute(f"UPDATE cyber_matches SET {field} = ? WHERE id = ?", (value, row_id))
            conn.commit()
        self.invalidate_cache()

    def invalidate_cache(self) -> None:
        self._enriched_cache = None
        self._aggregate_cache = {}
        self._tournament_avg = {}
        self._team_tournament_avg = {}
        self._tournament_predict_cache = {}

    @staticmethod
    def normalize_key(value: str) -> str:
        if value is None:
            return ""
        return " ".join(str(value).strip().lower().split())

    def get_enriched_dataframe(self) -> pd.DataFrame:
        if self._enriched_cache is not None:
            return self._enriched_cache.copy()

        df = self.get_dataframe()
        if df.empty:
            self._enriched_cache = df
            return df

        # Normalized keys for case-insensitive matching
        df["tournament_key"] = df["tournament"].fillna("").map(self.normalize_key)
        df["team_key"] = df["team"].fillna("").map(self.normalize_key)
        df["opponent_key"] = df["opponent"].fillna("").map(self.normalize_key)

        # Match points difference per pair
        diffs = []
        for i in range(0, len(df), 2):
            if i + 1 < len(df):
                diff = (df.iloc[i]["points"] or 0) - (df.iloc[i + 1]["points"] or 0)
                diffs.extend([diff, diff])
            else:
                diffs.append(0)
        df["pair_diff"] = diffs

        # Средние по турнирам для Controls и AttackKEF (фолбэк)
        tour_stats = df.groupby("tournament_key", dropna=True).agg(
            avg_controls=("controls", "mean"),
            avg_points=("points", "mean")
        )
        for t, row in tour_stats.iterrows():
            try:
                avg_controls = float(row.get("avg_controls") or 0)
            except Exception:
                avg_controls = 0.0
            try:
                avg_points = float(row.get("avg_points") or 0)
            except Exception:
                avg_points = 0.0
            attack_avg = (avg_points / avg_controls) if avg_controls else 0.0
            avg_controls = round(avg_controls, 2)
            attack_avg = round(attack_avg, 2)
            self._tournament_avg[str(t)] = (avg_controls, attack_avg)

        # Средние по командам в рамках турнира (для попарного расчёта INDEX)
        self._team_tournament_avg = {}
        team_tour_stats = df.groupby(["tournament_key", "team_key"], dropna=True).agg(
            avg_controls=("controls", "mean"),
            avg_points=("points", "mean")
        )
        for (t, team), row in team_tour_stats.iterrows():
            try:
                avg_c = float(row.get("avg_controls") or 0)
            except Exception:
                avg_c = 0.0
            try:
                avg_p = float(row.get("avg_points") or 0)
            except Exception:
                avg_p = 0.0
            self._team_tournament_avg[(str(t), str(team))] = (round(avg_c, 2), round(avg_p, 2))

        # Index
        def compute_index(row):
            tournament_key = str(row.get("tournament_key") or "")
            if tournament_key in self.excluded_tournaments:
                return 0.0
            status = str(row.get("status") or "").upper()
            controls = float(row.get("controls") or 0)
            attak = float(row.get("attak_kef") or 0)
            diff = float(row.get("pair_diff") or 0)
            idx = 10.0
            if status == "OT":
                idx -= 5.0
            if status == "FS":
                idx -= 3.0

            # Попарные средние: (среднее команды + среднее оппонента) / 2
            team_key = str(row.get("team_key") or "")
            opponent_key = str(row.get("opponent_key") or "")
            team_avg = self._team_tournament_avg.get((tournament_key, team_key), None)
            opp_avg = self._team_tournament_avg.get((tournament_key, opponent_key), None)

            if team_avg is not None and opp_avg is not None and team_avg[0] > 0 and opp_avg[0] > 0:
                avg_controls = round((team_avg[0] + opp_avg[0]) / 2, 2)
                avg_points = round((team_avg[1] + opp_avg[1]) / 2, 2)
                attack_avg = round(avg_points / avg_controls, 2) if avg_controls else 0.0
            else:
                # Фолбэк: среднее по турниру
                avg_controls, attack_avg = self._tournament_avg.get(tournament_key, (0.0, 0.0))

            if avg_controls > 0:
                low_ctrl = avg_controls * 0.9
                high_ctrl = avg_controls * 1.1
                if controls < low_ctrl:
                    idx -= 1.0
                if controls > high_ctrl:
                    idx -= 1.0
            else:
                if controls < 72.27:
                    idx -= 1.0
                if controls > 88.33:
                    idx -= 1.0

            if attack_avg > 0:
                high_att = attack_avg * 1.25
                low_att = attack_avg * 0.75
                if attak > high_att:
                    idx -= 2.0
                if attak < low_att:
                    idx -= 2.0
            else:
                if attak > 1.237:
                    idx -= 2.0
                if attak < 0.742:
                    idx -= 2.0

            if diff > 25:
                idx -= 2.0
            if diff < -25:
                idx -= 2.0
            return max(idx, 0.0)

        df["index"] = df.apply(compute_index, axis=1)

        # Time
        today = datetime.now().date()
        def parse_date_str(date_str: str) -> Optional[datetime]:
            if not date_str:
                return None
            s = date_str.strip()
            if " " in s:
                # сначала пробуем целиком, затем только дату
                candidates = [s, s.split(" ")[0]]
            else:
                candidates = [s]
            formats = [
                "%d.%m.%Y",
                "%Y-%m-%d",
                "%Y-%m-%d %H:%M:%S",
                "%d.%m.%Y %H:%M:%S",
            ]
            for cand in candidates:
                for fmt in formats:
                    try:
                        return datetime.strptime(cand, fmt)
                    except Exception:
                        continue
            return None

        def compute_time(row):
            date_str = str(row.get("date") or "").strip()
            if not date_str:
                return 10.0
            d = parse_date_str(date_str)
            if not d:
                return 10.0
            return (today - d.date()).days / 7.0

        df["time"] = df.apply(compute_time, axis=1)
        df["match_weight"] = (df["index"] * 40.0) / (df["time"] + 10.0)
        # В Excel эти значения хранятся с 7 знаками
        df["match_weight"] = df["match_weight"].round(7)

        # Weighted stats
        df["x_2pt_made"] = (df["two_pt_made"].astype(float) * df["match_weight"]).round(7)
        df["x_2pt_att"] = (df["two_pt_attempt"].astype(float) * df["match_weight"]).round(7)
        df["x_3pt_made"] = (df["three_pt_made"].astype(float) * df["match_weight"]).round(7)
        df["x_3pt_att"] = (df["three_pt_attempt"].astype(float) * df["match_weight"]).round(7)
        df["x_fta_made"] = (df["fta_made"].astype(float) * df["match_weight"]).round(7)
        df["x_fta_att"] = (df["fta_attempt"].astype(float) * df["match_weight"]).round(7)
        df["x_or"] = (df["off_rebound"].astype(float) * df["match_weight"]).round(7)
        df["x_to"] = (df["turnovers"].astype(float) * df["match_weight"]).round(7)
        df["x_controls"] = (df["controls"].astype(float) * df["match_weight"]).round(7)
        df["x_points"] = (df["points"].astype(float) * df["match_weight"]).round(7)
        df["x_attak"] = (df["attak_kef"].astype(float) * df["match_weight"]).round(7)

        self._enriched_cache = df.copy()
        return df

    def get_enriched_dataframe_filtered(self) -> pd.DataFrame:
        df = self.get_enriched_dataframe()
        if df.empty or not self.excluded_tournaments:
            return df
        return df[~df["tournament_key"].isin(self.excluded_tournaments)].copy()

    def get_aggregate(self, df: pd.DataFrame, col: str, value: str) -> Tuple[float, float]:
        if df.empty:
            return 0.0, 0.0
        subset = df[df[col] == value]
        if subset.empty:
            return 0.0, 0.0
        sum_v = subset["match_weight"].sum()
        if sum_v <= 0:
            return 0.0, 0.0
        avg_2pt_made = subset["x_2pt_made"].sum() / sum_v
        avg_2pt_att = subset["x_2pt_att"].sum() / sum_v
        avg_3pt_made = subset["x_3pt_made"].sum() / sum_v
        avg_3pt_att = subset["x_3pt_att"].sum() / sum_v
        avg_fta_made = subset["x_fta_made"].sum() / sum_v
        avg_fta_att = subset["x_fta_att"].sum() / sum_v
        avg_or = subset["x_or"].sum() / sum_v
        avg_to = subset["x_to"].sum() / sum_v

        controls = avg_2pt_att + avg_3pt_att + (avg_fta_att / 2.0) + avg_to - (avg_or / 2.0)
        points = (avg_2pt_made * 2.0) + (avg_3pt_made * 3.0) + avg_fta_made
        if controls <= 0:
            return 0.0, 0.0
        o_val = points / controls
        l_val = controls
        return o_val, l_val

    def build_aggregate_cache(self, df: pd.DataFrame) -> None:
        if df.empty:
            self._aggregate_cache = {"team": {}, "opponent": {}, "tournament": {}}
            return
        cache: Dict[str, Dict[str, Tuple[float, float]]] = {"team": {}, "opponent": {}, "tournament": {}}
        for col, key_col in (("team", "team_key"), ("opponent", "opponent_key"), ("tournament", "tournament_key")):
            grouped = df.groupby(key_col, dropna=True)
            for key, subset in grouped:
                if key is None or key == "":
                    continue
                sum_v = subset["match_weight"].sum()
                if sum_v <= 0:
                    cache[col][str(key)] = (0.0, 0.0)
                    continue
                avg_2pt_made = subset["x_2pt_made"].sum() / sum_v
                avg_2pt_att = subset["x_2pt_att"].sum() / sum_v
                avg_3pt_made = subset["x_3pt_made"].sum() / sum_v
                avg_3pt_att = subset["x_3pt_att"].sum() / sum_v
                avg_fta_made = subset["x_fta_made"].sum() / sum_v
                avg_fta_att = subset["x_fta_att"].sum() / sum_v
                avg_or = subset["x_or"].sum() / sum_v
                avg_to = subset["x_to"].sum() / sum_v
                # Controls считаем по попыткам (TM) для всех блоков.
                controls = avg_2pt_att + avg_3pt_att + (avg_fta_att / 2.0) + avg_to - (avg_or / 2.0)
                points = (avg_2pt_made * 2.0) + (avg_3pt_made * 3.0) + avg_fta_made
                if controls <= 0:
                    cache[col][str(key)] = (0.0, 0.0)
                else:
                    cache[col][str(key)] = (points / controls, controls)
        self._aggregate_cache = cache

    def _get_tournament_aggregates(self, tournament: str) -> Dict[str, Dict[str, Tuple[float, float]]]:
        """Build aggregate cache filtered to a single tournament."""
        tour_key = self.normalize_key(tournament)
        if tour_key in self._tournament_predict_cache:
            return self._tournament_predict_cache[tour_key]
        df = self.get_enriched_dataframe()
        if df.empty:
            return {"team": {}, "opponent": {}, "tournament": {}}
        df_t = df[df["tournament_key"] == tour_key].copy()
        if df_t.empty:
            return {"team": {}, "opponent": {}, "tournament": {}}
        cache: Dict[str, Dict[str, Tuple[float, float]]] = {"team": {}, "opponent": {}, "tournament": {}}
        for col, key_col in (("team", "team_key"), ("opponent", "opponent_key"), ("tournament", "tournament_key")):
            grouped = df_t.groupby(key_col, dropna=True)
            for key, subset in grouped:
                if key is None or key == "":
                    continue
                sum_v = subset["match_weight"].sum()
                if sum_v <= 0:
                    cache[col][str(key)] = (0.0, 0.0)
                    continue
                avg_2pt_made = subset["x_2pt_made"].sum() / sum_v
                avg_2pt_att = subset["x_2pt_att"].sum() / sum_v
                avg_3pt_made = subset["x_3pt_made"].sum() / sum_v
                avg_3pt_att = subset["x_3pt_att"].sum() / sum_v
                avg_fta_made = subset["x_fta_made"].sum() / sum_v
                avg_fta_att = subset["x_fta_att"].sum() / sum_v
                avg_or = subset["x_or"].sum() / sum_v
                avg_to = subset["x_to"].sum() / sum_v
                controls = avg_2pt_att + avg_3pt_att + (avg_fta_att / 2.0) + avg_to - (avg_or / 2.0)
                points = (avg_2pt_made * 2.0) + (avg_3pt_made * 3.0) + avg_fta_made
                if controls <= 0:
                    cache[col][str(key)] = (0.0, 0.0)
                else:
                    cache[col][str(key)] = (points / controls, controls)
        self._tournament_predict_cache[tour_key] = cache
        return cache

    def compute_predict(self, tournament: str, team1: str, team2: str) -> Tuple[float, float, float, float]:
        agg = self._get_tournament_aggregates(tournament)

        def get_agg(col: str, value: str) -> Tuple[float, float]:
            key = self.normalize_key(value)
            return agg.get(col, {}).get(key, (0.0, 0.0))

        o_team1, l_team1 = get_agg("team", team1)
        o_team2, l_team2 = get_agg("team", team2)
        o_opp_team1, l_opp_team1 = get_agg("opponent", team1)
        o_opp_team2, l_opp_team2 = get_agg("opponent", team2)
        o_tour, l_tour = get_agg("tournament", tournament)

        temp = ((l_team1 + l_team2 + l_opp_team1 + l_opp_team2) / 2.0) - l_tour
        it1 = temp * (o_team1 + o_opp_team2 - o_tour) + 2.0
        it2 = temp * (o_team2 + o_opp_team1 - o_tour) - 2.0
        predict = it1 + it2
        return predict, temp, it1, it2

    def get_debug_values(self, tournament: str, team1: str, team2: str) -> Dict[str, float]:
        agg = self._get_tournament_aggregates(tournament)

        def get_agg(col: str, value: str) -> Tuple[float, float]:
            key = self.normalize_key(value)
            return agg.get(col, {}).get(key, (0.0, 0.0))

        o_team1, l_team1 = get_agg("team", team1)
        o_team2, l_team2 = get_agg("team", team2)
        o_opp1, l_opp1 = get_agg("opponent", team1)
        o_opp2, l_opp2 = get_agg("opponent", team2)
        o_tour, l_tour = get_agg("tournament", tournament)

        temp = ((l_team1 + l_team2 + l_opp1 + l_opp2) / 2.0) - l_tour
        it1 = temp * (o_team1 + o_opp2 - o_tour) + 2.0
        it2 = temp * (o_team2 + o_opp1 - o_tour) - 2.0
        predict = it1 + it2

        return {
            "o_team1": o_team1, "l_team1": l_team1,
            "o_team2": o_team2, "l_team2": l_team2,
            "o_opp1": o_opp1, "l_opp1": l_opp1,
            "o_opp2": o_opp2, "l_opp2": l_opp2,
            "o_tour": o_tour, "l_tour": l_tour,
            "temp": temp, "it1": it1, "it2": it2, "predict": predict,
        }


class CybersTableDelegate(QStyledItemDelegate):
    """Фиксирует цвета редактора в таблице Cybers."""

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if isinstance(editor, QLineEdit):
            editor.setStyleSheet(
                "QLineEdit {"
                "background-color: #11244A;"
                "color: #E0E5EC;"
                "selection-background-color: #2EC4B6;"
                "selection-color: #0A192F;"
                "padding: 4px;"
                "}"
            )
            editor.setFrame(False)
            editor.setAutoFillBackground(True)
            editor.setMinimumHeight(24)
            palette = editor.palette()
            palette.setColor(QPalette.Base, QColor("#11244A"))
            palette.setColor(QPalette.Text, QColor("#E0E5EC"))
            palette.setColor(QPalette.Highlight, QColor("#2EC4B6"))
            palette.setColor(QPalette.HighlightedText, QColor("#0A192F"))
            editor.setPalette(palette)
        return editor

    def setEditorData(self, editor, index):
        super().setEditorData(editor, index)
        if isinstance(editor, QLineEdit):
            editor.selectAll()


class CybersBasesPage(QWidget):
    """Страница для загрузки базы Cybers из Excel (вставка)."""

    def __init__(self, db: CybersDatabase, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.db = db
        self.setup_ui()

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        self.base_tab = QWidget()
        base_layout = QVBoxLayout(self.base_tab)
        self.tabs.addTab(self.base_tab, "База")

        self.summary_tab = QWidget()
        summary_layout = QVBoxLayout(self.summary_tab)
        self.tabs.addTab(self.summary_tab, "Сводная статистика")

        info_label = QLabel(
            "Вставьте строки из Excel (17 столбцов, табуляция) и нажмите 'Импортировать'.\n"
            "Можно вставлять как TSV, так и HTML-таблицу из буфера.\n"
            "Один матч = 2 строки. Данные будут сохранены в базу Cybers."
        )
        info_label.setWordWrap(True)
        base_layout.addWidget(info_label)

        self.input_text = QTextEdit()
        self.input_text.setAcceptRichText(False)
        self.input_text.setPlaceholderText("Скопируйте строки из Excel и вставьте сюда...")
        self.input_text.setStyleSheet(
            """
            QTextEdit {
                color: #E0E0E0;
                background-color: #18181e;
                border: 1px solid #333339;
            }
            """
        )
        base_layout.addWidget(self.input_text)

        btn_layout = QHBoxLayout()
        self.paste_btn = QPushButton("Вставить из буфера")
        self.paste_btn.clicked.connect(self.paste_from_clipboard)
        btn_layout.addWidget(self.paste_btn)

        self.import_btn = QPushButton("Импортировать")
        self.import_btn.clicked.connect(self.on_import_clicked)
        btn_layout.addWidget(self.import_btn)

        self.clear_input_btn = QPushButton("Очистить ввод")
        self.clear_input_btn.clicked.connect(self.clear_input)
        btn_layout.addWidget(self.clear_input_btn)

        self.clear_btn = QPushButton("Очистить базу")
        self.clear_btn.clicked.connect(self.on_clear_clicked)
        btn_layout.addWidget(self.clear_btn)

        btn_layout.addStretch()
        base_layout.addLayout(btn_layout)

        self.count_label = QLabel("Загружено строк: 0")
        base_layout.addWidget(self.count_label)

        tools_layout = QHBoxLayout()
        tools_layout.addWidget(QLabel("Сортировка:"))
        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["По дате", "По турниру"])
        self.sort_combo.currentIndexChanged.connect(self.reload_table)
        tools_layout.addWidget(self.sort_combo)

        tools_layout.addSpacing(20)
        tools_layout.addWidget(QLabel("Поиск:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Текст для поиска...")
        self.search_input.returnPressed.connect(self.on_search)
        tools_layout.addWidget(self.search_input)
        self.search_btn = QPushButton("Найти")
        self.search_btn.clicked.connect(self.on_search)
        tools_layout.addWidget(self.search_btn)
        self.prev_btn = QPushButton("Назад")
        self.prev_btn.clicked.connect(self.on_search_prev)
        tools_layout.addWidget(self.prev_btn)
        self.next_btn = QPushButton("Вперёд")
        self.next_btn.clicked.connect(self.on_search_next)
        tools_layout.addWidget(self.next_btn)
        self.search_count_label = QLabel("")
        tools_layout.addWidget(self.search_count_label)
        tools_layout.addStretch()
        base_layout.addLayout(tools_layout)
        self.search_matches: List[Tuple[int, int]] = []
        self.search_index = -1

        self.table = QTableWidget()
        self.table.setColumnCount(17)
        self.table.setHorizontalHeaderLabels([
            "Дата", "Турнир", "Команда", "H/A",
            "2PTA", "2PTM", "3PTA", "3PTM",
            "FTA", "FTM", "OR", "TO",
            "Controls", "Points", "Opponent", "AttakKEF", "Status"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.MultiSelection)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setSortingEnabled(False)
        self.table.setAlternatingRowColors(False)
        self.table.setStyleSheet(
            """
            QTableWidget {
                color: #E0E0E0;
                background-color: #18181e;
                gridline-color: #333339;
            }
            QTableWidget::item {
                selection-background-color: #2a2a2e;
                selection-color: #E0E0E0;
            }
            QTableWidget::item:focus {
                background-color: #1f2f4a;
                color: #E0E5EC;
            }
            QTableWidget QLineEdit, QTableView QLineEdit {
                background-color: #11244A;
                color: #E0E5EC;
                selection-background-color: #2EC4B6;
                selection-color: #0A192F;
                border: none;
            }
            """
        )
        base_layout.addWidget(self.table)

        self.selection_count_label = QLabel("Выбрано строк: 0")
        base_layout.addWidget(self.selection_count_label, alignment=Qt.AlignRight)

        # Делегат, чтобы текст в редакторе был видимым
        delegate = CybersTableDelegate(self.table)
        self.table.setItemDelegate(delegate)
        for col in range(self.table.columnCount()):
            self.table.setItemDelegateForColumn(col, delegate)
        try:
            self.table.selectionModel().selectionChanged.connect(self.on_selection_changed)
        except Exception:
            pass

        # Удаление строк/турниров
        delete_layout = QHBoxLayout()
        self.delete_selected_btn = QPushButton("Удалить выбранные")
        self.delete_selected_btn.clicked.connect(self.delete_selected_rows)
        delete_layout.addWidget(self.delete_selected_btn)

        delete_layout.addWidget(QLabel("Турнир:"))
        self.tournament_delete_combo = QComboBox()
        self.tournament_delete_combo.setMinimumWidth(200)
        delete_layout.addWidget(self.tournament_delete_combo)

        self.delete_tournament_btn = QPushButton("Удалить турнир")
        self.delete_tournament_btn.clicked.connect(self.delete_selected_tournament)
        delete_layout.addWidget(self.delete_tournament_btn)

        self.replace_btn = QPushButton("Заменить")
        self.replace_btn.clicked.connect(self.replace_values_dialog)
        delete_layout.addWidget(self.replace_btn)

        self.open_tournament_btn = QPushButton("Открыть турнир")
        self.open_tournament_btn.clicked.connect(self.open_tournament_dialog)
        delete_layout.addWidget(self.open_tournament_btn)

        self.duplicates_btn = QPushButton("Найти дубли")
        self.duplicates_btn.clicked.connect(self.find_duplicates)
        delete_layout.addWidget(self.duplicates_btn)
        delete_layout.addStretch()
        base_layout.addLayout(delete_layout)

        # Сводная статистика
        summary_info = QLabel("Сводная статистика по турнирам (средние значения).")
        summary_info.setWordWrap(True)
        summary_layout.addWidget(summary_info)
        summary_filter_layout = QHBoxLayout()
        summary_filter_layout.addWidget(QLabel("Турнир:"))
        self.summary_tournament_combo = QComboBox()
        self.summary_tournament_combo.addItem("Все турниры")
        self.summary_tournament_combo.currentIndexChanged.connect(self.refresh_summary_table)
        summary_filter_layout.addWidget(self.summary_tournament_combo)
        summary_filter_layout.addStretch()
        summary_layout.addLayout(summary_filter_layout)
        self.refresh_summary_btn = QPushButton("Обновить")
        self.refresh_summary_btn.clicked.connect(self.refresh_summary_table)
        summary_layout.addWidget(self.refresh_summary_btn)
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(16)
        self.summary_table.setHorizontalHeaderLabels([
            "Турнир", "Игр", "2PTA", "2PTM", "3PTA", "3PTM", "FTA", "FTM",
            "OR", "TO", "Controls", "Points", "P/C", "2pt %", "3pt %", "FT %"
        ])
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.summary_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.summary_table.setSortingEnabled(True)
        summary_layout.addWidget(self.summary_table)

        self.reload_table()

    def paste_from_clipboard(self) -> None:
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if text:
            self.input_text.setPlainText(text)

    def clear_input(self) -> None:
        self.input_text.clear()

    def on_import_clicked(self) -> None:
        raw_text = self.input_text.toPlainText().strip()
        if raw_text:
            cells_rows = self.parse_tsv(raw_text)
        else:
            cells_rows = self.parse_clipboard()
            if not cells_rows:
                QMessageBox.warning(self, "Внимание", "Нет строк для импорта.")
                return

        rows, skipped, skipped_lines = self.build_rows(cells_rows)
        inserted = self.db.add_rows(rows)
        self.input_text.clear()
        self.reload_table()

        if skipped:
            QMessageBox.warning(
                self,
                "Импорт частично завершён",
                f"Загружено строк: {inserted}\nПропущено строк: {skipped}"
            )
            if skipped_lines:
                self.show_skipped_lines(skipped_lines)
        else:
            QMessageBox.information(self, "Готово", f"Загружено строк: {inserted}")

    def on_clear_clicked(self) -> None:
        reply = QMessageBox.question(
            self,
            "Очистить базу",
            "Вы уверены, что хотите удалить все записи?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.db.clear()
            self.reload_table()

    def parse_clipboard(self) -> List[Tuple[List[str], str]]:
        clipboard = QApplication.clipboard()
        mime = clipboard.mimeData()
        if mime is None:
            return []
        if mime.hasHtml():
            html = mime.html()
            return self.parse_html_table(html)
        text = clipboard.text()
        if text:
            return self.parse_tsv(text)
        return []

    def parse_html_table(self, html: str) -> List[Tuple[List[str], str]]:
        try:
            tables = pd.read_html(html, header=None)
        except Exception:
            return []
        if not tables:
            return []
        df = tables[0].fillna("")
        rows = []
        for _, row in df.iterrows():
            values = [str(v).strip() for v in row.tolist()]
            if any(values):
                rows.append((values, "\t".join(values)))
        return rows

    def parse_tsv(self, raw_text: str) -> List[Tuple[List[str], str]]:
        rows = []
        for line in raw_text.splitlines():
            if not line.strip():
                continue
            cells = [c.strip() for c in line.split("\t")]
            rows.append((cells, line))
        return rows

    def build_rows(self, cells_rows: List[Tuple[List[str], str]]) -> Tuple[List[dict], int, List[str]]:
        rows: List[dict] = []
        skipped = 0
        skipped_lines: List[str] = []

        def is_number_token(value: str) -> bool:
            v = str(value).strip().replace(",", ".")
            if not v:
                return False
            try:
                float(v)
                return True
            except Exception:
                return False

        def to_float(value: str) -> float:
            v = str(value).strip()
            if not v:
                return 0.0
            v = v.replace(",", ".")
            try:
                return float(v)
            except Exception:
                return 0.0

        def to_int(value: str) -> int:
            v = str(value).strip()
            if not v:
                return 0
            v = v.replace(",", ".")
            try:
                return int(round(float(v)))
            except Exception:
                return 0

        normalized: List[Tuple[List[str], str]] = []
        unstructured_lines: List[str] = []
        for cells, raw_line in cells_rows:
            if len(cells) == 16:
                normalized.append((cells + [""], raw_line))
            elif len(cells) >= 17:
                normalized.append((cells[:17], raw_line))
            elif len(cells) == 1 and str(cells[0]).strip():
                unstructured_lines.append(str(cells[0]))
            else:
                skipped += 1
                if raw_line:
                    skipped_lines.append(raw_line)

        i = 0
        def normalize_date_value(date_val: str) -> str:
            if not date_val:
                return date_val
            s = date_val.strip()
            if " " in s:
                s = s.split(" ")[0]
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
                try:
                    d = datetime.strptime(s, fmt)
                    return d.strftime("%d.%m.%Y")
                except Exception:
                    continue
            return date_val

        while i < len(normalized):
            if i + 1 >= len(normalized):
                skipped += 1
                break
            r1, raw1 = normalized[i]
            r2, raw2 = normalized[i + 1]
            t1 = str(r1[1]).strip()
            t2 = str(r2[1]).strip()
            ha1 = str(r1[3]).strip().upper()
            ha2 = str(r2[3]).strip().upper()
            if t1 and t2 and t1 != t2:
                skipped += 2
                skipped_lines.extend([raw1, raw2])
                i += 2
                continue
            if (ha1 and ha2) and not (ha1 == "H" and ha2 == "A"):
                skipped += 2
                skipped_lines.extend([raw1, raw2])
                i += 2
                continue

            for cells in (r1, r2):
                date_val = normalize_date_value(str(cells[0]).strip())

                row = {
                    "date": date_val,
                    "tournament": str(cells[1]).strip(),
                    "team": str(cells[2]).strip(),
                    "home_away": str(cells[3]).strip(),
                    "two_pt_made": to_int(cells[4]),
                    "two_pt_attempt": to_int(cells[5]),
                    "three_pt_made": to_int(cells[6]),
                    "three_pt_attempt": to_int(cells[7]),
                    "fta_made": to_int(cells[8]),
                    "fta_attempt": to_int(cells[9]),
                    "off_rebound": to_int(cells[10]),
                    "turnovers": to_int(cells[11]),
                    "controls": to_float(cells[12]),
                    "points": to_int(cells[13]),
                    "opponent": str(cells[14]).strip(),
                    "attak_kef": to_float(cells[15]),
                    "status": str(cells[16]).strip(),
                }
                rows.append(row)
            i += 2

        # Пытаемся разобрать строки без табуляции (умная валидация)
        if unstructured_lines:
            parsed_rows, skipped_unstructured = self.parse_unstructured_pairs(
                unstructured_lines, is_number_token, to_float
            )
            rows.extend(parsed_rows)
            skipped += skipped_unstructured

        return rows, skipped, skipped_lines

    def show_skipped_lines(self, skipped_lines: List[str]) -> None:
        dialog = QDialog(self)
        dialog.setWindowTitle("Пропущенные строки")
        dialog.setMinimumSize(600, 400)
        layout = QVBoxLayout(dialog)
        info = QLabel("Эти строки не были импортированы. Проверьте формат.")
        layout.addWidget(info)
        text = QPlainTextEdit()
        text.setReadOnly(True)
        text.setPlainText("\n".join(skipped_lines))
        layout.addWidget(text)
        btn = QDialogButtonBox(QDialogButtonBox.Close)
        btn.rejected.connect(dialog.reject)
        layout.addWidget(btn)
        dialog.exec_()

    def parse_unstructured_pairs(
        self,
        lines: List[str],
        is_number_token,
        to_float
    ) -> Tuple[List[dict], int]:
        parsed_rows: List[dict] = []
        skipped = 0
        i = 0

        def parse_line(line: str):
            tokens = [t for t in line.strip().split() if t]
            if not tokens:
                return None
            # Дата
            date_val = tokens[0] if tokens else ""
            rest = tokens[1:]
            if not rest:
                return None
            # Ищем H/A
            ha_idx = None
            for idx, tok in enumerate(rest):
                if tok.upper() in {"H", "A"}:
                    ha_idx = idx
                    break
            if ha_idx is None:
                return None
            before_ha = rest[:ha_idx]
            after_ha = rest[ha_idx + 1:]
            if len(after_ha) < 10:
                return None
            # Первые 10 чисел после H/A — статистика
            stats_tokens = []
            consumed = 0
            for tok in after_ha:
                if is_number_token(tok):
                    stats_tokens.append(tok)
                    consumed += 1
                    if consumed == 10:
                        break
                else:
                    # Если встретили текст до набора 10 чисел — формат неверный
                    return None
            if consumed < 10:
                return None
            remainder = after_ha[consumed:]
            status = ""
            if remainder and str(remainder[-1]).upper() in {"FS", "OT"}:
                status = str(remainder[-1]).upper()
                remainder = remainder[:-1]
            # AttakKEF — последнее числовое значение
            attak_idx = None
            for ridx in range(len(remainder) - 1, -1, -1):
                if is_number_token(remainder[ridx]):
                    attak_idx = ridx
                    break
            if attak_idx is None:
                return None
            opponent_tokens = remainder[:attak_idx]
            attak_kef = remainder[attak_idx]
            opponent = " ".join(opponent_tokens).strip()

            return {
                "date": date_val,
                "before_ha": before_ha,
                "home_away": rest[ha_idx].upper(),
                "stats": stats_tokens,
                "opponent": opponent,
                "attak_kef": attak_kef,
                "status": status,
            }

        while i < len(lines):
            if i + 1 >= len(lines):
                skipped += 1
                break
            l1 = parse_line(lines[i])
            l2 = parse_line(lines[i + 1])
            if not l1 or not l2:
                skipped += 2
                i += 2
                continue
            # Выделяем турнир как общий префикс до H/A
            b1 = l1["before_ha"]
            b2 = l2["before_ha"]
            prefix_len = 0
            for t1, t2 in zip(b1, b2):
                if t1 == t2:
                    prefix_len += 1
                else:
                    break
            if prefix_len == 0:
                skipped += 2
                i += 2
                continue
            tournament = " ".join(b1[:prefix_len]).strip()
            team1 = " ".join(b1[prefix_len:]).strip()
            team2 = " ".join(b2[prefix_len:]).strip()
            if not tournament or not team1 or not team2:
                skipped += 2
                i += 2
                continue
            if not (l1["home_away"] == "H" and l2["home_away"] == "A"):
                skipped += 2
                i += 2
                continue

            def build_row(line_data, team_name):
                stats = line_data["stats"]
                return {
                    "date": line_data["date"],
                    "tournament": tournament,
                    "team": team_name,
                    "home_away": line_data["home_away"],
                    "two_pt_made": to_int(stats[0]),
                    "two_pt_attempt": to_int(stats[1]),
                    "three_pt_made": to_int(stats[2]),
                    "three_pt_attempt": to_int(stats[3]),
                    "fta_made": to_int(stats[4]),
                    "fta_attempt": to_int(stats[5]),
                    "off_rebound": to_int(stats[6]),
                    "turnovers": to_int(stats[7]),
                    "controls": to_float(stats[8]),
                    "points": to_int(stats[9]),
                    "opponent": line_data["opponent"],
                    "attak_kef": to_float(line_data["attak_kef"]),
                    "status": line_data["status"],
                }

            parsed_rows.append(build_row(l1, team1))
            parsed_rows.append(build_row(l2, team2))
            i += 2

        return parsed_rows, skipped

    def reload_table(self) -> None:
        df = self.get_sorted_df(self.db.get_dataframe())
        self.table.setRowCount(len(df))
        for row_idx, (_, row) in enumerate(df.iterrows()):
            date_str = str(row.get("date") or "")
            values = [
                date_str,
                row.get("tournament", ""),
                row.get("team", ""),
                row.get("home_away", ""),
                row.get("two_pt_made", ""),
                row.get("two_pt_attempt", ""),
                row.get("three_pt_made", ""),
                row.get("three_pt_attempt", ""),
                row.get("fta_made", ""),
                row.get("fta_attempt", ""),
                row.get("off_rebound", ""),
                row.get("turnovers", ""),
                row.get("controls", ""),
                row.get("points", ""),
                row.get("opponent", ""),
                row.get("attak_kef", ""),
                row.get("status", ""),
            ]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem()
                sort_value = None
                if col_idx == 0:
                    item.setText(str(val))
                    try:
                        d = datetime.strptime(str(val), "%d.%m.%Y").date()
                        sort_value = QDate(d.year, d.month, d.day)
                    except Exception:
                        sort_value = str(val)
                    item.setData(Qt.UserRole, row.get("id"))
                elif col_idx in {12, 15}:  # controls, attak_kef
                    try:
                        num = float(val)
                        if col_idx == 12:
                            num = self.round_to_half(num)
                            item.setText(f"{num:.1f}")
                        else:
                            item.setText(f"{num:.2f}")
                        sort_value = num
                    except Exception:
                        item.setText(str(val))
                        sort_value = str(val)
                elif col_idx in {4, 5, 6, 7, 8, 9, 10, 11, 13}:
                    try:
                        num = int(round(float(val)))
                        item.setText(str(num))
                        sort_value = num
                    except Exception:
                        item.setText(str(val))
                        sort_value = str(val)
                else:
                    item.setText(str(val))
                    sort_value = str(val)
                if sort_value is not None:
                    item.setData(Qt.UserRole + 1, sort_value)
                self.table.setItem(row_idx, col_idx, item)

        self.count_label.setText(f"Загружено строк: {len(df)}")
        self.search_matches = []
        self.search_index = -1
        self.search_count_label.setText("")
        self.selection_count_label.setText("Выбрано строк: 0")
        if hasattr(self, "live_page") and self.live_page is not None:
            try:
                self.live_page.recompute_all()
            except Exception:
                pass
        self.update_tournament_delete_combo()
        self.refresh_summary_table()

    def round_to_half(self, value: float) -> float:
        return round(value * 2) / 2.0

    def get_sorted_df(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        sort_mode = "date"
        if hasattr(self, "sort_combo"):
            sort_mode = "tournament" if self.sort_combo.currentText() == "По турниру" else "date"
        rows = df.reset_index(drop=True)
        pairs = []
        for i in range(0, len(rows), 2):
            pair = rows.iloc[i:i + 2]
            if pair.empty:
                continue
            first = pair.iloc[0]
            date_str = str(first.get("date") or "")
            try:
                date_val = datetime.strptime(date_str, "%d.%m.%Y").date()
            except Exception:
                date_val = datetime.min.date()
            tournament = str(first.get("tournament") or "")
            key = (date_val, tournament.lower())
            if sort_mode == "tournament":
                key = (tournament.lower(), date_val)
            pairs.append((key, pair))
        pairs.sort(key=lambda x: x[0])
        sorted_pairs = [p for _, p in pairs]
        return pd.concat(sorted_pairs, ignore_index=True)

    def refresh_summary_table(self) -> None:
        df = self.db.get_dataframe()
        if df.empty:
            self.summary_table.setRowCount(0)
            return
        df["tournament"] = df["tournament"].fillna("")

        # Обновляем список турниров
        tournaments = sorted([t for t in df["tournament"].unique() if t])
        current = self.summary_tournament_combo.currentText()
        self.summary_tournament_combo.blockSignals(True)
        self.summary_tournament_combo.clear()
        self.summary_tournament_combo.addItem("Все турниры")
        self.summary_tournament_combo.addItems(tournaments)
        if current and current in tournaments:
            idx = self.summary_tournament_combo.findText(current)
            if idx >= 0:
                self.summary_tournament_combo.setCurrentIndex(idx)
        self.summary_tournament_combo.blockSignals(False)

        selected = self.summary_tournament_combo.currentText()
        if selected and selected != "Все турниры":
            df = df[df["tournament"] == selected]

        groups = df.groupby("tournament", dropna=True)

        self.summary_table.setRowCount(0)
        for tournament, g in groups:
            if not tournament:
                continue
            games = 0
            g_sorted = g.sort_values("id") if "id" in g.columns else g
            rows_list = g_sorted.to_dict("records")
            for i in range(0, len(rows_list), 2):
                if i + 1 < len(rows_list):
                    games += 1
            sum_2pta = g["two_pt_made"].astype(float).sum()
            sum_2ptm = g["two_pt_attempt"].astype(float).sum()
            sum_3pta = g["three_pt_made"].astype(float).sum()
            sum_3ptm = g["three_pt_attempt"].astype(float).sum()
            sum_fta = g["fta_made"].astype(float).sum()
            sum_ftm = g["fta_attempt"].astype(float).sum()
            sum_or = g["off_rebound"].astype(float).sum()
            sum_to = g["turnovers"].astype(float).sum()
            sum_controls = g["controls"].astype(float).sum()
            sum_points = g["points"].astype(float).sum()

            avg_2pta = sum_2pta / len(g)
            avg_2ptm = sum_2ptm / len(g)
            avg_3pta = sum_3pta / len(g)
            avg_3ptm = sum_3ptm / len(g)
            avg_fta = sum_fta / len(g)
            avg_ftm = sum_ftm / len(g)
            avg_or = sum_or / len(g)
            avg_to = sum_to / len(g)
            avg_controls = sum_controls / len(g)
            avg_points = sum_points / len(g)

            pc = (avg_points / avg_controls) if avg_controls else 0.0
            p2 = (sum_2pta / sum_2ptm * 100.0) if sum_2ptm else 0.0
            p3 = (sum_3pta / sum_3ptm * 100.0) if sum_3ptm else 0.0
            pft = (sum_fta / sum_ftm * 100.0) if sum_ftm else 0.0

            row_idx = self.summary_table.rowCount()
            self.summary_table.insertRow(row_idx)
            values = [
                tournament,
                games,
                avg_2pta, avg_2ptm, avg_3pta, avg_3ptm, avg_fta, avg_ftm,
                avg_or, avg_to, avg_controls, avg_points,
                pc, p2, p3, pft
            ]
            for col, val in enumerate(values):
                item = QTableWidgetItem()
                if col == 0:
                    item.setText(str(val))
                elif col == 1:
                    item.setText(str(int(val)))
                elif col in {13, 14, 15}:
                    item.setText(f"{val:.1f}%")
                elif col == 12:
                    item.setText(f"{val:.2f}")
                else:
                    if col == 10:
                        item.setText(f"{self.round_to_half(val):.1f}")
                    else:
                        item.setText(f"{val:.1f}")
                self.summary_table.setItem(row_idx, col, item)

    def on_search(self) -> None:
        query = self.search_input.text().strip()
        self.search_matches = []
        self.search_index = -1
        if not query:
            self.search_count_label.setText("")
            return
        q = query.lower()
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                if item and q in item.text().lower():
                    self.search_matches.append((r, c))
        if not self.search_matches:
            self.search_count_label.setText("Совпадений: 0")
            return
        self.search_index = 0
        self.focus_match(self.search_index)

    def on_search_next(self) -> None:
        if not self.search_matches:
            return
        self.search_index = (self.search_index + 1) % len(self.search_matches)
        self.focus_match(self.search_index)

    def on_search_prev(self) -> None:
        if not self.search_matches:
            return
        self.search_index = (self.search_index - 1) % len(self.search_matches)
        self.focus_match(self.search_index)

    def focus_match(self, index: int) -> None:
        if not self.search_matches:
            return
        row, col = self.search_matches[index]
        self.table.clearSelection()
        self.table.setCurrentCell(row, col)
        self.table.scrollToItem(self.table.item(row, col), QAbstractItemView.PositionAtCenter)
        self.search_count_label.setText(
            f"Совпадений: {len(self.search_matches)} | {index + 1}/{len(self.search_matches)} "
            f"(строка {row + 1}, столбец {col + 1})"
        )

    def on_selection_changed(self, selected, deselected) -> None:
        try:
            count = len(self.table.selectionModel().selectedRows())
        except Exception:
            count = 0
        self.selection_count_label.setText(f"Выбрано строк: {count}")

    def update_tournament_delete_combo(self) -> None:
        tournaments = self.db.get_tournaments()
        current = self.tournament_delete_combo.currentText() if hasattr(self, "tournament_delete_combo") else ""
        self.tournament_delete_combo.blockSignals(True)
        self.tournament_delete_combo.clear()
        self.tournament_delete_combo.addItem("Выберите турнир")
        self.tournament_delete_combo.addItems(tournaments)
        if current and current in tournaments:
            idx = self.tournament_delete_combo.findText(current)
            if idx >= 0:
                self.tournament_delete_combo.setCurrentIndex(idx)
        self.tournament_delete_combo.blockSignals(False)

    def update_tournament_checkboxes(self) -> None:
        if not hasattr(self, "no_layout"):
            return
        for i in reversed(range(self.no_layout.count())):
            widget = self.no_layout.itemAt(i).widget()
            if widget is not None:
                widget.setParent(None)

        tournaments = self.db.get_tournaments()
        if not tournaments:
            self.no_group.setVisible(False)
            return
        self.no_group.setVisible(True)
        for tournament in tournaments:
            norm = self.db.normalize_key(tournament)
            checkbox = QCheckBox(tournament)
            checkbox.setChecked(norm in self.db.excluded_tournaments)
            checkbox.setStyleSheet(
                """
                QCheckBox {
                    spacing: 8px;
                    font-size: 13px;
                    color: #E0E0E0;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                    border: 1px solid #6c6c6c;
                    background: #1f1f26;
                }
                QCheckBox::indicator:checked {
                    background: #2EC4B6;
                    border: 1px solid #2EC4B6;
                }
                """
            )
            def on_state_changed(state, t=tournament, cb=checkbox):
                if cb.isChecked():
                    self.db.excluded_tournaments.add(self.db.normalize_key(t))
                else:
                    self.db.excluded_tournaments.discard(self.db.normalize_key(t))
                self.db.invalidate_cache()
                if hasattr(self, "live_page") and self.live_page is not None:
                    try:
                        self.live_page.recompute_all()
                    except Exception:
                        pass
            checkbox.stateChanged.connect(on_state_changed)
            self.no_layout.addWidget(checkbox)

    def delete_selected_rows(self) -> None:
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.information(self, "Удаление", "Выберите строки для удаления.")
            return
        reply = QMessageBox.question(
            self,
            "Удалить строки",
            f"Удалить выбранные строки: {len(selected)}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        ids = []
        for idx in selected:
            item = self.table.item(idx.row(), 0)
            if item:
                row_id = item.data(Qt.UserRole)
                if row_id is not None:
                    ids.append(int(row_id))
        self.db.delete_rows(ids)
        self.reload_table()

    def delete_selected_tournament(self) -> None:
        tournament = self.tournament_delete_combo.currentText().strip()
        if not tournament or tournament == "Выберите турнир":
            QMessageBox.information(self, "Удаление", "Выберите турнир.")
            return
        reply = QMessageBox.question(
            self,
            "Удалить турнир",
            f"Удалить все записи турнира '{tournament}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        self.db.delete_tournament(tournament)
        self.reload_table()

    def open_tournament_dialog(self) -> None:
        dialog = CyberTournamentDialog(self.db, self)
        dialog.exec_()
        self.reload_table()

    def find_duplicates(self) -> None:
        duplicates = self.db.find_duplicate_pairs()
        if not duplicates:
            QMessageBox.information(self, "Дубли", "Дубли не найдены.")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Найденные дубли")
        dialog.setMinimumSize(700, 500)
        layout = QVBoxLayout(dialog)
        info = QLabel(f"Найдено дублирующихся матчей: {len(duplicates)}")
        layout.addWidget(info)
        text = QPlainTextEdit()
        text.setReadOnly(True)
        text_lines = []
        for idx, pair_ids in enumerate(duplicates, start=1):
            text_lines.append(f"{idx}. IDs: {pair_ids[0]} / {pair_ids[1]}")
        text.setPlainText("\n".join(text_lines))
        layout.addWidget(text)
        btn_layout = QHBoxLayout()
        delete_btn = QPushButton("Удалить дубли")
        close_btn = QPushButton("Закрыть")
        btn_layout.addWidget(delete_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        def on_delete():
            reply = QMessageBox.question(
                dialog,
                "Удалить дубли",
                f"Удалить все найденные дубли ({len(duplicates)} матчей)?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
            ids_to_delete = [i for pair in duplicates for i in pair]
            deleted = self.db.delete_rows(ids_to_delete)
            QMessageBox.information(
                dialog,
                "Удалено",
                f"Удалено строк: {deleted}\nIDs:\n" + ", ".join(map(str, ids_to_delete))
            )
            dialog.accept()
            self.reload_table()

        delete_btn.clicked.connect(on_delete)
        close_btn.clicked.connect(dialog.reject)
        dialog.exec_()

    def replace_values_dialog(self) -> None:
        if not self.db:
            QMessageBox.warning(self, "Ошибка", "База данных недоступна.")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Поиск и замена")
        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.addWidget(QLabel("Что заменить:"))
        old_edit = QLineEdit()
        dlg_layout.addWidget(old_edit)
        dlg_layout.addWidget(QLabel("На что заменить:"))
        new_edit = QLineEdit()
        dlg_layout.addWidget(new_edit)

        scope_layout = QHBoxLayout()
        scope_layout.addWidget(QLabel("Область:"))
        scope_combo = QComboBox()
        scope_combo.addItems(["Во всей таблице", "Только в выделенных строках", "Только в выделенных ячейках"])
        scope_layout.addWidget(scope_combo)
        scope_layout.addStretch()
        dlg_layout.addLayout(scope_layout)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        dlg_layout.addWidget(btn_box)

        def parse_value_for_column(col_idx: int, text: str):
            t = text.strip()
            if col_idx == 0:
                if not t:
                    return ""
                try:
                    return datetime.strptime(t, "%d.%m.%Y").strftime("%d.%m.%Y")
                except Exception:
                    return t
            if col_idx in {12, 15}:
                if not t:
                    return 0.0
                try:
                    return float(t.replace(",", "."))
                except Exception:
                    return 0.0
            if col_idx in {4, 5, 6, 7, 8, 9, 10, 11, 13}:
                if not t:
                    return 0
                try:
                    return int(round(float(t.replace(",", "."))))
                except Exception:
                    return 0
            return t

        def on_replace() -> None:
            old_val = old_edit.text()
            new_val = new_edit.text()
            if not old_val:
                QMessageBox.warning(dialog, "Внимание", "Введите текст для замены.")
                return

            indexes = []
            scope = scope_combo.currentText()
            if scope == "Только в выделенных ячейках":
                indexes = self.table.selectedIndexes()
                if not indexes:
                    QMessageBox.information(dialog, "Замена", "Нет выделенных ячеек.")
                    return
            elif scope == "Только в выделенных строках":
                selected_rows = self.table.selectionModel().selectedRows()
                if not selected_rows:
                    QMessageBox.information(dialog, "Замена", "Нет выделенных строк.")
                    return
                for row_idx in [idx.row() for idx in selected_rows]:
                    for c in range(self.table.columnCount()):
                        indexes.append(self.table.model().index(row_idx, c))
            else:
                for r in range(self.table.rowCount()):
                    for c in range(self.table.columnCount()):
                        indexes.append(self.table.model().index(r, c))

            replaced = 0
            field_map = {
                0: "date",
                1: "tournament",
                2: "team",
                3: "home_away",
                4: "two_pt_made",
                5: "two_pt_attempt",
                6: "three_pt_made",
                7: "three_pt_attempt",
                8: "fta_made",
                9: "fta_attempt",
                10: "off_rebound",
                11: "turnovers",
                12: "controls",
                13: "points",
                14: "opponent",
                15: "attak_kef",
                16: "status",
            }

            for idx in indexes:
                r = idx.row()
                c = idx.column()
                item = self.table.item(r, c)
                if item is None:
                    continue
                current_text = item.text()
                if old_val not in current_text:
                    continue
                new_text = current_text.replace(old_val, new_val)
                if new_text == current_text:
                    continue

                parsed_val = parse_value_for_column(c, new_text)
                item.setText(str(new_text))
                replaced += 1

                id_item = self.table.item(r, 0)
                row_id = id_item.data(Qt.UserRole) if id_item else None
                field_name = field_map.get(c)
                if row_id is None or not field_name:
                    continue
                try:
                    self.db.update_match_field(int(row_id), field_name, parsed_val)
                except Exception:
                    pass

            self.reload_table()
            QMessageBox.information(dialog, "Завершено", f"Заменено {replaced} ячеек.")
            dialog.accept()

        btn_box.accepted.connect(on_replace)
        btn_box.rejected.connect(dialog.reject)
        dialog.exec_()


class CyberTournamentDialog(QDialog):
    def __init__(self, db: CybersDatabase, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.db = db
        self._loading = False
        self.setup_ui()

    def setup_ui(self) -> None:
        self.setWindowTitle("Турнир — просмотр и редактирование")
        self.setMinimumSize(900, 600)
        layout = QVBoxLayout(self)

        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("Турнир:"))
        self.tournament_combo = QComboBox()
        top_layout.addWidget(self.tournament_combo)
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.clicked.connect(self.load_table)
        top_layout.addWidget(self.refresh_btn)
        top_layout.addStretch()
        layout.addLayout(top_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(17)
        self.table.setHorizontalHeaderLabels([
            "Дата", "Турнир", "Команда", "H/A",
            "2PTA", "2PTM", "3PTA", "3PTM",
            "FTA", "FTM", "OR", "TO",
            "Controls", "Points", "Opponent", "AttakKEF", "Status"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.MultiSelection)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        layout.addWidget(self.table)
        self.table.itemChanged.connect(self.on_item_changed)

        btn_layout = QHBoxLayout()
        self.delete_btn = QPushButton("Удалить выбранные")
        self.delete_btn.clicked.connect(self.delete_selected)
        btn_layout.addWidget(self.delete_btn)
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        self.load_tournaments()
        self.tournament_combo.currentIndexChanged.connect(self.load_table)
        self.load_table()

    def load_tournaments(self) -> None:
        tournaments = self.db.get_tournaments()
        self.tournament_combo.blockSignals(True)
        self.tournament_combo.clear()
        self.tournament_combo.addItems(tournaments)
        self.tournament_combo.blockSignals(False)

    def round_to_half(self, value: float) -> float:
        return round(value * 2) / 2.0

    def load_table(self) -> None:
        tournament = self.tournament_combo.currentText()
        if not tournament:
            self.table.setRowCount(0)
            return
        df = self.db.get_dataframe_for_tournament(tournament)
        self._loading = True
        self.table.setRowCount(len(df))
        for row_idx, (_, row) in enumerate(df.iterrows()):
            values = [
                row.get("date", ""),
                row.get("tournament", ""),
                row.get("team", ""),
                row.get("home_away", ""),
                row.get("two_pt_made", ""),
                row.get("two_pt_attempt", ""),
                row.get("three_pt_made", ""),
                row.get("three_pt_attempt", ""),
                row.get("fta_made", ""),
                row.get("fta_attempt", ""),
                row.get("off_rebound", ""),
                row.get("turnovers", ""),
                row.get("controls", ""),
                row.get("points", ""),
                row.get("opponent", ""),
                row.get("attak_kef", ""),
                row.get("status", ""),
            ]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem()
                if col_idx == 0:
                    item.setText(str(val))
                    item.setData(Qt.UserRole, row.get("id"))
                elif col_idx == 12:
                    try:
                        num = float(val)
                        num = self.round_to_half(num)
                        item.setText(f"{num:.1f}")
                    except Exception:
                        item.setText(str(val))
                elif col_idx == 15:
                    try:
                        num = float(val)
                        item.setText(f"{num:.2f}")
                    except Exception:
                        item.setText(str(val))
                elif col_idx in {4, 5, 6, 7, 8, 9, 10, 11, 13}:
                    try:
                        item.setText(str(int(round(float(val)))))
                    except Exception:
                        item.setText(str(val))
                else:
                    item.setText(str(val))
                self.table.setItem(row_idx, col_idx, item)
        self._loading = False

    def on_item_changed(self, item: QTableWidgetItem) -> None:
        if self._loading:
            return
        row = item.row()
        col = item.column()
        id_item = self.table.item(row, 0)
        row_id = id_item.data(Qt.UserRole) if id_item else None
        if row_id is None:
            return
        field_map = {
            0: "date",
            1: "tournament",
            2: "team",
            3: "home_away",
            4: "two_pt_made",
            5: "two_pt_attempt",
            6: "three_pt_made",
            7: "three_pt_attempt",
            8: "fta_made",
            9: "fta_attempt",
            10: "off_rebound",
            11: "turnovers",
            12: "controls",
            13: "points",
            14: "opponent",
            15: "attak_kef",
            16: "status",
        }
        field_name = field_map.get(col)
        if not field_name:
            return
        text = item.text().strip()
        if field_name == "date":
            value = text
        elif field_name in {"tournament", "team", "home_away", "opponent", "status"}:
            value = text
        elif field_name in {"controls", "attak_kef"}:
            try:
                value = float(text.replace(",", "."))
            except Exception:
                value = 0.0
        else:
            try:
                value = int(round(float(text.replace(",", "."))))
            except Exception:
                value = 0
        try:
            self.db.update_match_field(int(row_id), field_name, value)
        except Exception:
            pass

    def delete_selected(self) -> None:
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.information(self, "Удаление", "Выберите строки для удаления.")
            return
        reply = QMessageBox.question(
            self,
            "Удалить строки",
            f"Удалить выбранные строки: {len(selected)}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        ids = []
        for idx in selected:
            item = self.table.item(idx.row(), 0)
            if item:
                row_id = item.data(Qt.UserRole)
                if row_id is not None:
                    ids.append(int(row_id))
        self.db.delete_rows(ids)
        self.load_table()


class CyberLivePage(QWidget):
    """Страница Cyber LIVE (пока заглушка)."""

    def __init__(self, db: CybersDatabase, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.db = db
        self.setup_ui()

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        tabs = QTabWidget()
        layout.addWidget(tabs)

        self.lines_tab = QWidget()
        self.lines_layout = QVBoxLayout(self.lines_tab)
        self.setup_lines_tab()
        tabs.addTab(self.lines_tab, "Lines")

        self.predict_tab = QWidget()
        self.predict_layout = QVBoxLayout(self.predict_tab)
        self.setup_predict_tab()
        tabs.addTab(self.predict_tab, "Predict")

        self.load_saved_matches()

    def setup_lines_tab(self) -> None:
        info = QLabel(
            "Вставьте матчи (Турнир, Команда1, Команда2, Тотал) и нажмите 'Импортировать'.\n"
            "Формат: TSV или HTML из Excel."
        )
        info.setWordWrap(True)
        self.lines_layout.addWidget(info)

        self.lines_input = QTextEdit()
        self.lines_input.setAcceptRichText(False)
        self.lines_input.setPlaceholderText("Скопируйте строки из Excel и вставьте сюда...")
        self.lines_input.setStyleSheet(
            """
            QTextEdit {
                color: #E0E0E0;
                background-color: #18181e;
                border: 1px solid #333339;
            }
            """
        )
        self.lines_layout.addWidget(self.lines_input)

        btn_layout = QHBoxLayout()
        self.lines_paste_btn = QPushButton("Вставить из буфера")
        self.lines_paste_btn.clicked.connect(self.lines_paste_from_clipboard)
        btn_layout.addWidget(self.lines_paste_btn)

        self.lines_import_btn = QPushButton("Импортировать")
        self.lines_import_btn.clicked.connect(self.on_lines_import)
        btn_layout.addWidget(self.lines_import_btn)

        self.lines_clear_btn = QPushButton("Очистить")
        self.lines_clear_btn.clicked.connect(self.clear_lines)
        btn_layout.addWidget(self.lines_clear_btn)

        self.lines_add_btn = QPushButton("+")
        self.lines_add_btn.setMinimumWidth(40)
        self.lines_add_btn.setMaximumWidth(50)
        self.lines_add_btn.setToolTip("Добавить матч вручную")
        self.lines_add_btn.clicked.connect(self.add_manual_match)
        btn_layout.addWidget(self.lines_add_btn)

        btn_layout.addStretch()
        self.lines_layout.addLayout(btn_layout)

        self.lines_table = QTableWidget()
        self.lines_table.setColumnCount(11)
        self.lines_table.setHorizontalHeaderLabels([
            "Турнир", "Команда 1", "Команда 2", "Тотал", "TEMP",
            "Predict", "UNDER", "OVER", "CalcTEMP", "T2H", "T2H Predict"
        ])
        self.lines_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.lines_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.lines_table.setSelectionMode(QAbstractItemView.MultiSelection)
        self.lines_table.setEditTriggers(
            QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed
        )
        self.lines_table.setStyleSheet(
            "QTableWidget {"
            "color: #E0E0E0;"
            "background-color: #18181e;"
            "gridline-color: #333339;"
            "}"
            "QTableWidget::item {"
            "selection-background-color: #2a2a2e;"
            "selection-color: #E0E0E0;"
            "padding: 4px;"
            "}"
            "QTableWidget QLineEdit {"
            "color: #E0E0E0;"
            "background-color: #23232a;"
            "border: 1px solid #5a5aff;"
            "padding: 4px;"
            "}"
        )
        self.lines_table.itemChanged.connect(self.on_lines_item_changed)
        self.lines_layout.addWidget(self.lines_table)
        self._updating_lines = False

        lines_btn_layout = QHBoxLayout()
        self.lines_delete_btn = QPushButton("Удалить выбранные")
        self.lines_delete_btn.clicked.connect(self.delete_selected_lines)
        lines_btn_layout.addWidget(self.lines_delete_btn)
        self.lines_debug_btn = QPushButton("Показать расчёт")
        self.lines_debug_btn.clicked.connect(self.show_calc_debug)
        lines_btn_layout.addWidget(self.lines_debug_btn)
        lines_btn_layout.addStretch()
        self.lines_layout.addLayout(lines_btn_layout)

    def setup_predict_tab(self) -> None:
        info = QLabel("Результаты расчётов Predict/TEMP по текущим матчам из Lines.")
        info.setWordWrap(True)
        self.predict_layout.addWidget(info)

        self.predict_table = QTableWidget()
        self.predict_table.setColumnCount(7)
        self.predict_table.setHorizontalHeaderLabels([
            "Турнир", "Команда 1", "Команда 2", "TEMP", "Predict", "IT1", "IT2"
        ])
        self.predict_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.predict_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.predict_table.setStyleSheet(
            "QTableWidget {"
            "color: #E0E0E0;"
            "background-color: #18181e;"
            "gridline-color: #333339;"
            "}"
            "QTableWidget::item {"
            "selection-background-color: #2a2a2e;"
            "selection-color: #E0E0E0;"
            "}"
        )
        self.predict_layout.addWidget(self.predict_table)

    def add_manual_match(self) -> None:
        """Добавляет пустую строку для ручного ввода матча."""
        row_idx = self.lines_table.rowCount()
        self._updating_lines = True
        self.lines_table.blockSignals(True)
        self.lines_table.setRowCount(row_idx + 1)
        self.predict_table.setRowCount(row_idx + 1)
        self.set_lines_row(row_idx, "", "", "", "", 0.0)
        # Делаем турнир, команды и тотал редактируемыми
        for col in (0, 1, 2, 3):
            item = self.lines_table.item(row_idx, col)
            if item:
                item.setFlags(item.flags() | Qt.ItemIsEditable)
        self.lines_table.blockSignals(False)
        self._updating_lines = False
        self.save_current_matches()

    def lines_paste_from_clipboard(self) -> None:
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if text:
            self.lines_input.setPlainText(text)

    def clear_lines(self) -> None:
        reply = QMessageBox.question(
            self,
            "Очистить",
            "Вы уверены, что хотите очистить все матчи?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        reply2 = QMessageBox.question(
            self,
            "Подтвердите очистку",
            "Точно удалить все матчи из Cyber LIVE?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply2 != QMessageBox.Yes:
            return
        self.lines_input.clear()
        self.lines_table.setRowCount(0)
        self.predict_table.setRowCount(0)
        self.db.clear_live_matches()

    def on_lines_import(self) -> None:
        raw_text = self.lines_input.toPlainText().strip()
        if raw_text:
            rows = self.parse_lines_tsv(raw_text)
        else:
            rows = self.parse_lines_clipboard()
        if not rows:
            QMessageBox.warning(self, "Внимание", "Нет строк для импорта.")
            return
        self.append_lines_table(rows)
        self.lines_input.clear()

    def parse_lines_clipboard(self) -> List[List[str]]:
        clipboard = QApplication.clipboard()
        mime = clipboard.mimeData()
        if mime and mime.hasHtml():
            html = mime.html()
            return self.parse_lines_html(html)
        text = clipboard.text()
        if text:
            return self.parse_lines_tsv(text)
        return []

    def parse_lines_html(self, html: str) -> List[List[str]]:
        try:
            tables = pd.read_html(html, header=None)
        except Exception:
            return []
        if not tables:
            return []
        df = tables[0].fillna("")
        rows = []
        for _, row in df.iterrows():
            values = [str(v).strip() for v in row.tolist()]
            if any(values):
                rows.append(values)
        return rows

    def parse_lines_tsv(self, raw_text: str) -> List[List[str]]:
        rows = []
        for line in raw_text.splitlines():
            if not line.strip():
                continue
            cells = [c.strip() for c in line.split("\t")]
            rows.append(cells)
        return rows

    def append_lines_table(self, rows: List[List[str]]) -> None:
        parsed = []
        for cells in rows:
            cleaned = [c for c in cells if str(c).strip() != ""]
            if len(cleaned) < 3:
                continue
            tournament = str(cleaned[0]).strip()
            team1 = str(cleaned[1]).strip()
            team2 = str(cleaned[2]).strip()
            total = ""
            # 5 столбцов: Турнир Команда1 Команда2 Фора Тотал
            # Пропускаем 4-й (Фора, index 3), берём 5-й (Тотал, index 4)
            if len(cleaned) >= 5:
                try:
                    total = float(str(cleaned[4]).replace(",", "."))
                except Exception:
                    total = ""
            elif len(cleaned) >= 4:
                try:
                    total = float(str(cleaned[3]).replace(",", "."))
                except Exception:
                    total = ""
            parsed.append((tournament, team1, team2, total, 0.0))

        existing = self.get_lines_rows()
        combined = existing + parsed
        combined.sort(key=lambda x: str(x[0]).lower())

        self._updating_lines = True
        self.lines_table.blockSignals(True)
        self.lines_table.setRowCount(len(combined))
        self.predict_table.setRowCount(len(combined))
        for row_idx, (tournament, team1, team2, total, calc_temp) in enumerate(combined):
            self.set_lines_row(row_idx, tournament, team1, team2, total, calc_temp)
            self.update_predict_row(row_idx, tournament, team1, team2)
        self.lines_table.blockSignals(False)
        self._updating_lines = False
        self.save_current_matches()

    def set_lines_row(self, row_idx: int, tournament: str, team1: str, team2: str, total, calc_temp: float = 0.0):
        predict, temp, it1, it2 = self.db.compute_predict(tournament, team1, team2)
        pre_total = total if isinstance(total, (int, float)) and total != "" else predict
        under = (pre_total - predict) if (pre_total - predict) > 3 else ""
        over = (predict - pre_total) if (pre_total - predict) < -3 else ""
        if temp != 0:
            z = pre_total / (2.0 * temp)
            t2h = z * ((temp + calc_temp) / 2.0)
        else:
            t2h = 0.0

        # T2H Predict: корректировка T2H на процент отклонения Predict от Total
        t2h_predict = self._calc_t2h_predict(pre_total, predict, t2h)

        values = [
            tournament, team1, team2,
            self.format_num(pre_total),
            self.format_num(temp), self.format_num(predict), self.format_num(under),
            self.format_num(over),
            self.format_num(calc_temp), self.format_num(t2h),
            self._format_t2h_predict(t2h_predict)
        ]
        for col_idx, val in enumerate(values):
            item = QTableWidgetItem(str(val))
            if col_idx in (0, 1, 2, 3, 8):
                item.setFlags(item.flags() | Qt.ItemIsEditable)
            else:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if col_idx == 6:
                self.apply_under_style(item)
            if col_idx == 7:
                self.apply_over_style(item)
            if col_idx == 8:
                item.setBackground(QColor("#1a2b45"))
            if col_idx == 10:
                item.setBackground(QColor("#2b1a3d"))
            self.lines_table.setItem(row_idx, col_idx, item)

    def _calc_t2h_predict(self, pre_total, predict, t2h) -> str:
        """Рассчитывает T2H Predict: T2H скорректированный на % отклонения Predict от Total.

        Формула: percent = (predict - total) / total; t2h_predict = t2h * (1 + percent)
        Показывается только если |predict - total| >= 3.
        """
        try:
            pre_total_f = float(pre_total) if pre_total != "" else 0.0
            predict_f = float(predict) if predict != "" else 0.0
            t2h_f = float(t2h) if t2h != "" else 0.0
        except Exception:
            return ""
        if pre_total_f == 0 or predict_f == 0 or t2h_f == 0:
            return ""
        if abs(predict_f - pre_total_f) < 3:
            return ""
        pct = (predict_f - pre_total_f) / pre_total_f
        return round(t2h_f * (1 + pct), 2)

    def _format_t2h_predict(self, value) -> str:
        if value == "" or value is None:
            return ""
        try:
            return f"{float(value):.1f}"
        except Exception:
            return str(value)

    def on_lines_item_changed(self, item: QTableWidgetItem) -> None:
        if self._updating_lines:
            return
        row = item.row()
        try:
            tournament = self.lines_table.item(row, 0).text()
            team1 = self.lines_table.item(row, 1).text()
            team2 = self.lines_table.item(row, 2).text()
        except Exception:
            return

        try:
            pre_total = float(self.lines_table.item(row, 3).text().replace(",", "."))
        except Exception:
            pre_total = None
        try:
            calc_temp = float(self.lines_table.item(row, 8).text().replace(",", "."))
        except Exception:
            calc_temp = 0.0

        predict, temp, it1, it2 = self.db.compute_predict(tournament, team1, team2)
        if pre_total is None:
            pre_total = predict
        under = (pre_total - predict) if (pre_total - predict) > 3 else ""
        over = (predict - pre_total) if (pre_total - predict) < -3 else ""
        if temp != 0:
            z = pre_total / (2.0 * temp)
            t2h = z * ((temp + calc_temp) / 2.0)
        else:
            t2h = 0.0

        t2h_predict = self._calc_t2h_predict(pre_total, predict, t2h)

        self.lines_table.blockSignals(True)
        self.lines_table.item(row, 4).setText(self.format_num(temp))
        self.lines_table.item(row, 5).setText(self.format_num(predict))
        self.lines_table.item(row, 6).setText(self.format_num(under))
        self.lines_table.item(row, 7).setText(self.format_num(over))
        self.lines_table.item(row, 9).setText(self.format_num(t2h))
        # T2H Predict
        t2h_pred_item = self.lines_table.item(row, 10)
        if t2h_pred_item is None:
            t2h_pred_item = QTableWidgetItem()
            t2h_pred_item.setFlags(t2h_pred_item.flags() & ~Qt.ItemIsEditable)
            t2h_pred_item.setBackground(QColor("#2b1a3d"))
            self.lines_table.setItem(row, 10, t2h_pred_item)
        t2h_pred_item.setText(self._format_t2h_predict(t2h_predict))
        self.apply_under_style(self.lines_table.item(row, 6))
        self.apply_over_style(self.lines_table.item(row, 7))
        self.lines_table.blockSignals(False)
        self.update_predict_row(row, tournament, team1, team2)
        self.save_current_matches()

    def update_predict_row(self, row: int, tournament: str, team1: str, team2: str) -> None:
        predict, temp, it1, it2 = self.db.compute_predict(tournament, team1, team2)
        vals = [
            tournament, team1, team2,
            self.format_num(temp), self.format_num(predict), self.format_num(it1), self.format_num(it2)
        ]
        for c, v in enumerate(vals):
            item = self.predict_table.item(row, c)
            if item is None:
                item = QTableWidgetItem(str(v))
                self.predict_table.setItem(row, c, item)
            else:
                item.setText(str(v))

    def format_num(self, value) -> str:
        if value == "" or value is None:
            return ""
        try:
            return f"{float(value):.1f}"
        except Exception:
            return str(value)

    def apply_under_style(self, item: Optional[QTableWidgetItem]) -> None:
        if item is None:
            return
        if not item.text().strip():
            font = item.font()
            font.setBold(False)
            item.setFont(font)
            item.setForeground(QColor("#E0E0E0"))
            return
        font = item.font()
        font.setBold(True)
        item.setFont(font)
        item.setForeground(QColor("#ff4d4f"))

    def apply_over_style(self, item: Optional[QTableWidgetItem]) -> None:
        if item is None:
            return
        if not item.text().strip():
            font = item.font()
            font.setBold(False)
            item.setFont(font)
            item.setForeground(QColor("#E0E0E0"))
            return
        font = item.font()
        font.setBold(True)
        item.setFont(font)
        item.setForeground(QColor("#52c41a"))

    def get_lines_rows(self) -> List[Tuple[str, str, str, float, float]]:
        rows = []
        for r in range(self.lines_table.rowCount()):
            try:
                tournament = self.lines_table.item(r, 0).text()
                team1 = self.lines_table.item(r, 1).text()
                team2 = self.lines_table.item(r, 2).text()
            except Exception:
                continue
            try:
                total = float(self.lines_table.item(r, 3).text().replace(",", "."))
            except Exception:
                total = ""
            try:
                calc_temp = float(self.lines_table.item(r, 8).text().replace(",", "."))
            except Exception:
                calc_temp = 0.0
            rows.append((tournament, team1, team2, total, calc_temp))
        return rows

    def recompute_all(self) -> None:
        self._updating_lines = True
        self.lines_table.blockSignals(True)
        for r in range(self.lines_table.rowCount()):
            try:
                tournament = self.lines_table.item(r, 0).text()
                team1 = self.lines_table.item(r, 1).text()
                team2 = self.lines_table.item(r, 2).text()
            except Exception:
                continue
            try:
                pre_total = float(self.lines_table.item(r, 3).text().replace(",", "."))
            except Exception:
                pre_total = None
            try:
                calc_temp = float(self.lines_table.item(r, 8).text().replace(",", "."))
            except Exception:
                calc_temp = 0.0
            predict, temp, it1, it2 = self.db.compute_predict(tournament, team1, team2)
            if pre_total is None:
                pre_total = predict
            under = (pre_total - predict) if (pre_total - predict) > 3 else ""
            over = (predict - pre_total) if (pre_total - predict) < -3 else ""
            if temp != 0:
                z = pre_total / (2.0 * temp)
                t2h = z * ((temp + calc_temp) / 2.0)
            else:
                t2h = 0.0
            t2h_predict = self._calc_t2h_predict(pre_total, predict, t2h)
            self.lines_table.item(r, 4).setText(self.format_num(temp))
            self.lines_table.item(r, 5).setText(self.format_num(predict))
            self.lines_table.item(r, 6).setText(self.format_num(under))
            self.lines_table.item(r, 7).setText(self.format_num(over))
            self.lines_table.item(r, 9).setText(self.format_num(t2h))
            # T2H Predict
            t2h_pred_item = self.lines_table.item(r, 10)
            if t2h_pred_item is None:
                t2h_pred_item = QTableWidgetItem()
                t2h_pred_item.setFlags(t2h_pred_item.flags() & ~Qt.ItemIsEditable)
                t2h_pred_item.setBackground(QColor("#2b1a3d"))
                self.lines_table.setItem(r, 10, t2h_pred_item)
            t2h_pred_item.setText(self._format_t2h_predict(t2h_predict))
            self.apply_under_style(self.lines_table.item(r, 6))
            self.apply_over_style(self.lines_table.item(r, 7))
            self.update_predict_row(r, tournament, team1, team2)
        self.lines_table.blockSignals(False)
        self._updating_lines = False

    def delete_selected_lines(self) -> None:
        selected = self.lines_table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.information(self, "Удаление", "Выберите строки для удаления.")
            return
        reply = QMessageBox.question(
            self,
            "Удалить строки",
            f"Удалить выбранные матчи: {len(selected)}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        rows = sorted([idx.row() for idx in selected], reverse=True)
        for row in rows:
            self.lines_table.removeRow(row)
            self.predict_table.removeRow(row)
        self.save_current_matches()

    def save_current_matches(self) -> None:
        rows = self.get_lines_rows()
        self.db.save_live_matches(rows)

    def load_saved_matches(self) -> None:
        rows = self.db.load_live_matches()
        if not rows:
            return
        combined = [(t, t1, t2, total, calc) for (t, t1, t2, total, calc) in rows]
        combined.sort(key=lambda x: str(x[0]).lower())
        self._updating_lines = True
        self.lines_table.blockSignals(True)
        self.lines_table.setRowCount(len(combined))
        self.predict_table.setRowCount(len(combined))
        for row_idx, (tournament, team1, team2, total, calc_temp) in enumerate(combined):
            self.set_lines_row(row_idx, tournament, team1, team2, total, calc_temp)
            self.update_predict_row(row_idx, tournament, team1, team2)
        self.lines_table.blockSignals(False)
        self._updating_lines = False

    def show_calc_debug(self) -> None:
        selected = self.lines_table.selectionModel().selectedRows()
        if not selected:
            row = self.lines_table.currentRow()
            if row < 0:
                QMessageBox.information(self, "Расчёт", "Выберите строку матча.")
                return
            rows = [row]
        else:
            rows = [idx.row() for idx in selected][:1]

        row = rows[0]
        try:
            tournament = self.lines_table.item(row, 0).text()
            team1 = self.lines_table.item(row, 1).text()
            team2 = self.lines_table.item(row, 2).text()
        except Exception:
            QMessageBox.warning(self, "Расчёт", "Не удалось прочитать строку.")
            return

        debug = self.db.get_debug_values(tournament, team1, team2)
        excluded = sorted(list(self.db.excluded_tournaments))
        info_lines = [
            f"Турнир: {tournament}",
            f"Команда 1: {team1}",
            f"Команда 2: {team2}",
            "",
            f"O_team1: {debug['o_team1']:.6f} | L_team1: {debug['l_team1']:.6f}",
            f"O_team2: {debug['o_team2']:.6f} | L_team2: {debug['l_team2']:.6f}",
            f"O_opp1: {debug['o_opp1']:.6f} | L_opp1: {debug['l_opp1']:.6f}",
            f"O_opp2: {debug['o_opp2']:.6f} | L_opp2: {debug['l_opp2']:.6f}",
            f"O_tour: {debug['o_tour']:.6f} | L_tour: {debug['l_tour']:.6f}",
            "",
            f"TEMP: {debug['temp']:.6f}",
            f"IT1: {debug['it1']:.6f}",
            f"IT2: {debug['it2']:.6f}",
            f"Predict: {debug['predict']:.6f}",
            "",
            f"Исключённые турниры (No): {', '.join(excluded) if excluded else 'нет'}",
        ]

        dialog = QDialog(self)
        dialog.setWindowTitle("Промежуточные расчёты")
        dialog.setMinimumSize(600, 400)
        layout = QVBoxLayout(dialog)
        text = QPlainTextEdit()
        text.setReadOnly(True)
        text.setPlainText("\n".join(info_lines))
        layout.addWidget(text)
        btn = QDialogButtonBox(QDialogButtonBox.Close)
        btn.rejected.connect(dialog.reject)
        layout.addWidget(btn)
        dialog.exec_()


# Новый раздел: Анализ половин
class HalfsAnalysisPage(QWidget):
    """
    Интерфейс и логика для анализа матчей на основе тоталов первой половины.

    Пользователь выбирает Excel‑файл с матчами и дату матчей. Программа
    считывает значения первой половины для каждого матча (по указанным
    диапазонам столбцов), вычисляет коэффициенты over/under для первого тайма
    на основе данных из базы половин и выводит только те матчи, у которых
    коэффициент не больше 1.6. При желании пользователь может установить
    уведомления, которые появятся за 5 минут до начала выбранных матчей.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        # Импортируем базу половин при создании страницы
        try:
            from halfs_database import HalfsDatabase  # type: ignore
            self.db = HalfsDatabase()
        except Exception:
            self.db = None
        # Список информации о матчах для уведомлений
        self.matches: List[Dict[str, any]] = []
        # Инициализируем интерфейс
        self.setup_ui()

    def setup_ui(self) -> None:
        """Создаёт интерфейс раздела 'Анализ половин'."""
        layout = QVBoxLayout(self)
        # Заголовок
        title_label = QLabel("Анализ половин")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        layout.addSpacing(10)
        # Выбор файла
        file_layout = QHBoxLayout()
        self.file_line_edit = QLineEdit()
        self.file_line_edit.setPlaceholderText("Выберите файл .xlsx с матчами…")
        self.file_line_edit.setReadOnly(True)
        self.select_file_btn = QPushButton("Выбрать файл")
        self.select_file_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.file_line_edit)
        file_layout.addWidget(self.select_file_btn)
        layout.addLayout(file_layout)
        # Выбор даты матчей
        date_layout = QHBoxLayout()
        date_label = QLabel("Дата:")
        date_label.setMinimumWidth(40)
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        date_layout.addWidget(date_label)
        date_layout.addWidget(self.date_edit)
        date_layout.addStretch()
        layout.addLayout(date_layout)
        # Кнопка запуска анализа
        self.analyze_btn = QPushButton("Анализировать")
        self.analyze_btn.clicked.connect(self.analyze_matches)
        layout.addWidget(self.analyze_btn)
        # Таблица для отображения результатов
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels([
            "Дата", "Время", "Турнир", "Команда 1", "Команда 2",
            "Тотал 1‑й половины", "Тип", "Коэффициент"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.table)
        # Блок уведомлений
        notif_layout = QHBoxLayout()
        self.select_all_checkbox = QCheckBox("Выбрать все для уведомлений")
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
        notif_layout.addWidget(self.select_all_checkbox)
        self.notify_btn = QPushButton("Включить уведомления")
        self.notify_btn.clicked.connect(self.enable_notifications)
        notif_layout.addWidget(self.notify_btn)
        notif_layout.addStretch()
        layout.addLayout(notif_layout)
        # Системный трей для уведомлений
        self.tray_icon = QSystemTrayIcon(self)
        try:
            # Попробуем взять иконку основного окна
            parent_window = self.window()
            if hasattr(parent_window, "windowIcon"):
                self.tray_icon.setIcon(parent_window.windowIcon())
        except Exception:
            pass
        self.tray_icon.show()

    def select_file(self) -> None:
        """Открывает диалог выбора Excel‑файла с матчами."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите файл с матчами", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.file_line_edit.setText(file_path)
            self.file_path = file_path

    def analyze_matches(self) -> None:
        """
        Читает выбранный файл и выводит матчи, где коэффициент OVER или UNDER
        для первого тайма не превышает 1.6. Используются данные базы половин.
        """
        # Проверяем выбран файл
        if not hasattr(self, "file_path") or not self.file_path:
            QMessageBox.warning(self, "Файл не выбран", "Выберите файл для анализа.")
            return
        # Пытаемся прочитать файл
        try:
            df = pd.read_excel(self.file_path, header=0)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка чтения файла", f"Не удалось прочитать файл: {str(e)}")
            return
        # Очищаем таблицу и список матчей
        self.table.setRowCount(0)
        self.matches.clear()
        # Выбранная дата
        date_str = self.date_edit.date().toString("dd.MM.yyyy")
        # Определяем группы столбцов
        groups = [
            ("A", "B", "C", "D", "G"),
            ("I", "J", "K", "L", "O"),
            ("Q", "R", "S", "T", "W"),
            ("Y", "Z", "AA", "AB", "AE"),
            ("AG", "AH", "AI", "AJ", "AM"),
            ("AO", "AP", "AQ", "AR", "AU"),
            ("AW", "AX", "AY", "AZ", "BC"),
            ("BE", "BF", "BG", "BH", "BK"),
        ]
        # Функция преобразования буквы в номер столбца (нумерация с 0)
        def col_to_index(col: str) -> int:
            col = col.upper()
            idx = 0
            for c in col:
                if 'A' <= c <= 'Z':
                    idx = idx * 26 + (ord(c) - ord('A') + 1)
            return idx - 1
        # Проходим по всем строкам начиная со второй (первая строка содержит названия файлов)
        for i, row in df.iterrows():
            if i == 0:
                continue
            for grp in groups:
                time_col, tour_col, t1_col, t2_col, tot_col = grp
                try:
                    time_val = row.iloc[col_to_index(time_col)]
                    tour_val = row.iloc[col_to_index(tour_col)]
                    t1_val = row.iloc[col_to_index(t1_col)]
                    t2_val = row.iloc[col_to_index(t2_col)]
                    tot_val = row.iloc[col_to_index(tot_col)]
                except Exception:
                    # Если столбец отсутствует – пропускаем
                    continue
                # Пропускаем пустые или недопустимые тоталы
                if pd.isna(tot_val):
                    continue
                # Строковые значения очищаем от пробелов и заменяем запятую на точку
                if isinstance(tot_val, str):
                    if tot_val.strip() in {"-", "", "0"}:
                        continue
                    tot_str = tot_val.replace(",", ".").strip()
                else:
                    tot_str = str(tot_val)
                try:
                    tot_float = float(tot_str)
                except Exception:
                    continue
                if tot_float == 0:
                    continue
                # Остальные поля в строковый вид
                team1_name = str(t1_val).strip()
                team2_name = str(t2_val).strip()
                tournament_name = str(tour_val).strip()
                time_str = str(time_val).strip()
                # Без базы нет смысла считать коэффициенты
                if not self.db:
                    continue
                # Пороги: четверть, половина и матч
                q_threshold = tot_float / 2.0
                h_threshold = tot_float
                m_threshold = tot_float * 2.0
                coeffs = self.db.get_tot_coefficients(team1_name, team2_name, q_threshold, h_threshold, m_threshold, tournament_name)
                if not coeffs:
                    continue
                over_coeff = coeffs.get('over', {}).get('h1')
                under_coeff = coeffs.get('under', {}).get('h1')
                selected_type = None
                coeff_value = None
                if over_coeff is not None and over_coeff > 0 and over_coeff <= 1.6:
                    selected_type = "OVER"
                    coeff_value = over_coeff
                if under_coeff is not None and under_coeff > 0 and under_coeff <= 1.6:
                    # Выбираем меньший коэффициент, если оба подходят
                    if selected_type is None or (coeff_value is not None and under_coeff < coeff_value):
                        selected_type = "UNDER"
                        coeff_value = under_coeff
                if selected_type and coeff_value is not None:
                    row_idx = self.table.rowCount()
                    self.table.insertRow(row_idx)
                    self.table.setItem(row_idx, 0, QTableWidgetItem(date_str))
                    self.table.setItem(row_idx, 1, QTableWidgetItem(time_str))
                    self.table.setItem(row_idx, 2, QTableWidgetItem(tournament_name))
                    self.table.setItem(row_idx, 3, QTableWidgetItem(team1_name))
                    self.table.setItem(row_idx, 4, QTableWidgetItem(team2_name))
                    self.table.setItem(row_idx, 5, QTableWidgetItem(str(tot_float)))
                    self.table.setItem(row_idx, 6, QTableWidgetItem(selected_type))
                    self.table.setItem(row_idx, 7, QTableWidgetItem(f"{coeff_value:.2f}"))
                    # Запоминаем информацию о матче для уведомлений
                    try:
                        start_dt = datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M")
                    except Exception:
                        start_dt = None
                    self.matches.append({
                        'row': row_idx,
                        'datetime': start_dt,
                        'message': f"{date_str} {tournament_name} {team1_name} - {team2_name} Тотал {tot_float} {selected_type} {coeff_value:.2f}"
                    })
        # Если ничего не найдено, уведомляем пользователя
        if self.table.rowCount() == 0:
            QMessageBox.information(self, "Нет совпадений", "Не найдено игр с коэффициентом ≤ 1.6.")

    def toggle_select_all(self, state: int) -> None:
        """Обработчик переключения флажка 'Выбрать все'."""
        if state == Qt.Checked:
            self.table.selectAll()
        else:
            self.table.clearSelection()

    def enable_notifications(self) -> None:
        """Устанавливает уведомления за 5 минут до выбранных матчей."""
        selected_rows = {index.row() for index in self.table.selectionModel().selectedRows()}
        if not selected_rows:
            QMessageBox.warning(self, "Нет выбранных игр", "Выберите одну или несколько строк или используйте 'Выбрать все'.")
            return
        now = datetime.now()
        any_scheduled = False
        for match in self.matches:
            if match['row'] in selected_rows and match['datetime']:
                secs = (match['datetime'] - now).total_seconds() - 5 * 60
                if secs < 0:
                    secs = 0
                QTimer.singleShot(int(secs * 1000), lambda m=match: self.show_notification(m))
                any_scheduled = True
        if any_scheduled:
            QMessageBox.information(self, "Уведомления установлены", "Уведомления будут показаны за 5 минут до выбранных матчей.")

    def show_notification(self, match: Dict[str, any]) -> None:
        """Показывает всплывающее уведомление о матче."""
        message = match.get('message', '')
        try:
            self.tray_icon.showMessage(
                "Матч скоро начнётся",
                message,
                QSystemTrayIcon.Information,
                15000
            )
        except Exception:
            QMessageBox.information(self, "Матч скоро начнётся", message)


class SortHalvesThread(QThread):
    """Фоновый поток для переноса данных по четвертям"""
    error = pyqtSignal(str)
    log = pyqtSignal(str)
    # Сигнал для передачи сводки по количеству игр: {tournament: (inserted, normative)}
    summary = pyqtSignal(dict)
    def __init__(self, source_path: str, dest_path: str):
        super().__init__()
        self.source_path = source_path
        self.dest_path = dest_path
    def run(self):
        try:
            self.log.emit(f"Загрузка файлов: {self.source_path}, {self.dest_path}")
            import openpyxl
            # Загружаем исходный файл в режиме только для чтения, что уменьшает время загрузки
            src_wb = openpyxl.load_workbook(self.source_path, read_only=True, data_only=True)
            # Для исходного файла используем лист 'Четверти', если он есть, иначе второй лист
            if 'Четверти' in src_wb.sheetnames:
                src_ws = src_wb['Четверти']
            elif len(src_wb.worksheets) > 1:
                src_ws = src_wb.worksheets[1]
            else:
                src_ws = src_wb.worksheets[0]
            dst_wb = openpyxl.load_workbook(self.dest_path)
            # Группируем строки по турнирам для ускорения вставки
            grouped: Dict[str, List[List[any]]] = {}
            for row in src_ws.iter_rows(values_only=True):
                if not row or len(row) < 4:
                    continue
                tournament = row[1]
                if not tournament:
                    continue
                sheet_name = str(tournament).strip()
                if not sheet_name or sheet_name not in dst_wb.sheetnames:
                    continue
                data = list(row[2:14])
                grouped.setdefault(sheet_name, []).append(data)
            # Для сбора сводки по играм: {tournament: (inserted, normative)}
            games_summary: Dict[str, Tuple[int, int]] = {}
            from openpyxl.cell.cell import MergedCell
            from copy import copy
            # Обходим каждый турнир и вставляем данные с сохранением формата
            for sheet_name, rows_data in grouped.items():
                dest_ws = dst_wb[sheet_name]
                # Определяем строку‑шаблон для копирования стилей (берём вторую строку, либо первую если строк меньше)
                template_row = 2 if dest_ws.max_row >= 2 else 1
                max_len = len(rows_data[0]) if rows_data else 0
                template_styles: Dict[int, any] = {}
                for col_idx in range(1, max_len + 1):
                    tmpl_cell = dest_ws.cell(row=template_row, column=col_idx)
                    template_styles[col_idx] = copy(tmpl_cell._style)
                # Найдём первую строку для вставки, пропуская A1
                insert_row = dest_ws.max_row + 1
                for i, cell in enumerate(dest_ws["A"][1:], start=2):
                    value = cell.value
                    if (value is None or str(value).strip() == "") and not isinstance(cell, MergedCell):
                        row_has_merged = False
                        for col_idx in range(1, max_len + 1):
                            candidate = dest_ws.cell(row=i, column=col_idx)
                            if isinstance(candidate, MergedCell):
                                row_has_merged = True
                                break
                        if not row_has_merged:
                            insert_row = i
                            break
                inserted_count = 0
                for data in rows_data:
                    data_len = len(data)
                    # подбираем строку без mergedCell
                    while True:
                        row_has_merged = False
                        for col_idx in range(1, data_len + 1):
                            candidate = dest_ws.cell(row=insert_row, column=col_idx)
                            if isinstance(candidate, MergedCell):
                                row_has_merged = True
                                break
                        if row_has_merged:
                            insert_row += 1
                        else:
                            break
                    for col_idx, val in enumerate(data, start=1):
                        dest_cell = dest_ws.cell(row=insert_row, column=col_idx)
                        if col_idx in template_styles:
                            dest_cell._style = copy(template_styles[col_idx])
                        dest_cell.value = val
                    inserted_count += 1
                    insert_row += 1
                # Рассчитываем норматив: число уникальных команд // 2 (команды — первые две колонки)
                unique_teams: Set[str] = set()
                for row_vals in dest_ws.iter_rows(min_row=2, values_only=True):
                    if not row_vals:
                        continue
                    t1 = row_vals[0]
                    t2 = row_vals[1] if len(row_vals) > 1 else None
                    if isinstance(t1, str) and t1.strip():
                        unique_teams.add(t1.strip())
                    if isinstance(t2, str) and t2.strip():
                        unique_teams.add(t2.strip())
                normative = len(unique_teams) // 2
                games_summary[sheet_name] = (inserted_count, normative)
            import os
            try:
                # Пытаемся сохранить данные в исходный файл назначения
                dst_wb.save(self.dest_path)
                self.log.emit("Данные успешно перенесены")
            except PermissionError:
                # Если файл назначения открыт, уведомляем об ошибке и не создаём копию
                self.error.emit(
                    "Не удалось сохранить файл: доступ запрещён или файл уже открыт. "
                    "Пожалуйста, закройте файл назначения и повторите попытку."
                )
            # Отправляем сводку о количестве игр в UI вне зависимости от ошибок
            try:
                self.summary.emit(games_summary)
            except Exception:
                # Игнорируем ошибки при отправке сводки
                pass
        except Exception as e:
            self.error.emit(str(e))

class SortHalvesPage(QWidget):
    """Страница для сортировки половин (перенос данных между файлами)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()
    def setup_ui(self) -> None:
        # Используем вкладки для разделения сортировки и управления ссылками
        main_layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        # --- Вкладка для сортировки ---
        sort_widget = QWidget()
        sort_layout = QVBoxLayout(sort_widget)
        # Заголовок
        title_label = QLabel("Сортировка половин")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        sort_layout.addWidget(title_label)
        sort_layout.addSpacing(20)
        # Исходный файл
        src_group_label = QLabel("Исходный файл:")
        src_group_label.setFont(QFont("Arial", 12, QFont.Bold))
        sort_layout.addWidget(src_group_label)
        src_layout = QHBoxLayout()
        self.src_path_label = QLabel("Файл не выбран")
        self.src_path_label.setMinimumWidth(400)
        self.select_src_btn = QPushButton("Выбрать исходный файл")
        self.select_src_btn.setMinimumWidth(200)
        self.select_src_btn.clicked.connect(self.select_source_file)
        src_layout.addWidget(self.src_path_label)
        src_layout.addWidget(self.select_src_btn)
        sort_layout.addLayout(src_layout)
        sort_layout.addSpacing(20)
        # Файл назначения
        dst_group_label = QLabel("Файл назначения:")
        dst_group_label.setFont(QFont("Arial", 12, QFont.Bold))
        sort_layout.addWidget(dst_group_label)
        dst_layout = QHBoxLayout()
        self.dst_path_label = QLabel("Файл не выбран")
        self.dst_path_label.setMinimumWidth(400)
        self.select_dst_btn = QPushButton("Выбрать файл назначения")
        self.select_dst_btn.setMinimumWidth(200)
        self.select_dst_btn.clicked.connect(self.select_dest_file)
        dst_layout.addWidget(self.dst_path_label)
        dst_layout.addWidget(self.select_dst_btn)
        sort_layout.addLayout(dst_layout)
        sort_layout.addSpacing(20)
        # Кнопка запуска
        self.process_btn = QPushButton("Запустить сортировку")
        self.process_btn.setMinimumWidth(250)
        self.process_btn.setMinimumHeight(50)
        self.process_btn.setFont(QFont("Arial", 12))
        self.process_btn.clicked.connect(self.start_processing)
        self.process_btn.setEnabled(False)
        proc_layout = QHBoxLayout()
        proc_layout.addStretch()
        proc_layout.addWidget(self.process_btn)
        proc_layout.addStretch()
        sort_layout.addLayout(proc_layout)
        sort_layout.addSpacing(20)
        # Полоса прогресса
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setVisible(False)
        sort_layout.addWidget(self.progress_bar)
        # Статус
        self.status_label = QLabel("Готов к работе")
        self.status_label.setFont(QFont("Arial", 10))
        self.status_label.setAlignment(Qt.AlignCenter)
        sort_layout.addWidget(self.status_label)
        # Журнал логов
        log_label = QLabel("Журнал выполнения:")
        sort_layout.addWidget(log_label)
        self.log_list = QListWidget()
        self.log_list.setMinimumHeight(200)
        sort_layout.addWidget(self.log_list)
        # Добавляем вкладку сортировки
        self.tabs.addTab(sort_widget, "Сортировка")
        # --- Вкладка для ссылок на турниры ---
        links_widget = QWidget()
        links_layout = QVBoxLayout(links_widget)
        links_title = QLabel("Ссылки на турниры")
        links_title.setFont(QFont("Arial", 16, QFont.Bold))
        links_title.setAlignment(Qt.AlignCenter)
        links_layout.addWidget(links_title)
        links_layout.addSpacing(10)
        info_lbl = QLabel("Укажите ссылку на расписание для каждого турнира. Ссылки могут быть отредактированы.")
        info_lbl.setWordWrap(True)
        links_layout.addWidget(info_lbl)
        # Таблица для отображения турниров и ссылок
        self.links_table = QTableWidget()
        self.links_table.setColumnCount(2)
        self.links_table.setHorizontalHeaderLabels(["Турнир", "Ссылка на расписание"])
        self.links_table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.AnyKeyPressed)
        self.links_table.itemChanged.connect(self.on_link_item_changed)
        links_layout.addWidget(self.links_table)
        self.tabs.addTab(links_widget, "Ссылки на турниры")
        # Добавляем вкладки в основной макет
        main_layout.addWidget(self.tabs)

    def add_log(self, message: str) -> None:
        self.log_list.addItem(message)
        self.log_list.scrollToBottom()
    def update_process_button(self) -> None:
        src_selected = os.path.isfile(self.src_path_label.text())
        dst_selected = os.path.isfile(self.dst_path_label.text())
        self.process_btn.setEnabled(src_selected and dst_selected)
    def select_source_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите исходный файл", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.src_path_label.setText(file_path)
            self.add_log(f"Выбран исходный файл: {os.path.basename(file_path)}")
            self.update_process_button()
    def select_dest_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите файл назначения", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.dst_path_label.setText(file_path)
            self.add_log(f"Выбран файл назначения: {os.path.basename(file_path)}")
            self.update_process_button()
            # Обновляем список турниров и ссылок при выборе файла назначения
            try:
                self.populate_links_table()
            except Exception:
                pass
    def start_processing(self) -> None:
        self.process_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.status_label.setText("Сортировка данных...")
        source_path = self.src_path_label.text()
        dest_path = self.dst_path_label.text()
        self.thread = SortHalvesThread(source_path, dest_path)
        self.thread.finished.connect(self.on_finished)
        self.thread.error.connect(self.on_error)
        self.thread.log.connect(self.add_log)
        # Подключаем обработчик сводки игр
        try:
            self.thread.summary.connect(self.on_summary)
        except Exception:
            pass
        self.thread.start()

    # --- Методы для работы со ссылками на турниры ---
    def populate_links_table(self) -> None:
        """Заполняет таблицу ссылок на турниры на основе листов файла назначения.

        Если файл назначения не выбран или отсутствует, очищает таблицу. Для каждого
        листа (турнира) из книги сортирует названия по алфавиту и отображает
        существующие ссылки, сохранённые в QSettings. При добавлении или
        изменении ссылки данные сохраняются автоматически.
        """
        # Очищаем таблицу
        self.links_table.setRowCount(0)
        dest_path = self.dst_path_label.text()
        if not dest_path or not os.path.isfile(dest_path):
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(dest_path, read_only=True)
            tournaments = sorted(wb.sheetnames)
        except Exception:
            tournaments = []
        # Заполняем таблицу
        self.links_table.setRowCount(len(tournaments))
        for row_idx, tournament in enumerate(tournaments):
            # Название турнира
            item_t = QTableWidgetItem(tournament)
            # Название турнира не редактируемое
            flags = item_t.flags()
            item_t.setFlags(flags & ~Qt.ItemIsEditable)
            self.links_table.setItem(row_idx, 0, item_t)
            # Ссылка
            link_value = self.get_tournament_link(tournament)
            item_l = QTableWidgetItem(link_value)
            self.links_table.setItem(row_idx, 1, item_l)
        # Подгоняем ширину столбцов
        try:
            self.links_table.resizeColumnsToContents()
        except Exception:
            pass

    def get_tournament_link(self, tournament: str) -> str:
        """Возвращает сохранённую ссылку для турнира из QSettings."""
        settings = QSettings("ExcelAnalyzer", "TournamentLinks")
        return settings.value(tournament, "", type=str) or ""

    def set_tournament_link(self, tournament: str, link: str) -> None:
        """Сохраняет ссылку для турнира в QSettings."""
        settings = QSettings("ExcelAnalyzer", "TournamentLinks")
        settings.setValue(tournament, link.strip())

    def on_link_item_changed(self, item: QTableWidgetItem) -> None:
        """Обработчик изменения ячейки в таблице ссылок. Сохраняет изменённую ссылку."""
        # Только столбец 1 (ссылка) редактируется
        if item.column() != 1:
            return
        row = item.row()
        # Получаем название турнира из первой колонки
        tournament_item = self.links_table.item(row, 0)
        if not tournament_item:
            return
        tournament = tournament_item.text()
        new_link = item.text()
        # Сохраняем новую ссылку
        self.set_tournament_link(tournament, new_link)
    def on_finished(self) -> None:
        self.progress_bar.setVisible(False)
        self.process_btn.setEnabled(True)
        self.status_label.setText("Готово")
        self.add_log("Сортировка завершена")
    def on_error(self, message: str) -> None:
        self.progress_bar.setVisible(False)
        self.process_btn.setEnabled(True)
        self.status_label.setText("Ошибка")
        QMessageBox.critical(self, "Ошибка", message)
        self.add_log(f"Ошибка: {message}")

    def on_summary(self, summary: Dict[str, Tuple[int, int]]) -> None:
        """Показывает сводку по количеству игр, добавленных в каждый турнир.

        Данные сортируются так, что турниры с несоответствием норме отображаются первыми.
        В таблице имеется колонка с чекбоксами, позволяющая отметить турнир как
        проверенный. При установке галочки соответствующая строка удаляется из
        таблицы, что помогает отслеживать уже просмотренные турниры.

        Args:
            summary: словарь вида {tournament: (inserted_count, normative)}
        """
        if not summary:
            return
        # Создаём диалоговое окно для отображения сводки. Сохраняем ссылку
        # в локальной переменной, чтобы можно было переопределить closeEvent.
        # Используем модальный режим, чтобы перехватывать клавишу Enter
        dialog = QDialog(self)
        dialog.setWindowTitle("Сводка по играм")
        layout = QVBoxLayout(dialog)
        info_label = QLabel("Количество игр, добавленных в турниры:")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        # Создаём таблицу с колонкой чекбоксов и ссылкой
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["✓", "Турнир", "Добавлено", "Норма", "Ссылка"])
        # Сортируем турниры: сначала те, где количество игр не совпадает с нормой
        items = list(summary.items())
        items.sort(key=lambda x: (x[1][0] == x[1][1], x[0]))
        table.setRowCount(len(items))
        from PyQt5.QtCore import Qt
        from PyQt5.QtGui import QDesktopServices
        from PyQt5.QtCore import QUrl
        # Загружаем сохранённые ссылки из QSettings
        settings = QSettings("ExcelAnalyzer", "TournamentLinks")
        for row_idx, (tournament, counts) in enumerate(items):
            inserted_count, normative = counts
            # Чекбокс для отметки просмотренного турнира (не удаляется автоматически)
            chk = QCheckBox()
            table.setCellWidget(row_idx, 0, chk)
            # Название турнира
            item_t = QTableWidgetItem(str(tournament))
            # Количество игр
            item_ins = QTableWidgetItem(str(inserted_count))
            # Норматив
            item_norm = QTableWidgetItem(str(normative))
            # Показываем ссылку только если есть несоответствие
            link_text = ""
            link_data = ""
            if inserted_count != normative:
                link_data = settings.value(tournament, "", type=str) or ""
                link_text = link_data
            link_item = QTableWidgetItem(link_text)
            link_item.setData(Qt.UserRole, link_data)
            # Отображаем ссылку как синюю подчёркнутую строку, если она есть
            if link_data:
                fnt = link_item.font()
                fnt.setUnderline(True)
                link_item.setFont(fnt)
                link_item.setForeground(QBrush(QColor(42, 130, 218)))
            # Если несоответствие, выделяем красным и жирным
            if inserted_count != normative:
                for itm in (item_t, item_ins, item_norm, link_item):
                    fnt2 = itm.font()
                    fnt2.setBold(True)
                    itm.setFont(fnt2)
                    itm.setForeground(QBrush(QColor(220, 50, 50)))
            table.setItem(row_idx, 1, item_t)
            table.setItem(row_idx, 2, item_ins)
            table.setItem(row_idx, 3, item_norm)
            table.setItem(row_idx, 4, link_item)
        # Запрещаем редактирование таблицы; ссылки открываются по клику
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        # Обработчик клика по ссылке
        def handle_cell_clicked(row: int, column: int) -> None:
            if column == 4:
                itm = table.item(row, column)
                if itm:
                    url_str = itm.data(Qt.UserRole)
                    if url_str:
                        QDesktopServices.openUrl(QUrl(url_str))
        table.cellClicked.connect(handle_cell_clicked)
        # Подгоняем ширину колонок под содержимое
        try:
            table.resizeColumnsToContents()
        except Exception:
            pass
        layout.addWidget(table)
        # Кнопка для удаления отмеченных строк
        remove_btn = QPushButton("Удалить из списка")
        def remove_selected_rows() -> None:
            rows_to_remove = []
            for r in range(table.rowCount()):
                cb = table.cellWidget(r, 0)
                if cb and cb.isChecked():
                    rows_to_remove.append(r)
            for r in reversed(rows_to_remove):
                table.removeRow(r)
        remove_btn.clicked.connect(remove_selected_rows)
        layout.addWidget(remove_btn)
        # Кнопка закрытия окна
        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        # Функция для подтверждения закрытия
        def confirm_action() -> bool:
            reply = QMessageBox.question(
                dialog,
                "Закрыть окно",
                "Вы действительно хотите закрыть окно сводки?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            return reply == QMessageBox.Yes
        # Обработчики закрытия и клавиш
        def on_close_event(event):
            if confirm_action():
                event.accept()
            else:
                event.ignore()
        dialog.closeEvent = on_close_event
        def on_accept():
            if confirm_action():
                QDialog.accept(dialog)
        def on_reject():
            if confirm_action():
                QDialog.reject(dialog)
        dialog.accept = on_accept
        dialog.reject = on_reject
        # Перехватываем нажатие Enter
        def dialog_key_press(e):
            if e.key() in (Qt.Key_Return, Qt.Key_Enter):
                if confirm_action():
                    QDialog.accept(dialog)
            else:
                QDialog.keyPressEvent(dialog, e)
        dialog.keyPressEvent = dialog_key_press
        # Запускаем диалог
        dialog.exec_()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # Показываем путь к данным в заголовке, чтобы легко диагностировать проблемы
        data_dir = get_data_dir()
        self.setWindowTitle(f"Excel Analyzer Pro  —  data: {data_dir}")
        
        # Устанавливаем нормальный размер окна по умолчанию
        self.resize(1400, 800)
        
        # Центрируем окно на экране
        self.center_on_screen()
        
        # ВАЖНО: Устанавливаем минимальный размер окна
        self.setMinimumSize(800, 600)
        
        # Убираем ограничения на изменение размера
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Устанавливаем стандартный курсор для окна
        self.setCursor(Qt.ArrowCursor)
        
        # Инициализация компонентов UI
        self.theme_manager = ThemeManager(self)
        
        # Загружаем сохраненную тему
        self.current_theme = self.theme_manager.load_theme_preference()
        
        # Создание центрального виджета
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Основной layout - используем QHBoxLayout вместо splitter
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Создание боковой навигации
        self.sidebar = SidebarNavigation()
        self.sidebar.setFixedWidth(250)  # Фиксированная ширина для боковой панели
        
        # Создание контейнера для основного содержимого
        content_container = QWidget()
        content_layout = QVBoxLayout(content_container)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)
        
        # Создание заголовка с кнопкой темы
        header = self.create_header()
        content_layout.addWidget(header)
        
        # Создание стека страниц
        self.page_stack = QStackedWidget()
        content_layout.addWidget(self.page_stack)
        
        # Добавляем sidebar и content в main_layout
        main_layout.addWidget(self.sidebar)
        main_layout.addWidget(content_container)
        
        # Создание страниц — ленивая загрузка
        # Только первая страница (Ройка) создаётся сразу, остальные — по требованию
        self.royka_page = RoykaPage()
        self.cybers_db = None  # инициализируется лениво

        # Фабрики для создания страниц по требованию
        self._page_factories = {
            1: self._create_halfs_database_page,
            2: self._create_halfs_statistics_page,
            3: self._create_tournament_summary_page,
            4: self._create_halfs_analysis_page,
            5: self._create_cybers_bases_page,
            6: self._create_cyber_live_page,
            7: self._create_sort_halves_page,
        }
        self._pages_created = {0}  # Ройка уже создана

        # Добавляем страницы-заглушки в стек
        self.page_stack.addWidget(self.royka_page)               # 0 – Ройка
        for i in range(1, 8):
            self.page_stack.addWidget(QWidget())                 # заглушки

        # Инициализируем Telegram
        self.telegram = TelegramNotifier(
            bot_token="8300351888:AAG9myXyqGvkOEQYcqz_kXlw4tzjC8_644Q",
            chat_id=197670357
        )
        
        # Подключаем обработчик уведомлений, если страница мониторинга ставок доступна
        # В текущей версии страница мониторинга ставок удалена, поэтому не пытаемся подключиться
        # Оставляем этот блок на случай будущего расширения функционала
        if hasattr(self, "bets_monitor_page") and hasattr(self.bets_monitor_page, "match_notification"):
            try:
                self.bets_monitor_page.match_notification.connect(self.handle_match_notification)
            except Exception:
                pass

        # Подключение сигналов навигации
        self.sidebar.itemClicked.connect(self.on_navigation_clicked)
        
        # Применяем загруженную тему
        self.theme_manager.apply_theme(self.current_theme)
        
        # Выбираем первый пункт по умолчанию
        self.sidebar.setCurrentRow(0)
        self.page_stack.setCurrentIndex(0)
        
        # Устанавливаем иконку приложения
        self.setWindowIcon(self.create_app_icon())
        
    # ... остальные методы остаются без изменений ...
    
    def handle_match_notification(self, tournament, teams, notification_type):
        """Обработка уведомления о матче"""
        self.telegram.send_notification(tournament, teams, notification_type)

    def center_on_screen(self):
        """Центрирует окно на экране"""
        screen = QApplication.primaryScreen()
        if screen:
            screen_geometry = screen.availableGeometry()
            window_geometry = self.frameGeometry()
            window_geometry.moveCenter(screen_geometry.center())
            self.move(window_geometry.topLeft())
    
    def create_header(self):
        """Создание заголовка с кнопками управления"""
        header = QWidget()
        header.setFixedHeight(60)
        header.setObjectName("header")
        
        layout = QHBoxLayout(header)
        layout.setContentsMargins(20, 10, 20, 10)
        
        # Заголовок
        title = QLabel("Excel Analyzer Pro")
        title.setFont(QFont("Arial", 18, QFont.Bold))
        layout.addWidget(title)
        
        layout.addStretch()
        
        # Кнопка переключения темы
        self.theme_btn = QPushButton()
        self.theme_btn.setFixedSize(40, 40)
        self.theme_btn.setCursor(Qt.PointingHandCursor)
        self.theme_btn.clicked.connect(self.toggle_theme)
        self.update_theme_button()
        
        layout.addWidget(self.theme_btn)
        
        return header
    
    def create_app_icon(self):
        """Создание иконки приложения"""
        pixmap = QPixmap(32, 32)
        pixmap.fill(Qt.transparent)
        
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # Рисуем простую иконку
        gradient = QLinearGradient(0, 0, 32, 32)
        gradient.setColorAt(0, QColor("#4A90E2"))
        gradient.setColorAt(1, QColor("#357ABD"))
        
        painter.setBrush(QBrush(gradient))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(0, 0, 32, 32, 8, 8)
        
        # Добавляем текст
        painter.setPen(Qt.white)
        painter.setFont(QFont("Arial", 16, QFont.Bold))
        painter.drawText(QRect(0, 0, 32, 32), Qt.AlignCenter, "EA")
        
        painter.end()
        
        return QIcon(pixmap)
    
    def update_theme_button(self):
        """Обновление иконки кнопки темы"""
        if self.current_theme == "dark":
            self.theme_btn.setText("☀")
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3A3A3A;
                    border: 2px solid #555;
                    border-radius: 20px;
                    color: #FDB813;
                    font-size: 20px;
                }
                QPushButton:hover {
                    background-color: #4A4A4A;
                    border-color: #FDB813;
                }
            """)
        else:
            self.theme_btn.setText("🌙")
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #E8E8E8;
                    border: 2px solid #B0B0B0;
                    border-radius: 20px;
                    color: #4A5568;
                    font-size: 20px;
                }
                QPushButton:hover {
                    background-color: #D8D8D8;
                    border-color: #4A5568;
                }
            """)
    
    def toggle_theme(self):
        """Переключение темы"""
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.theme_manager.apply_theme(self.current_theme)
        self.update_theme_button()
        
        # Обновляем тему для страниц
        QApplication.instance().setProperty("theme", self.current_theme)
        
        # Обновляем таблицы если они отображены
        if hasattr(self.royka_page, 'current_table') and self.royka_page.current_table:
            self.royka_page.display_statistics(
                self.royka_page.last_tournament_name, 
                self.royka_page.last_stats
            )
        # После смены темы обновляем стили журналов
        try:
            self.update_debug_log_styles()
        except Exception:
            pass
    
    # --- Фабрики для ленивого создания страниц ---
    def _ensure_cybers_db(self):
        if self.cybers_db is None:
            self.cybers_db = CybersDatabase()

    def _create_halfs_database_page(self):
        return HalfsDatabasePage()

    def _create_halfs_statistics_page(self):
        return HalfsStatisticsPage()

    def _create_tournament_summary_page(self):
        return TournamentSummaryPage()

    def _create_halfs_analysis_page(self):
        return HalfsAnalysisPage()

    def _create_cybers_bases_page(self):
        self._ensure_cybers_db()
        page = CybersBasesPage(self.cybers_db)
        # Если Cyber LIVE уже создана — привязать
        if 6 in self._pages_created:
            page.live_page = self.page_stack.widget(6)
        return page

    def _create_cyber_live_page(self):
        self._ensure_cybers_db()
        page = CyberLivePage(self.cybers_db)
        # Если Cybers Bases уже создана — привязать
        if 5 in self._pages_created:
            self.page_stack.widget(5).live_page = page
        return page

    def _create_sort_halves_page(self):
        return SortHalvesPage()

    def _load_page_if_needed(self, index: int) -> None:
        """Создаёт страницу по требованию, заменяя заглушку."""
        if index in self._pages_created:
            return
        factory = self._page_factories.get(index)
        if factory is None:
            return
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            page = factory()
        except Exception as exc:
            logging.error(f"Ошибка при создании страницы (index={index}): {exc}", exc_info=True)
            # Показываем пользователю, что именно пошло не так
            error_page = QWidget()
            err_layout = QVBoxLayout(error_page)
            err_label = QLabel(
                f"Не удалось загрузить раздел.\n\n"
                f"Ошибка: {exc}\n\n"
                f"Путь к данным: {get_data_dir()}"
            )
            err_label.setWordWrap(True)
            err_label.setAlignment(Qt.AlignCenter)
            err_layout.addWidget(err_label)
            page = error_page
        finally:
            QApplication.restoreOverrideCursor()
        # Заменяем заглушку настоящей страницей
        old_widget = self.page_stack.widget(index)
        self.page_stack.removeWidget(old_widget)
        old_widget.deleteLater()
        self.page_stack.insertWidget(index, page)
        self._pages_created.add(index)
        # Сохраняем ссылки для обратной совместимости
        attr_map = {
            1: "halfs_database_page",
            2: "halfs_statistics_page",
            3: "tournament_summary_page",
            4: "halfs_analysis_page",
            5: "cybers_bases_page",
            6: "cyber_live_page",
            7: "sort_halves_page",
        }
        attr = attr_map.get(index)
        if attr:
            setattr(self, attr, page)

    def on_navigation_clicked(self, item):
        """Обработка клика по элементу навигации"""
        index = self.sidebar.row(item)
        self._load_page_if_needed(index)
        self.page_stack.setCurrentIndex(index)
        # Подгоняем ширину столбцов всех таблиц под содержимое при смене раздела
        try:
            QTimer.singleShot(0, self.adjust_table_columns)
        except Exception:
            pass
    
    def closeEvent(self, event):
        """Обработка закрытия окна"""
        # Сохраняем размер и позицию окна
        settings = QSettings("ExcelAnalyzer", "MainWindow")
        settings.setValue("geometry", self.saveGeometry())
        settings.setValue("windowState", self.saveState())
        event.accept()
    
    def showEvent(self, event):
        """Обработка показа окна"""
        super().showEvent(event)
        # Восстанавливаем размер и позицию окна
        settings = QSettings("ExcelAnalyzer", "MainWindow")
        geometry = settings.value("geometry")
        if geometry:
            self.restoreGeometry(geometry)
        state = settings.value("windowState")
        if state:
            self.restoreState(state)

    def adjust_table_columns(self):
        """Автоматически подгоняет ширину столбцов под содержимое для текущей страницы.

        Этот метод проходит по всем QTableWidget, найденным в окне, и вызывает
        resizeColumnsToContents(). Это обеспечивает, что столбцы всегда
        соответствуют ширине своего наибольшего элемента, делая таблицы
        удобными для чтения. Вызов осуществляется асинхронно через
        QTimer.singleShot в on_navigation_clicked, чтобы избежать
        конкуренции с обновлением содержимого.
        """
        current_widget = self.page_stack.currentWidget()
        if current_widget is None:
            return
        # Проходим по таблицам только в текущем разделе
        for table in current_widget.findChildren(QTableWidget):
            try:
                if table.property("skipAutoResize"):
                    continue
                # Избегаем дорогого ресайза больших таблиц
                if table.rowCount() > 500:
                    continue
                table.resizeColumnsToContents()
            except Exception:
                # На всякий случай игнорируем ошибки на отдельных таблицах
                pass

    def update_debug_log_styles(self):
        """
        Обновляет стили всех QPlainTextEdit (журналы) в соответствии с текущей темой.

        При тёмной теме используется тёмный фон и светлый текст, при светлой —
        светлый фон и тёмный текст. Функция проходит по всем QPlainTextEdit в окне
        и устанавливает им соответствующий stylesheet.
        """
        # Определяем CSS для каждой темы
        dark_css = """
        QPlainTextEdit {
            background-color: #11244A;
            color: #E0E5EC;
            border: 1px solid #1C3A6A;
            font-family: Consolas, 'Courier New', monospace;
            font-size: 12px;
            padding: 5px;
        }
        """
        light_css = """
        QPlainTextEdit {
            background-color: #F5FAF7;
            color: #0A192F;
            border: 1px solid #CDE7E4;
            font-family: Consolas, 'Courier New', monospace;
            font-size: 12px;
            padding: 5px;
        }
        """
        theme = getattr(self, 'current_theme', 'dark')
        css = dark_css if theme == 'dark' else light_css
        # Обновляем стиль каждого QPlainTextEdit
        for widget in self.findChildren(QPlainTextEdit):
            try:
                widget.setStyleSheet(css)
            except Exception:
                pass

# ---- Привязка методов HalfsPage к реализациям из HalfsDatabasePage ----
# При интеграции "Базы половин" часть методов класса HalfsPage оказалась
# определена внутри HalfsDatabasePage из‑за неверного уровня отступа.
# Чтобы вернуть работоспособность HalfsPage, привяжем эти реализации
# непосредственно из HalfsDatabasePage. Таким образом, HalfsPage вновь
# получит методы setup_ui, select_halfs_file и др., не требуя переопределять
# код целиком. Если класс HalfsDatabasePage недоступен (например, при
# ошибке импорта), привязка будет пропущена.
#
# Привязка методов HalfsPage к реализациям из HalfsDatabasePage была отменена,
# поскольку исходные методы были определены вне классов и вызывали исключения.
# Вместо этого HalfsPage использует собственный интерфейс или stub (см. класс HalfsPage).

# ---- Восстанавливаем раздел "Работа с половинами" ----
# В более поздних версиях интерфейс для HalfsPage (перенос данных между файлами
# "Cyber" и "Половины") оказался реализован как набор свободных функций
# (setup_ui, select_halfs_file, select_cyber_files, clear_cyber_files,
# update_process_button, process_files, update_progress, processing_finished,
# processing_error, add_log).  Чтобы HalfsPage снова стал полнофункциональным,
# привязываем эти функции к классу HalfsPage.  Теперь его __init__ вызовет
# правильный setup_ui вместо резервного stub.

# Переназначение методов HalfsPage к методам HalfsDatabasePage было удалено,
# чтобы HalfsPage использовал собственный интерфейс для работы с файлами.

def apply_dialog_style(dialog):
    """Применяет тёмный стиль к диалоговым окнам"""
    dialog.setStyleSheet("""
        QDialog {
            background-color: #12121A;
            color: #999999;
            border: 1px solid #333339;
        }
        QLabel {
            color: #999999;
        }
        QPushButton {
            background-color: #25252b;
            color: #999999;
            border: 1px solid #333339;
            border-radius: 4px;
            padding: 5px;
        }
        QPushButton:hover {
            background-color: #2d2d35;
            color: #b4b4b4;
        }
        QListWidget, QComboBox, QLineEdit {
            background-color: #18181e;
            color: #999999;
            border: 1px solid #333339;
            border-radius: 3px;
            padding: 4px;
        }
    """)

def main():
    """Главная функция запуска приложения"""
    try:
        print("Начало выполнения программы")
        logging.info("Начало выполнения программы")
        
        # Создание приложения
        app = QApplication(sys.argv)
        app.setApplicationName("Excel Analyzer Pro")
        app.setOrganizationName("ExcelAnalyzer")
        print("QApplication создан")
        
        # Настройка стиля приложения
        app.setStyle('Fusion')
        
        # Загружаем сохраненную тему
        temp_theme_manager = ThemeManager()
        saved_theme = temp_theme_manager.load_theme_preference()
        app.setProperty("theme", saved_theme)
        print(f"Загружена тема: {saved_theme}")
        
        # Создание и отображение главного окна
        print("Создание главного окна...")
        window = MainWindow()
        print("Отображение окна...")
        window.show()
        
        print("Запуск цикла обработки событий...")
        logging.info("Приложение успешно запущено")
        
        # Запуск приложения
        sys.exit(app.exec_())
        
    except Exception as e:
        print(f"Критическая ошибка: {str(e)}")
        print(traceback.format_exc())
        logging.critical(f"Критическая ошибка: {str(e)}\n{traceback.format_exc()}")
        sys.exit(1)

# Основной блок запуска приложения
if __name__ == "__main__":
    main()

# ===== LNBP / FLS INTEGRATION (added) =====
try:
    from scan_router import get_matches_by_url as _scan_router_get_matches_by_url
except Exception:
    _scan_router_get_matches_by_url = None

def get_matches_by_url(driver, url, target_date, logger=print):
    """
    Унифицированный вызов:
    • FLS → fibalivestats_handler
    • LNBP → lnbp_parser
    • Остальное → []
    """
    if _scan_router_get_matches_by_url is None:
        logger("scan_router не найден — возвращаю пусто")
        return []
    return _scan_router_get_matches_by_url(driver, url, target_date, logger=logger)
# ===== END INTEGRATION =====

# =====================================================================
# Новый раздел: Сортировка половин
def sort_polovin(source_file: str, dest_file: str, output_file: str | None = None) -> None:
    """
    Переносит данные по четвертям из `source_file` в соответствующие листы
    `dest_file`. Если указан `output_file`, результат сохраняется по данному
    пути, иначе изменения записываются непосредственно в `dest_file`.

    Аргументы:
        source_file (str): Путь к исходному файлу .xlsx с матчами по четвертям.
        dest_file (str): Путь к файлу назначения .xlsx с листами для каждого турнира.
        output_file (str | None, optional): Путь, по которому будет сохранён
            обновлённый файл. Если не указан, файл назначения изменяется на месте.

    Формат исходного файла:
        A: дата (игнорируется);
        B: название турнира;
        C: название команды 1;
        D: название команды 2;
        E–L: очки в 1–4 четвертях (парами);
        M–N: очки в овертайме (если были).

    Поведение:
        Для каждой строки исходного файла определяется название турнира (колонка B).
        Если лист с таким названием существует в файле назначения, то данные из
        колонок C–N копируются в конец этого листа (начиная с колонки A).
        Запись осуществляется без замены существующих данных. Лист "Чемпионаты"
        в исходном файле игнорируется.

    Оптимизации:
        * Исходный файл загружается в режиме только для чтения (`read_only=True`),
          что сокращает время на обработку больших таблиц.
        * Данные группируются по турнирам, что уменьшает количество обращений
          к листам назначения.
        * Перед вставкой выполняется поиск первой строки, свободной от данных
          и объединённых ячеек, чтобы корректно вставлять данные без замены
          существующих записей. Это также учитывает объединённые ячейки
          (merged cells) в листе назначения.
    """
    try:
        import openpyxl
        # Загружаем исходный файл в режиме только для чтения
        src_wb = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
        # Выбираем рабочий лист "Четверти" (если есть) или второй лист, иначе первый
        if 'Четверти' in src_wb.sheetnames:
            src_ws = src_wb['Четверти']
        elif len(src_wb.worksheets) > 1:
            src_ws = src_wb.worksheets[1]
        else:
            src_ws = src_wb.worksheets[0]
        # Загружаем файл назначения. Здесь режим только чтения не допускается, т.к. требуется запись.
        dst_wb = openpyxl.load_workbook(dest_file)
        # Сгруппируем строки по турнирам, чтобы сократить количество обращений к файлу назначения
        grouped: Dict[str, List[List[any]]] = {}
        for row in src_ws.iter_rows(values_only=True):
            # Пропускаем пустые строки и строки с недостаточным количеством столбцов
            if not row or len(row) < 4:
                continue
            tournament = row[1]  # Колонка B — название турнира
            if not tournament:
                continue
            sheet_name = str(tournament).strip()
            # Проверяем, существует ли лист с таким именем в файле назначения
            if not sheet_name or sheet_name not in dst_wb.sheetnames:
                continue
            # Берём данные из колонок C–N (индексация с нуля: элементы 2–13)
            data = list(row[2:14])
            grouped.setdefault(sheet_name, []).append(data)
        # Для каждого турнира вставляем строки в соответствующий лист
        for sheet_name, rows_data in grouped.items():
            dest_ws = dst_wb[sheet_name]
            # Импортируем здесь, чтобы не делать на уровне модуля, так как openpyxl
            # может отсутствовать в окружении тестирования
            from openpyxl.cell.cell import MergedCell
            # Находим первую строку, где мы можем безопасно писать значения. Мы
            # пропускаем строки, которые:
            #  1) содержат какие‑либо данные в колонке A
            #  2) или ячейка в колонке A является частью объединённого диапазона (MergedCell)
            #  3) или внутри диапазона записи (1..len(data)) находятся объединённые ячейки
            # Если ничего не найдено, используем строку после последней заполненной
            insert_row = dest_ws.max_row + 1
            # Пропускаем первую строку (ячейка A1), поскольку она всегда пустая
            # и её нужно игнорировать при поиске первой свободной строки
            for i, cell in enumerate(dest_ws["A"][1:], start=2):
                value = cell.value
                # Проверяем, что ячейка пуста и не объединена
                if (value is None or str(value).strip() == "") and not isinstance(cell, MergedCell):
                    # Дополнительно проверяем, что в строке нет объединённых ячеек в целевом диапазоне
                    row_has_merged = False
                    # Определяем максимальную длину данных в одной строке (13 столбцов). Если rows_data пусто,
                    # пропускаем проверку.
                    max_len = len(rows_data[0]) if rows_data else 0
                    for col_idx in range(1, max_len + 1):
                        candidate_cell = dest_ws.cell(row=i, column=col_idx)
                        if isinstance(candidate_cell, MergedCell):
                            row_has_merged = True
                            break
                    if not row_has_merged:
                        insert_row = i
                        break
            # Записываем каждую строку последовательно, начиная с найденной позиции
            for data in rows_data:
                for col_idx, val in enumerate(data, start=1):
                    # Получаем ячейку, которую будем заполнять. Если она часть mergedCell,
                    # openpyxl автоматически вернёт MergedCell, и попытка присвоить
                    # значение вызовет исключение. Проверяем это и при необходимости
                    # переносим запись на следующую строку.
                    cell = dest_ws.cell(row=insert_row, column=col_idx)
                    if isinstance(cell, MergedCell):
                        # Если хотя бы одна ячейка в этой строке объединена, выбираем следующую строку
                        insert_row += 1
                        cell = dest_ws.cell(row=insert_row, column=col_idx)
                    cell.value = val
                insert_row += 1
        # Сохраняем результаты. Если файл назначения занят (например, открыт в Excel) или
        # отсутствуют права на запись, будет выброшено PermissionError. В этом случае
        # выведем понятное сообщение.
        save_path = output_file if output_file else dest_file
        try:
            dst_wb.save(save_path)
        except PermissionError:
            # Если файл назначения открыт, пытаемся сохранить в новый файл
            import os
            base, ext = os.path.splitext(save_path)
            fallback_path = f"{base}_updated{ext or '.xlsx'}"
            dst_wb.save(fallback_path)
            # Сообщаем, что сохранение произведено в новый файл
            print(
                f"Предупреждение: файл назначения '{save_path}' был открыт и не мог быть изменён. "
                f"Данные были сохранены в '{fallback_path}'."
            )
            # Прерываем выполнение, чтобы вызывающий код мог обработать ситуацию
            return
    except Exception as e:
        # Перебрасываем исключения, чтобы вызывающий код мог корректно обработать их
        raise
